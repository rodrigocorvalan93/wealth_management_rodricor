# -*- coding: utf-8 -*-
"""
migrate_target_columns.py

Agrega las columnas `Precio Target`, `Stop Loss`, `Moneda Target` a la hoja
`blotter` de cada Excel master (multi-tenant) si no existen ya. También
agrega la fila `Distancia alerta target (bps)` a la hoja `config`.

Idempotente: re-ejecutar es seguro (skip si ya están).

USO:
    python migrate_target_columns.py                   # todos los users
    python migrate_target_columns.py --user rodricor   # solo uno
    python migrate_target_columns.py --dry-run
"""
from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent

NEW_BLOTTER_COLS = ["Precio Target", "Stop Loss", "Moneda Target"]
INSERT_AFTER = "Moneda Com"  # ubicación de las nuevas cols

NEW_CONFIG_ROW = ("Distancia alerta target (bps)", 10, "number",
                   "Distancia en bps que define 'cerca del target'. "
                   "Default 10. Subilo a 100 para alertas con margen.")
HEADER_ROW = 4


def find_user_xlsxs() -> list[tuple[str, Path]]:
    out = []
    inputs = HERE / "inputs"
    if not inputs.is_dir():
        return out
    for d in sorted(inputs.iterdir()):
        if not d.is_dir():
            continue
        for fname in ("wealth_management.xlsx", "wealth_management_rodricor.xlsx"):
            f = d / fname
            if f.is_file():
                out.append((d.name, f))
                break
    if not out:
        for fname in ("wealth_management.xlsx", "wealth_management_rodricor.xlsx"):
            f = inputs / fname
            if f.is_file():
                out.append(("default", f))
                break
    return out


def migrate_blotter(ws, dry_run: bool) -> list[str]:
    """Agrega cols nuevas a blotter después de INSERT_AFTER. Devuelve lista
    de cols agregadas."""
    headers = {}
    insert_after_col = None
    last_col = ws.max_column
    for c in range(1, last_col + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            v_str = str(v).strip()
            headers[v_str] = c
            if v_str == INSERT_AFTER:
                insert_after_col = c

    added = []
    for col_name in NEW_BLOTTER_COLS:
        if col_name in headers:
            continue
        added.append(col_name)

    if not added or dry_run:
        return added

    # Insertar TODAS las cols nuevas juntas, después de insert_after_col.
    # Si insert_after_col es None, las pongo al final.
    insert_at = (insert_after_col + 1) if insert_after_col else (last_col + 1)
    n = len(added)
    if insert_after_col is not None:
        ws.insert_cols(insert_at, amount=n)
    for offset, col_name in enumerate(added):
        cell = ws.cell(row=HEADER_ROW, column=insert_at + offset, value=col_name)
        # Heredar formato del header anterior (si hay)
        if insert_after_col:
            ref_cell = ws.cell(row=HEADER_ROW, column=insert_after_col)
            cell.font = ref_cell.font.copy()
            cell.fill = ref_cell.fill.copy()
            cell.alignment = ref_cell.alignment.copy()
    return added


def migrate_config(ws, dry_run: bool) -> bool:
    """Agrega fila NEW_CONFIG_ROW si no existe. True si la agregó."""
    # Header: Concepto en col 1
    target_concepto = NEW_CONFIG_ROW[0].strip().lower()
    last_data_row = HEADER_ROW
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            last_data_row = r
            if str(v).strip().lower() == target_concepto:
                return False  # ya existe
    if dry_run:
        return True
    new_row = last_data_row + 1
    for j, val in enumerate(NEW_CONFIG_ROW, start=1):
        ws.cell(row=new_row, column=j, value=val)
    return True


def migrate_xlsx(xlsx_path: Path, dry_run: bool) -> dict:
    """Migra un Excel master. Devuelve resumen de cambios."""
    if not xlsx_path.is_file():
        return {"error": f"no existe: {xlsx_path}"}
    summary = {"path": str(xlsx_path), "blotter_cols_added": [],
               "config_row_added": False, "backed_up": None}

    # Backup antes de tocar
    if not dry_run:
        ts = datetime.now().strftime("%Y%m%dT%H%M%S")
        bak = xlsx_path.with_suffix(f".pre-target-migration.{ts}.xlsx")
        shutil.copy2(xlsx_path, bak)
        summary["backed_up"] = str(bak.name)

    wb = load_workbook(filename=str(xlsx_path))
    if "blotter" in wb.sheetnames:
        summary["blotter_cols_added"] = migrate_blotter(wb["blotter"], dry_run)
    if "config" in wb.sheetnames:
        summary["config_row_added"] = migrate_config(wb["config"], dry_run)

    if not dry_run and (summary["blotter_cols_added"] or summary["config_row_added"]):
        wb.save(xlsx_path)
    elif not dry_run:
        # Si no hubo cambios, borrar el backup que ya creamos para no llenar
        if summary["backed_up"]:
            (xlsx_path.parent / summary["backed_up"]).unlink(missing_ok=True)
            summary["backed_up"] = None
    return summary


def main():
    p = argparse.ArgumentParser(description="Agrega cols target/stop al blotter")
    p.add_argument("--user", type=str, default=None,
                   help="Solo este user (default: todos)")
    p.add_argument("--dry-run", action="store_true",
                   help="No modifica, solo muestra qué haría")
    args = p.parse_args()

    masters = find_user_xlsxs()
    if not masters:
        print("[migrate-target] No se encontraron masters en inputs/")
        return 1
    if args.user:
        masters = [(u, p) for u, p in masters if u == args.user]
        if not masters:
            print(f"[migrate-target] User '{args.user}' no encontrado")
            return 1

    print(f"[migrate-target] {len(masters)} master(s){' (dry-run)' if args.dry_run else ''}")
    for user, xlsx in masters:
        print(f"\n--- {user} → {xlsx.name} ---")
        s = migrate_xlsx(xlsx, args.dry_run)
        if "error" in s:
            print(f"  ✗ {s['error']}"); continue
        if s["blotter_cols_added"]:
            print(f"  ✓ blotter: agregadas {s['blotter_cols_added']}")
        else:
            print(f"  · blotter: ya estaba migrado")
        if s["config_row_added"]:
            print(f"  ✓ config: agregada fila 'Distancia alerta target (bps)' = 10")
        else:
            print(f"  · config: ya tenía la fila")
        if s["backed_up"]:
            print(f"  ↩ backup: {s['backed_up']}")
    print("\n[migrate-target] Listo. Hacé refresh en la PWA o corré sync.py "
          "para que la DB tome los cambios.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
