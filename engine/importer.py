# -*- coding: utf-8 -*-
"""
engine/importer.py

Lee el Excel master `wealth_management_rodricor.xlsx` y popula la DB sqlite
con events + movements según el modelo de doble entrada.

Para cada hoja:
  - cuentas, monedas, especies → tablas maestras
  - blotter → events TRADE + 2 movements (activo en cuenta + cash en cuenta)
  - transferencias_cash → events TRANSFER_CASH + 2 movements
  - transferencias_activos → events TRANSFER_ASSET + 2 movements
  - ingresos → events INCOME + 2 movements (cuenta destino + external_income)
  - gastos → events EXPENSE / CARD_CHARGE + N movements (con expansión cuotas)
  - pasivos + pagos_pasivos → events LIABILITY_OPEN, LIABILITY_PAYMENT
  - asientos_contables → events ACCOUNTING_ADJUSTMENT + N movements
  - recurrentes → expande hasta hoy y genera events según frecuencia
"""

from __future__ import annotations

import sqlite3
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from .schema import (
    EventType,
    AccountKind,
    insert_account, insert_currency, insert_asset,
    insert_event, insert_movement,
)
from .fx import convert as fx_convert, FxError, auto_load_fx


# =============================================================================
# Helpers
# =============================================================================

def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _to_date_str(v: Any) -> Optional[str]:
    """Convierte cell value a string ISO YYYY-MM-DD."""
    if _is_empty(v):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s[:10]).isoformat()
        except ValueError:
            return None
    return None


def _to_float(v: Any) -> Optional[float]:
    if _is_empty(v):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(v: Any) -> Optional[int]:
    if _is_empty(v):
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _to_str(v: Any) -> Optional[str]:
    if _is_empty(v):
        return None
    return str(v).strip()


def _read_rows(ws, header_row: int = 4):
    """Itera filas de datos de una hoja, devuelve dict {header: value}.
    Saltea filas donde TODAS las columnas (excepto la primera) están vacías.
    """
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        headers.append(h)
    data_start = header_row + 1
    for r in range(data_start, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c + 1).value
                    for c in range(len(headers))]
        if all(_is_empty(v) for v in row_vals):
            continue
        yield r, dict(zip(headers, row_vals))


# =============================================================================
# Importadores por hoja
# =============================================================================

def import_monedas(conn, ws):
    """Hoja 'monedas': popula tabla currencies."""
    n = 0
    for r, row in _read_rows(ws):
        code = _to_str(row.get("Code"))
        if not code:
            continue
        insert_currency(
            conn,
            code=code,
            name=_to_str(row.get("Name")) or code,
            is_stable=bool(_to_int(row.get("Is Stable")) or 0),
            quote_vs=_to_str(row.get("Quote vs")),
            is_base=bool(_to_int(row.get("Is Base")) or 0),
            notes=_to_str(row.get("Notas")),
        )
        n += 1
    return n


def import_cuentas(conn, ws):
    """Hoja 'cuentas': popula tabla accounts.

    Columnas opcionales (Sprint B):
      - 'Investible': YES/NO/1/0 (default YES). Si NO, se excluye del
        "PN invertible" en reportes (ej cash de reserva no declarado).
      - 'Cash Purpose': texto libre ('OPERATIVO','RESERVA_NO_DECLARADO',...).
    """
    n = 0
    for r, row in _read_rows(ws):
        code = _to_str(row.get("Code"))
        if not code:
            continue

        # Investible: por default 1 (cuentas reales son invertibles).
        # Las cuentas técnicas (external_*, opening_balance, interest_*)
        # se marcan como 0 automáticamente abajo.
        investible_raw = _to_str(row.get("Investible"))
        if investible_raw is None:
            investible = 1
        else:
            investible = 1 if investible_raw.upper() in ("YES", "Y", "1", "TRUE", "SI", "SÍ") else 0

        # Cuentas técnicas siempre no-invertibles
        kind = _to_str(row.get("Kind")) or AccountKind.CASH_BANK
        if kind in (AccountKind.EXTERNAL, AccountKind.OPENING_BALANCE,
                    AccountKind.INTEREST_EXPENSE, AccountKind.INTEREST_INCOME):
            investible = 0

        insert_account(
            conn,
            code=code,
            name=_to_str(row.get("Name")) or code,
            kind=kind,
            institution=_to_str(row.get("Institution")),
            currency=_to_str(row.get("Currency")),
            card_cycle_kind=_to_str(row.get("Card Cycle")) or "NONE",
            card_close_day=_to_int(row.get("Close Day")),
            card_due_day=_to_int(row.get("Due Day")),
            card_currency=_to_str(row.get("Card Currency")),
            investible=investible,
            cash_purpose=_to_str(row.get("Cash Purpose")),
            notes=_to_str(row.get("Notes")),
        )
        n += 1
    return n


def import_aforos(conn, ws):
    """Hoja 'aforos': aforos BYMA por asset_class y/o ticker.

    Columnas: Scope Type ('CLASS'|'TICKER'), Scope Value, Aforo %,
              Source, Notes.

    Aforo % se acepta como 0..1 (0.85) o 0..100 (85) — se normaliza.
    """
    from .schema import insert_aforo
    n = 0
    for r, row in _read_rows(ws):
        scope_type = _to_str(row.get("Scope Type"))
        scope_value = _to_str(row.get("Scope Value"))
        aforo_pct = _to_float(row.get("Aforo %"))
        if not all([scope_type, scope_value]) or aforo_pct is None:
            continue
        if scope_type not in ("CLASS", "TICKER"):
            continue
        # Normalizar: si es > 1.5, asumimos que vino como porcentaje (ej 85.0)
        if aforo_pct > 1.5:
            aforo_pct = aforo_pct / 100.0
        if aforo_pct < 0 or aforo_pct > 1:
            continue
        insert_aforo(
            conn, scope_type=scope_type, scope_value=scope_value,
            aforo_pct=aforo_pct,
            source=_to_str(row.get("Source")),
            notes=_to_str(row.get("Notes")),
        )
        n += 1
    return n


def import_margin_config(conn, ws):
    """Hoja 'margin_config': configuración de leverage por cuenta (IBKR, etc).

    Columnas: Account, Mult. Overnight, Mult. Intraday,
              Funding Rate Annual, Funding Currency, Notes.
    """
    from .schema import insert_margin_config
    n = 0
    for r, row in _read_rows(ws):
        account = _to_str(row.get("Account"))
        if not account:
            continue
        mult_o = _to_float(row.get("Mult. Overnight")) or 1.0
        mult_i = _to_float(row.get("Mult. Intraday")) or mult_o
        funding = _to_float(row.get("Funding Rate Annual")) or 0.0
        # Aceptar 6.0 como 6% o 0.06 como 6%
        if funding > 1.5:
            funding = funding / 100.0
        insert_margin_config(
            conn, account=account,
            multiplier_overnight=mult_o,
            multiplier_intraday=mult_i,
            funding_rate_annual=funding,
            funding_currency=_to_str(row.get("Funding Currency")),
            notes=_to_str(row.get("Notes")),
        )
        n += 1
    return n


def import_config(conn, ws):
    """Hoja 'config': pares (Concepto, Valor) → settings table.

    Solo se almacenan keys que el engine reconoce. Otras filas se loguean.
    El mapping concepto → key es case-insensitive y normaliza espacios.
    """
    from .schema import set_setting

    # Mapping de "Concepto" en Excel → key canónica en settings table.
    KNOWN_KEYS = {
        "moneda base de reporte":          "anchor_currency",
        "fecha de arranque":                "start_date",
        "tolerancia qty (cero)":            "qty_zero_tolerance",
        "metodo pnl realizado":             "pnl_method",
        "método pnl realizado":             "pnl_method",
        "default plazo byma":               "byma_settle_default",
        "distancia alerta target (bps)":    "alert_distance_bps",
        "distancia alerta target":          "alert_distance_bps",
    }

    n = 0
    for r, row in _read_rows(ws):
        concepto = _to_str(row.get("Concepto"))
        valor = row.get("Valor")
        if not concepto:
            continue
        key = KNOWN_KEYS.get(concepto.strip().lower())
        if not key:
            continue
        if valor is None or (isinstance(valor, str) and not valor.strip()):
            continue
        set_setting(conn, key, valor)
        n += 1
    return n


def import_funding(conn, ws):
    """Hoja 'funding': cauciones, pases, préstamos de corto plazo.

    Cada fila genera:
      - 1 evento FUNDING_OPEN al iniciar (cash + sobre la cuenta)
      - Si Status='CLOSED': 1 evento FUNDING_CLOSE al cerrar (con interés)

    TOMA: TOMA dinero (recibo cash, debo capital + interés).
      - cocos: +cash
      - caucion_pasivo_<ccy> (LIABILITY): +deuda  ← se RESTA del PN
    COLOCA: COLOCO dinero (entrego cash, tengo crédito por cobrar).
      - cocos: -cash
      - caucion_activo_<ccy> (CASH_BROKER): +crédito  ← se SUMA al PN

    Esto netea correctamente: durante la vida de la caución, el PN no se
    infla — el cash entrante (TOMA) está compensado por la deuda generada.

    Sprint E: Si tiene 'Linked Trade ID', se usa para vincular costo/ingreso
    de funding al trade (visible en reportes de leverage).
    """
    n = 0
    for r, row in _read_rows(ws):
        fund_id = _to_str(row.get("Fund ID"))
        tipo = _to_str(row.get("Tipo"))            # TOMA | COLOCA
        subtipo = _to_str(row.get("Subtipo"))      # CAUCION, PASE, PRESTAMO_*
        cuenta = _to_str(row.get("Cuenta"))
        f_inicio = _to_date_str(row.get("Fecha Inicio"))
        f_fin = _to_date_str(row.get("Fecha Fin"))
        moneda = _to_str(row.get("Moneda"))
        monto = _to_float(row.get("Monto"))
        tna = _to_float(row.get("TNA")) or 0.0
        # 'Días' puede venir como fórmula evaluada por openpyxl con data_only=True.
        # Si llega 0 / None pero tenemos fechas válidas, lo derivamos de las
        # fechas — evita interés=0 espurio en cauciones cerradas.
        dias = _to_int(row.get("Días"))
        if (dias is None or dias == 0) and f_inicio and f_fin:
            try:
                computed = (date.fromisoformat(f_fin) - date.fromisoformat(f_inicio)).days
                if computed > 0:
                    dias = computed
            except (ValueError, TypeError):
                pass
        dias = dias or 0
        status = _to_str(row.get("Status")) or "OPEN"
        linked_trade_id = _to_str(row.get("Linked Trade ID"))
        description = _to_str(row.get("Description")) or f"{tipo} {subtipo} {monto} {moneda}"

        if not all([fund_id, tipo, cuenta, f_inicio, moneda, monto]):
            continue
        if tipo not in ("TOMA", "COLOCA"):
            continue

        # BUG 7: validar Linked Trade ID — warn si apunta a un trade que no existe.
        # No bloquea (puede ser data en orden de carga distinto), solo log.
        if linked_trade_id:
            cur = conn.execute(
                "SELECT 1 FROM events WHERE external_id=? AND event_type=? LIMIT 1",
                (linked_trade_id, EventType.TRADE),
            )
            if cur.fetchone() is None:
                import sys as _sys
                print(
                    f"[funding] WARN Fund ID={fund_id}: Linked Trade ID="
                    f"{linked_trade_id!r} no existe en blotter — la caución "
                    f"queda sin vinculación en /leverage.",
                    file=_sys.stderr,
                )

        # Normalizar TNA: si vino como 24 (asumimos %), convertir a 0.24
        if tna > 1.5:
            tna = tna / 100.0

        # Cuenta contracuenta del funding (auto-creada si no existe).
        # TOMA: deuda → kind=LIABILITY (resta del PN)
        # COLOCA: crédito por cobrar → kind=CASH_BROKER (suma al PN)
        if tipo == "TOMA":
            counter_account = f"caucion_pasivo_{moneda.lower()}"
            counter_kind = AccountKind.LIABILITY
            counter_name = f"Cauciones tomadas pendientes ({moneda})"
        else:
            counter_account = f"caucion_activo_{moneda.lower()}"
            counter_kind = AccountKind.CASH_BROKER
            counter_name = f"Cauciones colocadas pendientes ({moneda})"
        # Crear la cuenta si no existe (idempotente — INSERT OR REPLACE).
        # investible=1: estos pasivos/créditos son reales y siempre afectan PN
        # (en ambas vistas, "todo" y "solo invertible"). El flag investible
        # sigue aplicándose: el pasivo se acumula al PN invertible para
        # representar correctamente "lo que realmente tenés disponible".
        insert_account(
            conn, code=counter_account, name=counter_name, kind=counter_kind,
            currency=moneda,
            investible=1,
            cash_purpose="FUNDING",
            notes="Auto-creada por import_funding",
        )

        # FUNDING_OPEN
        sign_cash = 1 if tipo == "TOMA" else -1   # TOMA: gano cash; COLOCA: pierdo cash
        eid_open = insert_event(
            conn, EventType.FUNDING_OPEN,
            event_date=f_inicio,
            description=f"OPEN {description}",
            external_id=fund_id,
            source_row=r, source_sheet="funding",
            notes=(f"Linked Trade: {linked_trade_id}"
                   if linked_trade_id else None),
        )
        # Cash entra/sale de la cuenta
        insert_movement(
            conn, eid_open, account=cuenta, asset=moneda,
            qty=sign_cash * monto,
            notes=f"{tipo} {subtipo} {fund_id}",
        )
        # Contracuenta: deuda (TOMA) o crédito por cobrar (COLOCA).
        # Mismo signo que cash (positivo) para representar el saldo
        # acumulado de la deuda/crédito en esa cuenta.
        # Para TOMA: counter_account (LIABILITY) gana +monto = saldo deudor
        # Para COLOCA: counter_account (CASH_BROKER) gana +monto = receivable
        insert_movement(
            conn, eid_open,
            account=counter_account,
            asset=moneda,
            qty=monto,  # saldo positivo en la cuenta de funding
            notes=f"Contracuenta {tipo} {fund_id}",
        )

        # FUNDING_CLOSE: solo si status CLOSED y tenemos fecha fin
        if status == "CLOSED" and f_fin:
            interes = monto * tna * (dias / 365.0) if dias > 0 else 0.0
            # TOMA cierro: pago monto + interés (cash sale, deuda cancelada)
            # COLOCA cierro: recibo monto + interés (cash entra, crédito cancelado)
            sign_close = -sign_cash
            eid_close = insert_event(
                conn, EventType.FUNDING_CLOSE,
                event_date=f_fin,
                description=f"CLOSE {description} (int {interes:.2f})",
                external_id=fund_id,
                parent_event_id=eid_open,
                source_row=r, source_sheet="funding",
                notes=(f"Linked Trade: {linked_trade_id}"
                       if linked_trade_id else None),
            )
            # Capital de vuelta en la cuenta de cash
            insert_movement(
                conn, eid_close, account=cuenta, asset=moneda,
                qty=sign_close * monto,
                notes=f"Capital {fund_id}",
            )
            # Cancelar el saldo en la cuenta de funding (signo opuesto al OPEN)
            insert_movement(
                conn, eid_close,
                account=counter_account,
                asset=moneda,
                qty=-monto,
                notes=f"Cancelación {tipo} {fund_id}",
            )
            # Interés (si hay)
            if abs(interes) > 1e-9:
                if tipo == "TOMA":
                    # TOMA: pago interés → sale cash, va a interest_expense
                    insert_movement(
                        conn, eid_close, account=cuenta, asset=moneda,
                        qty=-interes,
                        notes=f"Interés pagado {fund_id} ({tna*100:.2f}% × {dias}d)",
                    )
                    insert_movement(
                        conn, eid_close, account="interest_expense", asset=moneda,
                        qty=interes,
                        notes=f"Interés caución {fund_id}",
                    )
                else:
                    # COLOCA: cobro interés → entra cash, va a interest_income
                    insert_movement(
                        conn, eid_close, account=cuenta, asset=moneda,
                        qty=interes,
                        notes=f"Interés cobrado {fund_id} ({tna*100:.2f}% × {dias}d)",
                    )
                    insert_movement(
                        conn, eid_close, account="interest_income", asset=moneda,
                        qty=-interes,
                        notes=f"Interés caución {fund_id}",
                    )
        n += 1
    return n


def import_especies(conn, ws):
    """Hoja 'especies': popula tabla assets."""
    n = 0
    for r, row in _read_rows(ws):
        ticker = _to_str(row.get("Ticker"))
        if not ticker:
            continue
        insert_asset(
            conn,
            ticker=ticker,
            name=_to_str(row.get("Name")) or ticker,
            asset_class=_to_str(row.get("Asset Class")) or "OTHER",
            currency=_to_str(row.get("Currency")) or "ARS",
            issuer=_to_str(row.get("Issuer")),
            sector=_to_str(row.get("Sector")),
            country=_to_str(row.get("Country")),
            notes=_to_str(row.get("Notes")),
        )
        n += 1
    return n


def _get_asset_currency(conn, ticker: str) -> Optional[str]:
    """Lee la moneda nativa de un asset desde la tabla `assets`."""
    cur = conn.execute("SELECT currency FROM assets WHERE ticker=?", (ticker,))
    row = cur.fetchone()
    if row is None:
        return None
    return row["currency"] if hasattr(row, "keys") else row[0]


def import_blotter(conn, ws):
    """Hoja 'blotter': cada fila → un event TRADE + 2 movements.

    Si la moneda del trade != moneda nativa del asset, convierte el unit_price
    a la moneda nativa usando FX del día. El cash sigue moviéndose en la
    moneda real del trade. Esto mantiene:
      - inventario en moneda nativa del bono (USB para AL30, USD para AAPL)
      - cash en la moneda real que se intercambió (ARS, USB, USDT, etc)

    Si el FX falla, NO se descarta el trade. Se registra con
    `price_currency = moneda Trade` y holdings.py se encarga de convertir
    a la moneda nativa al calcular avg_cost (con re-intento de FX).
    """
    n = 0
    fx_failed = []  # rows que tuvieron que caer al fallback (sin FX)
    for r, row in _read_rows(ws):
        ticker = _to_str(row.get("Ticker"))
        side = _to_str(row.get("Side"))
        qty = _to_float(row.get("Qty"))
        precio = _to_float(row.get("Precio"))
        moneda = _to_str(row.get("Moneda Trade"))
        cuenta = _to_str(row.get("Cuenta"))
        cuenta_cash = _to_str(row.get("Cuenta Cash")) or cuenta

        if not all([ticker, side, qty, precio, moneda, cuenta]):
            continue
        if side not in ("BUY", "SELL"):
            continue

        trade_date = _to_date_str(row.get("Trade Date"))
        settle_date = _to_date_str(row.get("Settle Date"))
        if not trade_date:
            continue

        external_id = _to_str(row.get("Trade ID"))
        description = (_to_str(row.get("Description"))
                       or f"{side} {qty} {ticker} @ {precio}")

        # Target / Stop-loss (solo se almacenan, alertas se computan en
        # holdings.py). Ambos opcionales.
        target_price = _to_float(row.get("Precio Target"))
        stop_loss_price = _to_float(row.get("Stop Loss"))
        target_currency = _to_str(row.get("Moneda Target")) or moneda

        # FX: si la moneda del trade != moneda nativa del asset, convertir
        # el unit_price a moneda nativa para que el inventario sea consistente.
        asset_ccy = _get_asset_currency(conn, ticker)
        if asset_ccy is None:
            # Asset no existe en `assets` — usar moneda del trade como fallback
            asset_ccy = moneda

        # Moneda en la que se almacena el unit_price del activo (para el
        # cálculo del avg_cost en holdings).
        # - Por default, asset_ccy (la nativa del activo)
        # - Si el FX falla, fallback a moneda Trade y holdings re-intenta
        #   la conversión al calcular tenencias.
        price_storage_ccy = asset_ccy
        if asset_ccy != moneda:
            try:
                # Convertir el precio unitario: precio (en moneda) → precio en asset_ccy
                precio_native = fx_convert(
                    conn, precio, moneda, asset_ccy, trade_date,
                )
            except FxError as e:
                # Fallback robusto: registrar el trade con price_currency =
                # moneda Trade. El cash se mueve correctamente y el avg_cost
                # se reconvierte al calcular holdings.
                print(f"[importer] WARN blotter row {r} ({ticker}): {e} — "
                      f"guardando precio en {moneda} (sin conversión)")
                precio_native = precio
                price_storage_ccy = moneda
                fx_failed.append({"row": r, "ticker": ticker,
                                   "from": moneda, "to": asset_ccy})
        else:
            precio_native = precio

        # Crear evento (con target/stop si fueron especificados)
        eid = insert_event(
            conn, EventType.TRADE,
            event_date=trade_date, settle_date=settle_date,
            description=description, external_id=external_id,
            source_row=r, source_sheet="blotter",
            notes=_to_str(row.get("Notes")),
            target_price=target_price,
            stop_loss_price=stop_loss_price,
            target_currency=target_currency if (target_price or stop_loss_price) else None,
        )

        # Movements:
        # BUY: cuenta GANA ticker, cuenta_cash PIERDE cash (qty × precio_trade)
        # SELL: cuenta PIERDE ticker, cuenta_cash GANA cash
        sign = 1 if side == "BUY" else -1

        # Cost basis y unit_price del INVENTARIO en moneda nativa del asset
        cost_basis_native = qty * precio_native
        # Cash en moneda real del trade
        cash_total = qty * precio  # siempre positivo

        # Movement 1: el activo (precio en moneda nativa o, si FX falla,
        # en la moneda del trade — holdings.py re-convierte si puede).
        notes_asset = None
        if price_storage_ccy != asset_ccy:
            notes_asset = (f"⚠ Sin FX {moneda}→{asset_ccy} en {trade_date}, "
                           f"precio en {moneda}")
        elif asset_ccy != moneda:
            notes_asset = f"FX: {moneda}→{asset_ccy} en {trade_date}"
        insert_movement(
            conn, eid,
            account=cuenta, asset=ticker,
            qty=sign * qty,
            unit_price=precio_native,
            price_currency=price_storage_ccy,
            cost_basis=sign * cost_basis_native,
            notes=notes_asset,
        )
        # Movement 2: el cash (en moneda del trade)
        insert_movement(
            conn, eid,
            account=cuenta_cash, asset=moneda,
            qty=-sign * cash_total,
        )

        # Comisión opcional
        comision = _to_float(row.get("Comisión"))
        moneda_com = _to_str(row.get("Moneda Com")) or moneda
        if comision and comision > 0:
            insert_movement(
                conn, eid,
                account=cuenta_cash, asset=moneda_com,
                qty=-comision,
                notes=f"Comisión {side}",
            )
            insert_movement(
                conn, eid,
                account="external_expense", asset=moneda_com,
                qty=comision,
                notes=f"Comisión {side}",
            )

        n += 1
    return {"n": n, "fx_failed": fx_failed}


def import_transferencias_cash(conn, ws):
    """Hoja 'transferencias_cash': cash entre cuentas propias."""
    n = 0
    for r, row in _read_rows(ws):
        monto = _to_float(row.get("Monto"))
        moneda = _to_str(row.get("Moneda"))
        origen = _to_str(row.get("Cuenta Origen"))
        destino = _to_str(row.get("Cuenta Destino"))
        fecha = _to_date_str(row.get("Fecha"))
        if not all([monto, moneda, origen, destino, fecha]):
            continue

        eid = insert_event(
            conn, EventType.TRANSFER_CASH,
            event_date=fecha,
            description=_to_str(row.get("Description")) or f"Transfer {origen}→{destino}",
            source_row=r, source_sheet="transferencias_cash",
            external_id=_to_str(row.get("Trade ID Externo")),
            notes=_to_str(row.get("Notes")),
        )
        insert_movement(conn, eid, account=origen,  asset=moneda, qty=-monto)
        insert_movement(conn, eid, account=destino, asset=moneda, qty=monto)
        n += 1
    return n


def import_transferencias_activos(conn, ws):
    """Hoja 'transferencias_activos': activos entre cuentas propias."""
    n = 0
    for r, row in _read_rows(ws):
        ticker = _to_str(row.get("Ticker"))
        qty = _to_float(row.get("Qty"))
        origen = _to_str(row.get("Cuenta Origen"))
        destino = _to_str(row.get("Cuenta Destino"))
        fecha = _to_date_str(row.get("Fecha"))
        if not all([ticker, qty, origen, destino, fecha]):
            continue

        eid = insert_event(
            conn, EventType.TRANSFER_ASSET,
            event_date=fecha,
            description=_to_str(row.get("Description")) or f"Transfer {ticker} {origen}→{destino}",
            source_row=r, source_sheet="transferencias_activos",
            notes=_to_str(row.get("Notes")),
        )
        insert_movement(conn, eid, account=origen,  asset=ticker, qty=-qty)
        insert_movement(conn, eid, account=destino, asset=ticker, qty=qty)
        n += 1
    return n


def import_ingresos(conn, ws):
    """Hoja 'ingresos': ingresos manuales (los recurrentes los maneja otra hoja)."""
    n = 0
    for r, row in _read_rows(ws):
        monto = _to_float(row.get("Monto"))
        moneda = _to_str(row.get("Moneda"))
        cuenta_dest = _to_str(row.get("Cuenta Destino"))
        fecha = _to_date_str(row.get("Fecha"))
        recurrente = _to_str(row.get("Recurrente?"))
        if recurrente == "YES":
            # Se carga desde 'recurrentes', skip
            continue
        if not all([monto, moneda, cuenta_dest, fecha]):
            continue

        eid = insert_event(
            conn, EventType.INCOME,
            event_date=fecha,
            description=_to_str(row.get("Concepto")) or "Ingreso",
            source_row=r, source_sheet="ingresos",
            notes=_to_str(row.get("Notes")),
        )
        insert_movement(conn, eid, account=cuenta_dest,    asset=moneda, qty=monto,
                        notes=_to_str(row.get("Categoría")))
        insert_movement(conn, eid, account="external_income", asset=moneda, qty=-monto)
        n += 1
    return n


def import_gastos(conn, ws):
    """Hoja 'gastos': gasto cash o tarjeta. Soporta cuotas (1 fila → N events).

    Si la cuenta destino es CARD_CREDIT y cuotas > 1:
      genera N CARD_CHARGE events, uno por cuota mensual a partir del cierre indicado.
    Si cuotas = 1:
      genera 1 EXPENSE (si cash) o 1 CARD_CHARGE (si tarjeta).
    """
    # Cargar lookup de cuentas para saber si es tarjeta
    accounts_kind = {}
    cur = conn.execute("SELECT code, kind, card_close_day FROM accounts")
    for a in cur.fetchall():
        accounts_kind[a["code"]] = (a["kind"], a["card_close_day"])

    n = 0
    for r, row in _read_rows(ws):
        monto = _to_float(row.get("Monto"))
        moneda = _to_str(row.get("Moneda"))
        cuenta_dest = _to_str(row.get("Cuenta Destino"))
        fecha = _to_date_str(row.get("Fecha"))
        cuotas = _to_int(row.get("Cuotas")) or 1
        recurrente = _to_str(row.get("Recurrente?"))
        if recurrente == "YES":
            continue
        if not all([monto, moneda, cuenta_dest, fecha]):
            continue

        kind, _close_day = accounts_kind.get(cuenta_dest, (None, None))
        is_card = (kind == AccountKind.CARD_CREDIT)
        concepto = _to_str(row.get("Concepto")) or "Gasto"
        categoria = _to_str(row.get("Categoría")) or ""
        tipo = _to_str(row.get("Tipo")) or "VARIABLE"  # FIJO/VARIABLE
        notes = _to_str(row.get("Notes"))

        if cuotas <= 1 or not is_card:
            # Caso simple: 1 evento
            event_type = EventType.CARD_CHARGE if is_card else EventType.EXPENSE
            eid = insert_event(
                conn, event_type,
                event_date=fecha,
                description=concepto,
                source_row=r, source_sheet="gastos",
                notes=f"{categoria} | {tipo}{' | ' + notes if notes else ''}",
            )
            if is_card:
                # CARD_CHARGE: incrementa pasivo de tarjeta (qty positiva en tarjeta)
                # contra external_expense
                insert_movement(
                    conn, eid, account=cuenta_dest, asset=moneda,
                    qty=monto,
                    notes=f"{categoria} | {tipo}",
                )
                insert_movement(
                    conn, eid, account="external_expense", asset=moneda,
                    qty=-monto,
                )
            else:
                # EXPENSE cash: cuenta destino paga (qty negativa)
                insert_movement(
                    conn, eid, account=cuenta_dest, asset=moneda,
                    qty=-monto,
                    notes=f"{categoria} | {tipo}",
                )
                insert_movement(
                    conn, eid, account="external_expense", asset=moneda,
                    qty=monto,
                )
            n += 1
        else:
            # Caso cuotas: N events CARD_INSTALLMENT
            cuota_monto = monto / cuotas
            # Parent event (la compra original como CARD_CHARGE diferida)
            parent_eid = insert_event(
                conn, EventType.CARD_CHARGE,
                event_date=fecha,
                description=f"{concepto} ({cuotas} cuotas)",
                source_row=r, source_sheet="gastos",
                notes=f"{categoria} | {tipo} | Cuotas={cuotas}",
            )
            # Generar N CARD_INSTALLMENT mensuales
            f0 = date.fromisoformat(fecha)
            for i in range(cuotas):
                # Cuota i va al mes i (mismo día)
                month_offset = i
                year = f0.year + (f0.month - 1 + month_offset) // 12
                month = (f0.month - 1 + month_offset) % 12 + 1
                try:
                    cuota_date = date(year, month, f0.day).isoformat()
                except ValueError:
                    # Día no válido para ese mes (ej 31 en febrero)
                    import calendar
                    last_day = calendar.monthrange(year, month)[1]
                    cuota_date = date(year, month, min(f0.day, last_day)).isoformat()
                eid = insert_event(
                    conn, EventType.CARD_INSTALLMENT,
                    event_date=cuota_date,
                    description=f"Cuota {i+1}/{cuotas}: {concepto}",
                    source_row=r, source_sheet="gastos",
                    parent_event_id=parent_eid,
                    notes=f"{categoria} | {tipo}",
                )
                insert_movement(
                    conn, eid, account=cuenta_dest, asset=moneda,
                    qty=cuota_monto,
                )
                insert_movement(
                    conn, eid, account="external_expense", asset=moneda,
                    qty=-cuota_monto,
                )
            n += 1
    return n


def import_recurrentes(conn, ws, fecha_corte: date):
    """Hoja 'recurrentes': expande hasta fecha_corte generando ingresos/gastos."""
    n = 0
    for r, row in _read_rows(ws):
        active = _to_str(row.get("Active"))
        if active != "YES":
            continue
        rule_name = _to_str(row.get("Rule Name"))
        event_type = _to_str(row.get("Event Type"))
        cuenta = _to_str(row.get("Cuenta"))
        asset = _to_str(row.get("Asset"))
        amount = _to_float(row.get("Amount"))
        description = _to_str(row.get("Description")) or rule_name
        categoria = _to_str(row.get("Categoría")) or ""
        tipo = _to_str(row.get("Tipo")) or ""
        start = _to_date_str(row.get("Start Date"))
        end = _to_date_str(row.get("End Date"))
        day_of_month = _to_int(row.get("Day of Month")) or 1

        if not all([rule_name, event_type, cuenta, asset, amount, start]):
            continue

        # Generar ocurrencias mensuales desde start hasta min(end, fecha_corte)
        d_start = date.fromisoformat(start)
        d_end = date.fromisoformat(end) if end else fecha_corte
        d_end = min(d_end, fecha_corte)

        # Iterar mes por mes
        y, m = d_start.year, d_start.month
        # Si d_start tiene día > day_of_month, arrancar el mes siguiente
        if d_start.day > day_of_month:
            m += 1
            if m > 12:
                m = 1; y += 1

        import calendar
        while True:
            last_day = calendar.monthrange(y, m)[1]
            d = date(y, m, min(day_of_month, last_day))
            if d > d_end:
                break
            if d >= d_start:
                # Generar evento
                if event_type == "INCOME":
                    eid = insert_event(
                        conn, EventType.INCOME,
                        event_date=d.isoformat(),
                        description=description,
                        source_row=r, source_sheet="recurrentes",
                        notes=f"Auto: {rule_name}",
                    )
                    insert_movement(conn, eid, account=cuenta, asset=asset, qty=amount,
                                    notes=categoria)
                    insert_movement(conn, eid, account="external_income", asset=asset, qty=-amount)
                elif event_type == "EXPENSE":
                    eid = insert_event(
                        conn, EventType.EXPENSE,
                        event_date=d.isoformat(),
                        description=description,
                        source_row=r, source_sheet="recurrentes",
                        notes=f"Auto: {rule_name} | {categoria} | {tipo}",
                    )
                    insert_movement(conn, eid, account=cuenta, asset=asset, qty=-amount,
                                    notes=f"{categoria} | {tipo}")
                    insert_movement(conn, eid, account="external_expense", asset=asset, qty=amount)
                elif event_type == "CARD_CHARGE":
                    eid = insert_event(
                        conn, EventType.CARD_CHARGE,
                        event_date=d.isoformat(),
                        description=description,
                        source_row=r, source_sheet="recurrentes",
                        notes=f"Auto: {rule_name} | {categoria} | {tipo}",
                    )
                    insert_movement(conn, eid, account=cuenta, asset=asset, qty=amount)
                    insert_movement(conn, eid, account="external_expense", asset=asset, qty=-amount)
                n += 1
            m += 1
            if m > 12:
                m = 1; y += 1

    return n


def import_pagos_pasivos(conn, ws):
    """Pagos a pasivos: cuotas de préstamos y cancelaciones de tarjeta."""
    n = 0
    for r, row in _read_rows(ws):
        fecha = _to_date_str(row.get("Fecha"))
        target = _to_str(row.get("Pasivo / Tarjeta"))
        monto_total = _to_float(row.get("Monto Total"))
        capital = _to_float(row.get("Capital")) or monto_total
        interes = _to_float(row.get("Interés")) or 0
        moneda = _to_str(row.get("Moneda"))
        cuenta_origen = _to_str(row.get("Cuenta Origen"))
        if not all([fecha, target, monto_total, moneda, cuenta_origen]):
            continue

        # Determinar si target es CARD_CREDIT o LIABILITY
        cur = conn.execute("SELECT kind FROM accounts WHERE code=?", (target,))
        row_acc = cur.fetchone()
        if row_acc is None:
            continue
        kind = row_acc["kind"]
        is_card = (kind == AccountKind.CARD_CREDIT)
        event_type = EventType.CARD_PAYMENT if is_card else EventType.LIABILITY_PAYMENT

        eid = insert_event(
            conn, event_type,
            event_date=fecha,
            description=_to_str(row.get("Description"))
                       or f"Pago {target} {monto_total:.2f} {moneda}",
            source_row=r, source_sheet="pagos_pasivos",
            notes=_to_str(row.get("Notes")),
        )
        # cash sale de cuenta_origen
        insert_movement(conn, eid, account=cuenta_origen, asset=moneda, qty=-monto_total)
        # tarjeta/pasivo se reduce: qty negativa (saldo decrece)
        insert_movement(conn, eid, account=target, asset=moneda, qty=-capital,
                        notes=f"Capital pagado")
        # interés (si > 0)
        if interes > 0:
            insert_movement(conn, eid, account="interest_expense", asset=moneda, qty=interes)
            # ajuste: total - capital - interés debería balancear
            # Si capital + interés ≠ total, generamos un movement de ajuste implícito
            # En el modelo simple, asumimos capital + interés = monto_total
        n += 1
    return n


def import_asientos(conn, ws):
    """Asientos contables manuales. Filas con mismo Event ID se agrupan en 1 evento."""
    grupos = {}  # event_id externo → list of rows
    for r, row in _read_rows(ws):
        eid_ext = _to_str(row.get("Event ID"))
        if not eid_ext:
            continue
        grupos.setdefault(eid_ext, []).append((r, row))

    n = 0
    for eid_ext, rows in grupos.items():
        # Tomar fecha y descripción de la primera fila
        first = rows[0][1]
        fecha = _to_date_str(first.get("Fecha"))
        desc = _to_str(first.get("Description"))
        if not fecha:
            continue
        eid = insert_event(
            conn, EventType.ACCOUNTING_ADJUSTMENT,
            event_date=fecha,
            description=desc or f"Asiento {eid_ext}",
            external_id=eid_ext,
            source_row=rows[0][0], source_sheet="asientos_contables",
        )
        # Generar todos los movements
        for r, row in rows:
            cuenta = _to_str(row.get("Cuenta"))
            asset = _to_str(row.get("Activo"))
            qty = _to_float(row.get("Qty (signada)"))
            unit_price = _to_float(row.get("Unit Price"))
            price_curr = _to_str(row.get("Price Currency"))
            if not all([cuenta, asset, qty]):
                continue
            cost_basis = qty * unit_price if unit_price else None
            insert_movement(
                conn, eid,
                account=cuenta, asset=asset, qty=qty,
                unit_price=unit_price, price_currency=price_curr,
                cost_basis=cost_basis,
                notes=_to_str(row.get("Notes")),
            )
        n += 1
    return n


# =============================================================================
# Importer principal
# =============================================================================

def import_all(db_path: str | Path, xlsx_path: str | Path,
               fecha_corte: date = None,
               fx_csv_path: str | Path = None,
               data_dir: str | Path = "data") -> dict:
    """Importa el master Excel completo a la DB. Devuelve estadísticas.

    fx_csv_path: si se pasa, importa ese CSV específicamente.
                 Si es None, intenta auto-cargar `<data_dir>/fx_historico.csv`.
    """
    from .schema import init_db
    if fecha_corte is None:
        fecha_corte = date.today()

    print(f"[importer] DB: {db_path}")
    print(f"[importer] XLSX: {xlsx_path}")
    print(f"[importer] Fecha de corte (para recurrentes): {fecha_corte}")

    # Recrear DB para idempotencia
    conn = init_db(db_path, drop_existing=True)
    wb = load_workbook(filename=str(xlsx_path), data_only=True)

    stats = {}

    # 1. Maestros (en orden de FK)
    if "monedas" in wb.sheetnames:
        stats["monedas"] = import_monedas(conn, wb["monedas"])
    if "cuentas" in wb.sheetnames:
        stats["cuentas"] = import_cuentas(conn, wb["cuentas"])
    if "especies" in wb.sheetnames:
        stats["especies"] = import_especies(conn, wb["especies"])
    if "aforos" in wb.sheetnames:
        stats["aforos"] = import_aforos(conn, wb["aforos"])
    if "margin_config" in wb.sheetnames:
        stats["margin_config"] = import_margin_config(conn, wb["margin_config"])
    if "config" in wb.sheetnames:
        stats["config"] = import_config(conn, wb["config"])

    conn.commit()

    # 1b. Cargar FX histórico ANTES de procesar trades
    #     (necesario para conversión cuando Moneda Trade != asset.currency)
    if fx_csv_path:
        from .fx import import_fx_csv
        stats["fx_rates"] = import_fx_csv(conn, fx_csv_path)
    else:
        stats["fx_rates"] = auto_load_fx(conn, data_dir)
    conn.commit()

    # 1c. Cargar precios de mercado (BYMA + CAFCI + cripto + yfinance)
    from .prices import auto_load_all as auto_load_prices
    price_stats = auto_load_prices(conn, data_dir)
    stats["prices"] = sum(price_stats.values())
    conn.commit()

    # 2. Eventos
    if "blotter" in wb.sheetnames:
        bl = import_blotter(conn, wb["blotter"])
        # Compat: import_blotter ahora devuelve dict {n, fx_failed}; tests
        # viejos pueden esperar un int — exponemos n en stats["blotter"]
        # y la lista de fallas en stats["blotter_fx_failed"].
        if isinstance(bl, dict):
            stats["blotter"] = bl["n"]
            if bl.get("fx_failed"):
                stats["blotter_fx_failed"] = bl["fx_failed"]
        else:
            stats["blotter"] = bl
    if "transferencias_cash" in wb.sheetnames:
        stats["transferencias_cash"] = import_transferencias_cash(conn, wb["transferencias_cash"])
    if "transferencias_activos" in wb.sheetnames:
        stats["transferencias_activos"] = import_transferencias_activos(conn, wb["transferencias_activos"])
    if "ingresos" in wb.sheetnames:
        stats["ingresos"] = import_ingresos(conn, wb["ingresos"])
    if "gastos" in wb.sheetnames:
        stats["gastos"] = import_gastos(conn, wb["gastos"])
    if "recurrentes" in wb.sheetnames:
        stats["recurrentes"] = import_recurrentes(conn, wb["recurrentes"], fecha_corte)
    if "pagos_pasivos" in wb.sheetnames:
        stats["pagos_pasivos"] = import_pagos_pasivos(conn, wb["pagos_pasivos"])
    if "asientos_contables" in wb.sheetnames:
        stats["asientos_contables"] = import_asientos(conn, wb["asientos_contables"])
    if "funding" in wb.sheetnames:
        stats["funding"] = import_funding(conn, wb["funding"])

    conn.commit()
    conn.close()
    return stats


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("USO: python -m engine.importer <xlsx_path> <db_path>")
        sys.exit(1)
    stats = import_all(sys.argv[2], sys.argv[1])
    for k, v in stats.items():
        print(f"  {k}: {v}")
