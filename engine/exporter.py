# -*- coding: utf-8 -*-
"""
engine/exporter.py

Generador de reportes Excel multi-sheet y HTML autocontenido del portfolio.

USO:
    from engine.exporter import export_excel, export_html

    export_excel(conn, "reports/2026-05-03.xlsx", fecha=date(2026, 5, 3))
    export_html(conn, "reports/2026-05-03.html", fecha=date(2026, 5, 3))

HOJAS DEL EXCEL:
    1. Dashboard
    2. Holdings
    3. PN por cuenta
    4. PN por asset class
    5. Cash Position
    6. Tarjetas
    7. PnL Realizado FIFO
    8. PnL No-Realizado
"""

from __future__ import annotations

import json
import sqlite3
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .holdings import (
    calculate_holdings, total_pn,
    by_asset_class, by_account, by_currency,
    filter_investible, filter_non_investible, by_cash_purpose,
)
from .pnl import (
    calculate_realized_pnl, aggregate_pnl_by_asset,
    aggregate_pnl_by_year, aggregate_pnl_by_year_currency,
    total_realized_pnl,
    calculate_unrealized_pnl_summary,
)
from .liabilities import all_card_snapshots
from .trade_stats import (
    calculate_trade_stats, trade_stats_by_asset, trade_stats_by_account,
)
from .snapshots import (
    record_snapshots, get_equity_curve, get_equity_curves_by_account,
    calculate_returns, TOTAL_KEY, TOTAL_INV_KEY,
)
from .buying_power import buying_power_summary


# =============================================================================
# Estilos Excel
# =============================================================================

NAVY = "1F3864"
LIGHT_NAVY = "2E5B9C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "00B050"
RED = "C00000"

FONT_TITLE = Font(name="Arial", size=18, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="Arial", size=11, italic=True, color="595959")
FONT_HEADER = Font(name="Arial", size=11, bold=True, color=WHITE)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_KPI = Font(name="Arial", size=22, bold=True, color=NAVY)
FONT_KPI_LABEL = Font(name="Arial", size=10, color="595959")

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_SUBTOTAL = PatternFill("solid", fgColor=LIGHT_GRAY)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

THIN = Side(style="thin", color="BFBFBF")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_NUMBER = '#,##0.00;[Red](#,##0.00);"-"'
FMT_NUMBER4 = '#,##0.0000;[Red](#,##0.0000);"-"'
FMT_PCT = '0.00%;[Red](0.00%);"-"'
FMT_INT = '#,##0;[Red](#,##0);"-"'
FMT_DATE = "yyyy-mm-dd"


# =============================================================================
# Helpers
# =============================================================================

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_headers(ws, row, headers, widths=None):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    if widths:
        _set_col_widths(ws, widths)


def _write_row(ws, row, values, formats=None, bold=False):
    formats = formats or [None] * len(values)
    for c, (val, fmt) in enumerate(zip(values, formats), start=1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = FONT_BOLD if bold else FONT_NORMAL
        cell.border = BORDER_THIN
        if fmt:
            cell.number_format = fmt
        if isinstance(val, (int, float)):
            cell.alignment = ALIGN_RIGHT


# =============================================================================
# Sheets
# =============================================================================

def _sheet_dashboard(wb, holdings, anchor_ccy, fecha):
    ws = wb.create_sheet("Dashboard", 0)

    # Título
    ws.cell(row=1, column=1, value=f"PORTFOLIO  —  {fecha.isoformat()}").font = FONT_TITLE
    ws.cell(row=2, column=1, value=f"Moneda ancla: {anchor_ccy}").font = FONT_SUBTITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # KPI: PN total
    tp = total_pn(holdings, anchor_ccy)
    ws.cell(row=4, column=1, value="PATRIMONIO NETO TOTAL").font = FONT_KPI_LABEL
    cell = ws.cell(row=5, column=1, value=tp["total_anchor"])
    cell.font = FONT_KPI
    cell.number_format = FMT_NUMBER
    ws.cell(row=6, column=1, value=anchor_ccy).font = FONT_SUBTITLE

    if tp["total_unconverted_count"] > 0:
        ws.cell(row=7, column=1,
                value=f"⚠ {tp['total_unconverted_count']} posiciones sin FX, no incluidas").font = FONT_SUBTITLE

    # PN por asset class
    ws.cell(row=9, column=1, value="POR ASSET CLASS").font = FONT_BOLD
    _write_headers(ws, 10, ["Clase", f"MV ({anchor_ccy})", "% Total"])
    cls_data = by_asset_class(holdings)
    cls_total = sum(cls_data.values())
    row = 11
    for cls, val in cls_data.items():
        pct = (val / cls_total) if cls_total else 0
        _write_row(ws, row, [cls, val, pct], formats=[None, FMT_NUMBER, FMT_PCT])
        row += 1

    # PN por moneda nativa (a la derecha)
    ws.cell(row=9, column=5, value="POR MONEDA NATIVA").font = FONT_BOLD
    _write_headers(ws, 10, ["Moneda", f"MV ({anchor_ccy})", "% Total"])
    # Reescribir headers en columnas 5-7
    for c, h in enumerate(["Moneda", f"MV ({anchor_ccy})", "% Total"], start=5):
        cell = ws.cell(row=10, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    cur_data = by_currency(holdings)
    cur_total = sum(cur_data.values())
    row = 11
    for ccy, val in cur_data.items():
        pct = (val / cur_total) if cur_total else 0
        for c, (v, fmt) in enumerate(zip([ccy, val, pct], [None, FMT_NUMBER, FMT_PCT]), start=5):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            if fmt: cell.number_format = fmt
            if isinstance(v, (int, float)): cell.alignment = ALIGN_RIGHT
        row += 1

    # Top 10 cuentas
    next_row = max(11 + len(cls_data), 11 + len(cur_data)) + 2
    ws.cell(row=next_row, column=1, value="TOP CUENTAS").font = FONT_BOLD
    _write_headers(ws, next_row + 1, ["Cuenta", f"MV ({anchor_ccy})", "% Total"])
    accs = by_account(holdings)
    accs_total = sum(accs.values())
    row = next_row + 2
    for acc, val in list(accs.items())[:10]:
        pct = (val / accs_total) if accs_total else 0
        _write_row(ws, row, [acc, val, pct], formats=[None, FMT_NUMBER, FMT_PCT])
        row += 1

    _set_col_widths(ws, [22, 18, 12, 4, 22, 18, 12, 4])
    ws.row_dimensions[5].height = 35


def _sheet_holdings(wb, holdings, anchor_ccy):
    ws = wb.create_sheet("Holdings")

    headers = [
        "Cuenta", "Activo", "Asset Class", "Qty",
        "Avg Cost", "Mkt Price", "Native Ccy",
        "MV (Native)", f"MV ({anchor_ccy})",
        "Unrealized PnL (Native)", "Unrealized %",
        "Price Source", "Price Date", "Fallback",
    ]
    widths = [22, 22, 14, 14, 12, 12, 9, 16, 16, 18, 12, 14, 12, 9]
    _write_headers(ws, 1, headers, widths)

    for i, h in enumerate(holdings, start=2):
        _write_row(ws, i, [
            h["account"],
            h["asset"],
            h["asset_class"] or "",
            h["qty"],
            h["avg_cost"],
            h["market_price"],
            h["native_currency"],
            h["mv_native"],
            h["mv_anchor"],
            h["unrealized_pnl_native"],
            h["unrealized_pct"],
            h["price_source"] or "",
            h["price_date"] or "",
            "Sí" if h["price_fallback"] else "",
        ], formats=[
            None, None, None, FMT_NUMBER4,
            FMT_NUMBER4, FMT_NUMBER4, None,
            FMT_NUMBER, FMT_NUMBER,
            FMT_NUMBER, FMT_PCT,
            None, None, None,
        ])

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(holdings) + 1}"


def _sheet_pn_by_account(wb, holdings, anchor_ccy):
    ws = wb.create_sheet("PN por cuenta")

    headers = ["Cuenta", f"MV ({anchor_ccy})", "% Total", "# Posiciones"]
    widths = [25, 18, 12, 14]
    _write_headers(ws, 1, headers, widths)

    accs = by_account(holdings)
    total = sum(accs.values())
    counts = defaultdict(int)
    for h in holdings:
        if h["mv_anchor"] is not None:
            counts[h["account"]] += 1

    for i, (acc, val) in enumerate(accs.items(), start=2):
        pct = (val / total) if total else 0
        _write_row(ws, i, [acc, val, pct, counts[acc]],
                   formats=[None, FMT_NUMBER, FMT_PCT, FMT_INT])

    # Total
    row = len(accs) + 2
    _write_row(ws, row, ["TOTAL", total, 1.0, sum(counts.values())],
               formats=[None, FMT_NUMBER, FMT_PCT, FMT_INT], bold=True)
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = FILL_SUBTOTAL


def _sheet_pn_by_class(wb, holdings, anchor_ccy):
    ws = wb.create_sheet("PN por asset class")

    headers = ["Asset Class", f"MV ({anchor_ccy})", "% Total", "# Posiciones"]
    widths = [22, 18, 12, 14]
    _write_headers(ws, 1, headers, widths)

    cls = by_asset_class(holdings)
    total = sum(cls.values())
    counts = defaultdict(int)
    for h in holdings:
        if h["mv_anchor"] is not None:
            counts[h["asset_class"] or "UNKNOWN"] += 1

    for i, (c_name, val) in enumerate(cls.items(), start=2):
        pct = (val / total) if total else 0
        _write_row(ws, i, [c_name, val, pct, counts[c_name]],
                   formats=[None, FMT_NUMBER, FMT_PCT, FMT_INT])

    row = len(cls) + 2
    _write_row(ws, row, ["TOTAL", total, 1.0, sum(counts.values())],
               formats=[None, FMT_NUMBER, FMT_PCT, FMT_INT], bold=True)
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = FILL_SUBTOTAL


def _sheet_cash_position(wb, holdings, anchor_ccy):
    ws = wb.create_sheet("Cash Position")

    headers = ["Cuenta", "Moneda", "Qty", f"MV ({anchor_ccy})"]
    widths = [25, 10, 18, 18]
    _write_headers(ws, 1, headers, widths)

    cash_only = [h for h in holdings if h["is_cash"]]
    cash_only.sort(key=lambda h: -(h["mv_anchor"] or 0))

    for i, h in enumerate(cash_only, start=2):
        _write_row(ws, i, [
            h["account"],
            h["native_currency"],
            h["qty"],
            h["mv_anchor"],
        ], formats=[None, None, FMT_NUMBER, FMT_NUMBER])

    # Subtotales por moneda
    by_ccy = defaultdict(float)
    for h in cash_only:
        if h["mv_anchor"] is not None:
            by_ccy[h["native_currency"]] += h["mv_anchor"]

    row = len(cash_only) + 3
    ws.cell(row=row, column=1, value="POR MONEDA").font = FONT_BOLD
    row += 1
    _write_headers(ws, row, ["Moneda", f"Total ({anchor_ccy})"])
    for ccy, val in sorted(by_ccy.items(), key=lambda kv: -kv[1]):
        row += 1
        _write_row(ws, row, [ccy, val], formats=[None, FMT_NUMBER])


def _sheet_tarjetas(wb, conn, fecha):
    ws = wb.create_sheet("Tarjetas")

    cards = all_card_snapshots(conn, fecha)

    headers = [
        "Tarjeta", "Code", "Moneda",
        "Saldo Actual", "Último Resumen", "Fecha Último Cierre",
        "Próximo Vencimiento", "Fecha Próximo Cierre", "Fecha Próximo Vto",
    ]
    widths = [25, 24, 8, 16, 16, 18, 18, 18, 18]
    _write_headers(ws, 1, headers, widths)

    for i, c in enumerate(cards, start=2):
        _write_row(ws, i, [
            c.card_name,
            c.card_code,
            c.currency,
            c.saldo_actual,
            c.saldo_ultimo_resumen,
            c.fecha_ultimo_cierre or "",
            c.saldo_proximo_vto,
            c.fecha_proximo_cierre or "",
            c.fecha_proximo_vto or "",
        ], formats=[
            None, None, None,
            FMT_NUMBER, FMT_NUMBER, None,
            FMT_NUMBER, None, None,
        ])

    # Totales por moneda
    if cards:
        from collections import defaultdict
        totals = defaultdict(float)
        for c in cards:
            totals[c.currency] += c.saldo_actual

        row = len(cards) + 3
        ws.cell(row=row, column=1, value="TOTAL POR MONEDA").font = FONT_BOLD
        row += 1
        for ccy, val in totals.items():
            ws.cell(row=row, column=1, value=ccy).font = FONT_BOLD
            cell = ws.cell(row=row, column=4, value=val)
            cell.font = FONT_BOLD
            cell.number_format = FMT_NUMBER
            row += 1


def _sheet_pnl_realizado(wb, fills, anchor_ccy):
    ws = wb.create_sheet("PnL Realizado FIFO")

    if not fills:
        ws.cell(row=1, column=1, value="No hay PnL realizado todavía.").font = FONT_SUBTITLE
        return

    headers = [
        "Fecha Venta", "Fecha Compra", "Holding (días)",
        "Cuenta", "Activo", "Qty",
        "Precio Compra", "Precio Venta", "Moneda",
        "PnL Realizado", "PnL %",
    ]
    widths = [12, 12, 12, 22, 22, 14, 14, 14, 8, 16, 10]
    _write_headers(ws, 1, headers, widths)

    for i, f in enumerate(fills, start=2):
        _write_row(ws, i, [
            f.fecha_venta,
            f.fecha_compra,
            f.holding_period_days,
            f.account,
            f.asset,
            f.qty,
            f.precio_compra,
            f.precio_venta,
            f.currency,
            f.pnl_realizado,
            f.pnl_pct,
        ], formats=[
            None, None, FMT_INT,
            None, None, FMT_NUMBER4,
            FMT_NUMBER4, FMT_NUMBER4, None,
            FMT_NUMBER, FMT_PCT,
        ])

    # Total al final
    row = len(fills) + 3
    ws.cell(row=row, column=1, value="TOTAL POR MONEDA").font = FONT_BOLD
    row += 1
    totals = total_realized_pnl(fills)
    for ccy, val in totals.items():
        ws.cell(row=row, column=1, value=ccy).font = FONT_BOLD
        cell = ws.cell(row=row, column=2, value=val)
        cell.font = FONT_BOLD
        cell.number_format = FMT_NUMBER
        row += 1

    ws.freeze_panes = "A2"


def _sheet_trade_stats(wb, fills, anchor_ccy):
    """Sprint A: métricas de trading (winners/losers, winrate, expectancy)."""
    ws = wb.create_sheet("Trade Stats")

    if not fills:
        ws.cell(row=1, column=1,
                value="Aún no hay trades cerrados para calcular métricas.").font = FONT_SUBTITLE
        return

    stats_by_ccy = calculate_trade_stats(fills)

    ws.cell(row=1, column=1, value="MÉTRICAS DE TRADING POR MONEDA").font = FONT_BOLD
    ws.cell(row=2, column=1,
            value="Performance histórica de los trades cerrados (FIFO).").font = FONT_SUBTITLE

    headers = [
        "Moneda", "# Trades", "# Ganadores", "# Perdedores", "# Scratch",
        "Winrate", "Gross Profit", "Gross Loss", "Net PnL",
        "Avg Winner", "Avg Loser", "Profit Factor", "Expectancy",
        "Best Trade", "Worst Trade", "Avg Holding (días)",
        "Racha Win Max", "Racha Loss Max",
    ]
    widths = [8, 10, 12, 12, 10, 10, 14, 14, 14, 12, 12, 12, 12, 14, 14, 14, 12, 12]
    _write_headers(ws, 4, headers, widths)

    row = 5
    for ccy, s in sorted(stats_by_ccy.items()):
        pf = s.profit_factor
        pf_display = "" if pf == float("inf") else pf
        _write_row(ws, row, [
            s.currency, s.n_trades, s.n_winners, s.n_losers, s.n_scratch,
            s.winrate, s.gross_profit, s.gross_loss, s.net_pnl,
            s.avg_winner, s.avg_loser, pf_display, s.expectancy,
            s.best_trade, s.worst_trade, s.avg_holding_days,
            s.largest_streak_wins, s.largest_streak_losses,
        ], formats=[
            None, FMT_INT, FMT_INT, FMT_INT, FMT_INT,
            FMT_PCT, FMT_NUMBER, FMT_NUMBER, FMT_NUMBER,
            FMT_NUMBER, FMT_NUMBER, FMT_NUMBER, FMT_NUMBER,
            FMT_NUMBER, FMT_NUMBER, FMT_NUMBER,
            FMT_INT, FMT_INT,
        ])
        row += 1

    # Subseción: por activo
    row += 2
    ws.cell(row=row, column=1, value="POR ACTIVO").font = FONT_BOLD
    row += 1
    h2 = ["Activo", "Moneda", "# Trades", "# Win", "# Loss", "Winrate",
          "Net PnL", "Avg PnL"]
    _write_headers(ws, row, h2, [22, 8, 10, 8, 8, 10, 14, 14])
    by_asset = trade_stats_by_asset(fills)
    for item in by_asset:
        row += 1
        _write_row(ws, row, [
            item["asset"], item["currency"], item["n_trades"],
            item["n_winners"], item["n_losers"], item["winrate"],
            item["net_pnl"], item["avg_pnl"],
        ], formats=[
            None, None, FMT_INT, FMT_INT, FMT_INT, FMT_PCT,
            FMT_NUMBER, FMT_NUMBER,
        ])

    # Subseción: por cuenta
    row += 3
    ws.cell(row=row, column=1, value="POR CUENTA").font = FONT_BOLD
    row += 1
    h3 = ["Cuenta", "Moneda", "# Trades", "# Win", "# Loss", "Winrate", "Net PnL"]
    _write_headers(ws, row, h3, [22, 8, 10, 8, 8, 10, 14])
    by_acc = trade_stats_by_account(fills)
    for item in by_acc:
        row += 1
        _write_row(ws, row, [
            item["account"], item["currency"], item["n_trades"],
            item["n_winners"], item["n_losers"], item["winrate"],
            item["net_pnl"],
        ], formats=[
            None, None, FMT_INT, FMT_INT, FMT_INT, FMT_PCT, FMT_NUMBER,
        ])


def _sheet_invertible(wb, holdings, anchor_ccy):
    """Sprint B: vista de PN invertible vs no-invertible."""
    ws = wb.create_sheet("PN Invertible")

    tp = total_pn(holdings, anchor_ccy)

    ws.cell(row=1, column=1, value="PN INVERTIBLE vs NO-INVERTIBLE").font = FONT_BOLD
    ws.cell(row=2, column=1,
            value="Excluye cuentas marcadas como Investible=NO en la hoja 'cuentas'.").font = FONT_SUBTITLE

    ws.cell(row=4, column=1, value="PN Total").font = FONT_NORMAL
    cell = ws.cell(row=4, column=2, value=tp["total_anchor"])
    cell.number_format = FMT_NUMBER
    ws.cell(row=4, column=3, value=anchor_ccy).font = FONT_SUBTITLE

    ws.cell(row=5, column=1, value="PN Invertible").font = FONT_BOLD
    cell = ws.cell(row=5, column=2, value=tp["total_investible"])
    cell.font = FONT_BOLD
    cell.number_format = FMT_NUMBER

    ws.cell(row=6, column=1, value="PN No-Invertible").font = FONT_NORMAL
    cell = ws.cell(row=6, column=2, value=tp["total_non_investible"])
    cell.number_format = FMT_NUMBER

    # Cash por purpose
    ws.cell(row=8, column=1, value="CASH POR PROPÓSITO").font = FONT_BOLD
    _write_headers(ws, 9, ["Cash Purpose", f"MV ({anchor_ccy})"], [28, 16])
    purposes = by_cash_purpose(holdings)
    row = 10
    for p, val in purposes.items():
        _write_row(ws, row, [p, val], formats=[None, FMT_NUMBER])
        row += 1

    # Detalle no-invertibles
    row += 2
    ws.cell(row=row, column=1, value="HOLDINGS NO-INVERTIBLES").font = FONT_BOLD
    row += 1
    _write_headers(ws, row, [
        "Cuenta", "Activo", "Asset Class", "Qty", f"MV ({anchor_ccy})", "Cash Purpose",
    ], [22, 22, 14, 14, 16, 24])
    non_inv = filter_non_investible(holdings)
    non_inv = [h for h in non_inv if h["mv_anchor"] is not None and abs(h["mv_anchor"]) > 1e-6]
    for h in non_inv:
        row += 1
        _write_row(ws, row, [
            h["account"], h["asset"], h["asset_class"] or "",
            h["qty"], h["mv_anchor"], h.get("cash_purpose") or "",
        ], formats=[None, None, None, FMT_NUMBER4, FMT_NUMBER, None])


def _sheet_buying_power(wb, conn, holdings, anchor_ccy):
    """Sprint D: poder de compra por cuenta (BYMA aforos + IBKR margin)."""
    ws = wb.create_sheet("Buying Power")

    ws.cell(row=1, column=1, value="PODER DE COMPRA POR CUENTA").font = FONT_BOLD
    ws.cell(row=2, column=1,
            value=("Cocos/Eco: aforos BYMA. IBKR: RegT margin (verificá parámetros). "
                   "Valores en moneda ancla.")).font = FONT_SUBTITLE

    summary = buying_power_summary(conn, holdings, anchor_ccy)

    if not summary:
        ws.cell(row=4, column=1,
                value="No hay cuentas broker con holdings/cash para calcular BP.").font = FONT_SUBTITLE
        return

    # Sección 1: resumen
    headers = [
        "Cuenta", "Tipo", f"Equity ({anchor_ccy})",
        f"Cash ({anchor_ccy})", f"Holdings MV ({anchor_ccy})",
        f"Garantía ({anchor_ccy})", f"Poder Compra ({anchor_ccy})",
        "Leverage Ratio", "Notas",
    ]
    widths = [22, 10, 16, 16, 18, 16, 18, 12, 35]
    _write_headers(ws, 4, headers, widths)
    row = 5

    for item in summary:
        if item["type"] == "BYMA":
            bp = item["result"]
            equity = bp.cash_total + bp.holdings_mv
            _write_row(ws, row, [
                bp.account, "BYMA", equity,
                bp.cash_total, bp.holdings_mv,
                bp.garantia_total, bp.poder_de_compra,
                bp.leverage_ratio, "Aforos BYMA aplicados",
            ], formats=[
                None, None, FMT_NUMBER,
                FMT_NUMBER, FMT_NUMBER,
                FMT_NUMBER, FMT_NUMBER,
                FMT_NUMBER, None,
            ])
            row += 1
        elif item["type"] == "MARGIN":
            bp_o = item["overnight"]
            bp_i = item["intraday"]
            _write_row(ws, row, [
                bp_o.account, "MARGIN (overnight)", bp_o.equity,
                "", "",
                bp_o.margin_disponible, bp_o.poder_de_compra,
                bp_o.multiplier,
                f"Mult x{bp_o.multiplier:.1f} | Funding {bp_o.funding_rate_annual*100:.2f}%/año",
            ], formats=[
                None, None, FMT_NUMBER,
                None, None,
                FMT_NUMBER, FMT_NUMBER,
                FMT_NUMBER, None,
            ])
            row += 1
            _write_row(ws, row, [
                bp_i.account, "MARGIN (intraday)", bp_i.equity,
                "", "",
                bp_i.margin_disponible, bp_i.poder_de_compra,
                bp_i.multiplier,
                f"Mult x{bp_i.multiplier:.1f} | Day-trading BP",
            ], formats=[
                None, None, FMT_NUMBER,
                None, None,
                FMT_NUMBER, FMT_NUMBER,
                FMT_NUMBER, None,
            ])
            row += 1

    # Sección 2: detalle de aforos para cuentas BYMA
    row += 2
    ws.cell(row=row, column=1,
            value="DETALLE GARANTÍAS POR HOLDING (BYMA)").font = FONT_BOLD
    row += 1
    _write_headers(ws, row, [
        "Cuenta", "Activo", "Asset Class",
        f"MV ({anchor_ccy})", "Aforo %", f"Garantía ({anchor_ccy})",
    ], [22, 18, 12, 16, 10, 16])
    for item in summary:
        if item["type"] != "BYMA":
            continue
        bp = item["result"]
        for d in bp.detalle_por_holding:
            row += 1
            _write_row(ws, row, [
                bp.account, d["asset"], d["asset_class"],
                d["mv_anchor"], d["aforo_pct"], d["garantia"],
            ], formats=[
                None, None, None,
                FMT_NUMBER, FMT_PCT, FMT_NUMBER,
            ])


def _sheet_equity_curve(wb, conn, anchor_ccy):
    """Sprint C: equity curve desde snapshots históricos."""
    ws = wb.create_sheet("Equity Curve")

    total_curve = get_equity_curve(conn, anchor_currency=anchor_ccy)
    inv_curve = get_equity_curve(conn, anchor_currency=anchor_ccy,
                                  investible_only=True)

    ws.cell(row=1, column=1, value="EQUITY CURVE").font = FONT_BOLD
    ws.cell(row=2, column=1,
            value="Snapshots históricos del PN. Se appendea uno por cada corrida del reporte.").font = FONT_SUBTITLE

    if not total_curve:
        ws.cell(row=4, column=1,
                value="Sin snapshots aún. Corré el reporte algunos días para ver evolución.").font = FONT_SUBTITLE
        return

    # Métricas
    rets = calculate_returns(total_curve)
    inv_rets = calculate_returns(inv_curve)

    ws.cell(row=4, column=1, value="MÉTRICAS PORTFOLIO").font = FONT_BOLD
    metrics = [
        ("PN Inicial", rets["first_value"], FMT_NUMBER),
        ("PN Actual", rets["last_value"], FMT_NUMBER),
        ("Retorno Absoluto", rets["total_return_abs"], FMT_NUMBER),
        ("Retorno %", rets["total_return_pct"], FMT_PCT),
        ("Max Drawdown %", rets["max_drawdown_pct"], FMT_PCT),
        ("# Snapshots", rets["n_periods"], FMT_INT),
    ]
    for i, (label, val, fmt) in enumerate(metrics, start=5):
        ws.cell(row=i, column=1, value=label).font = FONT_NORMAL
        c = ws.cell(row=i, column=2, value=val)
        c.number_format = fmt
        c.alignment = ALIGN_RIGHT
    set_col = 4
    ws.cell(row=4, column=set_col, value="MÉTRICAS PN INVERTIBLE").font = FONT_BOLD
    inv_metrics = [
        ("PN Inicial", inv_rets["first_value"], FMT_NUMBER),
        ("PN Actual", inv_rets["last_value"], FMT_NUMBER),
        ("Retorno Absoluto", inv_rets["total_return_abs"], FMT_NUMBER),
        ("Retorno %", inv_rets["total_return_pct"], FMT_PCT),
        ("Max Drawdown %", inv_rets["max_drawdown_pct"], FMT_PCT),
    ]
    for i, (label, val, fmt) in enumerate(inv_metrics, start=5):
        ws.cell(row=i, column=set_col, value=label).font = FONT_NORMAL
        c = ws.cell(row=i, column=set_col + 1, value=val)
        c.number_format = fmt
        c.alignment = ALIGN_RIGHT

    # Tabla de la curva total
    row = 13
    ws.cell(row=row, column=1, value="EQUITY CURVE (TOTAL)").font = FONT_BOLD
    row += 1
    _write_headers(ws, row, ["Fecha", f"PN Total ({anchor_ccy})", f"PN Invertible ({anchor_ccy})"],
                   [12, 18, 18])
    inv_by_date = {p["fecha"]: p["mv_anchor"] for p in inv_curve}
    for p in total_curve:
        row += 1
        _write_row(ws, row, [
            p["fecha"], p["mv_anchor"], inv_by_date.get(p["fecha"]),
        ], formats=[None, FMT_NUMBER, FMT_NUMBER])

    # Por cuenta
    by_acc = get_equity_curves_by_account(conn, anchor_currency=anchor_ccy)
    if by_acc:
        row += 3
        ws.cell(row=row, column=1, value="EQUITY CURVE POR CUENTA").font = FONT_BOLD
        row += 1
        accounts = sorted(by_acc.keys())
        _write_headers(ws, row, ["Fecha"] + accounts,
                       [12] + [16] * len(accounts))
        # Pivot por fecha
        all_dates = sorted({p["fecha"] for curve in by_acc.values() for p in curve})
        # Pre-calcular dict (account, fecha) -> mv
        lookup = {(acc, p["fecha"]): p["mv_anchor"]
                  for acc, curve in by_acc.items() for p in curve}
        for d in all_dates:
            row += 1
            vals = [d] + [lookup.get((acc, d)) for acc in accounts]
            fmts = [None] + [FMT_NUMBER] * len(accounts)
            _write_row(ws, row, vals, formats=fmts)


def _sheet_pnl_no_realizado(wb, holdings, anchor_ccy):
    ws = wb.create_sheet("PnL No-Realizado")

    summary = calculate_unrealized_pnl_summary(holdings)

    ws.cell(row=1, column=1, value="UNREALIZED PnL TOTAL").font = FONT_BOLD
    ws.cell(row=2, column=1, value=summary["total_unrealized_anchor"]).number_format = FMT_NUMBER
    ws.cell(row=3, column=1, value=anchor_ccy).font = FONT_SUBTITLE

    ws.cell(row=1, column=4, value=f"{summary['n_winners']} ganadoras / {summary['n_losers']} perdedoras").font = FONT_SUBTITLE

    headers = [
        "Cuenta", "Activo", "Qty",
        "Avg Cost", "Mkt Price", "Native Ccy",
        "Cost Basis", "Market Value",
        "Unrealized PnL", "Unrealized %",
    ]
    widths = [22, 22, 14, 12, 12, 9, 16, 16, 16, 12]
    _write_headers(ws, 5, headers, widths)

    # Solo activos no-cash, ordenados por unrealized_pnl desc
    items = [h for h in holdings if not h["is_cash"] and h["unrealized_pnl_native"] is not None]
    items.sort(key=lambda h: -(h["unrealized_pnl_native"] or 0))

    for i, h in enumerate(items, start=6):
        _write_row(ws, i, [
            h["account"], h["asset"], h["qty"],
            h["avg_cost"], h["market_price"], h["native_currency"],
            h["cost_basis_total"], h["mv_native"],
            h["unrealized_pnl_native"], h["unrealized_pct"],
        ], formats=[
            None, None, FMT_NUMBER4,
            FMT_NUMBER4, FMT_NUMBER4, None,
            FMT_NUMBER, FMT_NUMBER,
            FMT_NUMBER, FMT_PCT,
        ])

    ws.freeze_panes = "A6"


# =============================================================================
# Excel export
# =============================================================================

def export_excel(conn, output_path, fecha=None, anchor_currency="USD",
                 record_snapshot=True, investible_only=False):
    """Genera Excel multi-sheet del portfolio. Devuelve Path del archivo.

    Si record_snapshot=True, appendea un snapshot del PN a `pn_snapshots`
    (necesario para construir la equity curve histórica).

    Si investible_only=True, los breakdowns (asset class / cuenta / moneda /
    holdings) excluyen las cuentas marcadas como NO-INVERTIBLE.
    """
    if fecha is None:
        fecha = date.today()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    holdings_full = calculate_holdings(conn, fecha=fecha, anchor_currency=anchor_currency)
    fills = calculate_realized_pnl(conn, fecha_hasta=fecha)

    if record_snapshot:
        try:
            record_snapshots(conn, holdings_full, fecha, anchor_currency)
        except Exception as e:
            print(f"[exporter] WARN no se pudo guardar snapshot: {e}")

    # Vista filtrada vs completa
    if investible_only:
        holdings = filter_investible(holdings_full)
    else:
        holdings = holdings_full

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_dashboard(wb, holdings, anchor_currency, fecha)
    _sheet_holdings(wb, holdings, anchor_currency)
    _sheet_pn_by_account(wb, holdings, anchor_currency)
    _sheet_pn_by_class(wb, holdings, anchor_currency)
    # PN Invertible siempre usa la vista completa (compara ambos lados)
    _sheet_invertible(wb, holdings_full, anchor_currency)
    _sheet_cash_position(wb, holdings, anchor_currency)
    _sheet_buying_power(wb, conn, holdings_full, anchor_currency)
    _sheet_equity_curve(wb, conn, anchor_currency)
    _sheet_tarjetas(wb, conn, fecha)
    _sheet_pnl_realizado(wb, fills, anchor_currency)
    _sheet_pnl_no_realizado(wb, holdings, anchor_currency)
    _sheet_trade_stats(wb, fills, anchor_currency)

    wb.save(str(output_path))
    return output_path


# =============================================================================
# HTML export (autocontenido con chart.js inline)
# =============================================================================

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Portfolio — {fecha}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    background: #f5f5f7;
    color: #1d1d1f;
    margin: 0;
    padding: 20px;
  }}
  .container {{ max-width: 1280px; margin: 0 auto; }}
  h1 {{ color: #1F3864; font-size: 28px; margin: 0 0 8px 0; }}
  .subtitle {{ color: #595959; font-size: 14px; margin-bottom: 24px; }}
  .kpi-grid {{
    display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px;
    margin-bottom: 24px;
  }}
  .kpi {{
    background: white; border-radius: 12px; padding: 20px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
  }}
  .kpi-label {{ color: #595959; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .kpi-value {{ font-size: 28px; font-weight: bold; color: #1F3864; margin-top: 4px; }}
  .kpi-currency {{ font-size: 13px; color: #595959; }}
  .kpi.muted .kpi-value {{ color: #595959; font-size: 22px; }}
  .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }}
  .grid-3 {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; }}
  .card {{
    background: white; border-radius: 12px; padding: 20px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.05);
  }}
  h2 {{ color: #1F3864; font-size: 18px; margin: 0 0 16px 0; }}
  h3 {{ color: #1F3864; font-size: 14px; margin: 16px 0 8px 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{
    background: #1F3864; color: white; padding: 10px;
    text-align: left; font-weight: 600;
  }}
  td {{ padding: 8px 10px; border-bottom: 1px solid #e0e0e0; }}
  tr:hover {{ background: #f9f9f9; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .positive {{ color: #00B050; }}
  .negative {{ color: #C00000; }}
  .chart-container {{ position: relative; height: 300px; }}
  .chart-container.tall {{ height: 380px; }}
  .warn {{
    background: #FFF3CD; border-left: 4px solid #FFC107;
    padding: 12px; border-radius: 4px; margin-bottom: 16px; font-size: 13px;
  }}
  .info {{
    background: #E8F4F8; border-left: 4px solid #4F81BD;
    padding: 10px 12px; border-radius: 4px; margin: 12px 0; font-size: 12px;
    color: #595959;
  }}
  .stat-mini {{ display: flex; justify-content: space-between; padding: 4px 0;
                border-bottom: 1px dashed #e0e0e0; font-size: 13px; }}
  .stat-mini-label {{ color: #595959; }}
  .stat-mini-value {{ font-weight: 600; font-variant-numeric: tabular-nums; }}
  details summary {{ cursor: pointer; font-weight: 600; color: #1F3864; }}
  .view-toggle {{
    display: inline-flex; background: white; border-radius: 10px;
    padding: 4px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); margin: 0 0 20px 0;
    gap: 4px;
  }}
  .view-toggle button {{
    border: none; background: transparent; padding: 8px 18px;
    border-radius: 8px; cursor: pointer; font-size: 13px; font-weight: 600;
    color: #595959; transition: background .15s, color .15s;
  }}
  .view-toggle button.active {{
    background: #1F3864; color: white;
  }}
  .view-toggle button:hover:not(.active) {{
    background: #f0f0f0;
  }}
  .view-tag {{
    display: inline-block; background: #E8F4F8; color: #1F3864;
    padding: 2px 8px; border-radius: 6px; font-size: 11px; font-weight: 600;
    margin-left: 8px; vertical-align: middle;
  }}
  .empty-row {{ color: #999; font-style: italic; text-align: center; }}
</style>
</head>
<body>
<div class="container">
  <h1>📊 Portfolio</h1>
  <div class="subtitle">Fecha: {fecha}  ·  Moneda ancla: {anchor_ccy}</div>

  <div class="view-toggle" role="tablist" aria-label="Vista">
    <button data-view="all" id="btnViewAll">📦 Todo</button>
    <button data-view="investible" id="btnViewInv">💎 Solo invertible</button>
  </div>

  <div class="kpi-grid">
    <div class="kpi">
      <div class="kpi-label">Patrimonio Neto Total</div>
      <div class="kpi-value">{pn_formatted}</div>
      <div class="kpi-currency">{anchor_ccy}</div>
    </div>
    <div class="kpi">
      <div class="kpi-label">PN Invertible</div>
      <div class="kpi-value">{pn_invertible_formatted}</div>
      <div class="kpi-currency">{anchor_ccy} (excluye reserva)</div>
    </div>
    <div class="kpi muted">
      <div class="kpi-label">PN No-Invertible</div>
      <div class="kpi-value">{pn_non_invertible_formatted}</div>
      <div class="kpi-currency">{anchor_ccy}</div>
    </div>
  </div>

  {warn_html}

  <div class="grid">
    <div class="card">
      <h2>Por Asset Class <span class="view-tag" id="tagClass"></span></h2>
      <div class="chart-container"><canvas id="chartClass"></canvas></div>
    </div>
    <div class="card">
      <h2>Por Moneda Nativa <span class="view-tag" id="tagCcy"></span></h2>
      <div class="chart-container"><canvas id="chartCcy"></canvas></div>
    </div>
  </div>

  <div class="card" style="margin-top: 24px;">
    <h2>📈 Equity Curve</h2>
    {equity_curve_block}
  </div>

  <div class="card" style="margin-top: 24px;">
    <h2>🎯 Métricas de Trading (PnL Realizado FIFO)</h2>
    {trade_stats_block}
  </div>

  <div class="card" style="margin-top: 24px;">
    <h2>💪 Poder de Compra por Cuenta</h2>
    <div class="info">
      Cocos/Eco: aforos BYMA (caución). IBKR: RegT margin (verificá tus parámetros reales).
      Valores en {anchor_ccy}.
    </div>
    {buying_power_block}
  </div>

  <div class="card" style="margin-top: 24px;">
    <h2>Top 15 Posiciones <span class="view-tag" id="tagTop"></span></h2>
    <table>
      <thead>
        <tr>
          <th>Cuenta</th>
          <th>Activo</th>
          <th>Asset Class</th>
          <th class="num">Qty</th>
          <th class="num">Mkt Price</th>
          <th class="num">MV ({anchor_ccy})</th>
          <th class="num">Unrealized %</th>
        </tr>
      </thead>
      <tbody id="topHoldingsBody"></tbody>
    </table>
  </div>

  <div class="grid" style="margin-top: 24px;">
    <div class="card">
      <h2>PN por Cuenta <span class="view-tag" id="tagAcc"></span></h2>
      <table>
        <thead><tr><th>Cuenta</th><th class="num">MV ({anchor_ccy})</th><th class="num">%</th></tr></thead>
        <tbody id="byAccountBody"></tbody>
      </table>
    </div>
    <div class="card">
      <h2>PnL Realizado por Año + Moneda</h2>
      <table>
        <thead><tr><th>Año</th><th>Moneda</th><th class="num"># Trades</th><th class="num">PnL Total</th></tr></thead>
        <tbody>{pnl_year_rows}</tbody>
      </table>
    </div>
  </div>

  <details class="card" style="margin-top: 24px;">
    <summary>💵 Cash por Propósito (filtros)</summary>
    <table style="margin-top: 12px;">
      <thead><tr><th>Propósito</th><th class="num">MV ({anchor_ccy})</th></tr></thead>
      <tbody>{cash_purpose_rows}</tbody>
    </table>
  </details>

  <div class="subtitle" style="margin-top: 24px; text-align: center;">
    Generado por wm_engine
  </div>
</div>

<script>
const VIEWS = {views_json};
const DEFAULT_VIEW = "{default_view}";
const equityData = {equity_curve_json};
const ANCHOR = "{anchor_ccy}";

// Helpers
function fmtNum(v, dec) {{
  if (v === null || v === undefined) return "-";
  return Number(v).toLocaleString('es-AR', {{
    minimumFractionDigits: dec || 0,
    maximumFractionDigits: dec || 2,
  }});
}}
function fmtPct(v) {{
  if (v === null || v === undefined) return "-";
  const cls = v > 0 ? "positive" : v < 0 ? "negative" : "";
  const sign = v > 0 ? "+" : "";
  return `<span class="${{cls}}">${{sign}}${{(v * 100).toFixed(2)}}%</span>`;
}}

// Chart instances (re-renderables)
let chartClass = null, chartCcy = null;

function buildDoughnut(canvasId, data, palette) {{
  return new Chart(document.getElementById(canvasId), {{
    type: 'doughnut',
    data: {{
      labels: data.labels,
      datasets: [{{ data: data.values, backgroundColor: palette }}]
    }},
    options: {{ responsive: true, maintainAspectRatio: false }}
  }});
}}

function renderView(viewKey) {{
  const v = VIEWS[viewKey];
  if (!v) return;

  // 1. Asset class chart
  if (chartClass) chartClass.destroy();
  chartClass = buildDoughnut('chartClass', v.cls,
    ['#1F3864','#2E5B9C','#4F81BD','#8DB4E2','#B8CCE4','#DCE6F1','#F2F2F2','#7E7E7E']);

  // 2. Currency chart
  if (chartCcy) chartCcy.destroy();
  chartCcy = buildDoughnut('chartCcy', v.ccy,
    ['#1F3864','#00B050','#C00000','#FFC107','#9C27B0','#FF5722']);

  // 3. Top holdings table
  const topBody = document.getElementById('topHoldingsBody');
  if (!v.top_holdings || v.top_holdings.length === 0) {{
    topBody.innerHTML = '<tr><td colspan="7" class="empty-row">Sin posiciones en esta vista</td></tr>';
  }} else {{
    topBody.innerHTML = v.top_holdings.map(h => `
      <tr>
        <td>${{h.account}}</td>
        <td>${{h.asset}}</td>
        <td>${{h.asset_class || ''}}</td>
        <td class="num">${{fmtNum(h.qty, 4)}}</td>
        <td class="num">${{fmtNum(h.market_price, 4)}}</td>
        <td class="num">${{fmtNum(h.mv_anchor, 2)}}</td>
        <td class="num">${{fmtPct(h.unrealized_pct)}}</td>
      </tr>
    `).join('');
  }}

  // 4. PN por cuenta table
  const accBody = document.getElementById('byAccountBody');
  if (!v.accs || v.accs.length === 0) {{
    accBody.innerHTML = '<tr><td colspan="3" class="empty-row">Sin cuentas en esta vista</td></tr>';
  }} else {{
    accBody.innerHTML = v.accs.map(a => `
      <tr>
        <td>${{a.account}}</td>
        <td class="num">${{fmtNum(a.mv, 2)}}</td>
        <td class="num">${{(a.pct * 100).toFixed(1)}}%</td>
      </tr>
    `).join('');
  }}

  // 5. Tags
  const tagText = viewKey === 'investible' ? 'solo invertible' : 'todo';
  ['tagClass','tagCcy','tagTop','tagAcc'].forEach(id => {{
    const el = document.getElementById(id);
    if (el) el.textContent = tagText;
  }});

  // 6. Toggle buttons active state
  document.querySelectorAll('.view-toggle button').forEach(b => {{
    b.classList.toggle('active', b.dataset.view === viewKey);
  }});
}}

// Wire up toggle
document.querySelectorAll('.view-toggle button').forEach(btn => {{
  btn.addEventListener('click', () => renderView(btn.dataset.view));
}});

// Initial render
renderView(DEFAULT_VIEW);

// Equity curve (independiente del toggle — siempre muestra Total + Invertible)
if (equityData && equityData.labels && equityData.labels.length > 0) {{
  const palette = ['#1F3864','#00B050','#C00000','#FFC107','#9C27B0','#FF5722',
                   '#4F81BD','#8DB4E2','#2E5B9C','#595959'];
  const datasets = equityData.series.map((s, idx) => ({{
    label: s.label,
    data: s.values,
    borderColor: palette[idx % palette.length],
    backgroundColor: palette[idx % palette.length] + '20',
    borderWidth: s.bold ? 3 : 1.5,
    tension: 0.2,
    fill: false,
    spanGaps: true,
  }}));
  new Chart(document.getElementById('chartEquity'), {{
    type: 'line',
    data: {{ labels: equityData.labels, datasets: datasets }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{ legend: {{ position: 'bottom' }} }},
      scales: {{
        y: {{
          ticks: {{
            callback: v => v.toLocaleString('es-AR', {{ maximumFractionDigits: 0 }})
          }}
        }}
      }}
    }}
  }});
}}
</script>
</body>
</html>
"""


def _fmt_number(v, decimals=2):
    if v is None:
        return "-"
    return f"{v:,.{decimals}f}"


def _fmt_pct(v):
    if v is None:
        return "-"
    cls = "positive" if v > 0 else "negative" if v < 0 else ""
    return f'<span class="{cls}">{v*100:+.2f}%</span>'


def _html_trade_stats_block(fills, anchor_ccy):
    """Construye bloque HTML con métricas de trading."""
    if not fills:
        return '<div class="info">Sin trades cerrados aún. Vendé al menos una posición para ver métricas.</div>'

    stats = calculate_trade_stats(fills)
    if not stats:
        return '<div class="info">Sin métricas calculables.</div>'

    blocks = []
    for ccy in sorted(stats.keys()):
        s = stats[ccy]
        pf_str = "∞" if s.profit_factor == float("inf") else f"{s.profit_factor:.2f}"
        block = f"""
<div style="margin-bottom: 18px;">
  <h3>Moneda: {ccy} — {s.n_trades} trades cerrados</h3>
  <div class="grid-3">
    <div>
      <div class="stat-mini"><span class="stat-mini-label">✅ Ganadores</span><span class="stat-mini-value positive">{s.n_winners}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">❌ Perdedores</span><span class="stat-mini-value negative">{s.n_losers}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">⚪ Scratch</span><span class="stat-mini-value">{s.n_scratch}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">Winrate</span><span class="stat-mini-value">{s.winrate*100:.1f}%</span></div>
    </div>
    <div>
      <div class="stat-mini"><span class="stat-mini-label">Gross Profit</span><span class="stat-mini-value positive">{_fmt_number(s.gross_profit)}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">Gross Loss</span><span class="stat-mini-value negative">{_fmt_number(s.gross_loss)}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">Net PnL</span><span class="stat-mini-value {'positive' if s.net_pnl > 0 else 'negative' if s.net_pnl < 0 else ''}">{_fmt_number(s.net_pnl)}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">Profit Factor</span><span class="stat-mini-value">{pf_str}</span></div>
    </div>
    <div>
      <div class="stat-mini"><span class="stat-mini-label">Avg Winner</span><span class="stat-mini-value positive">{_fmt_number(s.avg_winner)}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">Avg Loser</span><span class="stat-mini-value negative">{_fmt_number(s.avg_loser)}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">Expectancy</span><span class="stat-mini-value">{_fmt_number(s.expectancy)}</span></div>
      <div class="stat-mini"><span class="stat-mini-label">Avg Holding (días)</span><span class="stat-mini-value">{s.avg_holding_days:.1f}</span></div>
    </div>
  </div>
  <div class="stat-mini" style="margin-top: 8px;">
    <span class="stat-mini-label">Best / Worst trade</span>
    <span class="stat-mini-value"><span class="positive">{_fmt_number(s.best_trade)}</span> / <span class="negative">{_fmt_number(s.worst_trade)}</span></span>
  </div>
  <div class="stat-mini">
    <span class="stat-mini-label">Racha máx wins / losses</span>
    <span class="stat-mini-value">{s.largest_streak_wins} / {s.largest_streak_losses}</span>
  </div>
</div>
"""
        blocks.append(block)
    return "".join(blocks)


def _html_buying_power_block(conn, holdings, anchor_ccy):
    """Construye bloque HTML del poder de compra."""
    summary = buying_power_summary(conn, holdings, anchor_ccy)
    if not summary:
        return '<div class="info">No hay cuentas broker con holdings/cash.</div>'

    rows = []
    for item in summary:
        if item["type"] == "BYMA":
            bp = item["result"]
            equity = bp.cash_total + bp.holdings_mv
            lev = f"{bp.leverage_ratio:.2f}x"
            rows.append(
                f'<tr>'
                f'<td><b>{bp.account}</b></td>'
                f'<td>BYMA (caución)</td>'
                f'<td class="num">{_fmt_number(equity)}</td>'
                f'<td class="num">{_fmt_number(bp.cash_total)}</td>'
                f'<td class="num">{_fmt_number(bp.garantia_holdings)}</td>'
                f'<td class="num"><b>{_fmt_number(bp.poder_de_compra)}</b></td>'
                f'<td class="num">{lev}</td>'
                f'</tr>'
            )
        else:
            bp_o = item["overnight"]
            bp_i = item["intraday"]
            funding_pct = bp_o.funding_rate_annual * 100
            rows.append(
                f'<tr>'
                f'<td><b>{bp_o.account}</b></td>'
                f'<td>MARGIN overnight</td>'
                f'<td class="num">{_fmt_number(bp_o.equity)}</td>'
                f'<td class="num">—</td>'
                f'<td class="num">{_fmt_number(bp_o.margin_disponible)}</td>'
                f'<td class="num"><b>{_fmt_number(bp_o.poder_de_compra)}</b></td>'
                f'<td class="num">x{bp_o.multiplier:.1f}</td>'
                f'</tr>'
            )
            rows.append(
                f'<tr>'
                f'<td></td>'
                f'<td>MARGIN intraday</td>'
                f'<td class="num">{_fmt_number(bp_i.equity)}</td>'
                f'<td class="num">—</td>'
                f'<td class="num">{_fmt_number(bp_i.margin_disponible)}</td>'
                f'<td class="num"><b>{_fmt_number(bp_i.poder_de_compra)}</b></td>'
                f'<td class="num">x{bp_i.multiplier:.1f}</td>'
                f'</tr>'
            )
            rows.append(
                f'<tr><td colspan="7" style="font-size: 11px; color: #595959; padding-left: 16px;">'
                f'Funding rate: {funding_pct:.2f}%/año ({bp_o.funding_currency or "?"}) · '
                f'Costo si usás todo el margen 1 día: {_fmt_number(bp_o.funding_cost_per_day)} {bp_o.funding_currency or anchor_ccy}</td></tr>'
            )

    return f"""
<table>
  <thead>
    <tr>
      <th>Cuenta</th><th>Tipo</th>
      <th class="num">Equity</th>
      <th class="num">Cash</th>
      <th class="num">Garantía / Margin Disp.</th>
      <th class="num">Poder de Compra</th>
      <th class="num">Lev</th>
    </tr>
  </thead>
  <tbody>{''.join(rows)}</tbody>
</table>
"""


def _html_equity_curve_block(conn, anchor_ccy):
    """Construye bloque HTML con gráfico equity curve."""
    total_curve = get_equity_curve(conn, anchor_currency=anchor_ccy)
    inv_curve = get_equity_curve(conn, anchor_currency=anchor_ccy,
                                  investible_only=True)
    if not total_curve:
        return ('<div class="info">Sin snapshots aún. Cada vez que generás el reporte se '
                'guarda un snapshot — corré el reporte algunos días para ver la evolución.</div>')

    rets = calculate_returns(total_curve)
    inv_rets = calculate_returns(inv_curve)

    metrics_html = f"""
<div class="grid-3" style="margin-bottom: 16px;">
  <div>
    <div class="stat-mini"><span class="stat-mini-label">PN inicial</span><span class="stat-mini-value">{_fmt_number(rets['first_value'])}</span></div>
    <div class="stat-mini"><span class="stat-mini-label">PN actual</span><span class="stat-mini-value">{_fmt_number(rets['last_value'])}</span></div>
  </div>
  <div>
    <div class="stat-mini"><span class="stat-mini-label">Retorno absoluto</span><span class="stat-mini-value {'positive' if rets['total_return_abs']>0 else 'negative' if rets['total_return_abs']<0 else ''}">{_fmt_number(rets['total_return_abs'])}</span></div>
    <div class="stat-mini"><span class="stat-mini-label">Retorno %</span><span class="stat-mini-value {'positive' if rets['total_return_pct']>0 else 'negative' if rets['total_return_pct']<0 else ''}">{rets['total_return_pct']*100:+.2f}%</span></div>
  </div>
  <div>
    <div class="stat-mini"><span class="stat-mini-label">Max Drawdown</span><span class="stat-mini-value negative">{rets['max_drawdown_pct']*100:.2f}%</span></div>
    <div class="stat-mini"><span class="stat-mini-label">Snapshots</span><span class="stat-mini-value">{rets['n_periods']}</span></div>
  </div>
</div>
<div class="chart-container tall"><canvas id="chartEquity"></canvas></div>
"""
    return metrics_html


def _equity_curve_chart_data(conn, anchor_ccy):
    """Prepara JSON para el chart Chart.js de equity curve."""
    total_curve = get_equity_curve(conn, anchor_currency=anchor_ccy)
    inv_curve = get_equity_curve(conn, anchor_currency=anchor_ccy,
                                  investible_only=True)
    by_acc = get_equity_curves_by_account(conn, anchor_currency=anchor_ccy)

    if not total_curve:
        return {"labels": [], "series": []}

    # Unir todas las fechas únicas
    all_dates = sorted({p["fecha"] for p in total_curve}
                       | {p["fecha"] for p in inv_curve}
                       | {p["fecha"] for c in by_acc.values() for p in c})

    def values_for(curve):
        m = {p["fecha"]: p["mv_anchor"] for p in curve}
        return [m.get(d) for d in all_dates]

    series = [
        {"label": "Total", "values": values_for(total_curve), "bold": True},
    ]
    if inv_curve:
        series.append({
            "label": "Invertible", "values": values_for(inv_curve), "bold": True,
        })
    # Por cuenta — solo top-N por último valor para no llenar el chart
    by_acc_sorted = sorted(
        by_acc.items(),
        key=lambda kv: -(kv[1][-1]["mv_anchor"] if kv[1] else 0),
    )[:8]
    for acc, curve in by_acc_sorted:
        series.append({
            "label": acc, "values": values_for(curve), "bold": False,
        })

    return {"labels": all_dates, "series": series}


def _compute_view_data(holdings, anchor_currency):
    """Calcula los datasets de un "view" para el HTML toggle.

    Devuelve dict con cls/ccy/accs/top_holdings listos para JSON o tablas.
    """
    cls_data = by_asset_class(holdings)
    ccy_data = by_currency(holdings)
    accs = by_account(holdings)
    accs_total = sum(accs.values())

    top_15 = [h for h in holdings[:15] if h.get("mv_anchor") is not None]

    return {
        "cls": {"labels": list(cls_data.keys()), "values": list(cls_data.values())},
        "ccy": {"labels": list(ccy_data.keys()), "values": list(ccy_data.values())},
        "accs": [
            {"account": acc, "mv": val,
             "pct": (val / accs_total) if accs_total else 0}
            for acc, val in list(accs.items())[:15]
        ],
        "top_holdings": [
            {
                "account": h["account"],
                "asset": h["asset"],
                "asset_class": h.get("asset_class", "") or "",
                "qty": h["qty"],
                "market_price": h["market_price"],
                "mv_anchor": h["mv_anchor"],
                "unrealized_pct": h.get("unrealized_pct"),
            }
            for h in top_15
        ],
    }


def export_html(conn, output_path, fecha=None, anchor_currency="USD",
                record_snapshot=True, default_view="all"):
    """Genera HTML autocontenido del portfolio. Devuelve Path.

    El HTML incluye un TOGGLE entre vista "Todo" (incluye no-invertibles) y
    "Solo invertible" (excluye reserva no declarada, etc.). Ambos datasets
    se incrustan en el HTML; el toggle JS hace el switch sin re-cargar.

    `default_view`: 'all' o 'investible' — qué vista se muestra al abrir.
    """
    if fecha is None:
        fecha = date.today()
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    holdings_full = calculate_holdings(conn, fecha=fecha, anchor_currency=anchor_currency)
    fills = calculate_realized_pnl(conn, fecha_hasta=fecha)

    if record_snapshot:
        try:
            record_snapshots(conn, holdings_full, fecha, anchor_currency)
        except Exception as e:
            print(f"[exporter] WARN no se pudo guardar snapshot: {e}")

    tp = total_pn(holdings_full, anchor_currency)
    holdings_inv = filter_investible(holdings_full)

    # Datasets para ambas vistas
    view_all = _compute_view_data(holdings_full, anchor_currency)
    view_inv = _compute_view_data(holdings_inv, anchor_currency)

    # PnL por año + moneda (mismo en ambas vistas — los trades cerrados
    # son cerrados, no dependen de qué cuentas son invertibles)
    pnl_yc = aggregate_pnl_by_year_currency(fills)
    pnl_year_rows = []
    for item in pnl_yc:
        cls = "positive" if item["pnl_total"] > 0 else "negative" if item["pnl_total"] < 0 else ""
        pnl_year_rows.append(
            f'<tr><td>{item["year"]}</td>'
            f'<td>{item["currency"]}</td>'
            f'<td class="num">{item["n_trades"]}</td>'
            f'<td class="num"><span class="{cls}">{_fmt_number(item["pnl_total"])}</span></td></tr>'
        )
    if not pnl_year_rows:
        pnl_year_rows.append('<tr><td colspan="4" style="text-align: center; color: #595959;">Sin trades cerrados aún</td></tr>')

    # Warn de FX faltante
    warn = ""
    if tp["total_unconverted_count"] > 0:
        warn = f'<div class="warn">⚠ {tp["total_unconverted_count"]} posiciones sin FX hacia {anchor_currency}, no incluidas en el total</div>'

    # Cash purpose (siempre vista completa)
    purposes = by_cash_purpose(holdings_full)
    cash_purpose_rows = []
    for p, val in purposes.items():
        cash_purpose_rows.append(
            f'<tr><td>{p}</td><td class="num">{_fmt_number(val)}</td></tr>'
        )
    if not cash_purpose_rows:
        cash_purpose_rows.append('<tr><td colspan="2" style="color: #595959; text-align: center;">Sin cash registrado</td></tr>')

    # Render datasets como JSON para el toggle JS
    views_json = json.dumps({
        "all": view_all,
        "investible": view_inv,
        "default": default_view if default_view in ("all", "investible") else "all",
    })

    html = HTML_TEMPLATE.format(
        fecha=fecha.isoformat(),
        anchor_ccy=anchor_currency,
        pn_formatted=_fmt_number(tp["total_anchor"]),
        pn_invertible_formatted=_fmt_number(tp["total_investible"]),
        pn_non_invertible_formatted=_fmt_number(tp["total_non_investible"]),
        warn_html=warn,
        pnl_year_rows="\n".join(pnl_year_rows),
        cash_purpose_rows="\n".join(cash_purpose_rows),
        trade_stats_block=_html_trade_stats_block(fills, anchor_currency),
        buying_power_block=_html_buying_power_block(conn, holdings_full, anchor_currency),
        equity_curve_block=_html_equity_curve_block(conn, anchor_currency),
        equity_curve_json=json.dumps(_equity_curve_chart_data(conn, anchor_currency)),
        views_json=views_json,
        default_view=default_view if default_view in ("all", "investible") else "all",
    )

    output_path.write_text(html, encoding="utf-8")
    return output_path
