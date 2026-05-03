# -*- coding: utf-8 -*-
"""
cli/cargar_iniciales.py

Lee la hoja `_carga_inicial` del Excel master y AGREGA filas a la hoja
`asientos_contables` con el formato de doble entrada que el motor espera.

Idempotente: identifica filas ya generadas (por Event ID `OPEN-AUTO-*`)
y las regenera limpias en cada corrida.

USO:
    python -m cli.cargar_iniciales [--xlsx PATH] [--fecha YYYY-MM-DD] [--dry-run]

EJEMPLO:
    # Carga las 80 filas de _carga_inicial → 160 filas en asientos_contables
    python -m cli.cargar_iniciales --fecha 2026-04-30
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from engine.fx import import_fx_csv, convert as fx_convert, FxError


# Prefijo de Event ID que identifica filas auto-generadas
AUTO_PREFIX = "OPEN-AUTO-"


def _build_fx_db(fx_csv_path):
    """Construye una DB sqlite temporal en memoria con el FX cargado.
    Devuelve (conn, n_rows).
    Si fx_csv_path no existe o no se puede cargar, devuelve (conn vacía, 0).
    """
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE fx_rates (
            fecha TEXT NOT NULL,
            moneda TEXT NOT NULL,
            rate REAL NOT NULL,
            base TEXT NOT NULL DEFAULT 'ARS',
            source TEXT,
            PRIMARY KEY (fecha, moneda, base)
        )
    """)
    conn.execute("CREATE INDEX idx_fx_moneda_fecha ON fx_rates(moneda, fecha)")
    conn.commit()

    n = 0
    if fx_csv_path and Path(fx_csv_path).is_file():
        try:
            n = import_fx_csv(conn, fx_csv_path)
        except Exception as e:
            print(f"[fx] WARN cargando FX: {e}")
    return conn, n


def _load_asset_currency_map(wb):
    """Lee la hoja 'especies' y devuelve dict {ticker: currency_native}.
    También considera las monedas mismas (ARS, USB, etc) como assets con su propia currency."""
    mapping = {}
    if "especies" in wb.sheetnames:
        ws = wb["especies"]
        # Headers en fila 4
        headers = []
        for c in range(1, ws.max_column + 1):
            h = ws.cell(row=4, column=c).value
            if isinstance(h, str):
                h = h.strip()
            headers.append(h)
        try:
            ticker_idx = headers.index("Ticker")
            currency_idx = headers.index("Currency")
        except ValueError:
            return mapping
        for r in range(5, ws.max_row + 1):
            ticker = ws.cell(row=r, column=ticker_idx + 1).value
            currency = ws.cell(row=r, column=currency_idx + 1).value
            if isinstance(ticker, str): ticker = ticker.strip()
            if isinstance(currency, str): currency = currency.strip()
            if ticker and currency:
                mapping[ticker] = currency

    # Las monedas mismas (cuando se usan como activo de cash) son su propia currency
    if "monedas" in wb.sheetnames:
        ws = wb["monedas"]
        for r in range(5, ws.max_row + 1):
            code = ws.cell(row=r, column=1).value
            if isinstance(code, str):
                code = code.strip()
                if code and code not in mapping:
                    mapping[code] = code  # ARS → ARS, USB → USB, etc

    return mapping


def _to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_date(v):
    """Convierte a date (no string) — necesario para que Excel lo formatee bien."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        try:
            return date.fromisoformat(v[:10])
        except ValueError:
            return None
    return None


def _read_carga_inicial(ws):
    """Lee la hoja _carga_inicial y devuelve lista de dicts."""
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=4, column=c).value
        if isinstance(h, str):
            h = h.strip()
        headers.append(h)

    rows = []
    for r in range(5, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c + 1).value for c in range(len(headers))]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
            continue
        d = dict(zip(headers, row_vals))

        # Strip strings
        for k, v in list(d.items()):
            if isinstance(v, str):
                d[k] = v.strip() or None

        rows.append((r, d))

    return rows


def _delete_auto_rows(ws):
    """Borra filas en asientos_contables cuyo Event ID empieza con OPEN-AUTO-."""
    headers = [c.value for c in ws[4]]
    try:
        eid_col = headers.index("Event ID") + 1
    except ValueError:
        return 0

    # Recolectar filas a borrar (de mayor a menor para no descalibrar)
    rows_to_delete = []
    for r in range(5, ws.max_row + 1):
        eid = ws.cell(row=r, column=eid_col).value
        if eid and isinstance(eid, str) and eid.strip().startswith(AUTO_PREFIX):
            rows_to_delete.append(r)

    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    return len(rows_to_delete)


def _append_asiento(ws, event_id, fecha, description,
                    cuenta, activo, qty, unit_price, price_currency, notes):
    """Agrega una fila al final de asientos_contables."""
    next_row = ws.max_row + 1
    # Si la fila max está vacía, usar esa
    while next_row > 5 and all(
        ws.cell(row=next_row - 1, column=c).value is None
        for c in range(1, 10)
    ):
        next_row -= 1

    cells = [
        event_id, fecha, description, cuenta, activo, qty,
        unit_price, price_currency, notes
    ]
    for c, val in enumerate(cells, start=1):
        ws.cell(row=next_row, column=c, value=val)


def generate_asientos(xlsx_path, fecha_default=None, dry_run=False,
                      fx_csv_path=None):
    """Lee _carga_inicial, regenera asientos auto-generados, devuelve stats.

    Si una fila tiene Price Currency != asset.currency (definida en `especies`),
    convierte unit_price a la moneda nativa del activo usando FX desde
    `data/fx_historico.csv`.
    """
    wb = load_workbook(filename=str(xlsx_path))

    if "_carga_inicial" not in wb.sheetnames:
        raise RuntimeError(
            f"Hoja '_carga_inicial' no existe en {xlsx_path}. "
            "Corré primero `python add_carga_inicial_sheet.py`."
        )
    if "asientos_contables" not in wb.sheetnames:
        raise RuntimeError("Hoja 'asientos_contables' no existe en el master")

    ws_in = wb["_carga_inicial"]
    ws_out = wb["asientos_contables"]

    # Construir mapping ticker → moneda nativa
    asset_ccy_map = _load_asset_currency_map(wb)

    # Construir DB temporal de FX
    if fx_csv_path is None:
        fx_csv_path = Path("data") / "fx_historico.csv"
    fx_conn, fx_n = _build_fx_db(fx_csv_path)
    print(f"[fx] cargado {fx_n} cotizaciones desde {fx_csv_path}")

    rows = _read_carga_inicial(ws_in)

    # Validaciones + conversión FX
    valid_rows = []
    errors = []
    fx_conversions = 0
    for r, d in rows:
        cuenta = _to_str(d.get("Cuenta"))
        activo = _to_str(d.get("Activo"))
        qty = _to_float(d.get("Qty"))

        if not cuenta:
            errors.append(f"  fila {r}: falta Cuenta")
            continue
        if not activo:
            errors.append(f"  fila {r}: falta Activo")
            continue
        if qty is None:
            errors.append(f"  fila {r}: falta Qty")
            continue
        if qty <= 0:
            errors.append(f"  fila {r}: Qty debe ser positivo (vos cargás siempre +, el script invierte el signo en opening_balance)")
            continue

        unit_price = _to_float(d.get("Unit Price"))
        price_currency = _to_str(d.get("Price Currency"))
        fecha = _to_date(d.get("Fecha")) or fecha_default

        # Conversión FX si Price Currency != moneda nativa del activo
        fx_note = None
        if unit_price is not None and price_currency:
            native_ccy = asset_ccy_map.get(activo)
            if native_ccy and native_ccy != price_currency:
                # Necesitamos convertir
                if fecha is None:
                    errors.append(
                        f"  fila {r}: necesita FX pero no hay Fecha (activo={activo}, "
                        f"price={price_currency}, native={native_ccy})"
                    )
                    continue
                if fx_n == 0:
                    errors.append(
                        f"  fila {r}: necesita FX pero no se cargó fx_historico.csv "
                        f"(activo={activo}, {price_currency}→{native_ccy})"
                    )
                    continue
                try:
                    converted = fx_convert(
                        fx_conn, unit_price, price_currency, native_ccy, fecha,
                        fallback_days=14,
                    )
                    fx_note = (f"FX: {price_currency}→{native_ccy} en {fecha}: "
                               f"{unit_price:.4f} {price_currency} = "
                               f"{converted:.6f} {native_ccy}")
                    unit_price = converted
                    price_currency = native_ccy
                    fx_conversions += 1
                except FxError as e:
                    errors.append(f"  fila {r}: {e} (activo={activo})")
                    continue

        valid_rows.append({
            "src_row": r,
            "cuenta": cuenta,
            "activo": activo,
            "qty": qty,
            "unit_price": unit_price,
            "price_currency": price_currency,
            "fecha": fecha,
            "description": _to_str(d.get("Description")) or f"Apertura: {qty} {activo} en {cuenta}",
            "notes": _to_str(d.get("Notes")),
            "strategy": _to_str(d.get("Strategy")),
            "fx_note": fx_note,
        })

    if errors:
        print("ERRORES en _carga_inicial:")
        for e in errors:
            print(e)
        if not valid_rows:
            return {"errores": len(errors), "generados": 0, "fx_conversions": fx_conversions}

    if dry_run:
        print(f"\n[dry-run] Se generarían {len(valid_rows) * 2} filas en asientos_contables")
        print(f"[dry-run] Conversiones FX aplicadas: {fx_conversions}")
        for i, r in enumerate(valid_rows[:5], start=1):
            extra = f" [{r['fx_note']}]" if r['fx_note'] else ""
            print(f"  {i}. {AUTO_PREFIX}{i:03d} | {r['fecha']} | {r['cuenta']} | "
                  f"{r['activo']} | qty=+{r['qty']} (y {-r['qty']} en opening_balance){extra}")
        if len(valid_rows) > 5:
            print(f"  ... ({len(valid_rows) - 5} más)")
        return {"errores": len(errors), "generados": 0,
                "dry_run": len(valid_rows), "fx_conversions": fx_conversions}

    # Borrar filas auto-generadas previas (idempotencia)
    deleted = _delete_auto_rows(ws_out)
    print(f"[carga] Borradas {deleted} filas auto-generadas previas")

    # Generar nuevas
    for i, r in enumerate(valid_rows, start=1):
        eid = f"{AUTO_PREFIX}{i:03d}"
        descr = r["description"]
        notes_active = r["notes"] or ""
        if r["strategy"]:
            notes_active = f"{r['strategy']} | {notes_active}".strip(" |")
        if r["fx_note"]:
            notes_active = f"{notes_active} | {r['fx_note']}".strip(" |") if notes_active else r["fx_note"]

        # Fila A: la cuenta gana el activo
        _append_asiento(
            ws_out,
            event_id=eid,
            fecha=r["fecha"],
            description=descr,
            cuenta=r["cuenta"],
            activo=r["activo"],
            qty=r["qty"],
            unit_price=r["unit_price"],
            price_currency=r["price_currency"],
            notes=notes_active,
        )
        # Fila B: opening_balance pierde el activo
        _append_asiento(
            ws_out,
            event_id=eid,
            fecha=r["fecha"],
            description=descr,
            cuenta="opening_balance",
            activo=r["activo"],
            qty=-r["qty"],
            unit_price=r["unit_price"],
            price_currency=r["price_currency"],
            notes="Contracuenta",
        )

    wb.save(str(xlsx_path))
    fx_conn.close()
    return {
        "errores": len(errors),
        "filas_carga_inicial": len(valid_rows),
        "filas_generadas_asientos": len(valid_rows) * 2,
        "deleted_previas": deleted,
        "fx_conversions": fx_conversions,
    }


def main():
    p = argparse.ArgumentParser(description="Carga inicial de saldos vía hoja temporal")
    p.add_argument("--xlsx", type=Path,
                   default=Path("inputs/wealth_management_rodricor.xlsx"))
    p.add_argument("--fecha", type=str, default=None,
                   help="Fecha por defecto si no se especifica en cada fila (YYYY-MM-DD)")
    p.add_argument("--dry-run", action="store_true",
                   help="Solo muestra lo que generaría, sin escribir")
    p.add_argument("--fx-csv", type=Path, default=None,
                   help="Path al fx_historico.csv (default: data/fx_historico.csv)")
    args = p.parse_args()

    fecha_default = date.fromisoformat(args.fecha) if args.fecha else None

    if not args.xlsx.is_file():
        print(f"[error] no se encontró {args.xlsx}", file=sys.stderr)
        return 1

    print(f"[carga] xlsx = {args.xlsx}")
    print(f"[carga] fecha default = {fecha_default}")
    print(f"[carga] dry_run = {args.dry_run}")
    print()

    stats = generate_asientos(args.xlsx, fecha_default, args.dry_run, fx_csv_path=args.fx_csv)
    print()
    print("=== Estadísticas ===")
    for k, v in stats.items():
        print(f"  {k}: {v}")

    if not args.dry_run and stats.get("filas_generadas_asientos"):
        print()
        print(f"OK: {stats['filas_generadas_asientos']} filas escritas a asientos_contables")
        print("Validá con: python -m cli.tarjetas")
    return 0


if __name__ == "__main__":
    sys.exit(main())
