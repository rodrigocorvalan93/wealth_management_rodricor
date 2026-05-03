# -*- coding: utf-8 -*-
"""
inspeccionar_planilla.py

Inspecciona la planilla v3.1 y dumpea su estructura: hojas, dimensiones,
headers, primeras filas. Solo lectura, NO modifica nada.

USO:
    python3 inspeccionar_planilla.py Planilla_Trading_Portfolio_Finanzas_v3_1.xlsx

Output: imprime en pantalla. Pegame TODA la salida (sin censura — son nombres
de hojas y columnas, no datos sensibles).
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta openpyxl. Instalá con:")
    print("  pip install openpyxl --break-system-packages")
    print("O en venv:")
    print("  source venv/bin/activate && pip install openpyxl")
    sys.exit(1)


HOJAS_RELEVANTES = [
    "Especies",
    "Saldo Inicial",
    "Blotter",
    "Holdings",
    "Config",
    "Funding",
]


def short(val, max_len=40):
    """Recorta el repr para que el output no se desborde."""
    if val is None:
        return "None"
    s = str(val)
    if len(s) > max_len:
        return s[:max_len] + "..."
    return s


def inspeccionar_hoja(wb, nombre):
    print(f"\n{'='*70}")
    print(f"HOJA: {nombre}")
    print(f"{'='*70}")
    if nombre not in wb.sheetnames:
        print(f"  ⚠ NO existe esta hoja en la planilla")
        return
    ws = wb[nombre]
    print(f"  Dimensiones: {ws.dimensions} ({ws.max_row} filas × {ws.max_column} cols)")
    print(f"  Hidden: {ws.sheet_state}")

    # Imprimir las primeras 5 filas (headers + algunas filas de data)
    rows_to_show = min(5, ws.max_row)
    print(f"\n  Primeras {rows_to_show} filas:")
    for row_idx in range(1, rows_to_show + 1):
        row_vals = []
        for col_idx in range(1, min(ws.max_column + 1, 16)):  # max 15 columnas
            cell = ws.cell(row=row_idx, column=col_idx)
            row_vals.append(short(cell.value, 25))
        print(f"    R{row_idx}: {row_vals}")

    # Buscar la fila de headers (la primera con muchos strings no vacíos)
    print(f"\n  Detectando fila de headers...")
    for row_idx in range(1, min(ws.max_row + 1, 10)):
        row_vals = [ws.cell(row=row_idx, column=c).value
                    for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in row_vals if v not in (None, "")]
        if len(non_empty) >= 3 and all(isinstance(v, str) for v in non_empty[:5]):
            print(f"    Probable fila headers = R{row_idx}: {[short(v, 20) for v in non_empty[:15]]}")
            break

    # Para hojas de datos: contar filas con contenido
    if nombre in ("Saldo Inicial", "Blotter", "Especies"):
        # Asumir headers en R1 o R2, datos arrancan en R2 o R3
        for header_row in (1, 2, 3):
            data_row_start = header_row + 1
            count = 0
            for r in range(data_row_start, ws.max_row + 1):
                # Buscar la primera columna con valor no nulo
                row_vals = [ws.cell(row=r, column=c).value
                            for c in range(1, min(ws.max_column + 1, 8))]
                if any(v not in (None, "") for v in row_vals):
                    count += 1
            if count > 0:
                print(f"  Si headers en R{header_row}: {count} filas de datos")


def listar_named_ranges(wb):
    print(f"\n{'='*70}")
    print(f"NAMED RANGES (rangos con nombre)")
    print(f"{'='*70}")
    if not wb.defined_names:
        print("  (ninguno)")
        return
    for name in wb.defined_names:
        defn = wb.defined_names[name]
        try:
            destinations = list(defn.destinations) if hasattr(defn, "destinations") else []
            for sheet, ref in destinations[:3]:
                print(f"  {name:<30} → {sheet}!{ref}")
        except Exception:
            print(f"  {name:<30} → {defn.value}")


def main():
    if len(sys.argv) != 2:
        print("USO: python3 inspeccionar_planilla.py <archivo.xlsx>")
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.is_file():
        print(f"No existe: {path}")
        sys.exit(1)

    print(f"Cargando {path}...")
    wb = load_workbook(filename=str(path), data_only=False, keep_links=False)

    print(f"\n{'='*70}")
    print(f"HOJAS EN EL WORKBOOK ({len(wb.sheetnames)} totales)")
    print(f"{'='*70}")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        marker = " [oculta]" if ws.sheet_state == "hidden" else ""
        print(f"  {i:2}. {name:<30} ({ws.max_row}×{ws.max_column}){marker}")

    for hoja in HOJAS_RELEVANTES:
        inspeccionar_hoja(wb, hoja)

    listar_named_ranges(wb)

    print(f"\n{'='*70}")
    print("FIN. Pegame TODA esta salida en el chat para que arme el parche.")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()