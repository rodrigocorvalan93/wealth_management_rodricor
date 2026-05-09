# -*- coding: utf-8 -*-
"""
add_carga_inicial_sheet.py

Agrega la hoja `_carga_inicial` a un master Excel existente sin destruir
las otras hojas. Si la hoja ya existe, no hace nada.

USO:
    python add_carga_inicial_sheet.py [path_xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


NAVY = "1F3864"
WHITE = "FFFFFF"
BLUE_INPUT = "0000FF"
YELLOW_INPUT = "FFF2CC"

FONT_TITLE = Font(name="Arial", size=14, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="Arial", size=10, italic=True, color="595959")
FONT_HEADER = Font(name="Arial", size=11, bold=True, color=WHITE)
FONT_INPUT = Font(name="Arial", size=11, color=BLUE_INPUT)
FONT_NORMAL = Font(name="Arial", size=11)

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_INPUT = PatternFill("solid", fgColor=YELLOW_INPUT)

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def add_hoja_carga_inicial(wb, with_examples: bool = True):
    """Agrega la hoja _carga_inicial al workbook.

    Si `with_examples` es True (default), la hoja viene con 9 filas de
    ejemplo para mostrar el formato. Para creación programática vía API
    conviene `with_examples=False` — sino esas filas se procesan como
    asientos reales y aparecen como ghost positions en el portfolio del
    usuario apenas corre el wizard de carga inicial.
    """
    if "_carga_inicial" in wb.sheetnames:
        return False

    ws = wb.create_sheet("_carga_inicial")

    # Banner
    ws.cell(row=1, column=1, value="CARGA INICIAL DE SALDOS").font = FONT_TITLE
    ws.cell(row=2, column=1,
            value=("Una fila por activo/saldo. Vos cargás simple. "
                   "El script `python -m cli.cargar_iniciales` genera los "
                   "asientos de doble entrada en `asientos_contables` "
                   "automáticamente con Event ID 'OPEN-AUTO-XXX'.")).font = FONT_SUBTITLE
    ws.cell(row=3, column=1,
            value=("Para activos: completá Unit Price + Price Currency. "
                   "Para cash (Activo=ARS/USB/USDT/etc): dejá Unit Price vacío. "
                   "Qty siempre POSITIVA (el script invierte el signo en opening_balance). "
                   "Si rerunneás el script, se regenera limpio.")).font = FONT_SUBTITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)

    # Headers
    headers = [
        "Fecha", "Cuenta", "Activo", "Qty", "Unit Price",
        "Price Currency", "Strategy", "Description", "Notes"
    ]
    widths = [12, 22, 16, 16, 14, 14, 14, 32, 25]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_THIN

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 30

    # Filas ejemplo (solo cuentas que existen en el template default)
    from datetime import date
    if not with_examples:
        # Validación + freeze + filter sin filas
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = "A4:I500"
        return True
    examples = [
        # Activos con cost basis
        (date(2026, 4, 30), "cocos",            "AL30D",            500,    65.50,    "USB",  "BH",       "1000 AL30D Cocos",       "ejemplo activo bono"),
        (date(2026, 4, 30), "cocos",            "GGAL",             1000,   5400,     "ARS",  "TRADING",  "GGAL Cocos",             ""),
        (date(2026, 4, 30), "delta_fci",        "DELTA_AHORRO_A",   5000,   3120.45,  "ARS",  "FCI",      "FCI Delta",              ""),
        # Cripto
        (date(2026, 4, 30), "binance",          "BTC",              0.05,   95000,    "USD",  "CRYPTO",   "BTC Binance",            ""),
        (date(2026, 4, 30), "binance",          "ETH",              2.5,    3500,     "USD",  "CRYPTO",   "ETH Binance",            ""),
        # Cash (sin unit_price)
        (date(2026, 4, 30), "cocos",            "ARS",              200000, None,     None,   "CASH",     "Cash ARS Cocos",         ""),
        (date(2026, 4, 30), "cocos",            "USB",              1500,   None,     None,   "CASH",     "Cash USB Cocos",         ""),
        (date(2026, 4, 30), "galicia_caja_ars", "ARS",              5000000,None,     None,   "CASH",     "Cash Galicia ARS",       ""),
        (date(2026, 4, 30), "binance",          "USDT",             2500,   None,     None,   "CASH",     "Stablecoin Binance",     ""),
    ]

    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                cell.font = FONT_INPUT
                cell.fill = FILL_INPUT
                cell.border = BORDER_THIN
            else:
                cell.font = FONT_NORMAL
                cell.border = BORDER_THIN
            if j == 1:
                cell.number_format = "yyyy-mm-dd"
            elif j in (4, 5):
                cell.number_format = '#,##0.0000;[Red](#,##0.0000)'

    # Validación de Strategy (libre, pero sugiere)
    dv_strat = DataValidation(
        type="list",
        formula1='"BH,TRADING,CORE,FCI,CRYPTO,CASH,DEBT"',
        allow_blank=True
    )
    dv_strat.add("G5:G500")
    ws.add_data_validation(dv_strat)

    # Freeze + filter
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:I500"

    return True


def main():
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    else:
        path = Path("inputs/wealth_management_rodricor.xlsx")

    if not path.is_file():
        print(f"[error] no se encontró {path}", file=sys.stderr)
        return 1

    print(f"[add] abriendo {path}...")
    wb = load_workbook(filename=str(path))

    if "_carga_inicial" in wb.sheetnames:
        print(f"[add] La hoja '_carga_inicial' YA EXISTE. No hago nada.")
        return 0

    added = add_hoja_carga_inicial(wb)
    if added:
        wb.save(str(path))
        print(f"[add] Hoja '_carga_inicial' agregada con 11 ejemplos. Cerrá y abrí Excel.")
        print(f"[add] Total hojas: {len(wb.sheetnames)}")
    else:
        print(f"[add] Nada que hacer.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
