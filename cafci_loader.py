# -*- coding: utf-8 -*-
"""
cafci_loader.py

Loader diario de cuotapartes (VCP) de FCIs desde la API de CAFCI.

CONVENCIÓN:
  - VCP normalizada = vcp_raw / 1000  (la API de CAFCI siempre devuelve en miles)
  - Moneda nativa del FCI sale del campo `moneda` del response (ARS/USD)
  - PERO si el ticker está en tu hoja `especies` del master Excel, se usa la
    moneda definida ahí (override). Esto resuelve casos donde CAFCI clasifica
    un fondo como USD pero vos sabés que es USB-MEP.

OUTPUT (formato compatible con planilla v3.1):
  data/precios_cafci.csv  →  Fecha, Ticker, Precio, Moneda, Fuente

Si el archivo destino ya existe, ANEXA filas nuevas (no pisa).
Si para una (Fecha, Ticker) ya hay una fila cargada, la actualiza con el
nuevo dato (último gana — útil si corrés el loader varias veces el mismo día).

USO:
    # último reporte (default — usa get_daily_report) con override de moneda
    python cafci_loader.py

    # un día específico
    python cafci_loader.py --fecha 2026-04-30

    # archivo de FCIs distinto
    python cafci_loader.py --fcis-file mis_fcis.txt

    # output a otra carpeta
    python cafci_loader.py --output-dir ./mi_data

    # NO usar override de moneda (respetar lo que dice CAFCI)
    python cafci_loader.py --no-xlsx-currency

    # Excel master alternativo (default: inputs/wealth_management_rodricor.xlsx)
    python cafci_loader.py --xlsx mi_master.xlsx

    # solo print, no escribir CSV (modo prueba)
    python cafci_loader.py --dry-run

REQUIERE en secrets.txt o env vars:
    CAFCI_TOKEN=Bearer eyJ...

ARCHIVO DE FCIs (default: fcis_cafci.txt):
    Cada línea: TICKER_INTERNO|NOMBRE_EXACTO_CAFCI
    El | es separador porque los nombres CAFCI a veces tienen comas.
    Líneas que empiezan con # son comentarios.
"""

from __future__ import annotations

import argparse
import os
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests


# =============================================================================
# Secrets (cargado al importar)
# =============================================================================

def load_secrets() -> int:
    """Carga secrets.txt (formato KEY=VALUE) a os.environ.

    Busca:
      1. ./secrets.txt (cwd)
      2. <script_dir>/secrets.txt
    No sobreescribe vars existentes. Devuelve cantidad de vars nuevas cargadas.
    """
    candidates = [
        Path.cwd() / "secrets.txt",
        Path(__file__).parent / "secrets.txt",
    ]
    for path in candidates:
        if not path.is_file():
            continue
        try:
            n_loaded = 0
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#") or "=" not in line:
                        continue
                    key, _, val = line.partition("=")
                    key = key.strip()
                    val = val.strip().strip('"').strip("'")
                    if key and key not in os.environ:
                        os.environ[key] = val
                        n_loaded += 1
            print(f"[secrets] cargado {path.name} ({n_loaded} vars nuevas)")
            return n_loaded
        except Exception as e:
            print(f"[secrets] error leyendo {path}: {e}", file=sys.stderr)
    return 0


load_secrets()


# =============================================================================
# Config
# =============================================================================

BASE_URL = "https://cloud.cafci.org.ar/api"
DEFAULT_TIMEOUT = 15  # segundos
_TOKEN_ENV = "CAFCI_TOKEN"
VCP_DIVISOR = 1000  # convención CAFCI: vcp viene en miles

# CAMBIO: archivo separado de BYMA (antes era "precios_historico.csv")
DEFAULT_OUTPUT_FILENAME = "precios_cafci.csv"

# Master Excel default (para leer overrides de moneda)
DEFAULT_XLSX = Path("inputs/wealth_management_rodricor.xlsx")


# =============================================================================
# API client (mínimo, autocontenido)
# =============================================================================

def _resolve_token(cafci_token: Optional[str] = None) -> str:
    """Resuelve el token CAFCI. Prioridad: arg > os.environ.

    El valor debe incluir el prefijo "Bearer ".
    """
    token = cafci_token or os.getenv(_TOKEN_ENV)
    if not token:
        raise RuntimeError(
            f"Falta el token de CAFCI. Definí {_TOKEN_ENV} en secrets.txt "
            f'(formato: {_TOKEN_ENV}=Bearer eyJ...) o pasalo como argumento.'
        )
    return token


def _get_json(url: str, cafci_token: Optional[str] = None) -> Dict[str, Any]:
    """GET genérico contra la API CAFCI.

    Tira RuntimeError con detalle si la respuesta no es 2xx o no es JSON válido.
    """
    headers = {"Authorization": _resolve_token(cafci_token)}
    try:
        r = requests.get(url, headers=headers, timeout=DEFAULT_TIMEOUT)
    except requests.RequestException as e:
        raise RuntimeError(f"CAFCI request falló: {type(e).__name__}: {e}") from e

    if r.status_code != 200:
        raise RuntimeError(
            f"CAFCI HTTP {r.status_code} en {url}: {r.text[:200]}"
        )
    try:
        return r.json()
    except ValueError as e:
        raise RuntimeError(
            f"CAFCI respondió con JSON inválido: {r.text[:200]}"
        ) from e


def get_daily_report(cafci_token: Optional[str] = None) -> pd.DataFrame:
    """Reporte diario completo de CAFCI (todos los fondos)."""
    data = _get_json(f"{BASE_URL}/reports/daily", cafci_token)
    return pd.json_normalize(data["records"])


def get_historic_report(
    fecha: str,
    cafci_token: Optional[str] = None,
) -> pd.DataFrame:
    """Reporte histórico de CAFCI para una fecha YYYY-MM-DD."""
    data = _get_json(
        f"{BASE_URL}/reports/historic?date={fecha}",
        cafci_token,
    )
    return pd.json_normalize(data["records"])


# =============================================================================
# VCP helpers
# =============================================================================

def vcp_normalizada(vcp_raw: Any) -> float:
    """Convierte VCP cruda (en miles) a unidades. NaN si no se puede parsear.

    Convención CAFCI: vcp viene siempre dividida por 1000 → para mostrarla en
    pesos por cuotaparte hay que dividir por 1000.
    """
    try:
        return float(vcp_raw) / VCP_DIVISOR
    except (TypeError, ValueError):
        return float("nan")


# =============================================================================
# Lectura del archivo de FCIs
# =============================================================================

@dataclass
class FCIMapping:
    """Mapping ticker_interno → nombre exacto en CAFCI."""
    ticker: str
    nombre_cafci: str


def parse_fcis_file(path: Path) -> List[FCIMapping]:
    """Lee fcis_cafci.txt. Cada línea: TICKER|NOMBRE_CAFCI.

    Ignora líneas vacías y comentarios (#).
    Tira ValueError si encuentra una línea malformada (sin |).
    """
    if not path.is_file():
        raise FileNotFoundError(path)
    out: List[FCIMapping] = []
    with open(path, "r", encoding="utf-8") as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "|" not in line:
                raise ValueError(
                    f"{path}:{lineno}: línea malformada (falta '|'): {line!r}\n"
                    f"  Formato esperado: TICKER|NOMBRE_EXACTO_CAFCI"
                )
            ticker, _, nombre = line.partition("|")
            ticker = ticker.strip()
            nombre = nombre.strip()
            if not ticker or not nombre:
                raise ValueError(
                    f"{path}:{lineno}: ticker o nombre vacío: {line!r}"
                )
            out.append(FCIMapping(ticker=ticker, nombre_cafci=nombre))
    return out


# =============================================================================
# NUEVO: Override de moneda desde hoja `especies` del Excel master
# =============================================================================

def load_currency_overrides_from_xlsx(xlsx_path: Path) -> Dict[str, str]:
    """Lee la hoja `especies` del master Excel y devuelve {ticker: currency}.

    Esto se usa para forzar la moneda de un FCI cuando vos sabés que la
    clasificación de CAFCI es incorrecta (ej fondos USB que CAFCI clasifica USD).

    Si el archivo no existe o no tiene la hoja, devuelve {} (silencioso, no error).
    """
    if not xlsx_path.is_file():
        return {}

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[xlsx] WARN: openpyxl no instalado, no puedo leer overrides de moneda")
        return {}

    try:
        wb = load_workbook(filename=str(xlsx_path), data_only=True)
    except Exception as e:
        print(f"[xlsx] WARN no pude abrir {xlsx_path}: {e}")
        return {}

    if "especies" not in wb.sheetnames:
        return {}

    ws = wb["especies"]
    # Headers en fila 4
    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=4, column=c).value
        if isinstance(h, str):
            h = h.strip()
        headers.append(h)

    try:
        ticker_idx = headers.index("Ticker")
        currency_idx = headers.index("Currency")
    except ValueError:
        return {}

    out: Dict[str, str] = {}
    for r in range(5, ws.max_row + 1):
        ticker = ws.cell(row=r, column=ticker_idx + 1).value
        currency = ws.cell(row=r, column=currency_idx + 1).value
        if isinstance(ticker, str): ticker = ticker.strip()
        if isinstance(currency, str): currency = currency.strip()
        if ticker and currency:
            out[ticker] = currency

    return out


# =============================================================================
# Lookup en el report
# =============================================================================

@dataclass
class FCISnapshot:
    """Snapshot de un FCI para una fecha."""
    ticker: str
    nombre_cafci: str
    fecha: str       # ISO YYYY-MM-DD (lo que viene de la API)
    vcp: float       # ya dividida por 1000
    moneda: str      # 'ARS' | 'USD' | 'USB' (después del override)
    moneda_cafci: str  # moneda original que dijo CAFCI (para auditoría)

    @property
    def is_valid(self) -> bool:
        import math
        return (
            self.vcp is not None
            and not math.isnan(self.vcp)
            and self.vcp > 0
            and bool(self.fecha)
        )


def lookup_fcis_in_report(
    df_report: pd.DataFrame,
    fcis: List[FCIMapping],
    currency_overrides: Optional[Dict[str, str]] = None,
) -> Tuple[List[FCISnapshot], List[str]]:
    """Busca cada FCI en el report. Devuelve (snapshots_validos, tickers_no_encontrados).

    Si `currency_overrides` está definido y el ticker está en el dict,
    usa esa moneda en lugar de la que dice CAFCI.
    """
    required_cols = {"nombreDeLaClaseDeFondo", "vcp", "fecha", "moneda"}
    missing_cols = required_cols - set(df_report.columns)
    if missing_cols:
        raise RuntimeError(
            f"El report de CAFCI no tiene las columnas esperadas. "
            f"Faltan: {sorted(missing_cols)}. "
            f"Columnas disponibles: {sorted(df_report.columns)[:20]}..."
        )

    # Index por nombre para lookup O(1)
    by_name: Dict[str, Dict[str, Any]] = {}
    for _, row in df_report.iterrows():
        nombre = row["nombreDeLaClaseDeFondo"]
        if isinstance(nombre, str):
            by_name[nombre] = row.to_dict()

    overrides = currency_overrides or {}
    snapshots: List[FCISnapshot] = []
    not_found: List[str] = []

    for fci in fcis:
        row = by_name.get(fci.nombre_cafci)
        if row is None:
            not_found.append(fci.ticker)
            continue

        moneda_cafci = str(row.get("moneda", "")).strip().upper() or "ARS"
        # Override si está definido para este ticker
        moneda_final = overrides.get(fci.ticker, moneda_cafci)

        snap = FCISnapshot(
            ticker=fci.ticker,
            nombre_cafci=fci.nombre_cafci,
            fecha=str(row.get("fecha", "")),
            vcp=vcp_normalizada(row.get("vcp")),
            moneda=moneda_final,
            moneda_cafci=moneda_cafci,
        )
        if snap.is_valid:
            snapshots.append(snap)
        else:
            not_found.append(fci.ticker)

    return snapshots, not_found


# =============================================================================
# CSV upsert (idéntico al de historico_byma_loader.py — mismo formato)
# =============================================================================

def upsert_csv(
    path: Path,
    new_rows: List[Dict[str, Any]],
    key_cols: List[str],
    column_order: List[str],
) -> Tuple[int, int]:
    """Anexa filas nuevas y reemplaza las que matcheen por key_cols. Devuelve (n_new, n_updated)."""
    if not new_rows:
        return 0, 0

    df_new = pd.DataFrame(new_rows)

    if path.is_file():
        try:
            df_old = pd.read_csv(path)
        except Exception as e:
            print(f"[upsert] error leyendo {path}: {e}. Lo recreo.", file=sys.stderr)
            df_old = pd.DataFrame(columns=column_order)
    else:
        df_old = pd.DataFrame(columns=column_order)

    # contar updates antes del merge
    if not df_old.empty and all(k in df_old.columns for k in key_cols):
        old_keys = df_old[key_cols].astype(str).agg("||".join, axis=1)
        new_keys = df_new[key_cols].astype(str).agg("||".join, axis=1)
        n_updated = int(old_keys.isin(new_keys).sum())
    else:
        n_updated = 0

    df_merged = pd.concat([df_old, df_new], ignore_index=True)
    df_merged = df_merged.drop_duplicates(subset=key_cols, keep="last")

    cols_existing = [c for c in column_order if c in df_merged.columns]
    cols_extra = [c for c in df_merged.columns if c not in column_order]
    df_merged = df_merged[cols_existing + cols_extra]

    if "Fecha" in df_merged.columns:
        df_merged = df_merged.sort_values("Fecha", kind="stable")

    path.parent.mkdir(parents=True, exist_ok=True)
    df_merged.to_csv(path, index=False, encoding="utf-8")
    n_new = len(df_new) - n_updated
    return n_new, n_updated


# =============================================================================
# Run
# =============================================================================

def run(
    fcis_file: Path,
    output_dir: Path,
    fecha: Optional[str] = None,
    dry_run: bool = False,
    use_xlsx_currency: bool = True,
    xlsx_path: Optional[Path] = None,
) -> int:
    """Punto de entrada principal. Devuelve exit code (0 OK, 1 error)."""

    # 1. Leer lista de FCIs
    try:
        fcis = parse_fcis_file(fcis_file)
    except FileNotFoundError:
        print(f"[error] archivo de FCIs no encontrado: {fcis_file}",
              file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"[error] {e}", file=sys.stderr)
        return 1

    if not fcis:
        print(f"[error] {fcis_file} no contiene ningún FCI", file=sys.stderr)
        return 1

    print(f"[fcis] {len(fcis)} FCIs a buscar (desde {fcis_file.name})")

    # 1b. NUEVO: Cargar overrides de moneda desde Excel
    overrides: Dict[str, str] = {}
    if use_xlsx_currency:
        xlsx_to_use = xlsx_path or DEFAULT_XLSX
        overrides = load_currency_overrides_from_xlsx(xlsx_to_use)
        if overrides:
            # Filtrar solo a los tickers de FCIs (el resto del Excel es ruido para esto)
            fci_tickers = {f.ticker for f in fcis}
            overrides = {t: c for t, c in overrides.items() if t in fci_tickers}
            print(f"[xlsx] cargado {len(overrides)} overrides de moneda desde {xlsx_to_use.name}")
        else:
            print(f"[xlsx] no se cargaron overrides (Excel no encontrado o sin hoja `especies`)")

    # 2. Descargar report
    try:
        if fecha:
            print(f"[cafci] descargando report histórico para {fecha}...")
            df_report = get_historic_report(fecha)
        else:
            print(f"[cafci] descargando reporte diario...")
            df_report = get_daily_report()
    except RuntimeError as e:
        print(f"[error] {e}", file=sys.stderr)
        return 1

    print(f"[cafci] report descargado: {len(df_report)} fondos en total")

    # 3. Lookup
    try:
        snapshots, not_found = lookup_fcis_in_report(df_report, fcis, overrides)
    except RuntimeError as e:
        print(f"[error] {e}", file=sys.stderr)
        return 1

    print(f"[lookup] encontrados: {len(snapshots)} | no encontrados: {len(not_found)}")
    if not_found:
        print(f"[lookup] sin datos: {', '.join(not_found)}")

    # 4. Mostrar snapshots (con override si aplica)
    print()
    print(f"  {'TICKER':<28} {'FECHA':<12} {'VCP':>14} {'MONEDA':<8} {'NOMBRE':<40}")
    print(f"  {'-'*28} {'-'*12} {'-'*14} {'-'*8} {'-'*40}")
    for s in snapshots:
        moneda_display = s.moneda
        if s.moneda != s.moneda_cafci:
            moneda_display = f"{s.moneda}*"  # marca override
        print(f"  {s.ticker:<28} {s.fecha:<12} {s.vcp:>14,.6f} {moneda_display:<8} {s.nombre_cafci}")
    overrides_aplicados = sum(1 for s in snapshots if s.moneda != s.moneda_cafci)
    if overrides_aplicados:
        print(f"\n  * = moneda override aplicado ({overrides_aplicados} FCIs)")
    print()

    if not snapshots:
        print("[done] no hay nada que escribir.")
        return 0

    # 5. Escribir CSV
    if dry_run:
        print("[dry-run] no se escribió CSV")
        return 0

    rows: List[Dict[str, Any]] = [
        {
            "Fecha": s.fecha,
            "Ticker": s.ticker,
            "Precio": round(s.vcp, 6),
            "Moneda": s.moneda,  # ya con override aplicado si corresponde
            "Fuente": "CAFCI daily VCP",
        }
        for s in snapshots
    ]

    output_path = output_dir / DEFAULT_OUTPUT_FILENAME
    n_new, n_upd = upsert_csv(
        path=output_path,
        new_rows=rows,
        key_cols=["Fecha", "Ticker"],
        column_order=["Fecha", "Ticker", "Precio", "Moneda", "Fuente"],
    )
    print(f"[csv] {output_path}: {len(rows)} filas → {n_new} nuevos, {n_upd} actualizados")
    print(f"[done] OK")
    return 0


# =============================================================================
# CLI
# =============================================================================

def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Loader de cuotapartes de FCIs (CAFCI) → CSV planilla v3.1",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument(
        "--fcis-file", type=Path, default=Path.cwd() / "fcis_cafci.txt",
        help="Archivo con lista de FCIs (default: ./fcis_cafci.txt)",
    )
    p.add_argument(
        "--output-dir", type=Path, default=Path.cwd() / "data",
        help="Carpeta destino del CSV (default: ./data)",
    )
    p.add_argument(
        "--fecha", type=str, default=None,
        help="Fecha YYYY-MM-DD (default: último reporte diario)",
    )
    p.add_argument(
        "--dry-run", action="store_true",
        help="No escribir CSV, solo mostrar resultado",
    )
    p.add_argument(
        "--no-xlsx-currency", action="store_true",
        help="NO usar override de moneda desde el Excel (default: usar override)",
    )
    p.add_argument(
        "--xlsx", type=Path, default=DEFAULT_XLSX,
        help=f"Excel master para leer overrides de moneda (default: {DEFAULT_XLSX})",
    )

    args = p.parse_args(argv)

    # Validar fecha si se pasó
    if args.fecha:
        try:
            date.fromisoformat(args.fecha)
        except ValueError:
            print(f"[error] fecha inválida: {args.fecha} (esperado YYYY-MM-DD)",
                  file=sys.stderr)
            return 1

    return run(
        fcis_file=args.fcis_file,
        output_dir=args.output_dir,
        fecha=args.fecha,
        dry_run=args.dry_run,
        use_xlsx_currency=not args.no_xlsx_currency,
        xlsx_path=args.xlsx,
    )


if __name__ == "__main__":
    sys.exit(main())
