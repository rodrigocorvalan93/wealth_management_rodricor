# -*- coding: utf-8 -*-
"""
migrate_master.py

Migra un Excel master EXISTENTE a la nueva estructura sin perder datos.

Cambios aplicados (idempotentes — re-ejecutar es seguro):

  1. Backup automático del archivo original a `<name>.backup-<ISO>.xlsx`.

  2. En hoja `cuentas`:
     - Agrega columnas `Investible` (default YES) y `Cash Purpose` (vacío).
     - Para cuentas técnicas (external_*, opening_balance, interest_*) y
       `cash_reserva`: investible=NO automático.

  3. En hoja `funding`:
     - Agrega columna `Linked Trade ID` (vacío).

  4. Crea hojas nuevas si no existen:
     - `aforos`         — aforos BYMA defaults (BOND_AR 85%, EQUITY_AR 70%, etc).
     - `margin_config`  — config de margen IBKR (x2 ON / x4 ID / 6% funding).

  5. Agrega columna `Row ID` al final de cada hoja de eventos. Genera IDs
     únicos estables para edición posterior vía API:
       - blotter           → BL-0001, BL-0002, ...
       - gastos            → GS-0001
       - ingresos          → IN-0001
       - transferencias_*  → TC-0001 / TA-0001
       - funding           → FN-0001
       - asientos_contables→ AS-0001 (un ID por GRUPO de Event ID)
       - recurrentes       → RC-0001
       - pagos_pasivos     → PP-0001

USO:
    python migrate_master.py inputs/wealth_management_rodricor.xlsx
    python migrate_master.py inputs/wealth_management_rodricor.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


# Estilos consistentes con build_master.py
NAVY = "1F3864"
WHITE = "FFFFFF"
BLUE_INPUT = "0000FF"
YELLOW_INPUT = "FFF2CC"

FONT_TITLE = Font(name="Arial", size=14, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="Arial", size=10, italic=True, color="595959")
FONT_HEADER = Font(name="Arial", size=11, bold=True, color=WHITE)
FONT_INPUT = Font(name="Arial", size=11, color=BLUE_INPUT)
FONT_NORMAL = Font(name="Arial", size=11)

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_INPUT = PatternFill("solid", fgColor=YELLOW_INPUT)

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# Sheet → prefijo de Row ID
ROW_ID_PREFIXES = {
    "blotter": "BL",
    "gastos": "GS",
    "ingresos": "IN",
    "transferencias_cash": "TC",
    "transferencias_activos": "TA",
    "funding": "FN",
    "asientos_contables": "AS",
    "recurrentes": "RC",
    "pagos_pasivos": "PP",
}

# Cuentas técnicas que SIEMPRE son no-invertibles
NON_INVESTIBLE_KINDS = {"EXTERNAL", "OPENING_BALANCE",
                        "INTEREST_EXPENSE", "INTEREST_INCOME"}
# Cuentas que opcionalmente quieras marcar no-invertibles por nombre
NON_INVESTIBLE_HINTS = ("reserva", "blanqueo", "no_decl", "nodecl", "guardado")


# =============================================================================
# Helpers
# =============================================================================

def _backup(xlsx_path: Path) -> Path:
    """Backup del Excel a <name>.backup-<ISO>.xlsx. Devuelve el path."""
    ts = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    backup = xlsx_path.with_name(f"{xlsx_path.stem}.backup-{ts}{xlsx_path.suffix}")
    shutil.copy2(xlsx_path, backup)
    return backup


def _find_header_row(ws: Worksheet, max_search: int = 10) -> Optional[int]:
    """Encuentra la fila de headers (busca el patrón del build_master).

    Convención: banner en filas 1-3, headers en fila 4.
    Si no, busca la primera fila con varios cells de texto contiguos.
    """
    # Caso típico: fila 4
    if ws.cell(row=4, column=1).value is not None and isinstance(ws.cell(row=4, column=1).value, str):
        # Verificar que hay al menos 2 headers
        n_headers = sum(1 for c in range(1, ws.max_column + 1)
                        if ws.cell(row=4, column=c).value)
        if n_headers >= 2:
            return 4
    # Búsqueda manual
    for r in range(1, max_search + 1):
        n = sum(1 for c in range(1, min(ws.max_column + 1, 20))
                if isinstance(ws.cell(row=r, column=c).value, str)
                and ws.cell(row=r, column=c).value)
        if n >= 3:
            return r
    return None


def _read_headers(ws: Worksheet, header_row: int) -> list[str]:
    """Lee los headers como lista (puede tener None en huecos)."""
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        out.append(str(v).strip() if v is not None else None)
    return out


def _last_data_row(ws: Worksheet, header_row: int) -> int:
    """Devuelve el último row con datos reales (no vacío)."""
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                last = r
                break
    return last


def _add_column_with_header(ws: Worksheet, header_row: int, name: str,
                             default_value=None, value_resolver=None,
                             width: int = 14) -> int:
    """Agrega una columna nueva al final con el header dado.

    - Si la columna ya existe, no la duplica.
    - default_value: valor para todas las filas de datos
    - value_resolver: función opcional (row_dict) -> value, override del default

    Devuelve el índice de columna (1-based).
    """
    headers = _read_headers(ws, header_row)
    # Ya existe?
    for i, h in enumerate(headers, start=1):
        if h and h.lower() == name.lower():
            return i
    # Añadir al final (después de la última columna no-None)
    last_col_with_header = max(
        (i for i, h in enumerate(headers, start=1) if h),
        default=0,
    )
    new_col = last_col_with_header + 1
    # Header
    cell = ws.cell(row=header_row, column=new_col, value=name)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_HEADER
    cell.border = BORDER_THIN
    ws.column_dimensions[get_column_letter(new_col)].width = width

    # Rellenar filas de datos
    last_data = _last_data_row(ws, header_row)
    for r in range(header_row + 1, last_data + 1):
        # Si la fila está completamente vacía, no agregar nada
        is_empty = all(
            ws.cell(row=r, column=c).value in (None, "")
            for c in range(1, last_col_with_header + 1)
        )
        if is_empty:
            continue
        # Resolver valor
        if value_resolver is not None:
            row_dict = {
                h: ws.cell(row=r, column=i + 1).value
                for i, h in enumerate(headers) if h
            }
            val = value_resolver(row_dict)
        else:
            val = default_value
        if val is not None:
            cell = ws.cell(row=r, column=new_col, value=val)
            cell.font = FONT_NORMAL
    return new_col


def _next_row_id(prefix: str, existing: set[str]) -> str:
    """Devuelve el próximo Row ID disponible para `prefix`."""
    n = 1
    while True:
        candidate = f"{prefix}-{n:04d}"
        if candidate not in existing:
            return candidate
        n += 1


# =============================================================================
# Migraciones por hoja
# =============================================================================

def migrate_cuentas(ws: Worksheet) -> dict:
    """Agrega columnas Investible + Cash Purpose."""
    header_row = _find_header_row(ws) or 4
    headers = _read_headers(ws, header_row)

    # Resolver Investible según Kind y Code
    def resolve_investible(row):
        kind = (row.get("Kind") or "").strip().upper()
        code = (row.get("Code") or "").strip().lower()
        if kind in NON_INVESTIBLE_KINDS:
            return "NO"
        if any(h in code for h in NON_INVESTIBLE_HINTS):
            return "NO"
        return "YES"

    def resolve_purpose(row):
        kind = (row.get("Kind") or "").strip().upper()
        code = (row.get("Code") or "").strip().lower()
        if kind in NON_INVESTIBLE_KINDS:
            return None
        if "reserva" in code or "no_decl" in code or "nodecl" in code:
            return "RESERVA_NO_DECLARADO"
        return "OPERATIVO"

    col_inv = _add_column_with_header(
        ws, header_row, "Investible",
        value_resolver=resolve_investible, width=11,
    )
    col_purpose = _add_column_with_header(
        ws, header_row, "Cash Purpose",
        value_resolver=resolve_purpose, width=22,
    )

    # Validación YES/NO en Investible
    try:
        col_letter = get_column_letter(col_inv)
        dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
        dv.add(f"{col_letter}{header_row + 1}:{col_letter}500")
        ws.add_data_validation(dv)
    except Exception:
        pass

    return {
        "investible_col": col_inv,
        "cash_purpose_col": col_purpose,
    }


def migrate_funding(ws: Worksheet) -> dict:
    """Agrega columna Linked Trade ID."""
    header_row = _find_header_row(ws) or 4
    col = _add_column_with_header(
        ws, header_row, "Linked Trade ID", default_value=None, width=16,
    )
    return {"linked_trade_id_col": col}


def add_row_id_to_sheet(ws: Worksheet, prefix: str) -> dict:
    """Agrega columna Row ID al final con valores únicos auto-generados.

    Para asientos_contables: agrupa por Event ID (mismo Event ID = mismo Row ID).
    """
    header_row = _find_header_row(ws) or 4
    headers = _read_headers(ws, header_row)

    # Si ya existe, retornar
    for i, h in enumerate(headers, start=1):
        if h and h.lower() == "row id":
            return {"row_id_col": i, "added": 0}

    last_data = _last_data_row(ws, header_row)

    # Asientos contables: usa Event ID para agrupar (mismo Event ID = mismo Row ID)
    is_grouped = (prefix == "AS")
    event_id_col = None
    if is_grouped:
        for i, h in enumerate(headers, start=1):
            if h and h.lower() == "event id":
                event_id_col = i
                break

    # Detectar IDs ya usados (en este caso, ninguno todavía, pero lo dejo robusto)
    existing_ids = set()

    # Generar nuevo Row ID
    last_col_with_header = max(
        (i for i, h in enumerate(headers, start=1) if h),
        default=0,
    )
    new_col = last_col_with_header + 1

    # Header
    cell = ws.cell(row=header_row, column=new_col, value="Row ID")
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_HEADER
    cell.border = BORDER_THIN
    ws.column_dimensions[get_column_letter(new_col)].width = 12

    # Para grupos: mapear event_id externo → Row ID
    group_map: dict = {}
    n_added = 0
    for r in range(header_row + 1, last_data + 1):
        is_empty = all(
            ws.cell(row=r, column=c).value in (None, "")
            for c in range(1, last_col_with_header + 1)
        )
        if is_empty:
            continue
        if is_grouped and event_id_col:
            ev_id = ws.cell(row=r, column=event_id_col).value
            if ev_id is None:
                continue
            ev_key = str(ev_id).strip()
            if ev_key not in group_map:
                rid = _next_row_id(prefix, existing_ids)
                existing_ids.add(rid)
                group_map[ev_key] = rid
            rid = group_map[ev_key]
        else:
            rid = _next_row_id(prefix, existing_ids)
            existing_ids.add(rid)
        cell = ws.cell(row=r, column=new_col, value=rid)
        cell.font = FONT_NORMAL
        n_added += 1

    return {"row_id_col": new_col, "added": n_added}


# =============================================================================
# Hojas nuevas (defaults aplicados)
# =============================================================================

def create_sheet_aforos(wb) -> Worksheet:
    """Crea hoja 'aforos' con defaults BYMA si no existe."""
    if "aforos" in wb.sheetnames:
        return wb["aforos"]
    ws = wb.create_sheet("aforos")

    # Banner
    ws.cell(row=1, column=1,
            value="AFOROS BYMA — Garantías de caución").font = FONT_TITLE
    ws.cell(row=2, column=1,
            value=("Aforo = % del valor de mercado aceptado como GARANTÍA por BYMA. "
                   "Define cuánto poder de compra ganás dejando el activo como margen. "
                   "Scope Type='CLASS' aplica a todo el asset_class; "
                   "'TICKER' override por instrumento.")
            ).font = FONT_SUBTITLE
    ws.merge_cells("A1:E1"); ws.merge_cells("A2:E2")

    headers = ["Scope Type", "Scope Value", "Aforo %", "Source", "Notes"]
    widths = [12, 22, 12, 14, 50]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER; cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(i)].width = w

    rows = [
        ("CLASS", "BOND_AR",    0.85, "BYMA approx", "Bonos soberanos AR"),
        ("CLASS", "EQUITY_AR",  0.70, "BYMA approx", "Acciones líderes Merval"),
        ("CLASS", "EQUITY_US",  0.70, "BYMA approx", "CEDEARs"),
        ("CLASS", "FCI",        0.90, "BYMA approx", "FCIs Money Market"),
        ("CLASS", "STABLECOIN", 0.50, "manual",      "USDT/USDC"),
        ("CLASS", "CRYPTO",     0.00, "manual",      "Cripto NO aceptado"),
        ("CLASS", "DERIVATIVE", 0.00, "BYMA",        "Derivados no garantizan"),
        ("TICKER","AL30D",      0.90, "BYMA",        "Bono soberano USD-MEP"),
        ("TICKER","GD30D",      0.90, "BYMA",        "Bono soberano USD-MEP"),
    ]
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONT_INPUT if val is not None else FONT_NORMAL
            cell.fill = FILL_INPUT
            cell.border = BORDER_THIN
            if j == 3:
                cell.number_format = "0.00%"

    dv = DataValidation(type="list", formula1='"CLASS,TICKER"', allow_blank=False)
    dv.add("A5:A500")
    ws.add_data_validation(dv)
    ws.freeze_panes = "A5"
    return ws


def create_sheet_margin_config(wb) -> Worksheet:
    """Crea hoja 'margin_config' con defaults IBKR si no existe."""
    if "margin_config" in wb.sheetnames:
        return wb["margin_config"]
    ws = wb.create_sheet("margin_config")

    ws.cell(row=1, column=1,
            value="MARGIN CONFIG (IBKR, etc)").font = FONT_TITLE
    ws.cell(row=2, column=1,
            value=("Configuración de leverage por cuenta para cuentas con margin "
                   "estilo RegT (IBKR: x2 overnight, x4 intraday). "
                   "VERIFICÁ los multiplicadores y la tasa de fondeo reales con tu broker. "
                   "Cocos/Eco usan la hoja 'aforos' (BYMA), no esta.")
            ).font = FONT_SUBTITLE
    ws.merge_cells("A1:F1"); ws.merge_cells("A2:F2")

    headers = ["Account", "Mult. Overnight", "Mult. Intraday",
               "Funding Rate Annual", "Funding Currency", "Notes"]
    widths = [22, 16, 16, 18, 14, 40]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER; cell.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(i)].width = w

    rows = [
        ("ibkr", 2.0, 4.0, 0.06, "USD",
         "RegT estándar: 50% margin overnight, 25% intraday. "
         "Funding ~6% anual USD aprox — VERIFICAR con tu cuenta."),
    ]
    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = FONT_INPUT if val is not None else FONT_NORMAL
            cell.fill = FILL_INPUT
            cell.border = BORDER_THIN
            if j in (2, 3):
                cell.number_format = "0.00"
            elif j == 4:
                cell.number_format = "0.00%"
    ws.freeze_panes = "A5"
    return ws


# =============================================================================
# Main
# =============================================================================

def migrate(xlsx_path: Path, dry_run: bool = False) -> dict:
    """Aplica todas las migraciones y devuelve un resumen."""
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"No existe: {xlsx_path}")

    summary = {
        "input": str(xlsx_path),
        "backup": None,
        "changes": [],
    }

    # Backup primero
    if not dry_run:
        backup = _backup(xlsx_path)
        summary["backup"] = str(backup)
        print(f"[migrate] backup creado: {backup}")

    wb = load_workbook(filename=str(xlsx_path))
    print(f"[migrate] {len(wb.sheetnames)} hojas: {', '.join(wb.sheetnames)}")

    # 1. cuentas
    if "cuentas" in wb.sheetnames:
        info = migrate_cuentas(wb["cuentas"])
        summary["changes"].append({"sheet": "cuentas", **info})
        print(f"[migrate] cuentas: cols Investible+Cash Purpose agregadas")

    # 2. funding
    if "funding" in wb.sheetnames:
        info = migrate_funding(wb["funding"])
        summary["changes"].append({"sheet": "funding", **info})
        print(f"[migrate] funding: col Linked Trade ID agregada")

    # 3. hojas nuevas
    if "aforos" not in wb.sheetnames:
        create_sheet_aforos(wb)
        summary["changes"].append({"sheet": "aforos", "created": True})
        print(f"[migrate] hoja 'aforos' creada con defaults BYMA")
    else:
        print(f"[migrate] hoja 'aforos' ya existe — sin cambios")

    if "margin_config" not in wb.sheetnames:
        create_sheet_margin_config(wb)
        summary["changes"].append({"sheet": "margin_config", "created": True})
        print(f"[migrate] hoja 'margin_config' creada con defaults IBKR")
    else:
        print(f"[migrate] hoja 'margin_config' ya existe — sin cambios")

    # 4. Row ID en hojas de eventos
    for sheet_name, prefix in ROW_ID_PREFIXES.items():
        if sheet_name not in wb.sheetnames:
            continue
        info = add_row_id_to_sheet(wb[sheet_name], prefix)
        summary["changes"].append({
            "sheet": sheet_name,
            "row_id_prefix": prefix,
            **info,
        })
        if info.get("added"):
            print(f"[migrate] {sheet_name}: {info['added']} Row IDs generados ({prefix}-XXXX)")
        else:
            print(f"[migrate] {sheet_name}: Row ID column ya presente")

    if dry_run:
        print("[migrate] DRY-RUN: no se modificó el archivo")
    else:
        wb.save(str(xlsx_path))
        print(f"[migrate] OK → {xlsx_path}")
        print(f"[migrate] Si algo salió mal, restaurá desde: {summary['backup']}")
    return summary


def main():
    p = argparse.ArgumentParser(description="Migra Excel master a la nueva estructura.")
    p.add_argument("xlsx", type=Path,
                   help="Path al Excel master (ej: inputs/wealth_management_rodricor.xlsx)")
    p.add_argument("--dry-run", action="store_true",
                   help="No escribe el archivo, solo muestra qué haría")
    args = p.parse_args()

    summary = migrate(args.xlsx, dry_run=args.dry_run)
    print()
    print(f"[migrate] Resumen: {len(summary['changes'])} hojas afectadas")
    return 0


if __name__ == "__main__":
    sys.exit(main())
