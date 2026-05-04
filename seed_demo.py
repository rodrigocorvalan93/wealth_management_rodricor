# -*- coding: utf-8 -*-
"""
seed_demo.py

Genera (o resetea) un Excel master de demo con datos hard-coded fijos.
Pensado para mostrar la app sin exponer datos reales.

Cuentas demo:
  - cocos_demo       (CASH_BROKER, ARS) — broker principal
  - ibkr_demo        (CASH_BROKER, USD) — cuenta US
  - binance_demo     (CASH_WALLET, sin moneda) — cripto
  - galicia_demo_ars (CASH_BANK, ARS)  — banco
  - galicia_demo_visa (CARD_CREDIT)    — tarjeta

Especies demo:
  - AL30D (BOND_AR, USB) — bono soberano AR
  - GGAL.BA (EQUITY_AR, ARS)
  - AAPL (EQUITY_US, USD), SPY (EQUITY_US, USD), TSLA (EQUITY_US, USD)
  - BTC (CRYPTO, USD), ETH (CRYPTO, USD), USDT (STABLECOIN, USD)
  - PEDLAR (FCI, ARS)

Trades históricos: ~10-12 con saldos abiertos + targets en algunos.
Ingresos: 2 sueldos. Gastos: 3 categorías. Saldos iniciales en
asientos_contables.

USO:
    python seed_demo.py inputs/demo/wealth_management.xlsx
    python seed_demo.py --user demo                  # resuelve path automático
    python seed_demo.py --user demo --reset          # borra master y rehace
"""

from __future__ import annotations

import argparse
import sys
from datetime import date
from pathlib import Path

HERE = Path(__file__).resolve().parent

# Datos hard-coded (fijos: cada reset da los mismos números)

CUENTAS_DEMO = [
    # (code, name, kind, institution, currency, cycle, close, due, card_ccy,
    #  investible, cash_purpose, notes)
    ("cocos_demo",       "Cocos Demo",            "CASH_BROKER",
     "Cocos Capital",    "ARS", "NONE", None, None, None,
     "YES", "OPERATIVO", "[DEMO] Broker AR principal"),
    ("ibkr_demo",        "Interactive Brokers Demo", "CASH_BROKER",
     "Interactive Brokers", "USD", "NONE", None, None, None,
     "YES", "OPERATIVO", "[DEMO] Cuenta US, equities + ETFs"),
    ("binance_demo",     "Binance Demo",          "CASH_WALLET",
     "Binance",          None,  "NONE", None, None, None,
     "YES", "OPERATIVO", "[DEMO] Wallet cripto"),
    ("galicia_demo_ars", "Banco Galicia ARS",     "CASH_BANK",
     "Banco Galicia",    "ARS", "NONE", None, None, None,
     "YES", "OPERATIVO", "[DEMO] Caja ARS"),
    ("galicia_demo_visa","Galicia Visa Demo",     "CARD_CREDIT",
     "Banco Galicia",    None,  "MONTHLY", 28, 10, "ARS",
     "YES", None, "[DEMO] Tarjeta de crédito"),
]

ESPECIES_DEMO = [
    # (ticker, name, asset_class, currency, issuer, sector, country, maturity, notes)
    ("AL30D",  "Bonar 2030 USD MEP",  "BOND_AR",   "USB",
     "Tesoro AR", "GOV",  "AR", "2030-07-09", "[DEMO] Soberano hard-dollar"),
    ("GGAL.BA","Grupo Galicia",       "EQUITY_AR", "ARS",
     "Galicia", "FIN",  "AR", None, "[DEMO] Banco mayor mkt cap"),
    ("AAPL",   "Apple Inc.",          "EQUITY_US", "USD",
     "Apple",   "TECH", "US", None, "[DEMO]"),
    ("SPY",    "SPDR S&P 500 ETF",    "EQUITY_US", "USD",
     "SSGA",    "ETF",  "US", None, "[DEMO] Index ETF"),
    ("TSLA",   "Tesla Inc.",          "EQUITY_US", "USD",
     "Tesla",   "AUTO", "US", None, "[DEMO]"),
    ("BTC",    "Bitcoin",             "CRYPTO",    "USD",
     None,      None,   None, None, "[DEMO]"),
    ("ETH",    "Ethereum",            "CRYPTO",    "USD",
     None,      None,   None, None, "[DEMO]"),
    ("USDT",   "Tether",              "STABLECOIN","USD",
     None,      None,   None, None, "[DEMO]"),
    ("PEDLAR", "Pellegrini Renta Fija","FCI",      "ARS",
     "Pellegrini", "RENTA_FIJA", "AR", None, "[DEMO] FCI ARS"),
]

# Saldos iniciales (fechas en 2026-01-01) → asientos_contables.
# Cada saldo inicial requiere 2 patas: la cuenta gana qty, opening_balance pierde qty.
# Tupla: (event_id, fecha, descripcion, cuenta, activo, qty_signada,
#        unit_price, price_ccy, notes).
def _build_asientos():
    raw = [
        # (cuenta_real, activo, qty) — dimensionados para que los trades
        # del demo dejen saldos positivos cómodos (ver cálculo abajo)
        ("cocos_demo",        "ARS",  8_000_000),  # alcanza GGAL+PEDLAR (≈2.5M)
        ("cocos_demo",        "USB",  120_000),     # alcanza AL30D 1500@65.5 ≈ 98k
        ("ibkr_demo",         "USD",  30_000),      # alcanza AAPL+SPY+TSLA ≈ 23k
        ("binance_demo",      "USD",  12_000),      # alcanza BTC+ETH ≈ 9.9k
        ("galicia_demo_ars",  "ARS",  1_500_000),
    ]
    out = []
    for i, (cuenta, activo, qty) in enumerate(raw, start=1):
        eid = f"DEMO-OPEN-{i:03d}"
        desc = f"[DEMO] Saldo inicial {cuenta} {activo}"
        out.append((eid, date(2026, 1, 1), desc, cuenta, activo, qty,
                     None, None, "[DEMO seed]"))
        out.append((eid, date(2026, 1, 1), desc, "opening_balance", activo,
                     -qty, None, None, "[DEMO seed]"))
    return out

ASIENTOS_DEMO = _build_asientos()

# Trades históricos (sin SELL para mantener posiciones abiertas con target)
TRADES_DEMO = [
    # (trade_id, trade_date, settle_date, cuenta, strategy, ticker, side,
    #  qty, precio, moneda, cuenta_cash, comision, moneda_com,
    #  precio_target, stop_loss, moneda_target, description, notes)

    # AR: AL30D + GGAL.BA en cocos
    ("D0001", date(2026, 1, 15), date(2026, 1, 17),
     "cocos_demo", "BH", "AL30D", "BUY", 1500, 65.50, "USB",
     "cocos_demo", 0, "USB",
     75.0, 60.0, "USB", "BUY AL30D — buy & hold", "[DEMO]"),
    ("D0002", date(2026, 2, 1), date(2026, 2, 5),
     "cocos_demo", "TRADING", "GGAL.BA", "BUY", 200, 6500, "ARS",
     "cocos_demo", 0, "ARS",
     8000, 5500, "ARS", "BUY GGAL.BA", "[DEMO] Pre balance"),

    # US: AAPL + SPY + TSLA en ibkr
    ("D0003", date(2026, 1, 20), date(2026, 1, 22),
     "ibkr_demo", "BH", "AAPL", "BUY", 30, 195.50, "USD",
     "ibkr_demo", 1.0, "USD",
     250.0, 170.0, "USD", "BUY AAPL", "[DEMO] Long term"),
    ("D0004", date(2026, 2, 10), date(2026, 2, 12),
     "ibkr_demo", "BH", "SPY", "BUY", 25, 580.20, "USD",
     "ibkr_demo", 1.0, "USD",
     None, None, None, "BUY SPY", "[DEMO] Index"),
    ("D0005", date(2026, 3, 5), date(2026, 3, 7),
     "ibkr_demo", "TRADING", "TSLA", "BUY", 10, 290.00, "USD",
     "ibkr_demo", 1.0, "USD",
     350.0, 250.0, "USD", "BUY TSLA", "[DEMO]"),

    # CRYPTO: BTC + ETH en binance — usamos USD para evitar dependencia
    # de FX USDT→USD que el demo standalone no tiene
    ("D0006", date(2026, 1, 10), date(2026, 1, 10),
     "binance_demo", "BH", "BTC", "BUY", 0.05, 95000, "USD",
     "binance_demo", 0.5, "USD",
     120000, 80000, "USD", "BUY BTC", "[DEMO]"),
    ("D0007", date(2026, 2, 15), date(2026, 2, 15),
     "binance_demo", "BH", "ETH", "BUY", 1.5, 3400, "USD",
     "binance_demo", 0.3, "USD",
     5000, 2800, "USD", "BUY ETH", "[DEMO]"),

    # FCI
    ("D0008", date(2026, 2, 20), date(2026, 2, 21),
     "cocos_demo", "CASH_PLUS", "PEDLAR", "BUY", 1000, 1200, "ARS",
     "cocos_demo", 0, "ARS",
     None, None, None, "BUY PEDLAR FCI", "[DEMO] Cash equivalent"),

    # Una venta parcial para que haya PnL realizado en el demo
    ("D0009", date(2026, 4, 1), date(2026, 4, 3),
     "ibkr_demo", "TRADING", "TSLA", "SELL", 5, 320.50, "USD",
     "ibkr_demo", 1.0, "USD",
     None, None, None, "SELL TSLA parcial", "[DEMO] Toma de ganancia parcial"),
]

# Ingresos demo (sueldos)
INGRESOS_DEMO = [
    # (fecha, concepto, categoria, monto, moneda, cuenta_destino, descripcion, notes)
    (date(2026, 3, 5),  "Sueldo Marzo",   "Sueldo", 800_000, "ARS",
     "galicia_demo_ars", "Liquidación marzo", "[DEMO]"),
    (date(2026, 4, 5),  "Sueldo Abril",   "Sueldo", 850_000, "ARS",
     "galicia_demo_ars", "Liquidación abril (aumento)", "[DEMO]"),
]

# Gastos demo
GASTOS_DEMO = [
    # (fecha, concepto, monto, moneda, cuenta, categoria, tipo, cuotas, notes)
    (date(2026, 4, 2), "Alquiler",          250_000, "ARS",
     "galicia_demo_ars", "Vivienda", "FIJO", 1, "[DEMO] Mensual"),
    (date(2026, 4, 10), "Supermercado",      85_000, "ARS",
     "galicia_demo_visa", "Comida", "VARIABLE", 1, "[DEMO]"),
    (date(2026, 4, 15), "Suscripción AWS",  50, "USD",
     "ibkr_demo",       "Servicios", "FIJO", 1, "[DEMO] Hosting"),
]


def _set_cell(ws, row, col, val):
    """Helper: escribe una celda. Para fechas usa formato yyyy-mm-dd."""
    c = ws.cell(row=row, column=col, value=val)
    if isinstance(val, date):
        c.number_format = "yyyy-mm-dd"
    return c


def _clear_data_rows(ws, header_row=4):
    """Borra todas las filas de datos (a partir de header_row+1)."""
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)


def _write_rows(ws, rows, headers, header_row=4):
    """Escribe filas debajo de header_row. headers es lista de col-names que
    se mapean a posición. rows es lista de tuplas mismo orden que headers."""
    for i, row in enumerate(rows, start=header_row + 1):
        for j, val in enumerate(row, start=1):
            _set_cell(ws, i, j, val)


def seed_demo(xlsx_path: Path) -> dict:
    """Sobreescribe (o crea) el master en xlsx_path con datos demo.

    Si el master ya existe, primero limpia las hojas relevantes (cuentas,
    especies, blotter, ingresos, gastos, asientos_contables) y reescribe.

    Devuelve un dict con stats de cuántas filas se escribieron.
    """
    from build_master import build_master

    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    # Si no existe, construir master vacío con build_master (template completo)
    if not xlsx_path.is_file():
        build_master(xlsx_path)

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)

    # 1. Cuentas (mantener cuentas técnicas que vienen de build_master + demo)
    ws = wb["cuentas"]
    # Conservar las cuentas técnicas (EXTERNAL, OPENING_BALANCE, etc.)
    # Estrategia: limpiar toda la data y rehacer con técnicas + demo
    HEADER_ROW = 4
    technicas = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        kind = ws.cell(row=r, column=3).value
        if kind in ("EXTERNAL", "OPENING_BALANCE", "INTEREST_EXPENSE",
                    "INTEREST_INCOME"):
            row = tuple(ws.cell(row=r, column=c).value
                         for c in range(1, 13))  # 12 cols
            technicas.append(row)
    _clear_data_rows(ws, HEADER_ROW)
    all_cuentas = list(CUENTAS_DEMO) + technicas
    _write_rows(ws, all_cuentas, [], HEADER_ROW)

    # 2. Especies — limpiar y reescribir solo las demo
    ws = wb["especies"]
    _clear_data_rows(ws, HEADER_ROW)
    _write_rows(ws, ESPECIES_DEMO, [], HEADER_ROW)

    # 3. Blotter — limpiar y reescribir
    ws = wb["blotter"]
    _clear_data_rows(ws, HEADER_ROW)
    _write_rows(ws, TRADES_DEMO, [], HEADER_ROW)

    # 4. Asientos contables (saldos iniciales emparejados con opening_balance)
    if "asientos_contables" in wb.sheetnames:
        ws = wb["asientos_contables"]
        _clear_data_rows(ws, HEADER_ROW)
        # Layout: Event ID | Fecha | Description | Cuenta | Activo | Qty |
        #         Unit Price | Price Currency | Notes
        _write_rows(ws, ASIENTOS_DEMO, [], HEADER_ROW)

    # 5. Ingresos
    if "ingresos" in wb.sheetnames:
        ws = wb["ingresos"]
        _clear_data_rows(ws, HEADER_ROW)
        # Layout: Fecha | Concepto | Categoría | Monto | Moneda | Cuenta | Description | Notes
        _write_rows(ws, INGRESOS_DEMO, [], HEADER_ROW)

    # 6. Gastos
    if "gastos" in wb.sheetnames:
        ws = wb["gastos"]
        _clear_data_rows(ws, HEADER_ROW)
        # Layout coincide con build_master hoja_gastos
        _write_rows(ws, GASTOS_DEMO, [], HEADER_ROW)

    # 7. Margin config: redirigir al ibkr_demo
    if "margin_config" in wb.sheetnames:
        ws = wb["margin_config"]
        _clear_data_rows(ws, HEADER_ROW)
        # (Account, Mult. Overnight, Mult. Intraday, Funding Rate Annual,
        #  Funding Currency, Notes)
        _write_rows(ws, [
            ("ibkr_demo", 2.0, 4.0, 0.06, "USD",
             "[DEMO] RegT estándar — 2x overnight, 4x intraday, ~6% funding USD"),
        ], [], HEADER_ROW)

    # 8. Aforos: limpiar (los demo no usan aforos)
    if "aforos" in wb.sheetnames:
        _clear_data_rows(wb["aforos"], HEADER_ROW)

    # 9. Limpiar otras hojas de eventos para que el demo arranque limpio
    for sheet_name in ("transferencias_cash", "transferencias_activos",
                        "funding", "pasivos", "pagos_pasivos", "recurrentes"):
        if sheet_name in wb.sheetnames:
            _clear_data_rows(wb[sheet_name], HEADER_ROW)

    wb.save(xlsx_path)
    return {
        "cuentas": len(CUENTAS_DEMO),
        "tecnicas_preservadas": len(technicas),
        "especies": len(ESPECIES_DEMO),
        "trades": len(TRADES_DEMO),
        "asientos_rows": len(ASIENTOS_DEMO),  # 2 patas por saldo
        "ingresos": len(INGRESOS_DEMO),
        "gastos": len(GASTOS_DEMO),
    }


def resolve_xlsx_for_user(user_id: str) -> Path:
    """Resuelve inputs/<user>/wealth_management.xlsx."""
    p = HERE / "inputs" / user_id / "wealth_management.xlsx"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def main():
    p = argparse.ArgumentParser(description="Genera master con datos demo fijos")
    p.add_argument("path", nargs="?", help="Path al xlsx (alternativa a --user)")
    p.add_argument("--user", type=str, default=None,
                   help="user_id (resuelve a inputs/<user>/wealth_management.xlsx)")
    p.add_argument("--reset", action="store_true",
                   help="Borra el master existente y rehace desde cero "
                        "(default: preserva otras hojas)")
    args = p.parse_args()

    if args.path:
        xlsx = Path(args.path)
    elif args.user:
        xlsx = resolve_xlsx_for_user(args.user)
    else:
        print("[seed_demo] Pasá un path o --user <id>")
        return 1

    if args.reset and xlsx.is_file():
        print(f"[seed_demo] --reset: borrando {xlsx}")
        xlsx.unlink()

    print(f"[seed_demo] Llenando {xlsx}...")
    stats = seed_demo(xlsx)
    print(f"[seed_demo] OK:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
