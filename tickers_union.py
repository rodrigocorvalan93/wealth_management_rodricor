# -*- coding: utf-8 -*-
"""
tickers_union.py

Escanea todos los Excel masters de todos los users (multi-tenant) y produce
un archivo `data/tickers_union.txt` con el set unión de Tickers de la hoja
`especies` de cada master.

Esto evita pedirles a los loaders que decidan qué tickers fetchear: corren
una sola vez para todos los users.

USO:
    python tickers_union.py                # genera data/tickers_union.txt
    python tickers_union.py --print        # solo imprime, no escribe
    python tickers_union.py --dry-run      # ídem

Layout esperado:
    inputs/<user_id>/wealth_management.xlsx
    data/tickers_union.txt
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError:
    print("[tickers_union] openpyxl no instalado. pip install openpyxl")
    sys.exit(1)


HERE = Path(__file__).resolve().parent
INPUTS_DIR = HERE / "inputs"
DATA_DIR = HERE / "data"
OUTPUT_FILE = DATA_DIR / "tickers_union.txt"
HEADER_ROW = 4  # convención del master


# Asset classes que SÍ van a loaders de precios (los demás se skipean).
# CASH y otros no tienen sentido fetchear.
LOADER_CLASSES = {
    "BOND_AR", "EQUITY_AR", "EQUITY_US", "FCI",
    "CRYPTO", "STABLECOIN",
}


def find_master_files() -> list[Path]:
    """Devuelve todos los wealth_management.xlsx en inputs/<user>/."""
    if not INPUTS_DIR.is_dir():
        return []
    out = []
    # Layout multi-tenant: inputs/<user>/wealth_management.xlsx
    for user_dir in sorted(INPUTS_DIR.iterdir()):
        if not user_dir.is_dir():
            continue
        for fname in ("wealth_management.xlsx",
                       "wealth_management_rodricor.xlsx"):
            f = user_dir / fname
            if f.is_file():
                out.append(f)
                break
    # Back-compat: layout legacy (un solo master en inputs/)
    if not out:
        for fname in ("wealth_management.xlsx",
                       "wealth_management_rodricor.xlsx"):
            f = INPUTS_DIR / fname
            if f.is_file():
                out.append(f)
                break
    return out


def extract_tickers_from_master(xlsx: Path) -> set[tuple[str, str]]:
    """Lee la hoja 'especies' de un master y devuelve {(ticker, asset_class)}."""
    out: set[tuple[str, str]] = set()
    try:
        wb = load_workbook(filename=str(xlsx), data_only=True, read_only=True)
    except Exception as e:
        print(f"[tickers_union] WARN no pude abrir {xlsx}: {e}")
        return out
    if "especies" not in wb.sheetnames:
        return out
    ws = wb["especies"]
    # Headers en HEADER_ROW (default 4)
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v:
            headers[str(v).strip()] = c
    ticker_col = headers.get("Ticker")
    class_col = headers.get("Asset Class")
    if not ticker_col:
        return out
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        ticker = ws.cell(row=r, column=ticker_col).value
        if not ticker:
            continue
        ticker = str(ticker).strip()
        if not ticker:
            continue
        cls = ws.cell(row=r, column=class_col).value if class_col else None
        cls = str(cls).strip() if cls else "OTHER"
        out.add((ticker, cls))
    return out


def union_tickers() -> dict[str, set[str]]:
    """Devuelve {asset_class: {tickers}} agregado de todos los masters."""
    by_class: dict[str, set[str]] = {}
    masters = find_master_files()
    for m in masters:
        for ticker, cls in extract_tickers_from_master(m):
            by_class.setdefault(cls, set()).add(ticker)
    return by_class


def write_union_file(by_class: dict[str, set[str]], path: Path) -> int:
    """Escribe tickers_union.txt con un ticker por línea (todos los classes
    relevantes para loaders). Devuelve cantidad de tickers escritos."""
    flat = set()
    for cls, ts in by_class.items():
        if cls in LOADER_CLASSES:
            flat.update(ts)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = sorted(flat)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return len(lines)


def get_tickers_by_class(target_class: str) -> set[str]:
    """Devuelve solo los tickers de un asset class particular. Útil para
    loaders específicos (ej cripto solo necesita CRYPTO + STABLECOIN)."""
    by_class = union_tickers()
    return by_class.get(target_class, set())


def main():
    p = argparse.ArgumentParser(description="Genera tickers_union.txt desde todos los masters")
    p.add_argument("--print", action="store_true",
                   help="Imprime tickers sin escribir archivo")
    p.add_argument("--dry-run", action="store_true",
                   help="Igual que --print, no escribe")
    p.add_argument("--by-class", action="store_true",
                   help="Imprime agrupado por asset class")
    args = p.parse_args()

    masters = find_master_files()
    print(f"[tickers_union] {len(masters)} master(s) encontrado(s):")
    for m in masters:
        print(f"  - {m.relative_to(HERE) if m.is_relative_to(HERE) else m}")

    by_class = union_tickers()
    total = sum(len(s) for s in by_class.values())
    print(f"[tickers_union] {total} tickers únicos en {len(by_class)} clases")

    if args.by_class:
        for cls in sorted(by_class.keys()):
            print(f"  [{cls}] ({len(by_class[cls])}):")
            for t in sorted(by_class[cls]):
                print(f"    {t}")

    if args.print or args.dry_run:
        flat = set()
        for cls in LOADER_CLASSES:
            flat.update(by_class.get(cls, set()))
        for t in sorted(flat):
            print(t)
        return 0

    n = write_union_file(by_class, OUTPUT_FILE)
    print(f"[tickers_union] OK → {OUTPUT_FILE} ({n} tickers de loaders)")


if __name__ == "__main__":
    sys.exit(main() or 0)
