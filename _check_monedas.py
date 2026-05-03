from openpyxl import load_workbook
wb = load_workbook("inputs/wealth_management_rodricor.xlsx", data_only=True)
ws = wb["monedas"]
print(f"{'Code':<8} {'Name':<25} {'Stable':<8} {'Quote vs':<10} {'Base':<6}")
print("-" * 60)
for r in range(5, ws.max_row + 1):
    code = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=2).value
    stable = ws.cell(row=r, column=3).value
    quote_vs = ws.cell(row=r, column=4).value
    base = ws.cell(row=r, column=5).value
    if code:
        print(f"{code:<8} {str(name):<25} {str(stable):<8} {str(quote_vs):<10} {str(base):<6}")
