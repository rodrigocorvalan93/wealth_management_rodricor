from openpyxl import load_workbook
wb = load_workbook('inputs/wealth_management_rodricor.xlsx')
filas = {
    'blotter':                 [5, 6],
    'transferencias_cash':     [5, 6, 7],
    'funding':                 [5],
    'ingresos':                [5, 6],
    'gastos':                  [5, 6, 7, 8],
    'recurrentes':             [5, 6, 7],
    'asientos_contables':      [5, 6, 7, 8],
}
for hoja, filas_ in filas.items():
    if hoja not in wb.sheetnames: continue
    ws = wb[hoja]
    for f in sorted(filas_, reverse=True):
        ws.delete_rows(f, 1)
wb.save('inputs/wealth_management_rodricor.xlsx')
print('Ejemplos del template borrados')
