from openpyxl import load_workbook
wb = load_workbook('inputs/wealth_management_rodricor.xlsx')

# Limpiar headers de TODAS las hojas (fila 4) y datos también
fixed = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.max_row < 4:
        continue
    # Strip headers en fila 4
    for cell in ws[4]:
        if isinstance(cell.value, str):
            stripped = cell.value.strip()
            if stripped != cell.value:
                cell.value = stripped
                fixed += 1
    # Strip TODOS los valores de texto en filas de datos
    for r in range(5, ws.max_row + 1):
        for cell in ws[r]:
            if isinstance(cell.value, str):
                stripped = cell.value.strip()
                if stripped != cell.value:
                    cell.value = stripped
                    fixed += 1

wb.save('inputs/wealth_management_rodricor.xlsx')
print(f'Headers/values trimmed: {fixed}')
