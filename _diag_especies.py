from openpyxl import load_workbook
wb = load_workbook('inputs/wealth_management_rodricor.xlsx', data_only=True)
ws = wb['especies']
print(f'Total filas en hoja: {ws.max_row}')
print(f'Total columnas: {ws.max_column}')
print()
print('Header (fila 4):')
for c, cell in enumerate(ws[4], start=1):
    print(f'  Col {c}: {cell.value!r}')
print()
print('Primeras 5 filas con datos:')
for r in range(5, min(15, ws.max_row + 1)):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    if any(v is not None for v in row):
        print(f'  Fila {r}: {row}')
