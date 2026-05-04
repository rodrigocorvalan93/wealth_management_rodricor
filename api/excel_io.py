# -*- coding: utf-8 -*-
"""
api/excel_io.py

Lectura y escritura del Excel master con `Row ID` como handle estable.

El Excel master es la fuente de verdad. La DB sqlite es una vista derivada.

Cada hoja de eventos (blotter, gastos, ingresos, ...) tiene una columna
`Row ID` (auto-generada por migrate_master.py o por este módulo cuando se
crea una fila vía API). Los Row IDs son inmutables — si una fila se borra,
la limpiamos pero no reusamos el ID.

Operaciones:
  list_rows(sheet)               — devuelve todas las filas como dicts
  get_row(sheet, row_id)         — devuelve una fila o None
  append_row(sheet, data)        — agrega fila nueva, devuelve Row ID asignado
  update_row(sheet, row_id, data)— modifica campos, devuelve fila actualizada
  delete_row(sheet, row_id)      — limpia los datos de la fila (Row ID queda)

Concurrencia: usar excel_lock() del módulo `state` antes de cualquier write.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# Estilos consistentes con build_master.py / migrate_master.py
NAVY = "1F3864"
BLUE_INPUT = "0000FF"
YELLOW_INPUT = "FFF2CC"
FONT_INPUT = Font(name="Arial", size=11, color=BLUE_INPUT)
FONT_NORMAL = Font(name="Arial", size=11)
FILL_INPUT = PatternFill("solid", fgColor=YELLOW_INPUT)
BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


# Sheet → prefijo de Row ID (auto-generado para hojas de eventos)
SHEET_PREFIX = {
    "blotter": "BL",
    "gastos": "GS",
    "ingresos": "IN",
    "transferencias_cash": "TC",
    "transferencias_activos": "TA",
    "funding": "FN",
    "asientos_contables": "AS",
    "recurrentes": "RC",
    "pagos_pasivos": "PP",
}

# Hojas maestras: usan natural key (no auto-genera Row ID).
# {sheet: column_with_natural_key}
MASTER_SHEETS = {
    "especies": "Ticker",
    "monedas":  "Code",
    "cuentas":  "Code",
    "aforos":   None,  # composite key, ver más abajo (skip por ahora)
    "margin_config": "Account",
}

# Todas las hojas soportadas por la API
ALLOWED_SHEETS = set(SHEET_PREFIX.keys()) | set(
    s for s, k in MASTER_SHEETS.items() if k is not None
)


def is_master_sheet(sheet: str) -> bool:
    """True si la hoja usa natural key en vez de Row ID auto-generado."""
    return sheet in MASTER_SHEETS and MASTER_SHEETS[sheet] is not None


def _key_column_for(sheet: str) -> str:
    """Devuelve la columna que se usa como handle de fila en esta hoja."""
    if is_master_sheet(sheet):
        return MASTER_SHEETS[sheet]
    return "Row ID"


HEADER_ROW = 4  # convención del master


# =============================================================================
# Helpers
# =============================================================================

def _read_headers(ws: Worksheet) -> dict:
    """Devuelve {header_name: column_index_1based} para la fila de headers."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v is not None:
            out[str(v).strip()] = c
    return out


def _last_data_row(ws: Worksheet, header_row: int = HEADER_ROW,
                    last_col_with_header: int = None) -> int:
    """Última fila con datos no-vacíos (en columnas de header)."""
    if last_col_with_header is None:
        headers = _read_headers(ws)
        last_col_with_header = max(headers.values()) if headers else 1
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(1, last_col_with_header + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                last = r
                break
    return last


def _next_row_id(prefix: str, used: set[str]) -> str:
    """Genera el próximo ID (BL-0001, BL-0002, ...) que no esté en `used`."""
    # Encontrar el max numérico actual y sumar 1
    max_n = 0
    pat = re.compile(r"^" + re.escape(prefix) + r"-(\d+)$")
    for s in used:
        m = pat.match(s or "")
        if m:
            max_n = max(max_n, int(m.group(1)))
    n = max_n + 1
    while True:
        candidate = f"{prefix}-{n:04d}"
        if candidate not in used:
            return candidate
        n += 1


def _coerce_value(value: Any) -> Any:
    """Convierte tipos comunes para escribir en Excel.

    - 'YYYY-MM-DD' → date
    - 'YYYY-MM-DDTHH:MM:SS' → datetime
    - 'true'/'false' → bool
    - números como string → float/int
    """
    if value is None:
        return None
    if isinstance(value, (int, float, bool, date, datetime)):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # Date
        if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            try:
                return date.fromisoformat(s)
            except ValueError:
                pass
        # Datetime
        if re.match(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}", s):
            try:
                return datetime.fromisoformat(s)
            except ValueError:
                pass
        return s
    return value


def _serialize_value(value: Any) -> Any:
    """Serializa un valor de cell para JSON."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


# =============================================================================
# Operaciones
# =============================================================================

def list_rows(xlsx_path: Path, sheet: str) -> list[dict]:
    """Lista todas las filas de una hoja como dicts.

    Para hojas de eventos: incluye 'row_id' tomado de la columna 'Row ID'.
    Para hojas maestras: incluye 'row_id' tomado de la columna natural
    (Ticker para especies, Code para monedas/cuentas, etc).

    Filas completamente vacías se filtran.
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = _read_headers(ws)
    if not headers:
        return []

    last_col = max(headers.values())
    last_row = _last_data_row(ws, last_col_with_header=last_col)
    key_col_name = _key_column_for(sheet)

    out = []
    for r in range(HEADER_ROW + 1, last_row + 1):
        row_dict = {}
        all_empty = True
        for h, c in headers.items():
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                all_empty = False
            row_dict[h] = _serialize_value(v)
        if all_empty:
            continue
        # Normalizar: row_id como key principal (puede ser "Row ID" o natural)
        rid = row_dict.get(key_col_name) or row_dict.get("row_id")
        if rid is not None:
            row_dict["row_id"] = str(rid).strip()
        # Excel row number (útil para debug, no para identificar)
        row_dict["_excel_row"] = r
        out.append(row_dict)
    return out


def get_row(xlsx_path: Path, sheet: str, row_id: str) -> Optional[dict]:
    """Devuelve la fila con ese ID o None.

    Para hojas maestras, `row_id` es el natural key (ej: ticker AL30D).
    """
    rows = list_rows(xlsx_path, sheet)
    target = str(row_id).strip()
    for row in rows:
        if str(row.get("row_id") or "").strip() == target:
            return row
    return None


def _used_row_ids(ws: Worksheet, headers: dict) -> set[str]:
    """Devuelve el set de Row IDs usados en la hoja."""
    if "Row ID" not in headers:
        return set()
    col = headers["Row ID"]
    out = set()
    last_col = max(headers.values())
    last_row = _last_data_row(ws, last_col_with_header=last_col)
    for r in range(HEADER_ROW + 1, last_row + 1):
        v = ws.cell(row=r, column=col).value
        if v:
            out.add(str(v).strip())
    return out


def _ensure_row_id_column(ws: Worksheet, headers: dict) -> int:
    """Si la columna Row ID no existe, la crea al final. Devuelve el índice."""
    if "Row ID" in headers:
        return headers["Row ID"]
    last_col = max(headers.values()) if headers else 0
    new_col = last_col + 1
    cell = ws.cell(row=HEADER_ROW, column=new_col, value="Row ID")
    cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER_THIN
    ws.column_dimensions[get_column_letter(new_col)].width = 12
    headers["Row ID"] = new_col
    return new_col


def append_row(xlsx_path: Path, sheet: str, data: dict) -> str:
    """Agrega una fila nueva. Devuelve el ID asignado.

    Para hojas de eventos: auto-genera Row ID (BL-XXXX, GS-XXXX, etc).
    Para hojas maestras: requiere que `data` contenga el natural key
    (Ticker / Code / Account según corresponda) y falla si ya existe.
    """
    if sheet not in ALLOWED_SHEETS:
        raise ValueError(f"Sheet '{sheet}' no soportada")

    wb = load_workbook(filename=str(xlsx_path))
    if sheet not in wb.sheetnames:
        raise ValueError(f"Hoja '{sheet}' no existe en {xlsx_path}")
    ws = wb[sheet]
    headers = _read_headers(ws)
    if not headers:
        raise ValueError(f"Hoja '{sheet}' sin headers")

    if is_master_sheet(sheet):
        # MASTER: usar natural key del data, no auto-genera
        key_col_name = _key_column_for(sheet)
        natural_key = data.get(key_col_name)
        if not natural_key:
            raise ValueError(
                f"Para crear en '{sheet}' tenés que enviar '{key_col_name}'"
            )
        natural_key = str(natural_key).strip()
        # Validar que no esté duplicado
        existing = get_row(xlsx_path, sheet, natural_key)
        if existing is not None:
            raise ValueError(
                f"Ya existe '{key_col_name}={natural_key}' en {sheet}. "
                f"Usá PUT /api/sheets/{sheet}/{natural_key} para modificar."
            )
        last_col = max(headers.values())
        last_row = _last_data_row(ws, last_col_with_header=last_col)
        new_row = last_row + 1
        for h, c in headers.items():
            if h in data:
                val = _coerce_value(data[h])
                if val is None and h == key_col_name:
                    continue  # ya validado arriba
                cell = ws.cell(row=new_row, column=c, value=val)
                cell.font = FONT_INPUT
                cell.fill = FILL_INPUT
                cell.border = BORDER_THIN
                if isinstance(val, (date, datetime)):
                    cell.number_format = "yyyy-mm-dd"
                elif isinstance(val, (int, float)):
                    cell.number_format = '#,##0.0000;[Red](#,##0.0000)'
        wb.save(str(xlsx_path))
        return natural_key

    # EVENT: auto-generar Row ID
    if sheet not in SHEET_PREFIX:
        raise ValueError(f"Sheet '{sheet}' no soportada para append (evento)")
    prefix = SHEET_PREFIX[sheet]
    row_id_col = _ensure_row_id_column(ws, headers)
    used = _used_row_ids(ws, headers)
    new_id = _next_row_id(prefix, used)

    last_col = max(headers.values())
    last_row = _last_data_row(ws, last_col_with_header=last_col)
    new_row = last_row + 1

    for h, c in headers.items():
        if h == "Row ID":
            cell = ws.cell(row=new_row, column=c, value=new_id)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            continue
        if h in data:
            val = _coerce_value(data[h])
            cell = ws.cell(row=new_row, column=c, value=val)
            cell.font = FONT_INPUT
            cell.fill = FILL_INPUT
            cell.border = BORDER_THIN
            if isinstance(val, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(val, (int, float)):
                cell.number_format = '#,##0.0000;[Red](#,##0.0000)'

    wb.save(str(xlsx_path))
    return new_id


def update_row(xlsx_path: Path, sheet: str, row_id: str, data: dict) -> dict:
    """Modifica los campos de una fila por ID. Devuelve la fila actualizada.

    Para hojas maestras: la columna del natural key NO se puede modificar
    desde update (filtramos data). Si querés cambiar el key, borrá y creá.
    """
    wb = load_workbook(filename=str(xlsx_path))
    if sheet not in wb.sheetnames:
        raise ValueError(f"Hoja '{sheet}' no existe")
    ws = wb[sheet]
    headers = _read_headers(ws)

    key_col_name = _key_column_for(sheet)
    if key_col_name not in headers:
        raise ValueError(f"Hoja '{sheet}' no tiene columna '{key_col_name}'")

    key_col = headers[key_col_name]
    last_col = max(headers.values())
    last_row = _last_data_row(ws, last_col_with_header=last_col)

    target_row = None
    target_key = str(row_id).strip()
    for r in range(HEADER_ROW + 1, last_row + 1):
        if str(ws.cell(row=r, column=key_col).value or "").strip() == target_key:
            target_row = r
            break
    if target_row is None:
        raise KeyError(f"'{row_id}' no encontrado en hoja '{sheet}'")

    for h, c in headers.items():
        if h == key_col_name:
            continue  # natural key / Row ID inmutables vía update
        if h in data:
            val = _coerce_value(data[h])
            cell = ws.cell(row=target_row, column=c, value=val)
            cell.font = FONT_INPUT
            cell.fill = FILL_INPUT
            cell.border = BORDER_THIN
            if isinstance(val, (date, datetime)):
                cell.number_format = "yyyy-mm-dd"
            elif isinstance(val, (int, float)):
                cell.number_format = '#,##0.0000;[Red](#,##0.0000)'

    wb.save(str(xlsx_path))
    return get_row(xlsx_path, sheet, row_id)


def delete_row(xlsx_path: Path, sheet: str, row_id: str) -> bool:
    """Borra una fila.

    - Hojas de eventos: SOFT delete (limpia campos, preserva Row ID tombstone)
      para mantener estabilidad de IDs subsiguientes.
    - Hojas maestras: HARD delete (remueve la fila entera). El natural key
      no debe quedar como tombstone porque colisionaría con futuros agregados.

    Devuelve True si encontró y borró la fila.
    """
    wb = load_workbook(filename=str(xlsx_path))
    if sheet not in wb.sheetnames:
        return False
    ws = wb[sheet]
    headers = _read_headers(ws)

    key_col_name = _key_column_for(sheet)
    if key_col_name not in headers:
        return False

    key_col = headers[key_col_name]
    last_col = max(headers.values())
    last_row = _last_data_row(ws, last_col_with_header=last_col)

    target_row = None
    target_key = str(row_id).strip()
    for r in range(HEADER_ROW + 1, last_row + 1):
        if str(ws.cell(row=r, column=key_col).value or "").strip() == target_key:
            target_row = r
            break
    if target_row is None:
        return False

    if is_master_sheet(sheet):
        # Hard delete: remueve la fila físicamente
        ws.delete_rows(target_row, 1)
    else:
        # Soft delete: limpia todos los campos excepto Row ID (tombstone)
        # IMPORTANTE: openpyxl ignora `ws.cell(..., value=None)` cuando ya hay
        # valor — hay que asignar via `cell.value = None`.
        for h, c in headers.items():
            if h == key_col_name:
                continue
            ws.cell(row=target_row, column=c).value = None

    wb.save(str(xlsx_path))
    return True


# =============================================================================
# Reverse mapping: para mostrar event_id de la DB junto al Row ID del Excel
# =============================================================================

def excel_row_to_event(conn, sheet: str, excel_row: int) -> Optional[dict]:
    """Encuentra el event_id en la DB que vino de (sheet, excel_row).

    Útil para enlazar la vista del Excel con resultados del engine.
    """
    cur = conn.execute(
        """SELECT event_id, event_type, event_date, description
           FROM events
           WHERE source_sheet = ? AND source_row = ?
           LIMIT 1""",
        (sheet, excel_row),
    )
    row = cur.fetchone()
    if not row:
        return None
    return dict(row)
