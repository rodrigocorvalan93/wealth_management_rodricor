# -*- coding: utf-8 -*-
"""
api/app.py

Backend Flask del wealth_management. Wraps el engine y expone:

  AUTH:        Bearer <WM_API_TOKEN> en header Authorization.

  GET    /api/health                       — ping (no requiere auth)
  GET    /api/summary                      — resumen ejecutivo del portfolio
  GET    /api/holdings                     — lista de holdings actuales
  GET    /api/equity-curve                 — serie temporal del PN
  GET    /api/buying-power                 — poder de compra por cuenta
  GET    /api/trade-stats                  — métricas de trading
  GET    /api/realized-pnl                 — fills de PnL realizado FIFO

  GET    /api/sheets/<sheet>               — lista filas de una hoja
  GET    /api/sheets/<sheet>/<row_id>      — una fila específica
  POST   /api/sheets/<sheet>               — agrega fila (devuelve row_id)
  PUT    /api/sheets/<sheet>/<row_id>      — modifica fila
  DELETE /api/sheets/<sheet>/<row_id>      — borra fila (soft, deja Row ID)

  POST   /api/upload/excel                 — sube nuevo Excel master (sobreescribe)
  GET    /api/download/excel               — baja el Excel master actual
  POST   /api/upload/prices                — sube CSV de precios al data dir
  POST   /api/refresh                      — fuerza re-import del Excel
  GET    /api/backups                      — lista los backups del Excel
  GET    /api/report/html                  — HTML autocontenido del reporte
  GET    /api/report/excel                 — Excel multi-sheet del reporte
  GET    /api/config                       — info de paths y config

USO LOCAL:
  WM_API_TOKEN=test python -m flask --app api.app run
"""

from __future__ import annotations

import io
import os
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

# Permitir importar engine/ aunque corra como script
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from flask import (
    Flask, request, jsonify, send_file, send_from_directory,
    abort, Response, render_template, g,
)

from engine.holdings import (
    calculate_holdings, total_pn, by_asset_class, by_account, by_currency,
    by_cash_purpose, filter_investible, filter_non_investible,
)
from engine.pnl import calculate_realized_pnl, total_realized_pnl
from engine.trade_stats import (
    calculate_trade_stats, trade_stats_by_asset, trade_stats_by_account,
)
from engine.snapshots import (
    record_snapshots, get_equity_curve, get_equity_curves_by_account,
    calculate_returns,
)
from engine.buying_power import buying_power_summary
from engine.exporter import export_excel, export_html

from .state import (
    get_settings, get_user_settings, list_user_ids,
    db_conn, excel_write_lock, backup_excel,
    list_backups, prune_backups, reimport_excel,
    DEFAULT_USER_ID,
)
from . import credentials as creds
from . import audit, ratelimit, jobs as bg_jobs
from . import auth as auth_mod
from .users import (
    load_users, resolve_user_by_token, get_active_user,
    admin_switch_to, admin_switch_clear, is_switched,
    add_user_to_config, remove_user_from_config, export_users_json,
    is_persistent, UserConfig,
)
from .excel_io import (
    SHEET_PREFIX, MASTER_SHEETS, ALLOWED_SHEETS, is_master_sheet,
    list_rows, get_row, append_row, update_row, delete_row,
)


def create_app() -> Flask:
    app = Flask(__name__,
                static_folder="static",
                template_folder="templates",
                static_url_path="/static")

    # Limite de tamaño de upload (50 MB). Flask aborta con 413 si se excede.
    # Aplica a todos los endpoints. Excel masters reales pesan <500 KB y
    # CSVs típicos <100 KB, así que 50 MB es generoso.
    app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

    @app.errorhandler(413)
    def upload_too_large(_e):
        from flask import jsonify
        return jsonify({"error": True, "code": 413,
                         "message": "Archivo demasiado grande (máx 50 MB)"}), 413

    # --- PWA: shell HTML ---
    @app.route("/")
    def root():
        return render_template("pwa.html")

    # Service worker debe servirse desde la raíz (no /static/) para tener
    # scope completo. Lo redirigimos.
    @app.route("/sw.js")
    def service_worker():
        resp = send_from_directory(app.static_folder, "sw.js")
        resp.headers["Service-Worker-Allowed"] = "/"
        resp.headers["Cache-Control"] = "no-cache"
        return resp

    # --- CORS abierto (single user, token-protected) ---
    @app.after_request
    def add_cors(resp):
        resp.headers["Access-Control-Allow-Origin"] = "*"
        resp.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
        resp.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        return resp

    @app.route("/api/<path:_>", methods=["OPTIONS"])
    def cors_preflight(_):
        return ("", 204)

    # --- Auth + user resolution (multi-tenant) ---

    # Endpoints de auth que NO requieren auth previa
    AUTH_PUBLIC_PATHS = {
        "/api/auth/signup", "/api/auth/login",
        "/api/auth/forgot-password", "/api/auth/reset-password",
        "/api/auth/verify-email",
    }

    @app.before_request
    def _resolve_user():
        """Resuelve el user activo desde el bearer token y lo setea en g.

        Acepta dos tipos de token:
          1. Session token de auth_db (preferido — emitido al login con
             email/password). Resuelve via api/auth.resolve_session.
          2. Legacy bearer token de WM_USERS_JSON / WM_API_TOKEN.

        - g.auth_user: el user dueño del token (UserConfig-like)
        - g.active_user_id: user_id efectivamente activo
        - g.is_admin: True si el auth_user es admin
        - g.is_superadmin: True si el auth_user es superadmin
        - g.is_switched: True si el admin está viendo datos de otro
        - g.auth_via: 'session' | 'legacy' | None
        """
        # Defaults siempre presentes para evitar AttributeError
        g.auth_user = None
        g.active_user_id = None
        g.is_admin = False
        g.is_superadmin = False
        g.is_switched = False
        g.user_token = None
        g.auth_via = None
        g.auth_email = None

        if not request.path.startswith("/api/"):
            return
        if request.method == "OPTIONS":
            return
        if request.path == "/api/health":
            return
        if request.path in AUTH_PUBLIC_PATHS:
            return

        auth = request.headers.get("Authorization", "")
        if not auth.startswith("Bearer "):
            return
        token = auth[len("Bearer "):].strip()
        g.user_token = token

        # 1) Session token (auth_db)
        try:
            session_info = auth_mod.resolve_session(token)
        except Exception:
            session_info = None
        if session_info:
            class _U:
                pass
            u = _U()
            u.user_id = session_info["user_id"]
            u.display_name = session_info["display_name"]
            u.is_admin = session_info["is_admin"] or session_info["is_superadmin"]
            u.token = token
            g.auth_user = u
            g.active_user_id = session_info["user_id"]
            g.is_admin = bool(session_info["is_admin"]
                                or session_info["is_superadmin"])
            g.is_superadmin = bool(session_info["is_superadmin"])
            g.auth_email = session_info.get("email")
            g.auth_email_verified = bool(session_info.get("email_verified"))
            g.auth_via = "session"
            # Switch state (compat) — superadmin puede usar el switch viejo
            from .users import is_switched as _is_switched, _switched
            if g.is_superadmin and token in _switched:
                target_id = _switched[token]
                # Validar que ese user_id realmente exista en alguno de los
                # backends
                target_legacy = next(
                    (lu for lu in load_users() if lu.user_id == target_id), None
                )
                if target_legacy:
                    g.active_user_id = target_legacy.user_id
                    g.is_switched = True
                else:
                    # Switch a un user de auth_db
                    p = auth_mod.get_user_profile(target_id)
                    if p:
                        g.active_user_id = p["user_id"]
                        g.is_switched = True
            return

        # 2) Legacy bearer token (WM_USERS_JSON / WM_API_TOKEN)
        active = get_active_user(token)
        if active is None:
            return
        auth_user, switched = active
        actual_token_owner = resolve_user_by_token(token)
        g.auth_user = actual_token_owner
        g.active_user_id = auth_user.user_id
        g.is_admin = bool(actual_token_owner and actual_token_owner.is_admin)
        # En legacy, admin es de facto superadmin (back-compat)
        g.is_superadmin = bool(actual_token_owner and actual_token_owner.is_admin)
        g.is_switched = switched
        g.auth_via = "legacy"

    @app.before_request
    def _rate_limit():
        """Rate limit simple por token. Mutations: 60/min. Reads: 240/min.
        Sin token: 30/min por IP (anti-bruteforce de login).

        Disable con WM_DISABLE_RATELIMIT=1 (útil en tests).
        """
        if os.environ.get("WM_DISABLE_RATELIMIT") == "1":
            return
        if not request.path.startswith("/api/"):
            return
        if request.method == "OPTIONS":
            return
        if request.path == "/api/health":
            return
        token = getattr(g, "user_token", None) or ""
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()
        key = token if token else f"ip:{ip}"
        is_mutation = request.method in ("POST", "PUT", "DELETE", "PATCH")
        if not token:
            limit, window = 30, 60   # anti-bruteforce
            action = "anon"
        elif is_mutation:
            limit, window = 60, 60
            action = "write"
        else:
            limit, window = 240, 60
            action = "read"
        ok, remaining = ratelimit.check(key, action, limit, window)
        if not ok:
            abort(429, f"Rate limit excedido ({limit} {action}/{window}s). "
                       f"Esperá un momento.")

    @app.after_request
    def _audit_request(resp):
        """Audit log para mutations. Best-effort, no rompe la response."""
        try:
            if (request.method in ("POST", "PUT", "DELETE", "PATCH")
                    and request.path.startswith("/api/")
                    and request.path != "/api/health"):
                user_id = getattr(g, "active_user_id", None)
                if user_id:
                    body = None
                    if request.is_json:
                        try:
                            body = request.get_json(silent=True)
                        except Exception:
                            body = None
                    audit.log(user_id, {
                        "ts": datetime.now().isoformat(timespec="seconds"),
                        "ip": request.headers.get("X-Forwarded-For",
                                                   request.remote_addr or ""),
                        "user_id": user_id,
                        "auth_user_id": (g.auth_user.user_id if g.auth_user
                                          else None),
                        "is_switched": bool(getattr(g, "is_switched", False)),
                        "method": request.method,
                        "path": request.path,
                        "status": resp.status_code,
                        "body_hash": audit.hash_body(body),
                    })
        except Exception as e:
            print(f"[audit] WARN: {e}")
        return resp

    def _require_auth():
        """Verifica que haya user resuelto en g."""
        if not getattr(g, "active_user_id", None):
            abort(401, "Token inválido o ausente")

    def _require_admin():
        """Verifica que el caller sea admin (independiente de switch state)."""
        _require_auth()
        if not getattr(g, "is_admin", False):
            abort(403, "Acción solo para admin")

    def _require_superadmin():
        """Solo superadmins (o admin legacy)."""
        _require_auth()
        if not getattr(g, "is_superadmin", False):
            abort(403, "Acción solo para superadmin")

    def _block_if_switched_mutation():
        """En vista 'switched', el admin no puede mutar datos del target user.
        Aborta 403 si la request es POST/PUT/DELETE y está switched."""
        if request.method in ("POST", "PUT", "DELETE", "PATCH"):
            if getattr(g, "is_switched", False):
                abort(403,
                      "Estás en modo 'switch user' (read-only). Volvé a tu user "
                      "para mutar datos.")

    # --- Helpers ---
    def _holdings(fecha=None, anchor=None):
        s = get_settings()
        anchor = (anchor or s.anchor).upper()
        if fecha is None:
            fecha = date.today()
        if isinstance(fecha, str):
            fecha = date.fromisoformat(fecha)
        conn = db_conn()
        try:
            return calculate_holdings(conn, fecha=fecha, anchor_currency=anchor), conn
        except Exception:
            conn.close()
            raise

    def _parse_query_fecha():
        f = request.args.get("fecha")
        return date.fromisoformat(f) if f else date.today()

    def _parse_query_anchor():
        return (request.args.get("anchor") or get_settings().anchor).upper()

    # =========================================================================
    # PUBLIC
    # =========================================================================

    @app.get("/api/health")
    def health():
        import os as _os
        users = load_users()
        is_multi = bool(_os.environ.get("WM_USERS_JSON"))
        # Back-compat: si single-tenant, exponer flags del user "default"
        # para que clientes viejos sigan funcionando.
        body = {
            "status": "ok",
            "version": "2.0",
            "auth_configured": len(users) > 0,
            "n_users": len(users),
            "multi_tenant": is_multi,
            "now": datetime.now().isoformat(),
        }
        if not is_multi and users:
            s = get_user_settings(users[0].user_id)
            body["xlsx_present"] = s.xlsx_path.is_file()
            body["db_present"] = s.db_path.is_file()
            body["anchor_default"] = s.anchor
        return jsonify(body)

    @app.get("/api/config")
    def config():
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        return jsonify({
            "user_id": g.active_user_id,
            "is_admin": g.is_admin,
            "is_superadmin": getattr(g, "is_superadmin", False),
            "is_switched": g.is_switched,
            "auth_user_id": g.auth_user.user_id if g.auth_user else None,
            "auth_user_display_name": g.auth_user.display_name if g.auth_user else None,
            "xlsx_path": str(s.xlsx_path),
            "xlsx_present": s.xlsx_path.is_file(),
            "db_present": s.db_path.is_file(),
            "anchor": s.anchor,
            "supported_sheets": sorted(ALLOWED_SHEETS),
            "event_sheets": list(SHEET_PREFIX.keys()),
            "master_sheets": {k: v for k, v in MASTER_SHEETS.items() if v is not None},
        })

    # =========================================================================
    # READ: portfolio analytics
    # =========================================================================

    @app.get("/api/summary")
    def summary():
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        holdings, conn = _holdings(fecha, anchor)
        try:
            tp = total_pn(holdings, anchor)
            return jsonify({
                "fecha": fecha.isoformat(),
                "anchor": anchor,
                "patrimonio_total": tp["total_anchor"],            # NETO
                "patrimonio_invertible": tp["total_investible"],
                "patrimonio_no_invertible": tp["total_non_investible"],
                "total_assets": tp["total_assets"],
                "total_liabilities": tp["total_liabilities"],
                "unconverted_count": tp["total_unconverted_count"],
                "by_asset_class": by_asset_class(holdings),
                "by_account": by_account(holdings),
                "by_currency": by_currency(holdings),
                "by_cash_purpose": by_cash_purpose(holdings),
                "n_positions": len(holdings),
            })
        finally:
            conn.close()

    @app.get("/api/holdings")
    def holdings_endpoint():
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        holdings, conn = _holdings(fecha, anchor)
        try:
            # Filtros opcionales
            account = request.args.get("account")
            asset_class = request.args.get("asset_class")
            investible_only = request.args.get("investible") == "true"
            data = holdings
            if investible_only:
                data = filter_investible(data)
            if account:
                data = [h for h in data if h["account"] == account]
            if asset_class:
                data = [h for h in data if h["asset_class"] == asset_class]
            # Limpiar valores no JSON-friendly
            return jsonify({
                "fecha": fecha.isoformat(),
                "anchor": anchor,
                "n": len(data),
                "items": [_clean_for_json(h) for h in data],
            })
        finally:
            conn.close()

    @app.get("/api/equity-curve")
    def equity_curve():
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        investible_only = request.args.get("investible") == "true"
        per_account = request.args.get("per_account") == "true"
        conn = db_conn()
        try:
            total = get_equity_curve(conn, anchor_currency=anchor,
                                      investible_only=investible_only)
            payload = {
                "anchor": anchor,
                "investible_only": investible_only,
                "total": total,
                "metrics": calculate_returns(total),
            }
            if per_account:
                by_acc = get_equity_curves_by_account(
                    conn, anchor_currency=anchor,
                    investible_only=investible_only,
                )
                payload["by_account"] = by_acc
            return jsonify(payload)
        finally:
            conn.close()

    @app.get("/api/asset/<path:ticker>/history")
    def asset_history_endpoint(ticker):
        """Detalle histórico de un activo: primera compra, todas las
        operaciones, evolución de precio + qty + market value, y métricas
        de retorno desde la incorporación.

        Query: ?account=cocos (opcional, filtra por cuenta)
        """
        _require_auth(); _block_if_switched_mutation()
        from engine.prices import get_price
        from engine.fx import convert as fx_convert, FxError
        anchor = _parse_query_anchor()
        account_filter = request.args.get("account")
        conn = db_conn()
        try:
            # 1) Movimientos del activo (ordenados ascendente)
            params = [ticker]
            sql = """
                SELECT m.movement_id, m.account, m.qty, m.unit_price,
                       m.price_currency, e.event_date, e.event_type, e.description
                FROM movements m
                JOIN events e ON e.event_id = m.event_id
                WHERE m.asset = ?
            """
            if account_filter:
                sql += " AND m.account = ?"
                params.append(account_filter)
            sql += " ORDER BY e.event_date ASC, m.movement_id ASC"
            cur = conn.execute(sql, params)
            movements = [dict(r) for r in cur.fetchall()]
            if not movements:
                return jsonify({"error": "asset sin movimientos"}), 404

            # 2) Datos del asset (currency, asset_class, name)
            cur = conn.execute(
                "SELECT name, asset_class, currency FROM assets WHERE ticker=?",
                (ticker,),
            )
            row = cur.fetchone()
            asset_info = dict(row) if row else {"name": ticker, "asset_class": "?",
                                                 "currency": None}

            first_date = movements[0]["event_date"][:10]

            # 3) Posición actual (qty + avg_cost) — sin filtrar account si
            #    queremos toda la posición consolidada
            qty = 0.0
            cost_acum = 0.0
            qty_buys = 0.0
            for m in movements:
                qty += m["qty"]
                if m["qty"] > 0 and m["unit_price"] is not None:
                    cost_acum += m["qty"] * m["unit_price"]
                    qty_buys += m["qty"]
            avg_cost = (cost_acum / qty_buys) if qty_buys > 0 else None

            # 4) Serie de precios desde first_date hasta hoy
            cur = conn.execute(
                """SELECT fecha, price, currency FROM prices
                   WHERE ticker=? AND fecha >= ?
                   ORDER BY fecha ASC""",
                (ticker, first_date),
            )
            prices = [{"fecha": r["fecha"], "price": r["price"],
                        "currency": r["currency"]} for r in cur.fetchall()]

            # 5) Construir evolución: para cada precio, calcular qty_held,
            #    market_value, unrealized_pnl
            qty_by_date = []
            running_qty = 0.0
            for m in movements:
                running_qty += m["qty"]
                qty_by_date.append((m["event_date"][:10], running_qty))

            def qty_at(fecha_iso):
                # qty acumulada al cierre de fecha_iso (incluye movimientos de ese día)
                q = 0.0
                for d, rq in qty_by_date:
                    if d <= fecha_iso:
                        q = rq
                    else:
                        break
                return q

            evolution = []
            for p in prices:
                q = qty_at(p["fecha"])
                if q == 0 and not evolution:
                    continue  # skip pre-position prices
                mv_native = q * p["price"]
                evolution.append({
                    "fecha": p["fecha"],
                    "price": p["price"],
                    "currency": p["currency"],
                    "qty": q,
                    "mv_native": mv_native,
                })

            # 6) Métricas: precio inicial (avg_cost), precio actual, %
            current_price = prices[-1]["price"] if prices else (avg_cost or 0)
            return_pct = None
            if avg_cost and avg_cost > 0:
                return_pct = (current_price - avg_cost) / avg_cost
            unrealized_native = (
                (current_price - (avg_cost or current_price)) * qty
                if qty != 0 else 0.0
            )

            # 7) PnL realizado del activo (desde fills)
            from engine.pnl import calculate_realized_pnl
            fills = [f for f in calculate_realized_pnl(conn) if f.asset == ticker]
            realized_pnl_total = sum(f.pnl_realizado for f in fills)
            realized_currency = fills[0].currency if fills else asset_info["currency"]

            # 8) Días desde primera compra
            try:
                from datetime import date as _date
                d0 = _date.fromisoformat(first_date)
                days_held = (_date.today() - d0).days
            except Exception:
                days_held = None

            return jsonify({
                "ticker": ticker,
                "name": asset_info["name"],
                "asset_class": asset_info["asset_class"],
                "native_currency": asset_info["currency"],
                "first_purchase_date": first_date,
                "days_held": days_held,
                "current_qty": qty,
                "avg_cost": avg_cost,
                "current_price": current_price,
                "unrealized_pnl_native": unrealized_native,
                "return_pct": return_pct,
                "realized_pnl_total": realized_pnl_total,
                "realized_currency": realized_currency,
                "n_trades": sum(1 for m in movements if m["qty"] != 0),
                "movements": [
                    {
                        "fecha": m["event_date"][:10],
                        "account": m["account"],
                        "qty": m["qty"],
                        "unit_price": m["unit_price"],
                        "currency": m["price_currency"],
                        "event_type": m["event_type"],
                        "description": m["description"],
                    }
                    for m in movements
                ],
                "evolution": evolution,
            })
        finally:
            conn.close()

    @app.get("/api/holdings-near-target")
    def holdings_near_target_endpoint():
        """Holdings cuya distancia al target o stop-loss está dentro del
        umbral configurado en `alert_distance_bps`. Query param `bps`
        opcional para override on-the-fly."""
        _require_auth(); _block_if_switched_mutation()
        from engine.holdings import filter_near_target
        from engine.schema import get_setting
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        holdings, conn = _holdings(fecha, anchor)
        try:
            override = request.args.get("bps")
            if override is not None:
                try:
                    bps = float(override)
                except ValueError:
                    abort(400, "bps debe ser numérico")
            else:
                bps = get_setting(conn, "alert_distance_bps",
                                   default=10.0, cast=float)
            alerts = filter_near_target(holdings, alert_distance_bps=bps)
            return jsonify({
                "alert_distance_bps": bps,
                "fecha": fecha.isoformat(),
                "anchor": anchor,
                "n_alerts": len(alerts),
                "alerts": alerts,
            })
        finally:
            conn.close()

    @app.get("/api/settings")
    def get_settings_endpoint():
        """Devuelve los settings persistidos en DB (key/value)."""
        _require_auth()
        from engine.schema import get_setting
        conn = db_conn()
        try:
            return jsonify({
                "alert_distance_bps": get_setting(conn, "alert_distance_bps",
                                                    default=10.0, cast=float),
                "anchor_currency": get_setting(conn, "anchor_currency",
                                                default=get_settings().anchor),
                "pnl_method": get_setting(conn, "pnl_method", default="FIFO"),
            })
        finally:
            conn.close()

    @app.put("/api/settings")
    def update_settings_endpoint():
        """Actualiza un setting. Body: {"key": "alert_distance_bps", "value": 50}.

        IMPORTANTE: el valor se persiste a la DB pero NO al Excel master,
        así que se va a perder en el próximo refresh. Para hacerlo permanente
        editá la hoja `config` del master.
        """
        _require_auth(); _block_if_switched_mutation()
        from engine.schema import set_setting
        body = request.get_json(silent=True) or {}
        key = body.get("key")
        value = body.get("value")
        if not key:
            abort(400, "key requerido")
        # Whitelist de keys editables
        allowed = {"alert_distance_bps"}
        if key not in allowed:
            abort(400, f"key '{key}' no editable. Permitidos: {sorted(allowed)}")
        conn = db_conn()
        try:
            set_setting(conn, key, value)
            conn.commit()
            return jsonify({"ok": True, "key": key, "value": value})
        finally:
            conn.close()

    @app.get("/api/buying-power")
    def buying_power_endpoint():
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        holdings, conn = _holdings(fecha, anchor)
        try:
            summary = buying_power_summary(conn, holdings, anchor)
            out = []
            for item in summary:
                if item["type"] == "BYMA":
                    out.append({
                        "account": item["account"],
                        "type": "BYMA",
                        "result": item["result"].to_dict(),
                    })
                else:
                    out.append({
                        "account": item["account"],
                        "type": "MARGIN",
                        "overnight": item["overnight"].to_dict(),
                        "intraday": item["intraday"].to_dict(),
                    })
            return jsonify({"anchor": anchor, "items": out})
        finally:
            conn.close()

    @app.get("/api/trade-stats")
    def trade_stats_endpoint():
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        conn = db_conn()
        try:
            fills = calculate_realized_pnl(conn, fecha_hasta=fecha)
            stats = calculate_trade_stats(fills)
            stats_dict = {ccy: s.to_dict() for ccy, s in stats.items()}
            # Sanear inf
            for s in stats_dict.values():
                if s["profit_factor"] == float("inf"):
                    s["profit_factor"] = None
            return jsonify({
                "fecha": fecha.isoformat(),
                "n_fills": len(fills),
                "by_currency": stats_dict,
                "by_asset": trade_stats_by_asset(fills),
                "by_account": trade_stats_by_account(fills),
            })
        finally:
            conn.close()

    @app.get("/api/asset-performance")
    def asset_performance_endpoint():
        """Tabla de retorno-desde-compra por holding actual.

        Para cada (account, asset) abierto: fecha primera compra, días held,
        avg_cost, precio actual, return % en moneda nativa, MV en ancla.
        Útil para "cómo viene rindiendo cada activo desde que lo incorporé".

        Query: ?investible=true para excluir cuentas no-invertibles.
        """
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        investible_only = request.args.get("investible") == "true"
        holdings, conn = _holdings(fecha, anchor)
        try:
            from datetime import date as _date
            today = _date.today()
            data = filter_investible(holdings) if investible_only else holdings
            # Filtrar a posiciones con avg_cost (excluye cash puro y cuentas
            # sin operaciones de compra)
            data = [h for h in data if h.get("avg_cost") and h.get("qty")
                    and not h.get("is_cash")
                    and h.get("market_price") is not None]

            rows = []
            for h in data:
                # Fecha primera compra para (account, asset)
                cur = conn.execute(
                    """SELECT MIN(e.event_date) AS d
                       FROM movements m
                       JOIN events e ON e.event_id = m.event_id
                       WHERE m.account=? AND m.asset=? AND m.qty>0""",
                    (h["account"], h["asset"]),
                )
                row = cur.fetchone()
                first = row["d"][:10] if row and row["d"] else None
                days = None
                if first:
                    try:
                        days = (today - _date.fromisoformat(first)).days
                    except Exception:
                        pass
                ret_pct = h.get("unrealized_pct")
                # Anualizado: (1+r)^(365/days)-1 cuando days>0
                ret_annual = None
                if ret_pct is not None and days and days > 0:
                    try:
                        base = 1.0 + ret_pct
                        if base > 0:
                            ret_annual = base ** (365.0 / days) - 1.0
                    except Exception:
                        pass
                rows.append({
                    "account": h["account"],
                    "asset": h["asset"],
                    "asset_class": h.get("asset_class"),
                    "name": h.get("name"),
                    "qty": h.get("qty"),
                    "native_currency": h.get("native_currency"),
                    "avg_cost": h.get("avg_cost"),
                    "market_price": h.get("market_price"),
                    "mv_anchor": h.get("mv_anchor"),
                    "unrealized_pnl_native": h.get("unrealized_pnl_native"),
                    "return_pct": ret_pct,
                    "return_annualized": ret_annual,
                    "first_purchase_date": first,
                    "days_held": days,
                })
            rows.sort(key=lambda r: -(r["return_pct"] or -1e9))
            return jsonify({
                "fecha": fecha.isoformat(),
                "anchor": anchor,
                "investible_only": investible_only,
                "n": len(rows),
                "items": rows,
            })
        finally:
            conn.close()

    @app.get("/api/cash-performance")
    def cash_performance_endpoint():
        """Retorno del cash en monedas no-anchor vs la moneda ancla.

        Para cada (account, currency) de cash en una moneda distinta al
        anchor, computa:
          - avg_fx_in   = rate ponderada por las entradas de cash (qty>0)
          - current_fx  = rate actual currency→anchor
          - return_pct  = current_fx / avg_fx_in - 1
          - mv_anchor   = qty actual * current_fx

        El cash en la moneda ancla siempre tiene return=0 (1 unit = 1 unit).
        El cash en monedas inestables (ARS si anchor=USD) tiene retorno
        TÍPICAMENTE NEGATIVO porque ARS se devalúa contra USD.
        """
        _require_auth(); _block_if_switched_mutation()
        from engine.holdings import (
            calculate_holdings, _load_currencies, SYSTEM_ACCOUNTS,
        )
        from engine.fx import get_rate
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        conn = db_conn()
        try:
            holdings = calculate_holdings(conn, fecha=fecha,
                                           anchor_currency=anchor)
            currencies_set = _load_currencies(conn)
            cash_holdings = [h for h in holdings
                             if h.get("is_cash") and h.get("qty")
                             and abs(h["qty"]) > 1e-6]

            today_iso = fecha.isoformat()
            rows = []
            for h in cash_holdings:
                ccy = h["native_currency"]
                qty = h["qty"]
                # Movimientos qty>0 (entradas) de ese (account, asset)
                cur = conn.execute(
                    """SELECT m.qty, e.event_date
                       FROM movements m JOIN events e ON e.event_id = m.event_id
                       WHERE m.account=? AND m.asset=? AND m.qty>0
                       ORDER BY e.event_date ASC""",
                    (h["account"], ccy),
                )
                inflows = [(r["event_date"][:10], r["qty"]) for r in cur.fetchall()]
                if not inflows:
                    continue
                first_date = inflows[0][0]

                # Para cada entrada, FX rate ccy→anchor en esa fecha.
                # Skip si falta FX (sin warning bloqueante: ese inflow no
                # entra en el promedio, los otros sí).
                weighted_rate_num = 0.0
                weighted_qty = 0.0
                for d, q in inflows:
                    if ccy == anchor:
                        rate = 1.0
                    else:
                        try:
                            rate = get_rate(conn, d, ccy, anchor, fallback_days=14)
                        except Exception:
                            rate = None
                    if rate is None or rate <= 0:
                        continue
                    weighted_rate_num += rate * q
                    weighted_qty += q
                if weighted_qty <= 0:
                    continue
                avg_fx_in = weighted_rate_num / weighted_qty

                # FX actual
                if ccy == anchor:
                    current_fx = 1.0
                else:
                    current_fx = get_rate(conn, today_iso, ccy, anchor,
                                            fallback_days=14)
                if current_fx is None or current_fx <= 0:
                    continue

                ret_pct = (current_fx / avg_fx_in) - 1.0
                mv_anchor = qty * current_fx
                cost_anchor = qty * avg_fx_in
                pnl_anchor = mv_anchor - cost_anchor

                rows.append({
                    "account": h["account"],
                    "currency": ccy,
                    "qty": qty,
                    "avg_fx_in": avg_fx_in,
                    "current_fx": current_fx,
                    "return_pct": ret_pct,
                    "mv_anchor": mv_anchor,
                    "cost_anchor": cost_anchor,
                    "pnl_anchor": pnl_anchor,
                    "first_inflow_date": first_date,
                    "n_inflows": len([1 for d, q in inflows]),
                })
            rows.sort(key=lambda r: -(r["return_pct"] or -1e9))
            return jsonify({
                "fecha": fecha.isoformat(),
                "anchor": anchor,
                "n": len(rows),
                "items": rows,
            })
        finally:
            conn.close()

    @app.get("/api/realized-pnl")
    def realized_pnl_endpoint():
        _require_auth(); _block_if_switched_mutation()
        fecha = _parse_query_fecha()
        conn = db_conn()
        try:
            fills = calculate_realized_pnl(conn, fecha_hasta=fecha)
            return jsonify({
                "n": len(fills),
                "totals_by_currency": total_realized_pnl(fills),
                "fills": [
                    {
                        "fecha_venta": f.fecha_venta,
                        "fecha_compra": f.fecha_compra,
                        "account": f.account,
                        "asset": f.asset,
                        "qty": f.qty,
                        "precio_compra": f.precio_compra,
                        "precio_venta": f.precio_venta,
                        "currency": f.currency,
                        "pnl_realizado": f.pnl_realizado,
                        "pnl_pct": f.pnl_pct,
                        "holding_period_days": f.holding_period_days,
                    } for f in fills
                ],
            })
        finally:
            conn.close()

    # =========================================================================
    # CRUD: hojas de eventos del Excel
    # =========================================================================

    @app.get("/api/sheets/<sheet>")
    def list_sheet(sheet):
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if sheet not in ALLOWED_SHEETS:
            abort(404, f"Sheet '{sheet}' no soportada. Disponibles: {list(SHEET_PREFIX.keys())}")
        if not s.xlsx_path.is_file():
            abort(404, "Excel master no presente")
        rows = list_rows(s.xlsx_path, sheet)
        return jsonify({"sheet": sheet, "n": len(rows), "items": rows})

    @app.get("/api/sheets/<sheet>/<row_id>")
    def get_sheet_row(sheet, row_id):
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if sheet not in ALLOWED_SHEETS:
            abort(404)
        row = get_row(s.xlsx_path, sheet, row_id)
        if row is None:
            abort(404, f"Row ID '{row_id}' no encontrado en '{sheet}'")
        return jsonify(row)

    @app.post("/api/sheets/<sheet>")
    def create_sheet_row(sheet):
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if sheet not in ALLOWED_SHEETS:
            abort(404)
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            abort(400, "Body debe ser un JSON object con {header: value}")
        with excel_write_lock():
            backup_excel()
            try:
                row_id = append_row(s.xlsx_path, sheet, data)
            except ValueError as e:
                abort(400, str(e))
            stats = reimport_excel()
            prune_backups(keep_last=50)
        return jsonify({
            "sheet": sheet, "row_id": row_id,
            "row": get_row(s.xlsx_path, sheet, row_id),
            "import_stats": stats,
        }), 201

    @app.put("/api/sheets/<sheet>/<row_id>")
    def update_sheet_row(sheet, row_id):
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if sheet not in ALLOWED_SHEETS:
            abort(404)
        data = request.get_json(silent=True) or {}
        if not isinstance(data, dict):
            abort(400, "Body debe ser un JSON object")
        with excel_write_lock():
            backup_excel()
            try:
                row = update_row(s.xlsx_path, sheet, row_id, data)
            except KeyError as e:
                abort(404, str(e))
            except ValueError as e:
                abort(400, str(e))
            stats = reimport_excel()
            prune_backups(keep_last=50)
        return jsonify({
            "sheet": sheet, "row_id": row_id, "row": row,
            "import_stats": stats,
        })

    @app.delete("/api/sheets/<sheet>/<row_id>")
    def delete_sheet_row(sheet, row_id):
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if sheet not in ALLOWED_SHEETS:
            abort(404)
        with excel_write_lock():
            backup_excel()
            ok = delete_row(s.xlsx_path, sheet, row_id)
            if not ok:
                abort(404, f"Row ID '{row_id}' no encontrado")
            stats = reimport_excel()
            prune_backups(keep_last=50)
        return jsonify({"deleted": True, "row_id": row_id, "import_stats": stats})

    # =========================================================================
    # File upload/download
    # =========================================================================

    @app.post("/api/upload/excel")
    def upload_excel():
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if "file" not in request.files:
            abort(400, "Subí el archivo en form-data como 'file'")
        f = request.files["file"]
        if not f.filename.lower().endswith((".xlsx", ".xls")):
            abort(400, "Solo .xlsx / .xls")
        with excel_write_lock():
            if s.xlsx_path.is_file():
                backup_excel()
            f.save(str(s.xlsx_path))
            stats = reimport_excel()
            prune_backups(keep_last=50)
        return jsonify({
            "saved": str(s.xlsx_path),
            "size_bytes": s.xlsx_path.stat().st_size,
            "import_stats": stats,
        })

    @app.get("/api/download/excel")
    def download_excel():
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if not s.xlsx_path.is_file():
            abort(404, "No hay Excel master")
        return send_file(
            str(s.xlsx_path),
            as_attachment=True,
            download_name=s.xlsx_path.name,
        )

    @app.post("/api/upload/prices")
    def upload_prices():
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if "file" not in request.files:
            abort(400, "Subí el CSV en form-data como 'file'")
        f = request.files["file"]
        # Nombre sugerido del CSV: usá el nombre original
        name = (request.form.get("name") or f.filename).strip()
        if not name.endswith(".csv"):
            abort(400, "Solo .csv")
        target = s.data_dir / name
        f.save(str(target))
        # Re-importar para refrescar precios
        with excel_write_lock():
            stats = reimport_excel()
        return jsonify({
            "saved": str(target),
            "size_bytes": target.stat().st_size,
            "import_stats": stats,
        })

    @app.post("/api/refresh")
    def refresh():
        _require_auth(); _block_if_switched_mutation()
        # Lock cubre TODO el flujo: re-import + record_snapshots. Sin esto,
        # otra request podría leer la DB durante init_db(drop_existing=True)
        # del reimport y obtener data parcial / corrupta.
        with excel_write_lock():
            stats = reimport_excel()
            # Grabar snapshot del PN del día. Necesario para que la equity
            # curve evolucione (antes solo ocurría al bajar un reporte).
            try:
                anchor = get_settings().anchor
                today = date.today()
                holdings, conn = _holdings(today, anchor)
                try:
                    n = record_snapshots(conn, holdings, today, anchor_currency=anchor)
                    conn.commit()
                finally:
                    conn.close()
                stats["snapshots"] = n
            except Exception as e:
                print(f"[refresh] WARN no pude grabar snapshot: {e}")
        return jsonify({"refreshed": True, "import_stats": stats})

    @app.get("/api/performance")
    def performance_endpoint():
        """Métricas full de performance: TWR + MWR + flows + curve.

        Query: ?anchor=USD&investible=1
        """
        _require_auth(); _block_if_switched_mutation()
        from engine.performance import performance_summary
        anchor = _parse_query_anchor()
        investible = request.args.get("investible") == "1"
        conn = db_conn()
        try:
            data = performance_summary(conn, anchor_currency=anchor,
                                         investible_only=investible)
            return jsonify(data)
        finally:
            conn.close()

    @app.get("/api/returns")
    def returns_endpoint():
        """Returns simples del PN para 1d/1w/1m/3m/ytd/1y.

        Query: ?anchor=USD&investible=1
        """
        _require_auth(); _block_if_switched_mutation()
        from engine.snapshots import returns_by_period
        anchor = _parse_query_anchor()
        investible = request.args.get("investible") == "1"
        conn = db_conn()
        try:
            data = returns_by_period(conn, anchor_currency=anchor,
                                       investible_only=investible)
            return jsonify({
                "anchor": anchor,
                "investible_only": investible,
                "as_of": date.today().isoformat(),
                "returns": data,
            })
        finally:
            conn.close()

    @app.post("/api/snapshots/backfill")
    def backfill_snapshots_endpoint():
        """Reconstruye la equity curve histórica computando holdings a fechas
        pasadas (semanal por default). Útil después de /api/snapshots DELETE
        o cuando arrancás con un baseline contaminado.

        Query:
          ?cadence=7         — días entre snapshots (default 7)
          ?from=YYYY-MM-DD   — fecha inicial (default: primer evento del user)
          ?to=YYYY-MM-DD     — fecha final (default: hoy)
        """
        _require_auth(); _block_if_switched_mutation()
        from engine.snapshots import backfill_snapshots
        anchor = _parse_query_anchor()
        try:
            cadence = int(request.args.get("cadence") or 7)
        except ValueError:
            cadence = 7
        cadence = max(1, min(cadence, 365))
        fecha_desde = request.args.get("from")
        fecha_hasta = request.args.get("to")
        conn = db_conn()
        try:
            result = backfill_snapshots(
                conn, anchor_currency=anchor, cadence_days=cadence,
                fecha_desde=fecha_desde, fecha_hasta=fecha_hasta,
            )
            return jsonify(result)
        finally:
            conn.close()

    @app.delete("/api/snapshots")
    def delete_snapshots():
        """Borra snapshots del PN. Útil para limpiar baseline contaminado
        (snapshots iniciales con PN parcial / sin FX que disparan TWR irreal).

        Query:
          ?before=YYYY-MM-DD  — borra todos los snapshots con fecha < ese día.
          ?all=1              — borra todo el histórico de snapshots.
        Sin parámetros: error.
        """
        _require_auth(); _block_if_switched_mutation()
        before = request.args.get("before")
        delete_all = request.args.get("all") == "1"
        if not before and not delete_all:
            return jsonify({"error": "missing 'before' o 'all=1'"}), 400
        conn = db_conn()
        try:
            if delete_all:
                cur = conn.execute("DELETE FROM pn_snapshots")
            else:
                cur = conn.execute(
                    "DELETE FROM pn_snapshots WHERE fecha < ?", (before,)
                )
            n = cur.rowcount
            conn.commit()
            return jsonify({"deleted": n, "before": before, "all": delete_all})
        finally:
            conn.close()

    @app.get("/api/backups")
    def backups():
        _require_auth(); _block_if_switched_mutation()
        return jsonify({
            "backups": [
                {
                    "name": b.name,
                    "size_bytes": b.stat().st_size,
                    "mtime": datetime.fromtimestamp(b.stat().st_mtime).isoformat(),
                }
                for b in list_backups(limit=50)
            ],
        })

    # =========================================================================
    # Cotizaciones (precios + FX) y cash por cuenta
    # =========================================================================

    @app.get("/api/prices")
    def prices_endpoint():
        """Devuelve la última cotización conocida por ticker.

        Query params:
          ?ticker=AL30D        — filtra por uno
          ?asset_class=BOND_AR — filtra por clase
        """
        _require_auth(); _block_if_switched_mutation()
        conn = db_conn()
        try:
            ticker = request.args.get("ticker")
            asset_class = request.args.get("asset_class")
            sql = """
                SELECT p.ticker, p.price, p.currency, p.fecha, p.source,
                       a.name, a.asset_class
                FROM prices p
                INNER JOIN (
                    SELECT ticker, MAX(fecha) AS max_fecha
                    FROM prices GROUP BY ticker
                ) m ON p.ticker = m.ticker AND p.fecha = m.max_fecha
                LEFT JOIN assets a ON a.ticker = p.ticker
            """
            params = []
            wheres = []
            if ticker:
                wheres.append("p.ticker = ?")
                params.append(ticker)
            if asset_class:
                wheres.append("a.asset_class = ?")
                params.append(asset_class)
            if wheres:
                sql += " WHERE " + " AND ".join(wheres)
            sql += " ORDER BY a.asset_class, p.ticker"
            rows = conn.execute(sql, params).fetchall()
            return jsonify({
                "n": len(rows),
                "items": [
                    {
                        "ticker": r["ticker"],
                        "name": r["name"] or "",
                        "asset_class": r["asset_class"] or "",
                        "price": r["price"],
                        "currency": r["currency"],
                        "fecha": r["fecha"],
                        "source": r["source"] or "",
                    } for r in rows
                ],
            })
        finally:
            conn.close()

    @app.get("/api/fx-rates")
    def fx_rates_endpoint():
        """Devuelve la última cotización FX conocida por (moneda, base)."""
        _require_auth(); _block_if_switched_mutation()
        conn = db_conn()
        try:
            sql = """
                SELECT fr.fecha, fr.moneda, fr.rate, fr.base, fr.source
                FROM fx_rates fr
                INNER JOIN (
                    SELECT moneda, base, MAX(fecha) AS max_fecha
                    FROM fx_rates GROUP BY moneda, base
                ) m ON fr.moneda = m.moneda AND fr.base = m.base
                       AND fr.fecha = m.max_fecha
                ORDER BY fr.base, fr.moneda
            """
            rows = conn.execute(sql).fetchall()
            return jsonify({
                "n": len(rows),
                "items": [
                    {
                        "fecha": r["fecha"],
                        "moneda": r["moneda"],
                        "rate": r["rate"],
                        "base": r["base"],
                        "source": r["source"] or "",
                    } for r in rows
                ],
            })
        finally:
            conn.close()

    # =========================================================================
    # Historial de precios + FX (lectura para todos, edición sólo superadmin)
    # =========================================================================
    #
    # Modelo:
    # - Los precios automáticos viven en data/precios_<source>.csv y se
    #   importan a la tabla `prices` del sqlite en cada reimport.
    # - Las correcciones manuales del superadmin se guardan en
    #   data/precios_manual.csv y se cargan ÚLTIMO en auto_load_all → pisan
    #   lo que vino del loader automático para esa (fecha, ticker).
    # - Mismo patrón para FX con data/fx_manual.csv.

    PRICES_MANUAL_CSV = "precios_manual.csv"
    FX_MANUAL_CSV = "fx_manual.csv"

    def _read_manual_csv(path, key_cols):
        import csv as _csv
        rows = []
        if not path.is_file():
            return rows
        with open(path, encoding="utf-8", newline="") as f:
            reader = _csv.DictReader(f)
            for r in reader:
                rows.append({k: (v or "").strip() for k, v in r.items()})
        return rows

    def _write_manual_csv(path, rows, header):
        import csv as _csv
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=header)
            w.writeheader()
            for r in rows:
                w.writerow({h: r.get(h, "") for h in header})

    def _upsert_manual_price(fecha, ticker, price, currency, source_note):
        s = get_settings()
        path = s.shared_data_dir / PRICES_MANUAL_CSV
        header = ["fecha", "ticker", "price", "currency", "source"]
        rows = _read_manual_csv(path, ("fecha", "ticker"))
        rows = [r for r in rows if not (r.get("fecha") == fecha
                                         and r.get("ticker") == ticker)]
        rows.append({
            "fecha": fecha,
            "ticker": ticker,
            "price": f"{float(price):.6f}",
            "currency": currency,
            "source": f"manual ({source_note})" if source_note else "manual",
        })
        _write_manual_csv(path, rows, header)

    def _delete_manual_price(fecha, ticker):
        s = get_settings()
        path = s.shared_data_dir / PRICES_MANUAL_CSV
        if not path.is_file():
            return False
        header = ["fecha", "ticker", "price", "currency", "source"]
        rows = _read_manual_csv(path, ("fecha", "ticker"))
        before = len(rows)
        rows = [r for r in rows if not (r.get("fecha") == fecha
                                         and r.get("ticker") == ticker)]
        _write_manual_csv(path, rows, header)
        return before != len(rows)

    def _upsert_manual_fx(fecha, moneda, rate, base, source_note):
        s = get_settings()
        path = s.shared_data_dir / FX_MANUAL_CSV
        header = ["fecha", "moneda", "rate", "base", "source"]
        rows = _read_manual_csv(path, ("fecha", "moneda", "base"))
        rows = [r for r in rows if not (r.get("fecha") == fecha
                                         and r.get("moneda") == moneda
                                         and (r.get("base") or "ARS") == base)]
        rows.append({
            "fecha": fecha,
            "moneda": moneda,
            "rate": f"{float(rate):.6f}",
            "base": base,
            "source": f"manual ({source_note})" if source_note else "manual",
        })
        _write_manual_csv(path, rows, header)

    def _delete_manual_fx(fecha, moneda, base):
        s = get_settings()
        path = s.shared_data_dir / FX_MANUAL_CSV
        if not path.is_file():
            return False
        header = ["fecha", "moneda", "rate", "base", "source"]
        rows = _read_manual_csv(path, ("fecha", "moneda", "base"))
        before = len(rows)
        rows = [r for r in rows if not (r.get("fecha") == fecha
                                         and r.get("moneda") == moneda
                                         and (r.get("base") or "ARS") == base)]
        _write_manual_csv(path, rows, header)
        return before != len(rows)

    @app.get("/api/prices/<ticker>/history")
    def prices_history(ticker):
        """Devuelve el historial de cotizaciones de un ticker (todas las
        fechas que tenga la DB). Query: ?n=180 (default) limita a las N
        más recientes.

        Cada fila incluye `is_manual: bool` para que la UI pueda diferenciar
        las que vienen del loader automático de las que el superadmin pisó.
        """
        _require_auth(); _block_if_switched_mutation()
        try:
            n = int(request.args.get("n", "180"))
        except ValueError:
            n = 180
        n = max(1, min(n, 5000))

        s = get_settings()
        # Cargar lo que hay en precios_manual.csv para flagear is_manual
        manual_keys = set()
        manual_path = s.shared_data_dir / PRICES_MANUAL_CSV
        if manual_path.is_file():
            for r in _read_manual_csv(manual_path, ("fecha", "ticker")):
                if r.get("ticker") == ticker:
                    manual_keys.add(r.get("fecha"))

        conn = db_conn()
        try:
            rows = conn.execute(
                """SELECT p.fecha, p.price, p.currency, p.source,
                          a.name, a.asset_class
                   FROM prices p
                   LEFT JOIN assets a ON a.ticker = p.ticker
                   WHERE p.ticker = ?
                   ORDER BY p.fecha DESC
                   LIMIT ?""",
                (ticker, n),
            ).fetchall()
            if not rows:
                return jsonify({"ticker": ticker, "items": [], "n": 0})
            return jsonify({
                "ticker": ticker,
                "name": rows[0]["name"] or "",
                "asset_class": rows[0]["asset_class"] or "",
                "n": len(rows),
                "items": [
                    {
                        "fecha": r["fecha"],
                        "price": r["price"],
                        "currency": r["currency"],
                        "source": r["source"] or "",
                        "is_manual": r["fecha"] in manual_keys,
                    } for r in rows
                ],
            })
        finally:
            conn.close()

    @app.put("/api/prices/<ticker>")
    def prices_upsert(ticker):
        """Crea o sobrescribe un precio (sólo superadmin).

        Body: {"fecha":"YYYY-MM-DD","price":123.45,"currency":"USD","note":"..."}

        Escribe a data/precios_manual.csv (loaded LAST en auto_load_all,
        así pisa al loader automático). Re-importa el master del caller.
        """
        _require_superadmin(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        fecha = (body.get("fecha") or "").strip()
        try:
            date.fromisoformat(fecha)
        except ValueError:
            abort(400, f"fecha inválida: {fecha} (YYYY-MM-DD)")
        try:
            price = float(body.get("price"))
        except (TypeError, ValueError):
            abort(400, "price debe ser numérico")
        if price <= 0:
            abort(400, "price debe ser > 0")
        currency = (body.get("currency") or "").strip().upper()
        if not currency:
            abort(400, "currency es requerido")
        note = (body.get("note") or "").strip()[:80]

        _upsert_manual_price(fecha, ticker, price, currency, note)
        s = get_settings()
        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "ok": True, "ticker": ticker, "fecha": fecha,
            "price": price, "currency": currency,
            "import_stats": stats,
        })

    @app.delete("/api/prices/<ticker>")
    def prices_delete(ticker):
        """Borra un precio MANUAL (sólo superadmin). Query: ?fecha=YYYY-MM-DD.

        Sólo se borra si la fila vino de precios_manual.csv. Si era de
        un loader automático, el siguiente loader la vuelve a poner —
        para suprimirla de verdad hay que correr la edición y poner
        un valor que se quiera mantener.
        """
        _require_superadmin(); _block_if_switched_mutation()
        fecha = (request.args.get("fecha") or "").strip()
        try:
            date.fromisoformat(fecha)
        except ValueError:
            abort(400, f"fecha inválida: {fecha} (YYYY-MM-DD)")
        ok = _delete_manual_price(fecha, ticker)
        s = get_settings()
        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "ok": True, "ticker": ticker, "fecha": fecha,
            "deleted_manual": ok, "import_stats": stats,
        })

    @app.get("/api/fx-rates/<moneda>/history")
    def fx_history(moneda):
        """Historial FX para una moneda. Query: ?base=ARS&n=180."""
        _require_auth(); _block_if_switched_mutation()
        base = (request.args.get("base") or "ARS").strip().upper()
        try:
            n = int(request.args.get("n", "180"))
        except ValueError:
            n = 180
        n = max(1, min(n, 5000))

        s = get_settings()
        manual_keys = set()
        manual_path = s.shared_data_dir / FX_MANUAL_CSV
        if manual_path.is_file():
            for r in _read_manual_csv(manual_path, ("fecha", "moneda", "base")):
                if r.get("moneda") == moneda and \
                        (r.get("base") or "ARS") == base:
                    manual_keys.add(r.get("fecha"))

        conn = db_conn()
        try:
            rows = conn.execute(
                """SELECT fecha, rate, source
                   FROM fx_rates
                   WHERE moneda = ? AND base = ?
                   ORDER BY fecha DESC
                   LIMIT ?""",
                (moneda, base, n),
            ).fetchall()
            return jsonify({
                "moneda": moneda, "base": base,
                "n": len(rows),
                "items": [
                    {
                        "fecha": r["fecha"],
                        "rate": r["rate"],
                        "source": r["source"] or "",
                        "is_manual": r["fecha"] in manual_keys,
                    } for r in rows
                ],
            })
        finally:
            conn.close()

    @app.put("/api/fx-rates/<moneda>")
    def fx_upsert(moneda):
        """Crea o sobrescribe un FX rate (sólo superadmin).

        Body: {"fecha":"YYYY-MM-DD","rate":1485.0,"base":"ARS","note":"..."}
        Escribe a data/fx_manual.csv. Re-importa el master del caller.
        """
        _require_superadmin(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        fecha = (body.get("fecha") or "").strip()
        try:
            date.fromisoformat(fecha)
        except ValueError:
            abort(400, f"fecha inválida: {fecha} (YYYY-MM-DD)")
        try:
            rate = float(body.get("rate"))
        except (TypeError, ValueError):
            abort(400, "rate debe ser numérico")
        if rate <= 0:
            abort(400, "rate debe ser > 0")
        base = (body.get("base") or "ARS").strip().upper()
        note = (body.get("note") or "").strip()[:80]

        _upsert_manual_fx(fecha, moneda, rate, base, note)
        s = get_settings()
        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "ok": True, "moneda": moneda, "base": base, "fecha": fecha,
            "rate": rate, "import_stats": stats,
        })

    @app.delete("/api/fx-rates/<moneda>")
    def fx_delete(moneda):
        """Borra un FX rate MANUAL (sólo superadmin)."""
        _require_superadmin(); _block_if_switched_mutation()
        fecha = (request.args.get("fecha") or "").strip()
        base = (request.args.get("base") or "ARS").strip().upper()
        try:
            date.fromisoformat(fecha)
        except ValueError:
            abort(400, f"fecha inválida: {fecha} (YYYY-MM-DD)")
        ok = _delete_manual_fx(fecha, moneda, base)
        s = get_settings()
        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "ok": True, "moneda": moneda, "base": base, "fecha": fecha,
            "deleted_manual": ok, "import_stats": stats,
        })

    @app.get("/api/cash")
    def cash_endpoint():
        """Saldos cash por cuenta (todas las monedas).

        Query: ?anchor=USD para incluir conversión a moneda ancla.
        """
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        holdings, conn = _holdings(fecha, anchor)
        try:
            cash_items = [h for h in holdings if h.get("is_cash")]
            # Excluir pasivos (los saldos deudores de tarjetas no son cash positivo)
            cash_items = [h for h in cash_items if not h.get("is_liability")]
            cash_items.sort(key=lambda h: -(h.get("mv_anchor") or 0))
            # Subtotales por moneda
            by_ccy = {}
            for h in cash_items:
                ccy = h["native_currency"]
                if ccy not in by_ccy:
                    by_ccy[ccy] = {"qty": 0.0, "mv_anchor": 0.0, "n": 0}
                by_ccy[ccy]["qty"] += h["qty"]
                if h.get("mv_anchor") is not None:
                    by_ccy[ccy]["mv_anchor"] += h["mv_anchor"]
                by_ccy[ccy]["n"] += 1
            total_anchor = sum(h.get("mv_anchor") or 0 for h in cash_items)
            return jsonify({
                "anchor": anchor,
                "fecha": fecha.isoformat(),
                "n": len(cash_items),
                "total_anchor": total_anchor,
                "by_currency": by_ccy,
                "items": [
                    {
                        "account": h["account"],
                        "account_name": h.get("account_name") or h["account"],
                        "account_kind": h.get("account_kind") or "",
                        "account_institution": h.get("account_institution") or "",
                        "currency": h["native_currency"],
                        "qty": h["qty"],
                        "mv_anchor": h.get("mv_anchor"),
                        "cash_purpose": h.get("cash_purpose") or "",
                        "investible": h.get("investible", True),
                    } for h in cash_items
                ],
            })
        finally:
            conn.close()

    # =========================================================================
    # Calendar — próximos eventos (cupones, vencimientos, cierres tarjetas, fundings)
    # =========================================================================

    @app.get("/api/calendar")
    def calendar_endpoint():
        """Devuelve eventos calendarizados de los próximos N días.

        Query: ?days=60 (default 60).

        Eventos incluidos:
          - Maturity de bonos (assets.maturity)
          - Cierres y vencimientos de tarjetas (próximo + siguiente)
          - Funding cierres (cauciones que vencen)
          - Recurrentes próximos (sueldo, alquiler, servicios)
        """
        _require_auth(); _block_if_switched_mutation()
        from datetime import timedelta
        import calendar as _cal
        days_ahead = int(request.args.get("days", 60))
        today = date.today()
        until = today + timedelta(days=days_ahead)
        events = []

        conn = db_conn()
        try:
            # 1. Maturity de bonos
            cur = conn.execute(
                """SELECT ticker, name, maturity, currency, asset_class
                   FROM assets WHERE maturity IS NOT NULL AND maturity != ''"""
            )
            for r in cur.fetchall():
                try:
                    md = date.fromisoformat(r["maturity"][:10])
                    if today <= md <= until:
                        events.append({
                            "fecha": md.isoformat(),
                            "tipo": "maturity_bono",
                            "icon": "📜",
                            "title": f"Vencimiento {r['ticker']}",
                            "subtitle": r["name"] or "",
                            "amount": None,
                            "currency": r["currency"],
                            "ref_id": r["ticker"],
                        })
                except (ValueError, TypeError):
                    pass

            # 2. Tarjetas: próximos cierres + vencimientos
            cur = conn.execute(
                """SELECT code, name, card_close_day, card_due_day, card_currency
                   FROM accounts
                   WHERE kind='CARD_CREDIT' AND card_close_day IS NOT NULL"""
            )
            for r in cur.fetchall():
                close_day = r["card_close_day"]
                due_day = r["card_due_day"]
                # Próximos N cierres
                cur_y, cur_m = today.year, today.month
                for _ in range(4):  # próximos 4 ciclos = ~4 meses
                    last = _cal.monthrange(cur_y, cur_m)[1]
                    close_d = date(cur_y, cur_m, min(close_day, last))
                    if close_d >= today and close_d <= until:
                        events.append({
                            "fecha": close_d.isoformat(),
                            "tipo": "card_close",
                            "icon": "💳",
                            "title": f"Cierre {r['name']}",
                            "subtitle": r["code"],
                            "amount": None,
                            "currency": r["card_currency"] or "ARS",
                            "ref_id": r["code"],
                        })
                    if due_day:
                        # Due es mes siguiente al cierre
                        due_y = cur_y + (1 if cur_m == 12 else 0)
                        due_m = 1 if cur_m == 12 else cur_m + 1
                        last_due = _cal.monthrange(due_y, due_m)[1]
                        due_d = date(due_y, due_m, min(due_day, last_due))
                        if due_d >= today and due_d <= until:
                            events.append({
                                "fecha": due_d.isoformat(),
                                "tipo": "card_due",
                                "icon": "💸",
                                "title": f"Vto {r['name']}",
                                "subtitle": r["code"],
                                "amount": None,
                                "currency": r["card_currency"] or "ARS",
                                "ref_id": r["code"],
                            })
                    # Avanzar mes
                    cur_m += 1
                    if cur_m > 12:
                        cur_m = 1; cur_y += 1

            # 3. Cauciones que vencen — leemos del Excel para tener
            #    Fecha Fin / Status / Monto sin tener que reconstruirlo
            #    desde events.
            try:
                fund_rows = list_rows(get_settings().xlsx_path, "funding")
            except Exception:
                fund_rows = []
            for fr in fund_rows:
                if (fr.get("Status") or "").upper() == "CLOSED":
                    continue
                if not fr.get("Fecha Fin"):
                    continue
                try:
                    fin = date.fromisoformat(str(fr["Fecha Fin"])[:10])
                except (ValueError, TypeError):
                    continue
                if today <= fin <= until:
                    events.append({
                        "fecha": fin.isoformat(),
                        "tipo": "funding_close",
                        "icon": "💰",
                        "title": f"Vence {fr.get('Tipo', '')} {fr.get('Subtipo', '')}",
                        "subtitle": f"{fr.get('Fund ID', '')} · {fr.get('Cuenta', '')}",
                        "amount": float(fr["Monto"]) if fr.get("Monto") else None,
                        "currency": fr.get("Moneda", ""),
                        "ref_id": fr.get("Fund ID"),
                    })

            # 4. Recurrentes activos en los próximos N días
            cur = conn.execute("SELECT * FROM recurring_rules WHERE active=1")
            rules = cur.fetchall()
            for rule in rules:
                # Generar próximas N ocurrencias dentro de la ventana
                day_of_month = rule["day_of_month"] or 1
                cur_y, cur_m = today.year, today.month
                for _ in range(3):
                    last = _cal.monthrange(cur_y, cur_m)[1]
                    occ = date(cur_y, cur_m, min(day_of_month, last))
                    if occ >= today and occ <= until:
                        events.append({
                            "fecha": occ.isoformat(),
                            "tipo": "recurring_" + rule["event_type"].lower(),
                            "icon": "🔁" if rule["event_type"] == "INCOME" else "📅",
                            "title": rule["rule_name"],
                            "subtitle": rule["description"] or "",
                            "amount": rule["amount"],
                            "currency": rule["asset"],
                            "ref_id": str(rule["rule_id"]),
                        })
                    cur_m += 1
                    if cur_m > 12:
                        cur_m = 1; cur_y += 1

            events.sort(key=lambda e: e["fecha"])
            return jsonify({"days": days_ahead, "n": len(events), "events": events})
        finally:
            conn.close()

    # =========================================================================
    # Admin: gestión de usuarios + switch view
    # =========================================================================

    @app.get("/api/admin/users")
    def admin_list_users():
        """Lista todos los users (sin tokens completos, solo preview)."""
        _require_admin()
        users = load_users()
        config_ids = set(u.user_id for u in users)
        disk_ids = set(list_user_ids())
        orphans = disk_ids - config_ids
        return jsonify({
            "users": [
                {
                    "user_id": u.user_id,
                    "display_name": u.display_name,
                    "is_admin": u.is_admin,
                    "token_preview": u.token[:8] + "..." if len(u.token) > 8 else "?",
                    "has_xlsx": (get_user_settings(u.user_id).xlsx_path).is_file(),
                    "has_db": (get_user_settings(u.user_id).db_path).is_file(),
                }
                for u in users
            ],
            "orphan_folders": sorted(orphans),
            "n_users": len(users),
            "persistent": is_persistent(),
        })

    @app.post("/api/admin/users")
    def admin_create_user():
        """Crea un nuevo user.

        Body JSON:
          { "user_id": "amigo", "display_name": "Marcos", "is_admin": false,
            "token": "auto"   ← si "auto", se genera random }

        Acciones:
          1. Genera token (si no se pasa)
          2. Crea folder inputs/<user_id>/
          3. Genera el Excel master (build_master + add_carga_inicial_sheet)
          4. Agrega user a WM_USERS_JSON in-memory
          5. Devuelve el token al admin para que lo comparta

        IMPORTANTE: el WSGI file de PythonAnywhere tiene que actualizarse
        para que el user persista entre reloads. La response incluye el
        snippet a copiar en el WSGI.
        """
        _require_admin()
        body = request.get_json(silent=True) or {}
        user_id = (body.get("user_id") or "").strip().lower()
        if not user_id or not user_id.replace("_", "").replace("-", "").isalnum():
            abort(400, "user_id inválido (solo letras/números/_/-, lowercase)")
        # Verificar duplicado
        existing = resolve_user_by_token("__never_match__")  # noop, just to ensure load
        for u in load_users():
            if u.user_id == user_id:
                abort(400, f"Ya existe un user con id '{user_id}'")

        display_name = body.get("display_name") or user_id
        is_admin = bool(body.get("is_admin", False))
        token = body.get("token") or ""
        if not token or token == "auto":
            import secrets
            token = secrets.token_hex(32)

        # Crear folders y master vacío
        new_settings = get_user_settings(user_id)
        new_settings.inputs_dir.mkdir(parents=True, exist_ok=True)
        new_settings.user_data_dir.mkdir(parents=True, exist_ok=True)
        new_settings.backups_dir.mkdir(parents=True, exist_ok=True)

        # Generar master con build_master + add_carga_inicial
        from build_master import build_master
        try:
            from add_carga_inicial_sheet import add_hoja_carga_inicial
        except ImportError:
            add_hoja_carga_inicial = None

        try:
            build_master(new_settings.xlsx_path)
            if add_hoja_carga_inicial:
                from openpyxl import load_workbook as _wb
                _wb_h = _wb(filename=str(new_settings.xlsx_path))
                add_hoja_carga_inicial(_wb_h, with_examples=False)
                _wb_h.save(str(new_settings.xlsx_path))
            # Limpiar las filas de ejemplo de TODAS las hojas de eventos
            # (blotter, ingresos, gastos, etc) para que el user nuevo arranque
            # con master vacío, listo para que cargue su data inicial.
            _blank_event_sheets(new_settings.xlsx_path)
        except Exception as e:
            abort(500, f"No se pudo crear el master para '{user_id}': {e}")

        # Auto-import: crear la DB del user para que /api/summary y similares
        # funcionen apenas hace login. Sin este import, switch read-only crashea
        # con FileNotFoundError porque la DB no existe.
        import_stats = None
        try:
            import_stats = reimport_excel(settings=new_settings)
        except Exception as e:
            print(f"[admin_create_user] WARN auto-import falló para "
                  f"'{user_id}': {e} — el user va a tener que hacer refresh manual")

        # Persistir en config in-memory (NO sobrevive un reload del web app)
        add_user_to_config(user_id, token, display_name=display_name,
                            is_admin=is_admin)

        persistent = is_persistent()
        response = {
            "created": True,
            "user_id": user_id,
            "display_name": display_name,
            "is_admin": is_admin,
            "token": token,
            "url": request.host_url.rstrip("/"),
            "import_stats": import_stats,
            "persistent": persistent,
        }
        if persistent:
            response["info"] = ("✓ User guardado en disk (WM_USERS_FILE). "
                                 "Sobrevive reloads sin tocar el WSGI.")
        else:
            response["wsgi_snippet"] = _build_wsgi_snippet()
            response["warning"] = (
                "⚠ Para que el user persista a reloads, agregá WM_USERS_FILE "
                "al WSGI (ver MULTITENANT.md). Mientras tanto, el user vive "
                "solo en memoria del server. Como workaround inmediato, copiá "
                "el snippet de wsgi_snippet al WSGI file."
            )
        return jsonify(response), 201

    def _blank_event_sheets(xlsx_path):
        """Borra las filas de DATOS (no headers) de las hojas de eventos del
        master de un user nuevo. Mantiene cuentas/monedas/especies/aforos
        completos como templates. Header en row 4."""
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(xlsx_path))
        EVENT_SHEETS = (
            "blotter", "ingresos", "gastos",
            "transferencias_cash", "transferencias_activos",
            "funding", "pasivos", "pagos_pasivos",
            "recurrentes", "asientos_contables",
        )
        HEADER_ROW = 4
        for sheet_name in EVENT_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            if ws.max_row > HEADER_ROW:
                ws.delete_rows(HEADER_ROW + 1, ws.max_row - HEADER_ROW)
        wb.save(str(xlsx_path))

    @app.post("/api/admin/users/<user_id>/seed-demo")
    def admin_seed_demo(user_id):
        """Sobreescribe el master del user con datos demo hard-coded fijos.

        Útil para tener un user 'demo' con datos repetibles para mostrar la app.
        Después de seed, también re-importa la DB.
        """
        _require_admin()
        if not any(u.user_id == user_id for u in load_users()):
            abort(404, f"User '{user_id}' no existe")
        # Hacer backup del master actual antes de overwritearlo
        target = get_user_settings(user_id)
        if target.xlsx_path.is_file():
            with excel_write_lock(settings=target):
                backup_excel(settings=target)
        # Seed
        from seed_demo import seed_demo
        try:
            stats = seed_demo(target.xlsx_path)
        except Exception as e:
            abort(500, f"Seed demo falló: {e}")
        # Re-importar
        try:
            import_stats = reimport_excel(settings=target)
        except Exception as e:
            return jsonify({
                "seeded": True, "seed_stats": stats,
                "import_failed": str(e),
            }), 200
        return jsonify({
            "seeded": True,
            "user_id": user_id,
            "seed_stats": stats,
            "import_stats": import_stats,
        })

    @app.delete("/api/admin/users/<user_id>")
    def admin_delete_user(user_id):
        """Borra un user del config in-memory. NO borra los archivos del disk
        (por seguridad). Si querés borrar también, usá ?delete_data=true."""
        _require_admin()
        if user_id == g.auth_user.user_id:
            abort(400, "No podés borrar tu propio user")
        if not any(u.user_id == user_id for u in load_users()):
            abort(404, f"User '{user_id}' no existe")

        delete_data = request.args.get("delete_data") == "true"
        remove_user_from_config(user_id)

        if delete_data:
            import shutil
            try:
                ud = get_user_settings(user_id).user_data_dir
                if ud.is_dir(): shutil.rmtree(ud)
                inp = get_user_settings(user_id).inputs_dir
                if inp.is_dir(): shutil.rmtree(inp)
            except OSError as e:
                return jsonify({"removed_config": True, "data_deleted": False,
                                 "error": str(e)})
        return jsonify({"removed_config": True, "data_deleted": delete_data})

    @app.post("/api/admin/switch")
    def admin_switch():
        """Switch view del admin a otro user (read-only).

        Body: {"user_id": "amigo"}  o {"user_id": null} para volver.
        """
        _require_admin()
        body = request.get_json(silent=True) or {}
        target = body.get("user_id")
        token = g.user_token
        if target is None or target == "" or target == g.auth_user.user_id:
            admin_switch_clear(token)
            return jsonify({"switched": False, "active_user_id": g.auth_user.user_id})
        try:
            admin_switch_to(token, target)
        except (PermissionError, ValueError) as e:
            abort(400, str(e))
        return jsonify({"switched": True, "active_user_id": target,
                         "warning": "Modo read-only. POST/PUT/DELETE están bloqueados."})

    def _build_wsgi_snippet():
        """Devuelve un fragmento de código para el WSGI file con el JSON
        actual de users."""
        import os, json
        return (
            "# Reemplazar tu WM_API_TOKEN viejo por esto:\n"
            f"os.environ['WM_USERS_JSON'] = '''{os.environ.get('WM_USERS_JSON', '{}')}'''\n"
            f"os.environ['WM_ADMIN_USER'] = '{os.environ.get('WM_ADMIN_USER', '')}'\n"
        )

    # =========================================================================
    # Auth: signup / login / forgot-password / verify-email / me
    # =========================================================================

    def _client_ctx() -> dict:
        ip = (request.headers.get("X-Forwarded-For",
                                    request.remote_addr or "")
              .split(",")[0].strip())
        return {"ip": ip,
                "user_agent": request.headers.get("User-Agent", "")[:200]}

    def _auth_err_response(e: auth_mod.AuthError):
        return jsonify({"error": True, "code": e.code,
                        "message": e.message}), e.http_status

    @app.post("/api/auth/signup")
    def auth_signup():
        body = request.get_json(silent=True) or {}
        email = body.get("email") or ""
        password = body.get("password") or ""
        display_name = body.get("display_name") or ""
        ctx = _client_ctx()
        try:
            r = auth_mod.signup(email, password, display_name=display_name,
                                ip=ctx["ip"], user_agent=ctx["user_agent"])
        except auth_mod.AuthError as e:
            return _auth_err_response(e)

        # Bootstrap del Excel master para el user nuevo. Sin esto, todas
        # las páginas tiran "Excel master no presente". Mismo flow que
        # /api/admin/users: build_master + add_carga_inicial + blank
        # event sheets + auto-import.
        bootstrap_warnings = []
        try:
            new_settings = get_user_settings(r.user_id)
            new_settings.inputs_dir.mkdir(parents=True, exist_ok=True)
            new_settings.user_data_dir.mkdir(parents=True, exist_ok=True)
            new_settings.backups_dir.mkdir(parents=True, exist_ok=True)
            if not new_settings.xlsx_path.is_file():
                from build_master import build_master
                try:
                    from add_carga_inicial_sheet import add_hoja_carga_inicial as _add_ci
                except ImportError:
                    _add_ci = None
                build_master(new_settings.xlsx_path)
                if _add_ci:
                    from openpyxl import load_workbook as _wb
                    _wb_h = _wb(filename=str(new_settings.xlsx_path))
                    _add_ci(_wb_h, with_examples=False)
                    _wb_h.save(str(new_settings.xlsx_path))
                _blank_event_sheets(new_settings.xlsx_path)
                # Auto-import para crear la DB
                try:
                    reimport_excel(settings=new_settings)
                except Exception as e:
                    bootstrap_warnings.append(
                        f"Auto-import falló: {e}. Hacé refresh manual."
                    )
        except Exception as e:
            print(f"[signup] WARN bootstrap del master falló para "
                  f"{r.user_id}: {e}")
            bootstrap_warnings.append(
                f"No pude crear el Excel master automáticamente: {e}. "
                f"Subí uno manualmente desde Settings."
            )

        # En modo dev sin SMTP, devolver el verify token para que el user
        # pueda verificarse sin email funcional.
        dev_token = r.verify_token_plain if (
            r.verify_email_via != "smtp"
            and os.environ.get("WM_AUTO_VERIFY_FIRST_SUPERADMIN") == "1"
            and r.is_superadmin
        ) else None
        return jsonify({
            "user_id": r.user_id,
            "email": r.email,
            "is_superadmin": r.is_superadmin,
            "is_admin": r.is_admin,
            "verify_email_sent_via": r.verify_email_via,
            "verify_token_dev": dev_token,
            "needs_verification": r.verify_token_plain is not None,
            "bootstrap_warnings": bootstrap_warnings,
        }), 201

    @app.post("/api/auth/login")
    def auth_login():
        body = request.get_json(silent=True) or {}
        email = body.get("email") or ""
        password = body.get("password") or ""
        ctx = _client_ctx()
        try:
            r = auth_mod.login(email, password, ip=ctx["ip"],
                               user_agent=ctx["user_agent"])
        except auth_mod.AuthError as e:
            return _auth_err_response(e)
        return jsonify({
            "session_token": r.session_token,
            "session_token_prefix": r.session_token_prefix,
            "expires_at": r.expires_at,
            "user": {
                "user_id": r.user_id,
                "email": r.email,
                "display_name": r.display_name,
                "is_admin": r.is_admin,
                "is_superadmin": r.is_superadmin,
                "email_verified": r.email_verified,
            },
        })

    @app.post("/api/auth/logout")
    def auth_logout():
        token = (request.headers.get("Authorization", "")
                 .replace("Bearer ", "").strip())
        ok = auth_mod.logout(token)
        return jsonify({"ok": ok})

    @app.post("/api/auth/forgot-password")
    def auth_forgot():
        body = request.get_json(silent=True) or {}
        email = body.get("email") or ""
        ctx = _client_ctx()
        # Siempre devolvemos OK aunque el email no exista (anti-enumeration)
        result = auth_mod.request_password_reset(email, ip=ctx["ip"])
        return jsonify({
            "ok": True,
            "message": "Si el email existe, te mandamos un link de "
                        "recuperación. Revisá spam también.",
            "delivery": result.get("via"),
        })

    @app.post("/api/auth/reset-password")
    def auth_reset():
        body = request.get_json(silent=True) or {}
        token = body.get("token") or ""
        new_password = body.get("new_password") or ""
        ctx = _client_ctx()
        try:
            r = auth_mod.reset_password_with_token(token, new_password,
                                                    ip=ctx["ip"])
        except auth_mod.AuthError as e:
            return _auth_err_response(e)
        return jsonify(r)

    @app.post("/api/auth/verify-email")
    def auth_verify_email():
        body = request.get_json(silent=True) or {}
        token = body.get("token") or ""
        try:
            r = auth_mod.verify_email_with_token(token)
        except auth_mod.AuthError as e:
            return _auth_err_response(e)
        return jsonify(r)

    @app.post("/api/auth/resend-verify")
    def auth_resend_verify():
        _require_auth()
        try:
            r = auth_mod.resend_verification(g.active_user_id)
        except auth_mod.AuthError as e:
            return _auth_err_response(e)
        return jsonify(r)

    @app.post("/api/auth/change-password")
    def auth_change_password():
        _require_auth()
        body = request.get_json(silent=True) or {}
        try:
            r = auth_mod.change_password(
                g.active_user_id,
                body.get("current_password") or "",
                body.get("new_password") or "",
                current_session_token=g.user_token,
            )
        except auth_mod.AuthError as e:
            return _auth_err_response(e)
        return jsonify(r)

    @app.get("/api/auth/me")
    def auth_me():
        _require_auth()
        # Si vino por session token, devolver el perfil del auth_users.
        if getattr(g, "auth_via", None) == "session":
            profile = auth_mod.get_user_profile(g.active_user_id)
            sessions = auth_mod.list_user_sessions(g.active_user_id)
            return jsonify({
                "auth_via": "session",
                "user": profile,
                "active_sessions": len(sessions),
                "sessions": sessions,
                "is_switched": bool(g.is_switched),
            })
        # Legacy
        return jsonify({
            "auth_via": "legacy",
            "user": {
                "user_id": g.active_user_id,
                "is_admin": g.is_admin,
                "is_superadmin": g.is_superadmin,
                "email": None,
                "email_verified": None,
                "display_name": (g.auth_user.display_name
                                  if g.auth_user else g.active_user_id),
            },
            "is_switched": bool(g.is_switched),
        })

    @app.put("/api/auth/me")
    def auth_update_me():
        _require_auth()
        body = request.get_json(silent=True) or {}
        if "display_name" in body:
            try:
                r = auth_mod.update_display_name(g.active_user_id,
                                                  body["display_name"])
            except auth_mod.AuthError as e:
                return _auth_err_response(e)
            return jsonify(r)
        return jsonify({"ok": True})

    @app.delete("/api/auth/sessions/<prefix>")
    def auth_revoke_session(prefix):
        _require_auth()
        ok = auth_mod.revoke_session(g.active_user_id, prefix)
        return jsonify({"revoked": ok})

    @app.delete("/api/auth/sessions")
    def auth_revoke_all_sessions():
        """Revoca todas las sesiones excepto la actual."""
        _require_auth()
        n = auth_mod.revoke_all_sessions(g.active_user_id,
                                          except_token=g.user_token)
        return jsonify({"revoked": n})

    # =========================================================================
    # Superadmin: gestión de auth_users
    # =========================================================================

    @app.get("/api/superadmin/users")
    def superadmin_list_users():
        _require_superadmin()
        return jsonify({
            "users": auth_mod.list_all_auth_users(),
            "self": g.active_user_id,
        })

    @app.put("/api/superadmin/users/<user_id>/admin")
    def superadmin_set_admin(user_id):
        _require_superadmin()
        body = request.get_json(silent=True) or {}
        is_admin = bool(body.get("is_admin", True))
        return jsonify(auth_mod.set_admin_flag(user_id, is_admin,
                                                 g.active_user_id))

    @app.put("/api/superadmin/users/<user_id>/superadmin")
    def superadmin_set_superadmin(user_id):
        _require_superadmin()
        body = request.get_json(silent=True) or {}
        is_super = bool(body.get("is_superadmin", True))
        if user_id == g.active_user_id and not is_super:
            abort(400, "No podés removerte a vos mismo el superadmin "
                       "(promové a otro user antes).")
        return jsonify(auth_mod.set_superadmin_flag(user_id, is_super,
                                                     g.active_user_id))

    @app.delete("/api/superadmin/users/<user_id>")
    def superadmin_delete_user(user_id):
        _require_superadmin()
        if user_id == g.active_user_id:
            abort(400, "No podés borrar tu propia cuenta desde acá. "
                       "Usá /api/account.")
        try:
            return jsonify(auth_mod.delete_auth_user(user_id, g.active_user_id))
        except auth_mod.AuthError as e:
            return _auth_err_response(e)

    # =========================================================================
    # Credenciales del user (BYMA, CAFCI, etc.) — encriptadas en disk
    # =========================================================================

    @app.get("/api/credentials")
    def get_credentials_endpoint():
        """Devuelve qué credenciales tiene seteadas el user (sin valores).

        El response es {"fields": [...], "configured": {key: True}}.
        Los valores secretos NUNCA se devuelven via API.
        Los campos marcados como superadmin_only sólo aparecen si el caller
        es superadmin (ej. cafci_token).
        """
        _require_auth(); _block_if_switched_mutation()
        is_super = bool(getattr(g, "is_superadmin", False))
        existing = creds.get_credentials(g.active_user_id)
        fields = creds.list_supported_fields(is_superadmin=is_super)
        visible_keys = {f["key"] for f in fields}
        return jsonify({
            "fields": fields,
            "configured": {k: True for k, v in existing.items()
                            if v and k in visible_keys},
            "is_superadmin": is_super,
        })

    @app.put("/api/credentials")
    def update_credentials_endpoint():
        """Actualiza credenciales. Body: {key: value, ...}.

        - Pasar value="" o null borra esa credencial.
        - Solo se aceptan keys de creds.CRED_FIELDS.
        - Las credenciales marcadas como superadmin_only (ej cafci_token)
          sólo pueden modificarlas los superadmins.
        """
        _require_auth(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        if not isinstance(body, dict):
            abort(400, "Body debe ser JSON object")
        is_super = bool(getattr(g, "is_superadmin", False))
        if not is_super:
            forbidden = creds.SUPERADMIN_ONLY_FIELDS & set(body.keys())
            if forbidden:
                abort(403,
                      f"Credenciales {sorted(forbidden)} sólo las puede "
                      "configurar el superadmin.")
        result = creds.set_credentials(g.active_user_id, body)
        return jsonify({"updated": True, "configured": result})

    @app.delete("/api/credentials")
    def delete_credentials_endpoint():
        """Borra TODAS las credenciales del user."""
        _require_auth(); _block_if_switched_mutation()
        ok = creds.delete_credentials(g.active_user_id)
        return jsonify({"deleted": ok})

    # =========================================================================
    # Loaders (precios) — disparables desde la app, usando credenciales del user
    # =========================================================================

    @app.post("/api/loaders/byma")
    def loader_byma_endpoint():
        """Corre el byma_loader.run() con las credenciales del user.

        Body opcional:
          {"tickers": ["AL30D","GD30D"]}   — override de tickers a pedir.
                                              Default: tickers_byma.txt si existe.

        El CSV resultante se escribe en data/precios_historico.csv (compartido).
        Después se hace re-import del Excel del user para que tome los precios
        nuevos.
        """
        _require_auth(); _block_if_switched_mutation()
        c = creds.get_credentials(g.active_user_id)
        if not c.get("byma_user") or not c.get("byma_pass"):
            abort(400, "Faltan credenciales BYMA. Cargalas en /settings.")

        # Override de la URL si la pusieron en credenciales
        if c.get("byma_api_url"):
            os.environ["BYMA_API_URL"] = c["byma_api_url"]

        body = request.get_json(silent=True) or {}
        tickers = body.get("tickers")
        if not tickers:
            tf = Path(__file__).resolve().parent.parent / "tickers_byma.txt"
            if tf.is_file():
                from byma_loader import parse_tickers_file
                try:
                    tickers = parse_tickers_file(tf)
                except Exception as e:
                    abort(400, f"No pude leer tickers_byma.txt: {e}")
        if not tickers:
            abort(400, "Lista de tickers vacía. Pasá 'tickers' en el body.")

        s = get_settings()
        try:
            from byma_loader import run as byma_run
            rc = byma_run(
                tickers=tickers,
                output_dir=s.data_dir,
                username=c["byma_user"],
                password=c["byma_pass"],
            )
        except Exception as e:
            abort(500, f"BYMA loader falló: {type(e).__name__}: {e}")
        if rc != 0:
            abort(500, "BYMA loader devolvió error (revisá logs)")

        # Re-importar para refrescar precios
        with excel_write_lock():
            stats = reimport_excel()
        return jsonify({"loader_rc": rc, "n_tickers": len(tickers),
                        "import_stats": stats})

    @app.post("/api/loaders/cafci")
    def loader_cafci_endpoint():
        """Corre el cafci_loader.run() — SOLO superadmin.

        El token de CAFCI se lee de las credenciales del superadmin.
        El CSV resultante (data/precios_cafci.csv) es compartido entre
        todos los users, así que cualquier user que re-importe su Excel
        después de esta llamada va a ver los precios actualizados.

        Body opcional:
          {"fecha": "2026-04-30"}  — backfill de un día específico
                                     (default: último reporte diario)
        """
        _require_superadmin(); _block_if_switched_mutation()
        c = creds.get_credentials(g.active_user_id)
        token = c.get("cafci_token")
        if not token:
            abort(400, "Falta el token de CAFCI. Cargalo en /credentials.")

        body = request.get_json(silent=True) or {}
        fecha = (body.get("fecha") or "").strip() or None
        if fecha:
            try:
                date.fromisoformat(fecha)
            except ValueError:
                abort(400, f"Fecha inválida: {fecha} (esperado YYYY-MM-DD)")

        repo_root = Path(__file__).resolve().parent.parent
        fcis_file = repo_root / "fcis_cafci.txt"
        if not fcis_file.is_file():
            abort(500, f"No encuentro {fcis_file}")

        s = get_settings()
        # cafci_loader.run() lee el token de os.environ["CAFCI_TOKEN"]
        prev_token = os.environ.get("CAFCI_TOKEN")
        os.environ["CAFCI_TOKEN"] = token
        try:
            from cafci_loader import run as cafci_run
            rc = cafci_run(
                fcis_file=fcis_file,
                output_dir=s.shared_data_dir,
                fecha=fecha,
                dry_run=False,
                use_xlsx_currency=True,
                xlsx_path=s.xlsx_path,
            )
        except Exception as e:
            abort(500, f"CAFCI loader falló: {type(e).__name__}: {e}")
        finally:
            if prev_token is None:
                os.environ.pop("CAFCI_TOKEN", None)
            else:
                os.environ["CAFCI_TOKEN"] = prev_token

        if rc != 0:
            abort(500, "CAFCI loader devolvió error (revisá logs)")

        # Re-importar el Excel del superadmin para que vea los nuevos
        # precios inmediatamente. El resto de los users tomará los
        # precios cuando ellos disparen un re-import.
        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "loader_rc": rc,
            "fecha": fecha or "daily",
            "import_stats": stats,
            "shared": True,
        })

    @app.post("/api/loaders/cripto")
    def loader_cripto_endpoint():
        """Refresca precios cripto desde CoinGecko (API pública, sin auth).

        Disponible para todos los users autenticados — el CSV resultante
        (data/precios_cripto.csv) es compartido.

        Body opcional:
          {"tickers": ["BTC","ETH","USDT"]}   — override de tickers
        """
        _require_auth(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        tickers = body.get("tickers")
        if tickers and not isinstance(tickers, list):
            abort(400, "tickers debe ser una lista")

        from cripto_loader import (
            upsert_current as cripto_upsert,
            DEFAULT_TICKERS as CRIPTO_DEFAULT,
            TICKER_TO_COINGECKO,
        )
        if not tickers:
            tickers = list(CRIPTO_DEFAULT)
        tickers = [t.strip().upper() for t in tickers if t and t.strip()]
        # Filtrar tickers desconocidos para CoinGecko
        unknown = [t for t in tickers if t not in TICKER_TO_COINGECKO]
        tickers = [t for t in tickers if t in TICKER_TO_COINGECKO]
        if not tickers:
            abort(400, "Ningún ticker reconocido por CoinGecko. "
                       f"Soportados: {sorted(TICKER_TO_COINGECKO.keys())}")

        s = get_settings()
        out_path = s.shared_data_dir / "precios_cripto.csv"
        try:
            n = cripto_upsert(out_path, tickers)
        except Exception as e:
            abort(502, f"CoinGecko falló: {type(e).__name__}: {e}")

        # Re-import para que el user actual vea los precios nuevos
        # (solo si tiene un master cargado — un user fresh todavía no)
        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "n_prices": n,
            "tickers": tickers,
            "unknown_skipped": unknown,
            "import_stats": stats,
            "shared": True,
        })

    @app.post("/api/loaders/fx")
    def loader_fx_endpoint():
        """Refresca cotizaciones FX (USD/CCL, USB/MEP, USD_OFICIAL/A3500)
        desde dolarapi.com + BCRA. Sin auth. Compartido entre todos los
        users.

        Body opcional:
          {"dias": 7}   — backfill de los últimos N días desde
                          argentinadatos. Default: solo el día.
        """
        _require_auth(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        dias = body.get("dias")
        try:
            dias = int(dias) if dias else None
        except (TypeError, ValueError):
            abort(400, "dias debe ser entero")

        s = get_settings()
        try:
            from fx_loader import run as fx_run
            rc = fx_run(
                output_dir=s.shared_data_dir,
                dias=dias,
                desde=None,
                dry_run=False,
            )
        except Exception as e:
            abort(502, f"FX loader falló: {type(e).__name__}: {e}")
        if rc != 0:
            abort(500, "FX loader devolvió error (revisá logs del server)")

        # FX foráneo (EUR, BRL, GBP, etc) desde Yahoo Finance — best effort
        yf_n = 0
        try:
            from yfinance_fx_loader import (
                upsert_current as yf_fx_upsert,
                DEFAULT_CURRENCIES as YF_FX_DEFAULT,
            )
            # Filtrar a las monedas que están en `monedas` del master del user
            wanted = set(YF_FX_DEFAULT)
            if s.xlsx_path.is_file():
                try:
                    from .excel_io import list_rows
                    user_currencies = {
                        str(r.get("Code") or "").strip().upper()
                        for r in list_rows(s.xlsx_path, "monedas")
                        if r.get("Code")
                    }
                    wanted &= user_currencies or wanted
                except Exception:
                    pass
            if wanted:
                yf_n = yf_fx_upsert(
                    s.shared_data_dir / "fx_historico.csv",
                    sorted(wanted),
                )
        except Exception as e:
            print(f"[fx] WARN yfinance_fx_loader falló: {e}")

        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "loader_rc": rc,
            "yf_fx_currencies": yf_n,
            "import_stats": stats,
            "shared": True,
        })

    @app.post("/api/loaders/yfinance")
    def loader_yfinance_endpoint():
        """Refresca precios de equities US (Yahoo Finance, sin auth).

        Toma los tickers de `especies` cuyo Asset Class es EQUITY_US,
        ETF, REIT, EQUITY_GLOBAL. Convierte ticker_US → ticker yahoo
        (ej AAPL_US → AAPL).

        Body opcional:
          {"tickers": ["AAPL","MSFT"]}   — override
        """
        _require_auth(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        tickers = body.get("tickers")
        if tickers and not isinstance(tickers, list):
            abort(400, "tickers debe ser una lista")

        s = get_settings()
        if not tickers:
            if not s.xlsx_path.is_file():
                abort(400, "No hay Excel master cargado. Pasá 'tickers' "
                           "explícito o cargá tu master primero.")
            try:
                from .excel_io import list_rows
                YF_CLASSES = {"EQUITY_US", "ETF", "REIT", "EQUITY_GLOBAL"}
                tickers = sorted({
                    str(r.get("Ticker") or "").strip()
                    for r in list_rows(s.xlsx_path, "especies")
                    if r.get("Ticker") and
                       (r.get("Asset Class") or "").strip() in YF_CLASSES
                })
                tickers = [t for t in tickers if t]
            except Exception as e:
                abort(500, f"No pude leer especies: {e}")
        if not tickers:
            return jsonify({
                "n_prices": 0,
                "tickers": [],
                "message": "No hay tickers US/ETF/REIT en `especies`",
                "shared": True,
            })

        try:
            from yfinance_loader import upsert_current as yf_upsert
            n = yf_upsert(
                s.shared_data_dir / "precios_us.csv",
                tickers,
            )
        except Exception as e:
            abort(502, f"yfinance falló: {type(e).__name__}: {e}")

        stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                stats = reimport_excel()
        return jsonify({
            "n_prices": n,
            "tickers": tickers,
            "import_stats": stats,
            "shared": True,
        })

    @app.post("/api/loaders/all")
    def loader_all_endpoint():
        """Corre todos los loaders sin auth (FX + cripto + yfinance) y,
        si el caller es superadmin con creds, también CAFCI.

        BYMA queda excluido porque depende de las creds personales del
        user — se sigue corriendo desde el botón individual.

        Devuelve un dict por loader con su resultado / error.
        """
        _require_auth(); _block_if_switched_mutation()
        is_super = bool(getattr(g, "is_superadmin", False))
        s = get_settings()
        results = {}

        # --- FX (dolarapi + BCRA) ---
        try:
            from fx_loader import run as fx_run
            rc = fx_run(output_dir=s.shared_data_dir, dias=None,
                         desde=None, dry_run=False)
            results["fx"] = {"ok": rc == 0, "rc": rc}
        except Exception as e:
            results["fx"] = {"ok": False, "error": f"{type(e).__name__}: {e}"}

        # --- FX foráneo (yfinance) ---
        try:
            from yfinance_fx_loader import (
                upsert_current as yf_fx_upsert,
                DEFAULT_CURRENCIES as YF_FX_DEFAULT,
            )
            wanted = set(YF_FX_DEFAULT)
            if s.xlsx_path.is_file():
                try:
                    from .excel_io import list_rows
                    user_currencies = {
                        str(r.get("Code") or "").strip().upper()
                        for r in list_rows(s.xlsx_path, "monedas")
                        if r.get("Code")
                    }
                    wanted &= user_currencies or wanted
                except Exception:
                    pass
            n = yf_fx_upsert(
                s.shared_data_dir / "fx_historico.csv",
                sorted(wanted),
            ) if wanted else 0
            results["yfinance_fx"] = {"ok": True, "n_currencies": n}
        except Exception as e:
            results["yfinance_fx"] = {"ok": False,
                                       "error": f"{type(e).__name__}: {e}"}

        # --- Cripto (CoinGecko) ---
        try:
            from cripto_loader import (
                upsert_current as cripto_upsert,
                DEFAULT_TICKERS as CRIPTO_DEFAULT,
                TICKER_TO_COINGECKO,
            )
            tickers = list(CRIPTO_DEFAULT)
            # Agregar los cripto que el user tenga en `especies`
            if s.xlsx_path.is_file():
                try:
                    from .excel_io import list_rows
                    CRYPTO_CLASSES = {"CRYPTO", "STABLECOIN"}
                    user_cripto = {
                        str(r.get("Ticker") or "").strip().upper()
                        for r in list_rows(s.xlsx_path, "especies")
                        if r.get("Ticker") and
                           (r.get("Asset Class") or "").strip() in CRYPTO_CLASSES
                    }
                    tickers = sorted(set(tickers) | user_cripto)
                except Exception:
                    pass
            tickers = [t for t in tickers if t in TICKER_TO_COINGECKO]
            n = cripto_upsert(
                s.shared_data_dir / "precios_cripto.csv",
                tickers,
            ) if tickers else 0
            results["cripto"] = {"ok": True, "n_prices": n}
        except Exception as e:
            results["cripto"] = {"ok": False,
                                  "error": f"{type(e).__name__}: {e}"}

        # --- yfinance equities US ---
        try:
            us_tickers = []
            if s.xlsx_path.is_file():
                from .excel_io import list_rows
                YF_CLASSES = {"EQUITY_US", "ETF", "REIT", "EQUITY_GLOBAL"}
                us_tickers = sorted({
                    str(r.get("Ticker") or "").strip()
                    for r in list_rows(s.xlsx_path, "especies")
                    if r.get("Ticker") and
                       (r.get("Asset Class") or "").strip() in YF_CLASSES
                })
                us_tickers = [t for t in us_tickers if t]
            if us_tickers:
                from yfinance_loader import upsert_current as yf_upsert
                n = yf_upsert(
                    s.shared_data_dir / "precios_us.csv",
                    us_tickers,
                )
                results["yfinance"] = {"ok": True, "n_prices": n,
                                        "n_tickers": len(us_tickers)}
            else:
                results["yfinance"] = {"ok": True, "n_prices": 0,
                                        "message": "Sin tickers US"}
        except Exception as e:
            results["yfinance"] = {"ok": False,
                                    "error": f"{type(e).__name__}: {e}"}

        # --- CAFCI (solo superadmin con token) ---
        if is_super:
            c = creds.get_credentials(g.active_user_id)
            if c.get("cafci_token"):
                try:
                    repo_root = Path(__file__).resolve().parent.parent
                    fcis_file = repo_root / "fcis_cafci.txt"
                    if fcis_file.is_file():
                        prev_token = os.environ.get("CAFCI_TOKEN")
                        os.environ["CAFCI_TOKEN"] = c["cafci_token"]
                        try:
                            from cafci_loader import run as cafci_run
                            rc = cafci_run(
                                fcis_file=fcis_file,
                                output_dir=s.shared_data_dir,
                                fecha=None,
                                dry_run=False,
                                use_xlsx_currency=True,
                                xlsx_path=s.xlsx_path,
                            )
                            results["cafci"] = {"ok": rc == 0, "rc": rc}
                        finally:
                            if prev_token is None:
                                os.environ.pop("CAFCI_TOKEN", None)
                            else:
                                os.environ["CAFCI_TOKEN"] = prev_token
                    else:
                        results["cafci"] = {"ok": False,
                                             "error": "fcis_cafci.txt no existe"}
                except Exception as e:
                    results["cafci"] = {"ok": False,
                                         "error": f"{type(e).__name__}: {e}"}
            else:
                results["cafci"] = {"ok": False,
                                     "error": "Sin token CAFCI configurado"}

        # Re-import si hay master
        import_stats = None
        if s.xlsx_path.is_file():
            with excel_write_lock():
                import_stats = reimport_excel()

        return jsonify({
            "results": results,
            "import_stats": import_stats,
            "is_superadmin": is_super,
        })

    # =========================================================================
    # Auto-import de tenencias desde brokers (Cocos / Binance / IBKR)
    # =========================================================================

    @app.get("/api/import/brokers")
    def import_list_brokers():
        """Devuelve los brokers soportados + qué credenciales tiene el user."""
        _require_auth()
        from engine import brokers
        brokers_meta = brokers.list_brokers()
        configured = creds.get_credentials(g.active_user_id)
        for b in brokers_meta:
            b["ready"] = all(configured.get(k) for k in b["needs"])
        return jsonify({"brokers": brokers_meta})

    @app.post("/api/import/<broker>/test")
    def import_test_credentials(broker):
        """Hace un ping a las credenciales del broker para validarlas
        SIN bajar todo el historial. Útil para feedback rápido al user."""
        _require_auth(); _block_if_switched_mutation()
        from engine import brokers
        mod = brokers.get(broker)
        if mod is None:
            abort(404, f"Broker '{broker}' no soportado")
        c = creds.get_credentials(g.active_user_id)
        # Algunos brokers exponen test_credentials, otros no
        test_fn = getattr(mod, "test_credentials", None)
        if test_fn is None:
            # Fallback: hacer fetch_positions y reportar
            try:
                r = mod.fetch_positions(c)
                return jsonify({"ok": True, "n_positions": len(r["positions"])})
            except Exception as e:
                return jsonify({"ok": False,
                                 "error": f"{type(e).__name__}: {e}"}), 200
        return jsonify(test_fn(c))

    def _fetch_and_annotate_preview(user_id: str, broker: str) -> dict:
        """Trabajo sincrónico de la preview: fetch al broker + match contra
        `especies`/`monedas` del Excel del user. Usado tanto por el endpoint
        sync como por el background job.

        No depende del request context — usa user_id explícito.
        """
        from engine import brokers
        from .state import get_user_settings
        mod = brokers.get(broker)
        if mod is None:
            raise ValueError(f"Broker '{broker}' no soportado")
        c = creds.get_credentials(user_id)
        result = mod.fetch_positions(c)

        # Match contra `especies` para flagear assets desconocidos
        s = get_user_settings(user_id)
        existing_tickers = set()
        existing_currencies = set()
        try:
            from .excel_io import list_rows
            for row in list_rows(s.xlsx_path, "especies"):
                t = row.get("Ticker")
                if t:
                    existing_tickers.add(str(t).strip())
            for row in list_rows(s.xlsx_path, "monedas"):
                code = row.get("Code")
                if code:
                    existing_currencies.add(str(code).strip())
        except Exception:
            pass
        for p in result["positions"]:
            t = p["ticker"]
            if p["is_cash"]:
                p["known"] = t in existing_currencies
            else:
                p["known"] = t in existing_tickers
        return result

    @app.post("/api/import/<broker>/preview")
    def import_preview(broker):
        """Trae las posiciones del broker y las devuelve SIN escribir nada.

        Endpoint SINCRÓNICO — para brokers rápidos (Cocos, Binance). Para
        IBKR que puede tardar > timeout del proxy, usar el endpoint async
        (preview/start + jobs/<id>).

        Body opcional:
          {"match_assets": true}  → si True, marca cuáles tickers ya
                                    existen en `especies` del Excel
                                    para que la UI los pueda diferenciar
                                    de los nuevos.
        """
        _require_auth(); _block_if_switched_mutation()
        from engine import brokers
        if brokers.get(broker) is None:
            abort(404, f"Broker '{broker}' no soportado")
        try:
            result = _fetch_and_annotate_preview(g.active_user_id, broker)
        except (ValueError, PermissionError) as e:
            abort(400, str(e))
        except Exception as e:
            abort(502, f"Error del broker: {type(e).__name__}: {e}")
        return jsonify(result)

    @app.post("/api/import/<broker>/preview/start")
    def import_preview_start(broker):
        """Arranca preview en background y devuelve job_id para pollear.

        Resuelve el problema de IBKR Flex tardando > 60s y matando el worker
        de gunicorn. El cliente pollea GET /api/import/jobs/<job_id> hasta
        que status="done" o "error".
        """
        _require_auth(); _block_if_switched_mutation()
        from engine import brokers
        if brokers.get(broker) is None:
            abort(404, f"Broker '{broker}' no soportado")
        user_id = g.active_user_id
        job_id = bg_jobs.create_job(
            user_id=user_id,
            kind=f"import_preview:{broker}",
            fn=lambda: _fetch_and_annotate_preview(user_id, broker),
        )
        return jsonify({"job_id": job_id, "status": "pending"})

    @app.get("/api/import/jobs/<job_id>")
    def import_job_status(job_id):
        """Devuelve el estado del job. Status: pending|done|error.

        Si done, `result` tiene la preview completa. Si error, `error` tiene
        el mensaje. Frontend hace polling hasta que cambie de pending.
        """
        _require_auth()
        job = bg_jobs.get_job(g.active_user_id, job_id)
        if job is None:
            abort(404, f"Job {job_id} no encontrado")
        return jsonify(job)

    @app.post("/api/import/<broker>/apply")
    def import_apply(broker):
        """Aplica una preview al Excel master.

        Body:
          {
            "account": "cocos",                     # cuenta destino (debe existir)
            "fecha": "2026-05-06",                   # fecha del saldo inicial (default hoy)
            "positions": [                           # subset que el user confirmó
              {"ticker":"AL30D","qty":1000,"avg_price":65.5,"currency":"USB",
               "asset_class":"BOND_AR","name":"Bonar 2030",
               "is_cash":false,"register_asset":true},
              ...
            ],
            "create_missing_assets": true            # crear assets nuevos en `especies`
          }

        Crea filas en la hoja `_carga_inicial` del Excel master, las
        materializa como OPENING_BALANCE en `asientos_contables`
        (vía cli.cargar_iniciales.generate_asientos) y re-importa.
        """
        _require_auth(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        account = (body.get("account") or "").strip()
        fecha = body.get("fecha") or date.today().isoformat()
        positions = body.get("positions") or []
        create_missing = bool(body.get("create_missing_assets", True))
        if not account or not positions:
            abort(400, "account y positions son requeridos")

        result = _write_carga_inicial(
            account=account, fecha=fecha, positions=positions,
            create_missing=create_missing, broker_label=broker,
        )
        result["broker"] = broker
        return jsonify(result)

    @app.post("/api/carga-inicial")
    def carga_inicial_manual():
        """Carga manual de tenencias viejas (saldos de apertura).

        Mismo contrato que /api/import/<broker>/apply pero sin broker:
        el user arma la lista de posiciones a mano (desde el wizard).

        Body:
          {
            "account": "cocos",
            "fecha": "2026-05-06",
            "positions": [
              {"ticker":"AL30D","qty":500,"unit_price":65.5,"currency":"USB",
               "asset_class":"BOND_AR","name":"Bonar 2030",
               "is_cash":false,"strategy":"BH"},
              {"ticker":"ARS","qty":200000,"is_cash":true,"currency":"ARS"},
              ...
            ],
            "create_missing_assets": true
          }
        """
        _require_auth(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        account = (body.get("account") or "").strip()
        fecha = body.get("fecha") or date.today().isoformat()
        positions = body.get("positions") or []
        create_missing = bool(body.get("create_missing_assets", True))
        if not account or not positions:
            abort(400, "account y positions son requeridos")

        # Aceptar 'unit_price' (manual) o 'avg_price' (broker preview)
        for p in positions:
            if "unit_price" in p and "avg_price" not in p:
                p["avg_price"] = p["unit_price"]

        result = _write_carga_inicial(
            account=account, fecha=fecha, positions=positions,
            create_missing=create_missing, broker_label="manual",
        )
        return jsonify(result)

    def _write_carga_inicial(account, fecha, positions, create_missing,
                              broker_label):
        """Escribe filas a `_carga_inicial`, las materializa como
        OPENING_BALANCE en `asientos_contables` (via cli.cargar_iniciales)
        y re-importa el Excel.

        Usado tanto por /api/import/<broker>/apply como por la carga manual.
        """
        s = get_settings()
        if not s.xlsx_path.is_file():
            abort(400, "No hay Excel master cargado todavía")

        from .excel_io import list_rows
        # Validar cuenta existe
        accounts = {r.get("Code"): r
                     for r in list_rows(s.xlsx_path, "cuentas")
                     if r.get("Code")}
        if account not in accounts:
            abort(400, f"Cuenta '{account}' no existe en `cuentas`")

        # Assets / monedas existentes
        existing_assets = {r.get("Ticker"): r
                            for r in list_rows(s.xlsx_path, "especies")
                            if r.get("Ticker")}
        existing_ccy = set(r.get("Code") for r in list_rows(s.xlsx_path, "monedas")
                            if r.get("Code"))

        from openpyxl import load_workbook as _load_wb

        # Headers de _carga_inicial:
        #   Fecha | Cuenta | Activo | Qty | Unit Price | Price Currency
        #   | Strategy | Description | Notes
        with excel_write_lock():
            backup_excel()
            n_assets = 0
            n_initial = 0
            warnings = []

            wb = _load_wb(filename=str(s.xlsx_path))
            if "_carga_inicial" not in wb.sheetnames:
                # La hoja no existe — la creamos sin filas de ejemplo (las
                # filas de ejemplo del template default se procesarían como
                # asientos reales y aparecerían como ghost positions)
                try:
                    from add_carga_inicial_sheet import add_hoja_carga_inicial
                    add_hoja_carga_inicial(wb, with_examples=False)
                except Exception as e:
                    abort(500, f"No pude crear hoja _carga_inicial: {e}")
            ws_in = wb["_carga_inicial"]

            for p in positions:
                ticker = (p.get("ticker") or "").strip()
                qty = float(p.get("qty") or 0)
                ccy = (p.get("currency") or "ARS").strip().upper()
                avg = p.get("avg_price")
                is_cash = bool(p.get("is_cash"))
                cls = p.get("asset_class") or "OTHER"
                name = p.get("name") or ticker
                strategy = (p.get("strategy") or "").strip() or None
                description = (p.get("description")
                                 or f"Saldo inicial {ticker} ({broker_label})")

                if not ticker or abs(qty) < 1e-9:
                    continue

                # Crear asset si no existe (solo no-cash)
                if not is_cash and ticker not in existing_assets:
                    if create_missing:
                        try:
                            # Cerrar el wb actual para que append_row no
                            # sobreescriba lo que tenemos en memoria.
                            wb.save(str(s.xlsx_path))
                            from .excel_io import append_row as _append_row
                            _append_row(s.xlsx_path, "especies", {
                                "Ticker": ticker,
                                "Name": name or ticker,
                                "Asset Class": cls,
                                "Currency": ccy,
                                "Notes": f"Auto-importado de {broker_label}",
                            })
                            existing_assets[ticker] = True
                            n_assets += 1
                            wb = _load_wb(filename=str(s.xlsx_path))
                            ws_in = wb["_carga_inicial"]
                        except Exception as e:
                            warnings.append(
                                f"No pude crear asset {ticker}: {e}"
                            )
                            continue
                    else:
                        warnings.append(
                            f"Asset {ticker} no existe — skipped"
                        )
                        continue

                # Cash: validar que la moneda exista
                if is_cash and ticker not in existing_ccy:
                    warnings.append(
                        f"Moneda {ticker} no existe en `monedas` — skipped"
                    )
                    continue

                unit_price = avg if avg else (1.0 if is_cash else None)
                price_currency = ccy if not is_cash else None
                # Cash: el motor usa Activo=ARS/USB/etc y no necesita
                # unit_price ni price_currency (es la propia moneda).
                row_vals = [
                    fecha,           # A: Fecha
                    account,         # B: Cuenta
                    ticker,          # C: Activo
                    qty,             # D: Qty
                    unit_price,      # E: Unit Price
                    price_currency,  # F: Price Currency
                    strategy or ("CASH" if is_cash else None),  # G: Strategy
                    description,     # H: Description
                    None,            # I: Notes
                ]
                # Append en la primera fila vacía después del header
                next_row = ws_in.max_row + 1
                # Buscar la primera fila realmente vacía (algunas hojas
                # tienen filas con valores None que dejan max_row alto)
                for r in range(5, ws_in.max_row + 2):
                    if all(ws_in.cell(row=r, column=c).value in (None, "")
                            for c in range(1, 10)):
                        next_row = r
                        break
                for c, val in enumerate(row_vals, start=1):
                    ws_in.cell(row=next_row, column=c, value=val)
                n_initial += 1

            wb.save(str(s.xlsx_path))

            # Materializar las filas a `asientos_contables` como
            # OPENING_BALANCE — sin esto, el motor nunca las ve.
            asientos_stats = {}
            try:
                from cli.cargar_iniciales import generate_asientos
                fx_csv = s.shared_data_dir / "fx_historico.csv"
                asientos_stats = generate_asientos(
                    s.xlsx_path,
                    fecha_default=date.fromisoformat(fecha),
                    dry_run=False,
                    fx_csv_path=fx_csv,
                )
            except Exception as e:
                warnings.append(
                    f"generate_asientos falló: {type(e).__name__}: {e}"
                )

            # Re-importar para que el motor procese la carga inicial
            stats = reimport_excel()
            prune_backups(keep_last=50)

        return {
            "ok": True,
            "account": account,
            "fecha": fecha,
            "n_assets_created": n_assets,
            "n_positions_written": n_initial,
            "n_asientos_generated": (asientos_stats or {})
                .get("filas_generadas_asientos", 0),
            "warnings": warnings,
            "import_stats": stats,
        }

    # =========================================================================
    # Conciliación de cash bancario
    # =========================================================================
    #
    # Modelo: el motor calcula saldo de cada (cuenta_cash, moneda) sumando
    # los movements del ledger. El user puede reportar el saldo REAL del
    # banco/wallet/broker, y la API genera un asiento ACCOUNTING_ADJUSTMENT
    # que cierra el delta. Idempotente por Event ID
    # `RECON-{cuenta}-{moneda}-{fecha}` — re-cargar el mismo (cuenta, moneda,
    # fecha) reemplaza el asiento previo.
    #
    # La contracuenta del ajuste es `opening_balance` (boundary, no entra en
    # P&L ni en holdings). Eso evita inflar/deflar las cuentas de
    # external_income / external_expense con ajustes técnicos.

    CASH_KINDS = ("CASH_BANK", "CASH_BROKER", "CASH_WALLET", "CASH_PHYSICAL")

    @app.get("/api/conciliacion/cash")
    def conciliacion_get():
        """Lista cuentas de cash con su saldo calculado por moneda y la
        última conciliación registrada (si existe).

        Response:
          {
            "accounts": [
              {
                "cuenta": "cocos",
                "name": "Cocos broker",
                "kind": "CASH_BROKER",
                "currency": "ARS",   # default currency (puede haber otras)
                "saldos": [
                  {"moneda": "ARS", "saldo": 12345.67,
                   "last_recon_date": "2026-04-30",
                   "last_recon_real": 12500.00},
                  ...
                ]
              }, ...
            ]
          }
        """
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if not s.db_path.is_file():
            return jsonify({"accounts": []})
        conn = db_conn()
        try:
            placeholders = ",".join("?" for _ in CASH_KINDS)
            cur = conn.execute(
                f"SELECT code, name, kind, currency FROM accounts "
                f"WHERE kind IN ({placeholders}) "
                f"ORDER BY kind, code",
                CASH_KINDS,
            )
            cuentas = [dict(r) for r in cur.fetchall()]

            accounts = []
            for acc in cuentas:
                # Saldos por asset (filtrar solo monedas — los assets no-cash
                # no aplican para conciliación de banco)
                cur2 = conn.execute(
                    """
                    SELECT m.asset, COALESCE(SUM(m.qty), 0) as saldo
                    FROM movements m
                    WHERE m.account = ?
                      AND m.asset IN (SELECT code FROM currencies)
                    GROUP BY m.asset
                    """,
                    (acc["code"],),
                )
                saldos_by_moneda = {}
                for r in cur2.fetchall():
                    saldo = float(r["saldo"] or 0)
                    saldos_by_moneda[r["asset"]] = saldo

                # Asegurar que la moneda nativa de la cuenta aparezca aún si
                # tiene saldo 0 — el user puede querer iniciar conciliación
                # desde cero.
                if acc.get("currency"):
                    saldos_by_moneda.setdefault(acc["currency"], 0.0)

                saldos = []
                for moneda, saldo in sorted(saldos_by_moneda.items()):
                    # Saltar saldos cero EXCEPTO la moneda nativa (para que
                    # cuentas vacías sigan apareciendo).
                    if abs(saldo) < 1e-6 and moneda != acc.get("currency"):
                        continue
                    cur3 = conn.execute(
                        """
                        SELECT e.event_date
                        FROM events e
                        JOIN movements m2 ON m2.event_id = e.id
                        WHERE m2.account = ? AND m2.asset = ?
                          AND e.type = 'ACCOUNTING_ADJUSTMENT'
                          AND e.external_id LIKE 'RECON-%'
                        ORDER BY e.event_date DESC, e.id DESC
                        LIMIT 1
                        """,
                        (acc["code"], moneda),
                    )
                    last = cur3.fetchone()
                    saldos.append({
                        "moneda": moneda,
                        "saldo_calculado": round(saldo, 6),
                        "last_recon_date": last["event_date"] if last else None,
                    })
                if saldos:
                    accounts.append({
                        "cuenta": acc["code"],
                        "name": acc["name"],
                        "kind": acc["kind"],
                        "currency": acc["currency"],
                        "saldos": saldos,
                    })
            return jsonify({"accounts": accounts})
        finally:
            conn.close()

    @app.post("/api/conciliacion/cash")
    def conciliacion_post():
        """Carga un saldo real de una cuenta cash y genera/actualiza el
        asiento de ajuste necesario para cerrar el delta.

        Body:
          {
            "cuenta": "cocos",
            "moneda": "ARS",
            "fecha": "2026-05-09",          # default: hoy
            "saldo_real": 12500.00,
            "notes": "..."                   # opcional
          }
        """
        _require_auth(); _block_if_switched_mutation()
        body = request.get_json(silent=True) or {}
        cuenta = (body.get("cuenta") or "").strip()
        moneda = (body.get("moneda") or "").strip().upper()
        fecha = body.get("fecha") or date.today().isoformat()
        saldo_raw = body.get("saldo_real")
        notes = (body.get("notes") or "").strip()

        if not cuenta or not moneda or saldo_raw is None:
            abort(400, "cuenta, moneda y saldo_real son requeridos")
        try:
            saldo_real = float(saldo_raw)
        except (TypeError, ValueError):
            abort(400, "saldo_real debe ser numérico")
        try:
            date.fromisoformat(fecha)
        except ValueError:
            abort(400, f"fecha inválida: {fecha} (YYYY-MM-DD)")

        s = get_settings()
        if not s.xlsx_path.is_file():
            abort(400, "No hay Excel master cargado todavía")

        # Validaciones contra la DB actual + saldo calculado
        conn = db_conn()
        try:
            cur = conn.execute("SELECT kind FROM accounts WHERE code=?", (cuenta,))
            row = cur.fetchone()
            if not row:
                abort(400, f"Cuenta '{cuenta}' no existe")
            if row["kind"] not in CASH_KINDS:
                abort(400, f"Cuenta '{cuenta}' no es de cash (kind={row['kind']})")
            cur = conn.execute(
                "SELECT 1 FROM accounts WHERE code='opening_balance' LIMIT 1"
            )
            if not cur.fetchone():
                abort(400, "Falta la cuenta técnica `opening_balance` en "
                           "tu hoja `cuentas`. Agregala (kind=OPENING_BALANCE) "
                           "antes de conciliar.")
            cur = conn.execute(
                "SELECT 1 FROM currencies WHERE code=? LIMIT 1", (moneda,)
            )
            if not cur.fetchone():
                abort(400, f"Moneda '{moneda}' no existe en `monedas`")

            # Saldo calculado al fecha — sumamos qty de movements cuyo
            # event.event_date <= fecha. Excluimos asientos previos del
            # mismo Event ID porque estamos por reemplazarlos.
            event_id = f"RECON-{cuenta}-{moneda}-{fecha}".upper()
            cur = conn.execute(
                """
                SELECT COALESCE(SUM(m.qty), 0) AS saldo
                FROM movements m
                JOIN events e ON e.id = m.event_id
                WHERE m.account = ? AND m.asset = ?
                  AND e.event_date <= ?
                  AND COALESCE(e.external_id, '') <> ?
                """,
                (cuenta, moneda, fecha, event_id),
            )
            saldo_calculado = float(cur.fetchone()["saldo"] or 0)
        finally:
            conn.close()

        delta = round(saldo_real - saldo_calculado, 6)
        if abs(delta) < 1e-6:
            return jsonify({
                "ok": True,
                "cuenta": cuenta, "moneda": moneda, "fecha": fecha,
                "saldo_calculado": saldo_calculado,
                "saldo_real": saldo_real,
                "delta": 0.0,
                "asiento_generado": False,
                "message": "Sin diferencias — no se generó asiento.",
            })

        description = (
            f"Conciliación {cuenta} {moneda}: real={saldo_real:.2f} "
            f"vs calc={saldo_calculado:.2f} (Δ={delta:+.2f})"
        )
        if notes:
            description = f"{description} — {notes[:140]}"

        from openpyxl import load_workbook as _load_wb
        with excel_write_lock():
            backup_excel()
            wb = _load_wb(filename=str(s.xlsx_path))
            if "asientos_contables" not in wb.sheetnames:
                abort(500, "Hoja asientos_contables no existe en el master")
            ws = wb["asientos_contables"]
            headers = [c.value for c in ws[4]]
            col = {str(h).strip(): i + 1
                    for i, h in enumerate(headers) if h}
            required = ("Event ID", "Fecha", "Description",
                         "Cuenta", "Activo", "Qty")
            missing = [h for h in required if h not in col]
            if missing:
                abort(500, f"asientos_contables sin columnas: {missing}")

            # Borrar filas previas con el mismo Event ID (idempotencia)
            rows_to_delete = []
            for r in range(5, ws.max_row + 1):
                eid = ws.cell(row=r, column=col["Event ID"]).value
                if eid and isinstance(eid, str) and \
                        eid.strip().upper() == event_id:
                    rows_to_delete.append(r)
            for r in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(r, 1)

            # Apendear las dos filas (cuenta + opening_balance)
            for cuenta_mov, qty_mov, mov_notes in [
                (cuenta, delta, "Ajuste por saldo bancario reportado"),
                ("opening_balance", -delta, "Contracuenta de conciliación"),
            ]:
                # Buscar primera fila vacía (en lugar de usar max_row,
                # que puede ser alto si hay celdas con None pero borradas)
                next_row = ws.max_row + 1
                for r in range(5, ws.max_row + 2):
                    if all(ws.cell(row=r, column=c).value in (None, "")
                            for c in range(1, len(headers) + 1)):
                        next_row = r
                        break
                fields = {
                    "Event ID": event_id,
                    "Fecha": fecha,
                    "Description": description,
                    "Cuenta": cuenta_mov,
                    "Activo": moneda,
                    "Qty": qty_mov,
                    "Notes": mov_notes,
                }
                for h, val in fields.items():
                    if h in col:
                        ws.cell(row=next_row, column=col[h], value=val)

            wb.save(str(s.xlsx_path))
            stats = reimport_excel()
            prune_backups(keep_last=50)

        return jsonify({
            "ok": True,
            "cuenta": cuenta, "moneda": moneda, "fecha": fecha,
            "saldo_calculado": saldo_calculado,
            "saldo_real": saldo_real,
            "delta": delta,
            "asiento_generado": True,
            "event_id": event_id,
            "import_stats": stats,
        })

    # =========================================================================
    # Audit log per-user (read-only)
    # =========================================================================

    @app.get("/api/audit-log")
    def audit_log_endpoint():
        """Devuelve las últimas 100 entradas del audit log del user."""
        _require_auth()
        try:
            n = int(request.args.get("n", "100"))
        except ValueError:
            n = 100
        n = max(1, min(n, 1000))
        return jsonify({
            "user_id": g.active_user_id,
            "entries": audit.tail(g.active_user_id, n=n),
        })

    # =========================================================================
    # Account self-service: export + delete (Ley 25.326 / GDPR-equivalent)
    # =========================================================================

    @app.get("/api/account/export")
    def account_export_endpoint():
        """Devuelve un ZIP con los datos del user:
          - wealth_management.xlsx (Excel master)
          - wealth.db              (DB sqlite derivada)
          - audit.log              (log de mutations)
          - credentials.enc        NO se incluye por seguridad (cliente las re-setea)

        Útil para: backup local, migrar a otro deploy, cumplir derecho de
        portabilidad (Ley 25.326).
        """
        _require_auth(); _block_if_switched_mutation()
        import io as _io
        import zipfile as _zip
        s = get_settings()
        buf = _io.BytesIO()
        with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED) as zf:
            if s.xlsx_path.is_file():
                zf.write(s.xlsx_path, "wealth_management.xlsx")
            if s.db_path.is_file():
                zf.write(s.db_path, "wealth.db")
            audit_path = s.user_data_dir / "audit.log"
            if audit_path.is_file():
                zf.write(audit_path, "audit.log")
            # README explicativo
            zf.writestr(
                "README.txt",
                "Export de tu cuenta WM Wealth Management.\n"
                f"User ID: {g.active_user_id}\n"
                f"Generado: {datetime.now().isoformat()}\n\n"
                "Archivos:\n"
                "  wealth_management.xlsx  — tus inputs (cuentas, trades, gastos, ...)\n"
                "  wealth.db               — vista derivada en sqlite\n"
                "  audit.log               — log de mutations vía API\n\n"
                "NO se incluye credentials.enc por seguridad.\n",
            )
        buf.seek(0)
        from flask import Response as _Resp
        fname = f"wm-export-{g.active_user_id}-{date.today().isoformat()}.zip"
        return _Resp(
            buf.read(), mimetype="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.post("/api/account/bootstrap")
    def account_bootstrap_endpoint():
        """Crea el Excel master inicial para el user actual si no existe.

        Útil para users registrados antes de que el signup hiciera
        bootstrap automático. Idempotente: si el Excel ya existe, no
        toca nada y devuelve already=True.
        """
        _require_auth(); _block_if_switched_mutation()
        s = get_settings()
        if s.xlsx_path.is_file():
            return jsonify({"already": True, "xlsx_path": str(s.xlsx_path)})
        s.inputs_dir.mkdir(parents=True, exist_ok=True)
        s.user_data_dir.mkdir(parents=True, exist_ok=True)
        s.backups_dir.mkdir(parents=True, exist_ok=True)
        warnings = []
        try:
            from build_master import build_master
            try:
                from add_carga_inicial_sheet import add_hoja_carga_inicial as _add_ci
            except ImportError:
                _add_ci = None
            build_master(s.xlsx_path)
            if _add_ci:
                from openpyxl import load_workbook as _wb
                _wb_h = _wb(filename=str(s.xlsx_path))
                _add_ci(_wb_h, with_examples=False)
                _wb_h.save(str(s.xlsx_path))
            _blank_event_sheets(s.xlsx_path)
            try:
                stats = reimport_excel()
            except Exception as e:
                warnings.append(f"Auto-import falló: {e}")
                stats = None
        except Exception as e:
            abort(500, f"No pude generar el master: {type(e).__name__}: {e}")
        return jsonify({
            "ok": True,
            "xlsx_path": str(s.xlsx_path),
            "import_stats": stats,
            "warnings": warnings,
        })

    @app.delete("/api/account")
    def account_delete_endpoint():
        """Borra TODOS los datos del user actual (excepto la entrada de
        users.json — eso lo maneja el admin si quiere reusar el user_id).

        Requiere header `X-Confirm-Delete: yes` para evitar accidentes.
        Borra: xlsx, db, backups, credentials.enc, audit.log.
        """
        _require_auth(); _block_if_switched_mutation()
        if request.headers.get("X-Confirm-Delete") != "yes":
            abort(400, "Falta header X-Confirm-Delete: yes")
        import shutil as _sh
        s = get_settings()
        deleted = []
        for path in (s.xlsx_path, s.db_path,
                     s.user_data_dir / "credentials.enc",
                     s.user_data_dir / "audit.log"):
            try:
                if path.is_file():
                    path.unlink(); deleted.append(str(path.name))
            except OSError:
                pass
        try:
            if s.backups_dir.is_dir():
                _sh.rmtree(s.backups_dir); deleted.append("excel_backups/")
        except OSError:
            pass
        return jsonify({"deleted": deleted, "user_id": g.active_user_id,
                        "warning": "Tu user sigue existiendo. Pedile al admin"
                                    " que lo elimine para liberar el user_id."})

    # =========================================================================
    # Reports
    # =========================================================================

    @app.get("/api/report/html")
    def report_html():
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        s = get_settings()
        out = s.data_dir / f"_tmp_report_{fecha.isoformat()}.html"
        conn = db_conn()
        try:
            export_html(conn, out, fecha=fecha, anchor_currency=anchor,
                        record_snapshot=True)
        finally:
            conn.close()
        html = out.read_text(encoding="utf-8")
        try:
            out.unlink()
        except OSError:
            pass
        return Response(html, mimetype="text/html")

    @app.get("/api/report/excel")
    def report_excel():
        _require_auth(); _block_if_switched_mutation()
        anchor = _parse_query_anchor()
        fecha = _parse_query_fecha()
        s = get_settings()
        out = s.data_dir / f"_tmp_report_{fecha.isoformat()}.xlsx"
        conn = db_conn()
        try:
            export_excel(conn, out, fecha=fecha, anchor_currency=anchor,
                         record_snapshot=True)
        finally:
            conn.close()
        return send_file(
            str(out), as_attachment=True,
            download_name=f"{fecha.isoformat()}_portfolio.xlsx",
        )

    # =========================================================================
    # Error handlers
    # =========================================================================

    @app.errorhandler(400)
    @app.errorhandler(401)
    @app.errorhandler(403)
    @app.errorhandler(404)
    @app.errorhandler(500)
    @app.errorhandler(502)
    @app.errorhandler(503)
    @app.errorhandler(504)
    def err(e):
        return jsonify({
            "error": True,
            "code": getattr(e, "code", 500),
            "message": getattr(e, "description", str(e)),
        }), getattr(e, "code", 500)

    return app


def _clean_for_json(d: dict) -> dict:
    """Remueve campos que no son JSON-serializable y formatea None/inf."""
    import math
    out = {}
    for k, v in d.items():
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            out[k] = None
        elif isinstance(v, (date, datetime)):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out


# Para `flask --app api.app run` y para WSGI:
app = create_app()
