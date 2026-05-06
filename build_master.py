# -*- coding: utf-8 -*-
"""
build_master.py

Construye el Excel master `wealth_management_rodricor.xlsx` con las 13 hojas
necesarias para Sub-sprint 1A: blotter, gastos, pasivos, asientos contables,
y todas las hojas de soporte.

Genera headers, validaciones (DataValidation), filas ejemplo, estilos.

USO:
    python3 build_master.py [output_path]
    python3 build_master.py inputs/wealth_management_rodricor.xlsx
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


# =============================================================================
# Estilos
# =============================================================================

NAVY = "1F3864"
WHITE = "FFFFFF"
GRAY_LIGHT = "F2F2F2"
BLUE_INPUT = "0000FF"
YELLOW_INPUT = "FFF2CC"
GREEN_LINK = "008000"

FONT_TITLE = Font(name="Arial", size=14, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="Arial", size=10, italic=True, color="595959")
FONT_HEADER = Font(name="Arial", size=11, bold=True, color=WHITE)
FONT_INPUT = Font(name="Arial", size=11, color=BLUE_INPUT)
FONT_NORMAL = Font(name="Arial", size=11)
FONT_FORMULA = Font(name="Arial", size=11, color="000000")

FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_INPUT = PatternFill("solid", fgColor=YELLOW_INPUT)
FILL_GRAY = PatternFill("solid", fgColor=GRAY_LIGHT)

ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

BORDER_THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_banner(ws, title, subtitle, n_cols):
    """Pone banner con título en R1 y subtítulo en R2."""
    ws.cell(row=1, column=1, value=title).font = FONT_TITLE
    ws.cell(row=2, column=1, value=subtitle).font = FONT_SUBTITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)


def style_headers(ws, headers, header_row=4):
    """Pone headers estilizados en la fila indicada."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_THIN
    ws.row_dimensions[header_row].height = 30


def set_widths(ws, widths):
    """Setea anchos de columna."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_freeze(ws, header_row):
    """Freeze panes en la fila después del header."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def add_filter(ws, header_row, n_cols, last_row=2000):
    """Habilita AutoFilter."""
    last_col = get_column_letter(n_cols)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"


def style_input_cell(cell):
    """Estilo para celdas de input manual."""
    cell.font = FONT_INPUT
    cell.fill = FILL_INPUT
    cell.border = BORDER_THIN


# =============================================================================
# Hojas
# =============================================================================

def hoja_config(wb):
    """Hoja de configuración general."""
    ws = wb.create_sheet("config")
    style_banner(
        ws, "CONFIGURACIÓN",
        "Parámetros generales del sistema. Cargá moneda base, fecha de arranque, etc.",
        4,
    )

    style_headers(ws, ["Concepto", "Valor", "Tipo", "Notas"], header_row=4)
    set_widths(ws, [25, 25, 12, 50])

    rows = [
        ("Moneda base de reporte", "USD", "string",
         "Moneda en la que se reporta el PN total. Típicamente USD (CCL)."),
        ("Fecha de arranque", "2026-04-01", "date",
         "Fecha desde la que el motor reconstruye holdings/PnL."),
        ("Tolerancia qty (cero)", 1e-6, "number",
         "Saldos con |qty| menor que esto se consideran cero."),
        ("Método PnL realizado", "FIFO", "string",
         "FIFO | WAC | LIFO. Por defecto FIFO."),
        ("Default plazo BYMA", "24hs", "string",
         "Plazo de liquidación por defecto."),
        ("Distancia alerta target (bps)", 10, "number",
         "Distancia en bps (1 bp = 0.01%) que define 'cerca del target'. "
         "Default 10 (≈ precio en target). Subilo a 100 para alertas con margen."),
    ]
    for i, (concepto, valor, tipo, notas) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=concepto).font = FONT_NORMAL
        cell = ws.cell(row=i, column=2, value=valor)
        style_input_cell(cell)
        ws.cell(row=i, column=3, value=tipo).font = FONT_SUBTITLE
        ws.cell(row=i, column=4, value=notas).font = FONT_SUBTITLE

    add_freeze(ws, 4)
    return ws


def hoja_monedas(wb):
    """Hoja de monedas."""
    ws = wb.create_sheet("monedas")
    style_banner(
        ws, "MONEDAS",
        ("Monedas y stablecoins que aparecen en el ledger. "
         "Cargá una fila por cada moneda. Indicá contra qué se cotiza (quote_vs)."),
        6,
    )

    headers = ["Code", "Name", "Is Stable", "Quote vs", "Is Base", "Notas"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [10, 24, 10, 12, 10, 40])

    examples = [
        ("ARS", "Peso Argentino", 0, None, 1, "Moneda doméstica. Es base."),
        ("USD", "Dólar CCL (offshore)", 0, "ARS", 0, "Cable. Ancla de valuación."),
        ("USB", "Dólar MEP (local)", 0, "ARS", 0, "Bolsa. Lo que liquida Cocos."),
        ("USD_OFICIAL", "Dólar Mayorista BCRA", 0, "ARS", 0,
         "A3500. Para reportes regulatorios."),
        ("USDT", "Tether", 1, "USD", 0, "Stablecoin USD."),
        ("USDC", "USD Coin", 1, "USD", 0, "Stablecoin USD."),
        ("BTC", "Bitcoin", 0, "USD", 0, ""),
        ("ETH", "Ethereum", 0, "USD", 0, ""),
        ("EUR", "Euro", 0, "USD", 0, ""),
        ("BRL", "Real brasileño", 0, "USD", 0, ""),
        ("UYU", "Peso uruguayo", 0, "USD", 0, ""),
        ("PEN", "Sol peruano", 0, "USD", 0, ""),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if j in (1, 4):  # Code, Quote vs
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL

    # Validaciones: Is Stable y Is Base = 0/1
    dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    dv.add(f"C5:C200")
    dv.add(f"E5:E200")
    ws.add_data_validation(dv)

    add_freeze(ws, 4)
    add_filter(ws, 4, 6, last_row=200)

    # Named range para que otras hojas lo usen
    return ws


def hoja_cuentas(wb):
    """Hoja de cuentas (brokers, bancos, wallets, tarjetas)."""
    ws = wb.create_sheet("cuentas")
    style_banner(
        ws, "CUENTAS",
        ("Cuentas en bancos, brokers, wallets cripto, y tarjetas de crédito. "
         "Para tarjetas, completá los días de cierre y vencimiento. "
         "'Investible' = NO excluye la cuenta del 'PN invertible' (ej cash de reserva no declarado)."),
        12,
    )

    headers = [
        "Code", "Name", "Kind", "Institution", "Currency",
        "Card Cycle", "Close Day", "Due Day", "Card Currency",
        "Investible", "Cash Purpose", "Notes"
    ]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [22, 26, 14, 18, 10, 12, 10, 10, 12, 11, 22, 30])

    # Cada fila: (code, name, kind, institution, currency, cycle, close, due,
    #             card_ccy, investible, cash_purpose, notes)
    examples = [
        ("cocos",            "Cocos Capital",          "CASH_BROKER",
         "Cocos Capital",    "ARS", "NONE", None, None, None,
         "YES", "OPERATIVO", "Broker principal"),
        ("eco",              "Eco Valores",            "CASH_BROKER",
         "Eco Valores",      "ARS", "NONE", None, None, None,
         "YES", "OPERATIVO", "Trading"),
        ("delta_fci",        "Delta FCI",              "CASH_BROKER",
         "Delta AM",         "ARS", "NONE", None, None, None,
         "YES", "OPERATIVO", "FCIs Delta"),
        ("ibkr",             "Interactive Brokers",    "CASH_BROKER",
         "Interactive Brokers","USD","NONE", None, None, None,
         "YES", "OPERATIVO", "Cuenta margen RegT (verificar parámetros reales)"),
        ("galicia_caja_ars", "Banco Galicia ARS",      "CASH_BANK",
         "Banco Galicia",    "ARS", "NONE", None, None, None,
         "YES", "OPERATIVO", "Caja ahorro ARS"),
        ("galicia_caja_usd", "Banco Galicia USD",      "CASH_BANK",
         "Banco Galicia",    "USD", "NONE", None, None, None,
         "YES", "OPERATIVO", "Caja ahorro USD"),
        ("santander_ars",    "Banco Santander ARS",    "CASH_BANK",
         "Santander",        "ARS", "NONE", None, None, None,
         "YES", "OPERATIVO", ""),
        ("binance",          "Binance Wallet",         "CASH_WALLET",
         "Binance",          None,  "NONE", None, None, None,
         "YES", "OPERATIVO", "Wallet cripto"),
        ("cash_transac",     "Cash transaccional",     "CASH_PHYSICAL",
         "Personal",         "ARS", "NONE", None, None, None,
         "YES", "OPERATIVO", "Efectivo del día a día"),
        ("cash_reserva",     "Cash reserva (no decl.)","CASH_PHYSICAL",
         "Personal",         "ARS", "NONE", None, None, None,
         "NO",  "RESERVA_NO_DECLARADO",
         "Cash NO declarado: excluido del PN invertible"),
        ("galicia_visa_ars", "Galicia Visa ARS",       "CARD_CREDIT",
         "Banco Galicia",    None,  "MONTHLY", 28, 10, "ARS",
         "YES", None, "Cierra día 28"),
        ("galicia_visa_usd", "Galicia Visa USD",       "CARD_CREDIT",
         "Banco Galicia",    None,  "MONTHLY", 28, 10, "USD",
         "YES", None, "Compras en USD"),
        # Cuentas técnicas (no-invertibles)
        ("external_income",  "[Ext] Ingresos externos","EXTERNAL",
         None,               None,  "NONE", None, None, None,
         "NO",  None, "Contracuenta de ingresos (sueldos, etc)"),
        ("external_expense", "[Ext] Gastos externos",  "EXTERNAL",
         None,               None,  "NONE", None, None, None,
         "NO",  None, "Contracuenta de gastos"),
        ("opening_balance",  "[Sys] Saldo inicial",    "OPENING_BALANCE",
         None,               None,  "NONE", None, None, None,
         "NO",  None, "Contracuenta del asiento de apertura"),
        ("interest_expense", "[Rdo] Intereses pagados","INTEREST_EXPENSE",
         None,               None,  "NONE", None, None, None,
         "NO",  None, "Cuenta de resultado por intereses"),
        ("interest_income",  "[Rdo] Intereses cobrados","INTEREST_INCOME",
         None,               None,  "NONE", None, None, None,
         "NO",  None, "Cuenta de resultado por intereses cobrados"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if j <= 11 and val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL

    # Validaciones: Kind, Card Cycle, Investible
    kinds = ("CASH_BANK,CASH_BROKER,CASH_WALLET,CASH_PHYSICAL,CARD_CREDIT,"
             "LIABILITY,EXTERNAL,OPENING_BALANCE,INTEREST_EXPENSE,INTEREST_INCOME")
    dv_kind = DataValidation(type="list", formula1=f'"{kinds}"', allow_blank=False)
    dv_kind.add(f"C5:C200")
    ws.add_data_validation(dv_kind)

    dv_cycle = DataValidation(type="list", formula1='"MONTHLY,NONE"', allow_blank=True)
    dv_cycle.add(f"F5:F200")
    ws.add_data_validation(dv_cycle)

    dv_inv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv_inv.add(f"J5:J200")
    ws.add_data_validation(dv_inv)

    add_freeze(ws, 4)
    add_filter(ws, 4, 12, last_row=200)
    return ws


def hoja_aforos(wb):
    """Hoja de aforos BYMA para garantías de caución."""
    ws = wb.create_sheet("aforos")
    style_banner(
        ws, "AFOROS BYMA — Garantías de caución",
        ("Aforo = % del valor de mercado aceptado como GARANTÍA por BYMA. "
         "Define cuánto poder de compra ganás dejando el activo como margen. "
         "Scope Type='CLASS' aplica a todo el asset_class; 'TICKER' override por instrumento."),
        5,
    )
    headers = ["Scope Type", "Scope Value", "Aforo %", "Source", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 22, 12, 14, 50])

    # Defaults aproximados — usuario debe verificar con tabla BYMA vigente.
    examples = [
        ("CLASS", "BOND_AR",    0.85, "BYMA approx", "Bonos soberanos AR (AL30, GD30, BPC...)"),
        ("CLASS", "EQUITY_AR",  0.70, "BYMA approx", "Acciones líderes Merval"),
        ("CLASS", "EQUITY_US",  0.70, "BYMA approx", "CEDEARs"),
        ("CLASS", "FCI",        0.90, "BYMA approx", "FCIs Money Market"),
        ("CLASS", "STABLECOIN", 0.50, "manual",      "USDT/USDC — no típicamente aceptado"),
        ("CLASS", "CRYPTO",     0.00, "manual",      "Cripto NO aceptado"),
        ("CLASS", "DERIVATIVE", 0.00, "BYMA",        "Derivados no garantizan"),
        # Overrides por ticker (ejemplo)
        ("TICKER", "AL30D",     0.90, "BYMA",        "Bono soberano USD-MEP, alto aforo"),
        ("TICKER", "GD30D",     0.90, "BYMA",        "Bono soberano USD-MEP"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j == 3:
                cell.number_format = "0.00%"

    dv_scope = DataValidation(type="list", formula1='"CLASS,TICKER"', allow_blank=False)
    dv_scope.add("A5:A500")
    ws.add_data_validation(dv_scope)

    add_freeze(ws, 4)
    add_filter(ws, 4, 5, last_row=500)
    return ws


def hoja_margin_config(wb):
    """Hoja de configuración de margin/leverage por cuenta (IBKR, etc)."""
    ws = wb.create_sheet("margin_config")
    style_banner(
        ws, "MARGIN CONFIG (IBKR, etc)",
        ("Configuración de leverage por cuenta para cuentas con margin estilo "
         "RegT (Reg-T en IBKR: x2 overnight, x4 intraday day-trade). "
         "VERIFICÁ los multiplicadores y la tasa de fondeo reales con tu broker. "
         "Cocos/Eco usan la hoja 'aforos' (BYMA), no esta."),
        6,
    )

    headers = ["Account", "Mult. Overnight", "Mult. Intraday",
               "Funding Rate Annual", "Funding Currency", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [22, 16, 16, 18, 14, 40])

    examples = [
        ("ibkr", 2.0, 4.0, 0.06, "USD",
         "RegT estándar: 50% margin overnight, 25% intraday. Funding ~6% anual USD aprox — verificar."),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j in (2, 3):
                cell.number_format = "0.00"
            elif j == 4:
                cell.number_format = "0.00%"

    add_freeze(ws, 4)
    add_filter(ws, 4, 6, last_row=200)
    return ws


def hoja_especies(wb):
    """Hoja de assets/instrumentos."""
    ws = wb.create_sheet("especies")
    style_banner(
        ws, "ESPECIES",
        ("Master de instrumentos. Todo ticker que aparezca en blotter o "
         "transferencias_activos debe estar acá primero."),
        8,
    )

    headers = ["Ticker", "Name", "Asset Class", "Currency",
               "Issuer", "Sector", "Country", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [16, 32, 16, 10, 16, 18, 14, 30])

    examples = [
        ("AL30",      "Bono AL30 ARS",          "BOND_AR",  "ARS",
         "Tesoro AR", "Soberano AR", "Argentina", "Bono Globales 2030 - peso"),
        ("AL30D",     "Bono AL30 USD-MEP",      "BOND_AR",  "USB",
         "Tesoro AR", "Soberano AR", "Argentina", "Globales 2030 USD MEP"),
        ("AL30C",     "Bono AL30 USD-CCL",      "BOND_AR",  "USD",
         "Tesoro AR", "Soberano AR", "Argentina", "Globales 2030 USD CCL"),
        ("GD30",      "Bono GD30 ARS",          "BOND_AR",  "ARS",
         "Tesoro AR", "Soberano AR", "Argentina", ""),
        ("TX26",      "Bono BONCER TX26",       "BOND_AR",  "ARS",
         "Tesoro AR", "Soberano AR", "Argentina", "BONCER TX26"),
        ("TXMJ9",     "Lecap TXMJ9",            "BOND_AR",  "ARS",
         "Tesoro AR", "Soberano AR", "Argentina", "Lecap"),
        ("GGAL",      "Grupo Galicia",          "EQUITY_AR","ARS",
         "GGAL",      "Banking", "Argentina", ""),
        ("AAPL",      "Apple CEDEAR",           "EQUITY_US","ARS",
         "Apple Inc", "Technology", "USA", "CEDEAR"),
        ("DELTA_AHORRO_A", "Delta Ahorro Plus - Clase A", "FCI", "ARS",
         "Delta AM",  "FCI", "Argentina", "FCI Money Market"),
        ("BTC",       "Bitcoin",                "CRYPTO",   "USD",
         None,        "Crypto", "Global", ""),
        ("ETH",       "Ethereum",               "CRYPTO",   "USD",
         None,        "Crypto", "Global", ""),
        ("USDT",      "Tether (asset)",         "STABLECOIN","USD",
         "Tether",    "Crypto", "Global", "Cuando lo tradeás como activo"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if j <= 7 and val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL

    # Validación Asset Class
    classes = ("CASH,BOND_AR,BOND_CORP_AR,BOND_US,"
               "EQUITY_AR,EQUITY_US,EQUITY_GLOBAL,ETF,REIT,"
               "FCI,CRYPTO,STABLECOIN,"
               "DERIVATIVE,COMMODITY,REAL_ESTATE,PRIVATE,OTHER")
    dv = DataValidation(type="list", formula1=f'"{classes}"', allow_blank=False)
    dv.add("C5:C500")
    ws.add_data_validation(dv)

    add_freeze(ws, 4)
    add_filter(ws, 4, 8, last_row=500)
    return ws


def hoja_blotter(wb):
    """Hoja de trades (BUY/SELL)."""
    ws = wb.create_sheet("blotter")
    style_banner(
        ws, "BLOTTER",
        ("Trades de activos. Una fila por leg (compra o venta). "
         "Saldos iniciales: usá hoja 'asientos_contables' o cargá BUY virtual con cuenta='opening_balance'."),
        15,
    )

    headers = [
        "Trade ID", "Trade Date", "Settle Date",
        "Cuenta", "Strategy", "Ticker",
        "Side", "Qty", "Precio", "Moneda Trade",
        "Cuenta Cash", "Comisión", "Moneda Com",
        "Precio Target", "Stop Loss", "Moneda Target",
        "Description", "Notes"
    ]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 12, 12, 18, 14, 12, 8, 14, 12, 10, 18, 10, 10,
                     12, 12, 12, 30, 25])

    # Ejemplos: trade TXMJ9 con target / stop
    examples = [
        ("T0001-A", date(2026, 4, 28), date(2026, 4, 30),
         "cocos", "TRADING", "TXMJ9", "BUY", 250000000, 0.80, "ARS",
         "cocos", 0, "ARS",
         0.85, 0.78, "ARS",                    # target / stop / moneda target
         "BUY TXMJ9", "Compra inicial trading"),
        ("T0001-B", date(2026, 4, 30), date(2026, 5, 4),
         "cocos", "TRADING", "TXMJ9", "SELL", 250000000, 0.835, "ARS",
         "cocos", 0, "ARS",
         None, None, None,                      # SELL típicamente sin target nuevo
         "SELL TXMJ9", "Cierre con +3.5pp"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j in (2, 3):  # dates
                cell.number_format = "yyyy-mm-dd"
            elif j in (8, 9, 12, 14, 15):  # qty, precio, comisión, target, stop
                cell.number_format = '#,##0.0000;[Red](#,##0.0000)'

    # Validación Side
    dv_side = DataValidation(type="list", formula1='"BUY,SELL"', allow_blank=False)
    dv_side.add("G5:G2000")
    ws.add_data_validation(dv_side)

    add_freeze(ws, 4)
    add_filter(ws, 4, 18, last_row=2000)
    return ws


def hoja_transferencias_cash(wb):
    """Transferencias de cash entre cuentas propias."""
    ws = wb.create_sheet("transferencias_cash")
    style_banner(
        ws, "TRANSFERENCIAS DE CASH",
        ("Movimientos de efectivo entre cuentas propias. NO afecta P&L. "
         "Ej: Galicia → Santander, retiro de cajero, depósito."),
        9,
    )

    headers = ["Fecha", "Monto", "Moneda", "Cuenta Origen",
               "Cuenta Destino", "Tipo de Cambio", "Description", "Trade ID Externo", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 16, 10, 22, 22, 14, 30, 14, 25])

    examples = [
        (date(2026, 5, 2), 1000000, "ARS", "galicia_caja_ars", "santander_ars",
         None, "Transferencia entre bancos", None, ""),
        (date(2026, 5, 2), 200000, "ARS", "galicia_caja_ars", "cash_transac",
         None, "Retiro cajero - bolsillo", None, "Para gastos del mes"),
        (date(2026, 5, 2), 800000, "ARS", "galicia_caja_ars", "cash_reserva",
         None, "Retiro - reserva", None, ""),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j == 1:
                cell.number_format = "yyyy-mm-dd"
            elif j in (2, 6):
                cell.number_format = '#,##0.00;[Red](#,##0.00)'

    add_freeze(ws, 4)
    add_filter(ws, 4, 9, last_row=2000)
    return ws


def hoja_transferencias_activos(wb):
    """Movimientos de activos entre cuentas propias."""
    ws = wb.create_sheet("transferencias_activos")
    style_banner(
        ws, "TRANSFERENCIAS DE ACTIVOS",
        ("Movimientos de activos entre cuentas propias (custodia). "
         "Ej: AL30 de Cocos a Eco. NO genera PnL. Para casos con costo, usá blotter."),
        7,
    )

    headers = ["Fecha", "Ticker", "Qty", "Cuenta Origen",
               "Cuenta Destino", "Description", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 14, 16, 22, 22, 30, 25])

    add_freeze(ws, 4)
    add_filter(ws, 4, 7, last_row=2000)
    return ws


def hoja_funding(wb):
    """Funding (cauciones, préstamos)."""
    ws = wb.create_sheet("funding")
    style_banner(
        ws, "FUNDING",
        ("Cauciones, pases y préstamos de corto plazo. "
         "TOMA: pagás interés. COLOCA: cobrás interés."),
        13,
    )

    headers = ["Fund ID", "Tipo", "Subtipo", "Cuenta", "Fecha Inicio",
               "Fecha Fin", "Moneda", "Monto", "TNA", "Días",
               "Status", "Linked Trade ID", "Description", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 10, 12, 18, 12, 12, 10, 16, 10, 8, 12, 16, 25, 25])

    examples = [
        ("F0001", "TOMA", "CAUCION", "cocos", date(2026, 4, 30),
         date(2026, 5, 4), "ARS", 200000000, 0.24,
         "=IF(F5=\"\",\"\",F5-E5)", "OPEN", "T0001-A",
         "Caución TOMA cubre TXMJ9", "4 días"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                if isinstance(val, str) and val.startswith("="):
                    cell.font = FONT_FORMULA
                else:
                    style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j in (5, 6):
                cell.number_format = "yyyy-mm-dd"
            elif j in (8, 9):
                cell.number_format = '#,##0.0000;[Red](#,##0.0000)'

    # Validaciones
    dv_tipo = DataValidation(type="list", formula1='"TOMA,COLOCA"', allow_blank=False)
    dv_tipo.add("B5:B2000")
    ws.add_data_validation(dv_tipo)
    dv_subtipo = DataValidation(
        type="list",
        formula1='"CAUCION,PASE,PRESTAMO_FRANCES,PRESTAMO_ALEMAN,PRESTAMO_BULLET"',
        allow_blank=False)
    dv_subtipo.add("C5:C2000")
    ws.add_data_validation(dv_subtipo)
    dv_status = DataValidation(type="list", formula1='"OPEN,CLOSED"', allow_blank=False)
    dv_status.add("K5:K2000")
    ws.add_data_validation(dv_status)

    add_freeze(ws, 4)
    add_filter(ws, 4, 14, last_row=2000)
    return ws


def hoja_ingresos(wb):
    """Ingresos."""
    ws = wb.create_sheet("ingresos")
    style_banner(
        ws, "INGRESOS",
        ("Sueldos, dividendos, cupones, premios, otros ingresos. "
         "Para recurrentes (sueldo mensual), usá hoja 'recurrentes'."),
        9,
    )

    headers = ["Fecha", "Concepto", "Categoría", "Monto", "Moneda",
               "Cuenta Destino", "Recurrente?", "Description", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 24, 16, 16, 8, 22, 12, 30, 25])

    examples = [
        (date(2026, 5, 1), "Sueldo abril 2026", "Sueldo", 5000000, "ARS",
         "galicia_caja_ars", "NO", "Pago mensual", ""),
        (date(2026, 4, 30), "Cupón AL30D", "Cupón", 850, "USD",
         "cocos", "NO", "Cupón semestral", "Auto-bajada de precio dirty"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j == 1:
                cell.number_format = "yyyy-mm-dd"
            elif j == 4:
                cell.number_format = '#,##0.00;[Red](#,##0.00)'

    dv_rec = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    dv_rec.add("G5:G2000")
    ws.add_data_validation(dv_rec)

    add_freeze(ws, 4)
    add_filter(ws, 4, 9, last_row=2000)
    return ws


def hoja_gastos(wb):
    """Gastos (cash o tarjeta)."""
    ws = wb.create_sheet("gastos")
    style_banner(
        ws, "GASTOS",
        ("Gastos del mes. Cuenta Destino determina si fue cash o tarjeta. "
         "Cuotas: poné número en columna 'Cuotas' (default 1)."),
        12,
    )

    headers = ["Fecha", "Concepto", "Categoría", "Tipo", "Monto", "Moneda",
               "Cuenta Destino", "Cuotas", "Recurrente?", "Cierre Card", "Description", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 24, 16, 10, 14, 8, 22, 8, 12, 12, 25, 20])

    examples = [
        (date(2026, 5, 5), "Alquiler", "Vivienda", "FIJO", 800000, "ARS",
         "galicia_caja_ars", 1, "YES", None, "Alquiler mensual", ""),
        (date(2026, 5, 2), "Supermercado", "Alimentación", "VARIABLE", 80000, "ARS",
         "galicia_caja_ars", 1, "NO", None, "Día 1 del mes", ""),
        (date(2026, 5, 3), "Restaurante", "Entretenimiento", "VARIABLE", 35000, "ARS",
         "galicia_visa_ars", 1, "NO", date(2026, 5, 28), "Cena con amigos", ""),
        (date(2026, 5, 1), "Vuelo a Punta del Este", "Viajes", "VARIABLE", 1200, "USD",
         "galicia_visa_usd", 6, "NO", date(2026, 5, 28), "Viaje verano 2027", "6 cuotas"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j in (1, 10):
                cell.number_format = "yyyy-mm-dd"
            elif j == 5:
                cell.number_format = '#,##0.00;[Red](#,##0.00)'

    dv_tipo = DataValidation(type="list", formula1='"FIJO,VARIABLE"', allow_blank=False)
    dv_tipo.add("D5:D2000")
    ws.add_data_validation(dv_tipo)
    dv_rec = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    dv_rec.add("I5:I2000")
    ws.add_data_validation(dv_rec)

    add_freeze(ws, 4)
    add_filter(ws, 4, 12, last_row=3000)
    return ws


def hoja_pasivos(wb):
    """Pasivos (préstamos personales, hipoteca)."""
    ws = wb.create_sheet("pasivos")
    style_banner(
        ws, "PASIVOS",
        ("Préstamos personales, hipoteca. Las tarjetas de crédito NO van acá "
         "(se manejan automáticamente desde la hoja 'cuentas' + 'gastos')."),
        12,
    )

    headers = ["Liab ID", "Tipo", "Acreedor", "Cuenta", "Fecha Inicio",
               "Fecha Fin", "Moneda", "Capital", "TNA", "Cuotas Total",
               "Status", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 12, 24, 18, 12, 12, 10, 16, 10, 12, 10, 25])

    dv_tipo = DataValidation(
        type="list",
        formula1='"BULLET,FRENCH,GERMAN,AMERICAN,CAUCION"',
        allow_blank=False)
    dv_tipo.add("B5:B500")
    ws.add_data_validation(dv_tipo)
    dv_status = DataValidation(type="list",
                                formula1='"OPEN,CLOSED,DEFAULT"',
                                allow_blank=False)
    dv_status.add("K5:K500")
    ws.add_data_validation(dv_status)

    add_freeze(ws, 4)
    add_filter(ws, 4, 12, last_row=500)
    return ws


def hoja_pagos_pasivos(wb):
    """Pagos a pasivos (cuotas de préstamo, cancelación de tarjeta)."""
    ws = wb.create_sheet("pagos_pasivos")
    style_banner(
        ws, "PAGOS A PASIVOS",
        ("Pagos a pasivos: cuotas de préstamos personales (capital + interés) "
         "y cancelación de tarjetas de crédito."),
        9,
    )

    headers = ["Fecha", "Pasivo / Tarjeta", "Monto Total", "Capital",
               "Interés", "Moneda", "Cuenta Origen", "Description", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 22, 16, 16, 14, 8, 22, 30, 25])

    add_freeze(ws, 4)
    add_filter(ws, 4, 9, last_row=2000)
    return ws


def hoja_recurrentes(wb):
    """Reglas de recurrencia (sueldo, alquiler, etc)."""
    ws = wb.create_sheet("recurrentes")
    style_banner(
        ws, "RECURRENTES",
        ("Reglas de auto-repetición. El motor genera ingresos/gastos "
         "automáticamente según la frecuencia, hasta la fecha de hoy."),
        12,
    )

    headers = ["Rule Name", "Event Type", "Cuenta", "Asset", "Amount",
               "Description", "Categoría", "Tipo", "Start Date", "End Date",
               "Day of Month", "Active"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [22, 12, 22, 8, 16, 22, 16, 10, 12, 12, 8, 8])

    examples = [
        ("Sueldo mensual", "INCOME", "galicia_caja_ars", "ARS", 5000000,
         "Sueldo regular", "Sueldo", "FIJO", date(2026, 1, 1), None, 1, "YES"),
        ("Alquiler", "EXPENSE", "galicia_caja_ars", "ARS", 800000,
         "Alquiler depto", "Vivienda", "FIJO", date(2026, 1, 1), None, 5, "YES"),
        ("Servicios", "EXPENSE", "galicia_caja_ars", "ARS", 150000,
         "Luz + gas + internet", "Servicios", "FIJO", date(2026, 1, 1), None, 15, "YES"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j in (9, 10):
                cell.number_format = "yyyy-mm-dd"
            elif j == 5:
                cell.number_format = '#,##0.00;[Red](#,##0.00)'

    dv_type = DataValidation(type="list",
                              formula1='"INCOME,EXPENSE,CARD_CHARGE"',
                              allow_blank=False)
    dv_type.add("B5:B500")
    ws.add_data_validation(dv_type)
    dv_active = DataValidation(type="list", formula1='"YES,NO"', allow_blank=False)
    dv_active.add("L5:L500")
    ws.add_data_validation(dv_active)

    add_freeze(ws, 4)
    add_filter(ws, 4, 12, last_row=500)
    return ws


def hoja_asientos(wb):
    """Asientos contables manuales (ajustes, previsiones, amortizaciones)."""
    ws = wb.create_sheet("asientos_contables")
    style_banner(
        ws, "ASIENTOS CONTABLES",
        ("Ajustes manuales: previsiones, amortizaciones, pase de inventario, "
         "asientos de apertura. Cargá las patas como filas (mínimo 2 por evento)."),
        9,
    )

    headers = ["Event ID", "Fecha", "Description", "Cuenta", "Activo",
               "Qty (signada)", "Unit Price", "Price Currency", "Notes"]
    style_headers(ws, headers, header_row=4)
    set_widths(ws, [12, 12, 30, 22, 14, 18, 14, 10, 25])

    examples = [
        # Asiento de apertura: tenés 1000 AL30D al 30 abril
        ("OPEN-001", date(2026, 4, 30), "Apertura: 1000 AL30D en Cocos",
         "cocos", "AL30D", 1000, 65.50, "USB", "Saldo inicial al arrancar tracking"),
        ("OPEN-001", date(2026, 4, 30), "Apertura: 1000 AL30D en Cocos",
         "opening_balance", "AL30D", -1000, 65.50, "USB",
         "Contracuenta del asiento de apertura"),
        # Otro ejemplo: previsión de cobro
        ("ADJ-002", date(2026, 5, 1), "Previsión cobro pendiente",
         "external_income", "ARS", -50000, None, None, "Devengado pero no cobrado"),
        ("ADJ-002", date(2026, 5, 1), "Previsión cobro pendiente",
         "galicia_caja_ars", "ARS", 50000, None, None, "Pendiente de acreditar"),
    ]
    for i, row in enumerate(examples, start=5):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if val is not None:
                style_input_cell(cell)
            else:
                cell.font = FONT_NORMAL
            if j == 2:
                cell.number_format = "yyyy-mm-dd"
            elif j in (6, 7):
                cell.number_format = '#,##0.0000;[Red](#,##0.0000)'

    add_freeze(ws, 4)
    add_filter(ws, 4, 9, last_row=2000)
    return ws


def hoja_index(wb):
    """Hoja índice con explicación de cada hoja del archivo."""
    ws = wb.create_sheet("INDEX", 0)  # primera hoja
    style_banner(
        ws,
        "WEALTH MANAGEMENT — Master input",
        "Sistema de wealth management: cargá acá tus operaciones del día a día.",
        4,
    )

    style_headers(ws, ["#", "Hoja", "Para qué sirve", "Cuándo cargar"],
                  header_row=4)
    set_widths(ws, [4, 22, 60, 30])

    rows = [
        (1, "config",          "Parámetros generales", "Una vez al inicio"),
        (2, "monedas",         "Monedas y stablecoins (ARS, USD, USB, USDT, BTC...)",
         "Una vez al inicio"),
        (3, "cuentas",         "Brokers, bancos, wallets, tarjetas (incluye flag Investible)",
         "Una vez al inicio"),
        (4, "especies",        "Master de instrumentos (bonos, acciones, FCI, cripto)",
         "Antes de cada nuevo ticker"),
        (5, "aforos",          "Aforos BYMA para cálculo de poder de compra (Cocos/Eco)",
         "Setup + cuando cambia BYMA"),
        (6, "margin_config",   "Margin/leverage por cuenta (IBKR x2/x4 + fondeo)",
         "Setup por cuenta"),
        (7, "blotter",         "Trades de activos (BUY/SELL)", "Cada operación"),
        (8, "transferencias_activos",  "Movimiento de activos entre cuentas (sin PnL)",
         "Cuando movés"),
        (9, "transferencias_cash",     "Movimiento de cash entre cuentas (sin PnL)",
         "Cuando transferís"),
        (10, "funding",         "Cauciones, pases, préstamos de corto plazo (con Linked Trade ID)",
         "Cada operación"),
        (11, "ingresos",        "Sueldos, dividendos, cupones, premios",
         "Cuando cobrás"),
        (12, "gastos",         "Gastos del mes (cash o tarjeta)", "Cuando ocurre"),
        (13, "pasivos",        "Préstamos personales, hipoteca", "Una vez por pasivo"),
        (14, "pagos_pasivos",  "Cuotas de préstamos, cancelación tarjetas",
         "En cada pago"),
        (15, "recurrentes",    "Sueldo/alquiler/servicios — auto-repetición",
         "Una vez por concepto"),
        (16, "asientos_contables", "Asientos manuales: apertura, previsiones, ajustes",
         "Cuando aplique"),
    ]
    for i, (n, hoja, qparahace, cuando) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=n).alignment = ALIGN_CENTER
        ws.cell(row=i, column=2, value=hoja).font = Font(bold=True)
        ws.cell(row=i, column=3, value=qparahace).alignment = ALIGN_LEFT
        ws.cell(row=i, column=4, value=cuando).font = FONT_SUBTITLE

    # Notas finales
    ws.cell(row=22, column=1,
            value="LEYENDA DE COLORES").font = Font(bold=True, color=NAVY)
    ws.cell(row=23, column=1, value="• Celda azul + amarillo: vos cargás").font = FONT_INPUT
    ws.cell(row=24, column=1, value="• Celda negra: fórmula automática").font = FONT_NORMAL
    ws.cell(row=25, column=1, value="• Celda gris: header / no editar").font = FONT_NORMAL

    add_freeze(ws, 4)
    return ws


# =============================================================================
# Main
# =============================================================================

def build_master(output_path: Path):
    """Construye el master Excel completo."""
    wb = Workbook()
    # Borrar el sheet default
    wb.remove(wb.active)

    print(f"[build] generando {output_path.name}...")

    # Orden importa: INDEX primero (después de creadas las otras quedan según orden)
    hoja_config(wb)
    hoja_monedas(wb)
    hoja_cuentas(wb)
    hoja_especies(wb)
    hoja_aforos(wb)
    hoja_margin_config(wb)
    hoja_blotter(wb)
    hoja_transferencias_cash(wb)
    hoja_transferencias_activos(wb)
    hoja_funding(wb)
    hoja_ingresos(wb)
    hoja_gastos(wb)
    hoja_pasivos(wb)
    hoja_pagos_pasivos(wb)
    hoja_recurrentes(wb)
    hoja_asientos(wb)
    # INDEX al final pero la pongo en posición 0
    hoja_index(wb)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"[build] OK → {output_path}")
    print(f"[build] {len(wb.sheetnames)} hojas: {', '.join(wb.sheetnames)}")


def main():
    if len(sys.argv) > 1:
        out_path = Path(sys.argv[1])
    else:
        out_path = Path("inputs/wealth_management_rodricor.xlsx")
    build_master(out_path)


if __name__ == "__main__":
    main()
