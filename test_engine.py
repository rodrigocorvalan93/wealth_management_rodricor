# -*- coding: utf-8 -*-
"""
test_engine.py

Tests end-to-end del motor de wealth management.

USO:
    python3 test_engine.py
"""

from __future__ import annotations

import sqlite3
import sys
import tempfile
from datetime import date
from pathlib import Path

# Permitir import desde la carpeta padre
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from build_master import build_master
from engine.importer import import_all
from engine.liabilities import card_snapshot, all_card_snapshots
from engine.schema import init_db, AccountKind


def test_1_build_master():
    """build_master genera archivo con 15 hojas."""
    print("\nTest 1 (build_master):")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    try:
        build_master(path)
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(path))
        assert len(wb.sheetnames) >= 14, f"Esperaba ≥14 hojas, got {len(wb.sheetnames)}"
        assert "blotter" in wb.sheetnames
        assert "gastos" in wb.sheetnames
        assert "cuentas" in wb.sheetnames
        assert "asientos_contables" in wb.sheetnames
        print(f"  ✓ {len(wb.sheetnames)} hojas: {', '.join(wb.sheetnames)}")
    finally:
        path.unlink()


def test_2_import_round_trip():
    """build_master → import → verificar contadores."""
    print("\nTest 2 (import round-trip):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        stats = import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        # Las hojas con ejemplos pre-cargados:
        assert stats["monedas"] >= 12
        assert stats["cuentas"] >= 16
        assert stats["especies"] >= 12
        assert stats["blotter"] == 2  # 2 trades de TXMJ9
        assert stats["transferencias_cash"] == 3
        assert stats["ingresos"] == 2
        assert stats["gastos"] == 3
        assert stats["recurrentes"] >= 1
        assert stats["asientos_contables"] == 2  # 2 grupos
        print(f"  ✓ stats: {stats}")


def test_3_blotter_balance():
    """Test que el TXMJ9 trade neteado da qty=0 (compra+venta misma cantidad)."""
    print("\nTest 3 (blotter qty net):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row
        cur = conn.execute(
            """SELECT SUM(qty) AS net FROM movements
               WHERE account='cocos' AND asset='TXMJ9'"""
        )
        net = cur.fetchone()["net"]
        # BUY 250M + SELL 250M = 0
        assert abs(net) < 1e-3, f"Expected ~0, got {net}"
        print(f"  ✓ TXMJ9 net qty: {net}")

        # PnL del trade: SELL 0.835 - BUY 0.80 = +0.035 × 250M = 8.75M
        cur = conn.execute(
            """SELECT SUM(qty) AS cash_net FROM movements
               WHERE account='cocos' AND asset='ARS'
               AND event_id IN (SELECT event_id FROM events WHERE event_type='TRADE')"""
        )
        cash_net = cur.fetchone()["cash_net"]
        # BUY: -200M, SELL: +208.75M → net +8.75M
        expected = 250000000 * (0.835 - 0.80)
        assert abs(cash_net - expected) < 1, f"Expected {expected}, got {cash_net}"
        print(f"  ✓ TXMJ9 cash net: ARS {cash_net:,.0f} (esperado {expected:,.0f})")
        conn.close()


def test_4_cuotas_expansion():
    """Test que 6 cuotas → 6 events CARD_INSTALLMENT."""
    print("\nTest 4 (cuotas expansion):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        cur = conn.execute(
            """SELECT COUNT(*) AS n FROM events WHERE event_type='CARD_INSTALLMENT'"""
        )
        n = cur.fetchone()["n"]
        # El vuelo es 1 fila con 6 cuotas → 6 CARD_INSTALLMENT
        assert n == 6, f"Esperaba 6 cuotas, got {n}"
        print(f"  ✓ {n} CARD_INSTALLMENT events generados")

        # Verificar que cada cuota es 200 USD
        cur = conn.execute(
            """SELECT m.qty FROM movements m
               JOIN events e ON e.event_id = m.event_id
               WHERE e.event_type='CARD_INSTALLMENT'
               AND m.account='galicia_visa_usd'"""
        )
        qtys = [row["qty"] for row in cur.fetchall()]
        assert all(abs(q - 200) < 1e-3 for q in qtys), f"Cuotas no son 200: {qtys}"
        print(f"  ✓ Cada cuota = USD 200")
        conn.close()


def test_5_card_snapshot_galicia_usd():
    """Galicia Visa USD: 1 cuota cayó en mayo 1, próximo cierre 28 mayo."""
    print("\nTest 5 (card_snapshot Galicia USD):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        snap = card_snapshot(conn, "galicia_visa_usd", ref_date=date(2026, 5, 2))
        assert snap is not None
        # 2 mayo: ya pasó 1 cuota (1 mayo) → saldo actual = 200
        assert abs(snap.saldo_actual - 200) < 1e-3, \
            f"saldo actual: {snap.saldo_actual}"
        # Último cierre: 28 abril → 0 (la cuota cayó después del 28/4)
        assert abs(snap.saldo_ultimo_resumen) < 1e-3, \
            f"último resumen: {snap.saldo_ultimo_resumen}"
        # Próximo cierre: 28 mayo
        assert snap.fecha_proximo_cierre == "2026-05-28", \
            f"próximo cierre: {snap.fecha_proximo_cierre}"
        # Próximo vto: la 1 cuota va al cierre de 28/5 → USD 200
        assert abs(snap.saldo_proximo_vto - 200) < 1e-3, \
            f"próximo vto: {snap.saldo_proximo_vto}"
        print(f"  ✓ Saldo actual: USD {snap.saldo_actual}")
        print(f"  ✓ Último resumen ({snap.fecha_ultimo_cierre}): USD {snap.saldo_ultimo_resumen}")
        print(f"  ✓ Próximo vto ({snap.fecha_proximo_vto}): USD {snap.saldo_proximo_vto}")
        conn.close()


def test_6_card_snapshot_galicia_ars_post_cierre():
    """Si pongo ref_date después del cierre del 28/5, el resumen ya cerró."""
    print("\nTest 6 (card snapshot post-cierre):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        # Fecha = 1 junio 2026 (después del cierre 28/5)
        import_all(db, xlsx, fecha_corte=date(2026, 6, 1))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        snap = card_snapshot(conn, "galicia_visa_ars", ref_date=date(2026, 6, 1))
        assert snap is not None
        # Restaurante de 35k cargado el 3/5 → cae en cierre 28/5 → último resumen
        assert abs(snap.saldo_ultimo_resumen - 35000) < 1e-3, \
            f"último resumen: {snap.saldo_ultimo_resumen}"
        # Saldo actual igual al último (no hay más gastos después)
        assert abs(snap.saldo_actual - 35000) < 1e-3, \
            f"saldo actual: {snap.saldo_actual}"
        print(f"  ✓ Saldo actual al 1/6: ARS {snap.saldo_actual:,.0f}")
        print(f"  ✓ Último resumen ({snap.fecha_ultimo_cierre}): ARS {snap.saldo_ultimo_resumen:,.0f}")
        conn.close()


def test_7_recurrentes_expansion():
    """Recurrentes: sueldo enero 2026 hasta hoy = 5 meses."""
    print("\nTest 7 (recurrentes expansion):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        # Sueldo: enero a mayo (5 meses, día 1 cada mes)
        cur = conn.execute(
            """SELECT COUNT(*) AS n FROM events
               WHERE event_type='INCOME'
               AND description='Sueldo regular'"""
        )
        n_sueldos = cur.fetchone()["n"]
        # Desde 1/1/2026 hasta 1/5/2026 = 5 ocurrencias el día 1
        assert n_sueldos == 5, f"Sueldos: {n_sueldos}"
        print(f"  ✓ Sueldos generados: {n_sueldos}")

        # Verificar que cada uno mueve 5M ARS a galicia_caja_ars
        cur = conn.execute(
            """SELECT SUM(m.qty) AS total FROM movements m
               JOIN events e ON e.event_id=m.event_id
               WHERE e.description='Sueldo regular'
               AND m.account='galicia_caja_ars'"""
        )
        total = cur.fetchone()["total"]
        # 5 sueldos × 5M = 25M
        assert abs(total - 25_000_000) < 1, f"Total sueldos: {total}"
        print(f"  ✓ Total ingresos: ARS {total:,.0f}")
        conn.close()


def test_8_asientos_apertura():
    """Asientos contables: apertura de 1000 AL30D balancea con opening_balance."""
    print("\nTest 8 (asientos apertura):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        # Verificar que cocos tiene 1000 AL30D
        cur = conn.execute(
            """SELECT SUM(qty) AS bal FROM movements
               WHERE account='cocos' AND asset='AL30D'"""
        )
        bal = cur.fetchone()["bal"]
        assert abs(bal - 1000) < 1e-3, f"Cocos AL30D bal: {bal}"
        print(f"  ✓ Saldo Cocos AL30D: {bal}")

        # Verificar que opening_balance tiene -1000
        cur = conn.execute(
            """SELECT SUM(qty) AS bal FROM movements
               WHERE account='opening_balance' AND asset='AL30D'"""
        )
        bal = cur.fetchone()["bal"]
        assert abs(bal + 1000) < 1e-3, f"opening_balance AL30D bal: {bal}"
        print(f"  ✓ Saldo opening_balance AL30D: {bal}")

        # Total AL30D en sistema = 0 (balance contable)
        cur = conn.execute(
            "SELECT SUM(qty) AS total FROM movements WHERE asset='AL30D'"
        )
        total = cur.fetchone()["total"]
        assert abs(total) < 1e-3, f"AL30D total no balancea: {total}"
        print(f"  ✓ AL30D balance global: {total}")
        conn.close()


def test_9_v_balances():
    """La vista v_balances filtra los saldos = 0."""
    print("\nTest 9 (v_balances view):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        cur = conn.execute("SELECT * FROM v_balances ORDER BY account, asset")
        rows = list(cur.fetchall())
        # No debería haber saldos exactamente 0 (la vista los filtra)
        assert all(abs(r["balance"]) >= 1e-6 for r in rows)
        print(f"  ✓ {len(rows)} balances (todos != 0):")
        for r in rows[:10]:
            print(f"    {r['account']:<22} {r['asset']:<10} {r['balance']:>16,.4f}")
        if len(rows) > 10:
            print(f"    ... ({len(rows)-10} más)")
        conn.close()


# =============================================================================
# Main
# =============================================================================

if __name__ == "__main__":
    tests = [
        test_1_build_master,
        test_2_import_round_trip,
        test_3_blotter_balance,
        test_4_cuotas_expansion,
        test_5_card_snapshot_galicia_usd,
        test_6_card_snapshot_galicia_ars_post_cierre,
        test_7_recurrentes_expansion,
        test_8_asientos_apertura,
        test_9_v_balances,
    ]
    for t in tests:
        t()
    print("\n" + "=" * 70)
    print(f"✓ Todos los {len(tests)} tests pasaron")
    print("=" * 70)
