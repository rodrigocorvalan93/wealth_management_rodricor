# -*- coding: utf-8 -*-
"""
migrate_adr_to_us.py

Renombra tickers con convención antigua `_ADR` → `_US` (mercado extranjero
vía Yahoo Finance). Aplica a todos los user masters multi-tenant.

Convención nueva:
  AAPL_US  → equity US (vía Yahoo Finance, mercado extranjero)
  AAPL_AR  → CEDEAR local (vía BYMA, mercado AR)
  AAPL_ADR → DEPRECATED, alias de AAPL_US

Renombra en:
  - Hoja `especies`: columna Ticker
  - Hoja `blotter`: columna Ticker (todos los trades)
  - Hoja `transferencias_activos`: columnas Activo
  - Hoja `asientos_contables`: columna Activo

NO toca CSVs en data/ (esos se regeneran al correr sync.py).

USO:
    python migrate_adr_to_us.py                # todos los users
    python migrate_adr_to_us.py --user rodricor
    python migrate_adr_to_us.py --dry-run      # solo muestra qué cambiaría
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
HEADER_ROW = 4

# Hoja → lista de columnas de header donde aparecen tickers
TICKER_COLUMNS = {
    "especies":              ["Ticker"],
    "blotter":               ["Ticker"],
    "transferencias_activos":["Activo"],
    "asientos_contables":    ["Activo"],
}


def find_user_xlsxs() -> list[tuple[str, Path]]:
    out = []
    inputs = HERE / "inputs"
    if not inputs.is_dir():
        return out
    for d in sorted(inputs.iterdir()):
        if not d.is_dir():
            continue
        for fname in ("wealth_management.xlsx", "wealth_management_rodricor.xlsx"):
            f = d / fname
            if f.is_file():
                out.append((d.name, f))
                break
    if not out:
        for fname in ("wealth_management.xlsx", "wealth_management_rodricor.xlsx"):
            f = inputs / fname
            if f.is_file():
                out.append(("default", f))
                break
    return out


def find_header_col(ws, header_name: str) -> int | None:
    """Busca el número de columna con header_name. None si no existe."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v and str(v).strip() == header_name:
            return c
    return None


def rename_ticker(value: str) -> str | None:
    """Devuelve el nuevo nombre si hay que renombrar, None si no aplica."""
    if not isinstance(value, str):
        return None
    v = value.strip().upper()
    if v.endswith("_ADR"):
        return v[:-4] + "_US"
    return None


def migrate_xlsx(xlsx_path: Path, dry_run: bool) -> dict:
    if not xlsx_path.is_file():
        return {"error": f"no existe: {xlsx_path}"}

    summary = {"path": str(xlsx_path), "renames": [], "backed_up": None}

    if not dry_run:
        ts = datetime.now().strftime("%Y%m%dT%H%M%S")
        bak = xlsx_path.with_suffix(f".pre-adr-migration.{ts}.xlsx")
        shutil.copy2(xlsx_path, bak)
        summary["backed_up"] = bak.name

    wb = load_workbook(filename=str(xlsx_path))
    n_changes = 0

    for sheet_name, header_names in TICKER_COLUMNS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for header in header_names:
            col = find_header_col(ws, header)
            if col is None:
                continue
            for r in range(HEADER_ROW + 1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                old = cell.value
                new = rename_ticker(old) if old else None
                if new and new != old:
                    summary["renames"].append({
                        "sheet": sheet_name, "row": r, "col": header,
                        "old": old, "new": new,
                    })
                    if not dry_run:
                        cell.value = new
                    n_changes += 1

    if not dry_run:
        if n_changes > 0:
            wb.save(xlsx_path)
        elif summary["backed_up"]:
            (xlsx_path.parent / summary["backed_up"]).unlink(missing_ok=True)
            summary["backed_up"] = None

    summary["n_changes"] = n_changes
    return summary


def main():
    p = argparse.ArgumentParser(description="Renombra _ADR → _US en masters")
    p.add_argument("--user", type=str, default=None,
                   help="Solo este user (default: todos)")
    p.add_argument("--dry-run", action="store_true",
                   help="No modifica, solo muestra qué cambiaría")
    args = p.parse_args()

    masters = find_user_xlsxs()
    if not masters:
        print("[migrate-adr] No se encontraron masters en inputs/")
        return 1
    if args.user:
        masters = [(u, p) for u, p in masters if u == args.user]
        if not masters:
            print(f"[migrate-adr] User '{args.user}' no encontrado")
            return 1

    print(f"[migrate-adr] {len(masters)} master(s){' (dry-run)' if args.dry_run else ''}")
    for user, xlsx in masters:
        print(f"\n--- {user} → {xlsx.name} ---")
        s = migrate_xlsx(xlsx, args.dry_run)
        if "error" in s:
            print(f"  ✗ {s['error']}"); continue
        if s["n_changes"] == 0:
            print(f"  · sin cambios (ningún _ADR encontrado)")
        else:
            print(f"  ✓ {s['n_changes']} celdas renombradas:")
            for r in s["renames"][:20]:
                print(f"    {r['sheet']}.{r['col']} (fila {r['row']}): {r['old']} → {r['new']}")
            if len(s["renames"]) > 20:
                print(f"    ... y {len(s['renames']) - 20} más")
        if s["backed_up"]:
            print(f"  ↩ backup: {s['backed_up']}")

    if not args.dry_run:
        print("\n[migrate-adr] Listo. Próximos pasos:")
        print("  1. Hacé refresh en la PWA (Settings → ⟳) para que el motor")
        print("     re-importe el master con los nuevos tickers.")
        print("  2. Si querés agregar el equivalente CEDEAR (mercado AR), ")
        print("     creá una nueva fila en `especies` con AAPL_AR (asset_class")
        print("     EQUITY_US o EQUITY_AR según preferencia).")
        print("  3. Corré `python sync.py` para fetchear precios de los")
        print("     nuevos _US y _AR.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
