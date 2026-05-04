# -*- coding: utf-8 -*-
"""
test_target_alerts.py — Tests del feature Precio Target / Stop-Loss + alertas.

Cubre:
  - Schema con cols target_price, stop_loss_price, target_currency
  - Importer lee cols del blotter
  - holdings.py expone target/stop + dist_to_target_bps + dist_to_stop_bps
  - filter_near_target con umbral configurable
  - Settings table get/set
  - Migration script idempotente
  - API: /api/holdings-near-target, GET/PUT /api/settings
"""

from __future__ import annotations

import os
import sqlite3
import sys
import tempfile
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


def _build_master_with_targets(xlsx_path):
    """Crea un master + lo modifica para tener un BUY abierto con TP+SL."""
    from build_master import build_master
    from openpyxl import load_workbook
    build_master(xlsx_path)
    wb = load_workbook(xlsx_path)
    ws = wb["blotter"]
    # El default tiene BUY+SELL del mismo qty (TXMJ9). Reemplazo el SELL por
    # uno parcial (qty=100M) para dejar 150M abiertos con target del BUY.
    # Headers en row 4: localizar Qty
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=c).value
        if v: headers[str(v).strip()] = c
    qty_col = headers["Qty"]
    # row 6 es el SELL → ajustar qty a 100M (deja 150M abierto)
    ws.cell(row=6, column=qty_col, value=100000000)
    wb.save(xlsx_path)


def test_1_schema_has_target_cols():
    print("\n[T 1] schema events tiene target_price / stop_loss / target_currency:")
    with tempfile.TemporaryDirectory() as tmp:
        from engine.schema import init_db
        db = Path(tmp) / "w.db"
        conn = init_db(db)
        cols = [r[1] for r in conn.execute("PRAGMA table_info(events)").fetchall()]
        assert "target_price" in cols, f"Falta target_price; cols={cols}"
        assert "stop_loss_price" in cols
        assert "target_currency" in cols
        print(f"  ✓ events tiene cols target ({sorted(c for c in cols if 'target' in c or 'stop' in c)})")
        # Settings table
        conn.execute("SELECT * FROM settings LIMIT 0")
        print(f"  ✓ settings table existe")


def test_2_set_get_setting():
    print("\n[T 2] set_setting / get_setting con cast:")
    with tempfile.TemporaryDirectory() as tmp:
        from engine.schema import init_db, set_setting, get_setting
        conn = init_db(Path(tmp) / "w.db")
        set_setting(conn, "alert_distance_bps", 50)
        assert get_setting(conn, "alert_distance_bps", cast=float) == 50.0
        assert get_setting(conn, "alert_distance_bps", cast=int) == 50
        # Default cuando no existe
        assert get_setting(conn, "missing_key", default=99, cast=int) == 99
        # Cast bool
        set_setting(conn, "feature_x", "true")
        assert get_setting(conn, "feature_x", cast=bool) is True
        set_setting(conn, "feature_y", "0")
        assert get_setting(conn, "feature_y", cast=bool) is False
        print(f"  ✓ get/set/default/bool/cast OK")


def test_3_importer_reads_target_from_blotter():
    print("\n[T 3] importer lee 'Precio Target' / 'Stop Loss' del blotter:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "m.xlsx"
        _build_master_with_targets(xlsx)
        from engine.importer import import_all
        import_all(tmp / "w.db", xlsx)
        conn = sqlite3.connect(tmp / "w.db")
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            "SELECT external_id, target_price, stop_loss_price, target_currency "
            "FROM events WHERE external_id LIKE 'T0001%'"
        ).fetchall()
        rows = {r["external_id"]: dict(r) for r in rows}
        # T0001-A es el BUY con target=0.85, stop=0.78
        assert rows["T0001-A"]["target_price"] == 0.85
        assert rows["T0001-A"]["stop_loss_price"] == 0.78
        assert rows["T0001-A"]["target_currency"] == "ARS"
        # T0001-B es el SELL sin target
        assert rows["T0001-B"]["target_price"] is None
        print(f"  ✓ BUY tiene target/stop, SELL no")


def test_4_importer_reads_config_sheet():
    print("\n[T 4] importer lee 'Distancia alerta target (bps)' de config sheet:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        from build_master import build_master
        xlsx = tmp / "m.xlsx"
        build_master(xlsx)
        from engine.importer import import_all
        from engine.schema import get_setting
        import_all(tmp / "w.db", xlsx)
        conn = sqlite3.connect(tmp / "w.db"); conn.row_factory = sqlite3.Row
        v = get_setting(conn, "alert_distance_bps", cast=float)
        assert v == 10.0, f"got {v}"
        # También el resto
        assert get_setting(conn, "anchor_currency") == "USD"
        assert get_setting(conn, "pnl_method") == "FIFO"
        print(f"  ✓ alert_distance_bps={v}, anchor=USD, pnl=FIFO")


def test_5_holdings_exposes_target_and_dist():
    print("\n[T 5] calculate_holdings expone target + dist_to_target_bps:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "m.xlsx"
        _build_master_with_targets(xlsx)
        from engine.importer import import_all
        from engine.holdings import calculate_holdings
        import_all(tmp / "w.db", xlsx)
        conn = sqlite3.connect(tmp / "w.db"); conn.row_factory = sqlite3.Row
        hs = calculate_holdings(conn)
        txm = next((h for h in hs if h["asset"] == "TXMJ9"), None)
        assert txm is not None, "TXMJ9 holding missing (debería tener 150M qty open)"
        assert txm["qty"] == 150000000  # 250M BUY - 100M SELL
        assert txm["target_price"] == 0.85
        assert txm["stop_loss_price"] == 0.78
        # market_price es el precio del SELL (0.835) si no hay otra cotización
        # → como fallback usa avg_cost (0.80)
        # dist_to_target = (0.80 - 0.85) / 0.85 * 10000 = -588 bps
        assert txm["dist_to_target_bps"] is not None
        # dist_to_stop = (0.80 - 0.78) / 0.78 * 10000 = +256 bps
        assert txm["dist_to_stop_bps"] is not None
        assert txm["dist_to_stop_bps"] > 0  # arriba del stop, safe
        print(f"  ✓ TXMJ9 qty={txm['qty']}, mp={txm['market_price']:.4f}, "
              f"tgt={txm['target_price']}, dist_tp={txm['dist_to_target_bps']:.0f}bp, "
              f"dist_sl={txm['dist_to_stop_bps']:.0f}bp")


def test_6_filter_near_target():
    print("\n[T 6] filter_near_target detecta TP y SL alerts:")
    from engine.holdings import filter_near_target
    holdings = [
        # Caso 1: precio justo en target
        {"asset": "X1", "dist_to_target_bps": 0, "dist_to_stop_bps": 1000},
        # Caso 2: precio +50 bps PASÓ el target
        {"asset": "X2", "dist_to_target_bps": 50, "dist_to_stop_bps": 1500},
        # Caso 3: precio -50 bps cerca del target (dentro de 100bps de threshold)
        {"asset": "X3", "dist_to_target_bps": -50, "dist_to_stop_bps": 800},
        # Caso 4: precio cerca del stop (dist_sl=+5bps, dentro)
        {"asset": "X4", "dist_to_target_bps": -2000, "dist_to_stop_bps": 5},
        # Caso 5: precio PERFORÓ el stop (dist_sl negativo)
        {"asset": "X5", "dist_to_target_bps": -3000, "dist_to_stop_bps": -100},
        # Caso 6: lejos de ambos
        {"asset": "X6", "dist_to_target_bps": -5000, "dist_to_stop_bps": 5000},
        # Caso 7: sin target ni stop
        {"asset": "X7", "dist_to_target_bps": None, "dist_to_stop_bps": None},
    ]
    alerts = filter_near_target(holdings, alert_distance_bps=100)
    by_asset = {a["asset"]: a["alert"] for a in alerts}
    assert by_asset.get("X1") == "TP"  # exacto en target
    assert by_asset.get("X2") == "TP"  # pasó target
    assert by_asset.get("X3") == "TP"  # -50 bps cerca, dentro de 100
    assert by_asset.get("X4") == "SL"  # cerca del stop
    assert by_asset.get("X5") == "SL"  # perforó stop
    assert "X6" not in by_asset
    assert "X7" not in by_asset
    print(f"  ✓ alerts: {by_asset}")


def test_7_filter_near_target_strict_10bps():
    print("\n[T 7] filter_near_target con 10 bps (default) es estricto:")
    from engine.holdings import filter_near_target
    holdings = [
        # +5 bps pasó target → TP
        {"asset": "A", "dist_to_target_bps": 5, "dist_to_stop_bps": 200},
        # -50 bps falta para target → no alert (está fuera de 10 bps)
        {"asset": "B", "dist_to_target_bps": -50, "dist_to_stop_bps": 200},
        # -10 bps justo en threshold → alert
        {"asset": "C", "dist_to_target_bps": -10, "dist_to_stop_bps": 200},
    ]
    alerts = filter_near_target(holdings, alert_distance_bps=10)
    by_asset = {a["asset"]: a["alert"] for a in alerts}
    assert by_asset.get("A") == "TP"
    assert "B" not in by_asset
    assert by_asset.get("C") == "TP"
    print(f"  ✓ a 10 bps solo alerta lo que está realmente cerca: {by_asset}")


def test_8_active_target_uses_most_recent_buy():
    print("\n[T 8] _get_active_target usa el BUY más reciente con target:")
    with tempfile.TemporaryDirectory() as tmp:
        from engine.schema import init_db, insert_event, insert_movement
        from engine.holdings import _get_active_target
        conn = init_db(Path(tmp) / "w.db")
        # Primero insertar accounts y assets minimos
        conn.execute("INSERT INTO currencies(code,name) VALUES('ARS','peso')")
        conn.execute("INSERT INTO accounts(code,name,kind) VALUES('cocos','cocos','CASH_BROKER')")
        conn.execute("INSERT INTO assets(ticker,name,asset_class,currency) VALUES('XYZ','XYZ','BOND_AR','ARS')")
        # BUY 1: target=10
        e1 = insert_event(conn, "TRADE", "2026-01-01", target_price=10, target_currency="ARS")
        insert_movement(conn, e1, account="cocos", asset="XYZ", qty=100, unit_price=8)
        # BUY 2 (más reciente): target=12
        e2 = insert_event(conn, "TRADE", "2026-02-01", target_price=12, target_currency="ARS")
        insert_movement(conn, e2, account="cocos", asset="XYZ", qty=50, unit_price=11)
        # BUY 3 (más reciente todavía, pero SIN target)
        e3 = insert_event(conn, "TRADE", "2026-03-01")
        insert_movement(conn, e3, account="cocos", asset="XYZ", qty=20, unit_price=11.5)
        conn.commit()
        target = _get_active_target(conn, "cocos", "XYZ")
        assert target is not None
        # Debe devolver el BUY 2 (último con target no-null)
        assert target["target_price"] == 12, f"got {target}"
        print(f"  ✓ tomó target del BUY más reciente con target no-null: {target['target_price']}")


def test_9_migration_idempotent():
    print("\n[T 9] migrate_target_columns es idempotente:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        from build_master import build_master
        xlsx = tmp / "m.xlsx"
        build_master(xlsx)
        # build_master YA tiene las cols nuevas (porque actualizamos build_master).
        # Ejecutar la migration debe ser no-op.
        import migrate_target_columns as mtc
        s = mtc.migrate_xlsx(xlsx, dry_run=False)
        assert s["blotter_cols_added"] == [], f"esperaba [] (ya migrado), got {s['blotter_cols_added']}"
        assert s["config_row_added"] is False
        print(f"  ✓ master ya migrado (build_master incluye cols) → no-op")
        # Re-ejecutar también es no-op
        s2 = mtc.migrate_xlsx(xlsx, dry_run=False)
        assert s2["blotter_cols_added"] == []
        print(f"  ✓ segunda corrida también no-op")


def test_10_migration_on_legacy_master():
    print("\n[T 10] migrate_target_columns funciona sobre master VIEJO sin las cols:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        from build_master import build_master
        from openpyxl import load_workbook
        xlsx = tmp / "m.xlsx"
        build_master(xlsx)
        # Simular master "legacy" removiendo las cols de target del blotter
        wb = load_workbook(xlsx)
        ws = wb["blotter"]
        # Buscar y eliminar cols target nuevas
        cols_to_remove = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=4, column=c).value
            if v in ("Precio Target", "Stop Loss", "Moneda Target"):
                cols_to_remove.append(c)
        # Borrar de derecha a izquierda
        for c in sorted(cols_to_remove, reverse=True):
            ws.delete_cols(c)
        # Borrar también la config row
        wsc = wb["config"]
        for r in range(5, wsc.max_row + 1):
            v = wsc.cell(row=r, column=1).value
            if v and "Distancia alerta target" in str(v):
                wsc.delete_rows(r)
                break
        wb.save(xlsx)

        # Verificar que el master quedó "legacy"
        wb2 = load_workbook(xlsx)
        ws_blot = wb2["blotter"]
        headers = [str(ws_blot.cell(row=4, column=c).value).strip()
                    for c in range(1, ws_blot.max_column + 1) if ws_blot.cell(row=4, column=c).value]
        assert "Precio Target" not in headers, f"should be removed; got {headers}"

        # Correr migration
        import migrate_target_columns as mtc
        s = mtc.migrate_xlsx(xlsx, dry_run=False)
        assert "Precio Target" in s["blotter_cols_added"]
        assert "Stop Loss" in s["blotter_cols_added"]
        assert "Moneda Target" in s["blotter_cols_added"]
        assert s["config_row_added"] is True
        # Verificar que ahora SÍ están
        wb3 = load_workbook(xlsx)
        headers2 = [str(wb3["blotter"].cell(row=4, column=c).value).strip()
                     for c in range(1, wb3["blotter"].max_column + 1)
                     if wb3["blotter"].cell(row=4, column=c).value]
        for col in ("Precio Target", "Stop Loss", "Moneda Target"):
            assert col in headers2, f"{col} should be back in {headers2}"
        # E importer todavía funciona
        from engine.importer import import_all
        stats = import_all(tmp / "w.db", xlsx)
        assert "blotter" in stats
        print(f"  ✓ legacy master migrado + importer funciona: blotter rows={stats['blotter']}")


def test_11_api_endpoints():
    print("\n[T 11] API: /api/holdings-near-target + GET/PUT /api/settings:")
    import json
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        os.environ["WM_BASE_DIR"] = str(tmp)
        os.environ["WM_USERS_JSON"] = json.dumps({
            "u1": {"token": "tok-u1", "is_admin": True}
        })
        os.environ["WM_ADMIN_USER"] = "u1"
        os.environ.pop("WM_API_TOKEN", None)
        from api.state import reset_settings
        reset_settings()

        # Crear master
        (tmp / "inputs" / "u1").mkdir(parents=True)
        xlsx = tmp / "inputs" / "u1" / "wealth_management.xlsx"
        _build_master_with_targets(xlsx)

        from api.app import create_app
        app = create_app()
        client = app.test_client()
        H = {"Authorization": "Bearer tok-u1"}

        # Refresh primero
        r = client.post("/api/refresh", headers=H)
        assert r.status_code == 200, r.data

        # GET settings
        r = client.get("/api/settings", headers=H)
        assert r.status_code == 200
        body = r.get_json()
        assert body["alert_distance_bps"] == 10.0
        print(f"  ✓ GET /api/settings → {body}")

        # PUT settings
        r = client.put("/api/settings", headers=H,
                        json={"key": "alert_distance_bps", "value": 200})
        assert r.status_code == 200
        # Ahora con 200 bps el alert debería disparar
        r = client.get("/api/holdings-near-target", headers=H)
        body = r.get_json()
        assert body["alert_distance_bps"] == 200.0
        print(f"  ✓ /api/holdings-near-target a 200 bps → {body['n_alerts']} alerts")

        # PUT setting con key no permitido
        r = client.put("/api/settings", headers=H,
                        json={"key": "foo", "value": 1})
        assert r.status_code == 400
        # Override on-the-fly
        r = client.get("/api/holdings-near-target?bps=5000", headers=H)
        body = r.get_json()
        assert body["alert_distance_bps"] == 5000.0
        # Con 5000 bps (50%) debería caer prácticamente todo
        print(f"  ✓ override ?bps=5000 → {body['n_alerts']} alerts")


# =============================================================================
# Main
# =============================================================================

if __name__ == "__main__":
    tests = [
        test_1_schema_has_target_cols,
        test_2_set_get_setting,
        test_3_importer_reads_target_from_blotter,
        test_4_importer_reads_config_sheet,
        test_5_holdings_exposes_target_and_dist,
        test_6_filter_near_target,
        test_7_filter_near_target_strict_10bps,
        test_8_active_target_uses_most_recent_buy,
        test_9_migration_idempotent,
        test_10_migration_on_legacy_master,
        test_11_api_endpoints,
    ]
    failed = []
    for t in tests:
        try:
            t()
        except Exception as e:
            print(f"  ✗ FAIL: {e}")
            import traceback; traceback.print_exc()
            failed.append(t.__name__)
    print("\n" + "=" * 70)
    if failed:
        print(f"✗ {len(failed)}/{len(tests)} tests FALLARON: {failed}")
        sys.exit(1)
    else:
        print(f"✓ Todos los {len(tests)} tests pasaron")
    print("=" * 70)
