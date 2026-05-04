# -*- coding: utf-8 -*-
"""
test_migration.py

Tests del script migrate_master.py:
  - Simula un Excel con la estructura ANTIGUA (sin aforos/margin_config,
    sin Investible/Cash Purpose, sin Linked Trade ID, sin Row ID).
  - Corre la migración.
  - Verifica que:
    * Backup creado
    * Nuevas hojas y columnas presentes
    * Datos originales NO se perdieron
    * Re-migrar es idempotente
    * El importer puede leer el resultado correctamente
"""

from __future__ import annotations

import shutil
import sqlite3
import sys
import tempfile
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from build_master import build_master
from migrate_master import migrate, ROW_ID_PREFIXES
from engine.importer import import_all


def _downgrade_to_old_format(xlsx: Path):
    """Convierte un Excel "nuevo" (con aforos, etc) en uno "viejo".

    Borra hojas nuevas y columnas nuevas, simulando el estado pre-migración.
    """
    wb = load_workbook(filename=str(xlsx))
    # Borrar hojas nuevas
    for sheet in ("aforos", "margin_config"):
        if sheet in wb.sheetnames:
            del wb[sheet]
    # Borrar columnas nuevas en cuentas
    if "cuentas" in wb.sheetnames:
        ws = wb["cuentas"]
        # Headers en row 4. Borramos Investible y Cash Purpose si existen
        headers = [ws.cell(row=4, column=c).value
                   for c in range(1, ws.max_column + 1)]
        cols_to_delete = []
        for i, h in enumerate(headers, start=1):
            if h in ("Investible", "Cash Purpose"):
                cols_to_delete.append(i)
        # Borrar de derecha a izquierda
        for c in reversed(cols_to_delete):
            ws.delete_cols(c)
    # Borrar columna nueva en funding
    if "funding" in wb.sheetnames:
        ws = wb["funding"]
        headers = [ws.cell(row=4, column=c).value
                   for c in range(1, ws.max_column + 1)]
        for i, h in enumerate(headers, start=1):
            if h == "Linked Trade ID":
                ws.delete_cols(i)
                break
    wb.save(str(xlsx))


def _has_column(ws, header_row, name) -> bool:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and str(v).strip().lower() == name.lower():
            return True
    return False


def _column_index(ws, header_row, name) -> int:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v and str(v).strip().lower() == name.lower():
            return c
    return -1


def test_1_migration_creates_backup():
    print("\n[mig 1] backup creado:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "wm.xlsx"
        build_master(xlsx)
        _downgrade_to_old_format(xlsx)
        original_size = xlsx.stat().st_size

        summary = migrate(xlsx)
        backup = Path(summary["backup"])
        assert backup.is_file()
        assert backup.stat().st_size == original_size  # es copia idéntica del original
        print(f"  ✓ backup: {backup.name} ({backup.stat().st_size:,} bytes)")


def test_2_migration_adds_new_sheets():
    print("\n[mig 2] hojas nuevas creadas:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "wm.xlsx"
        build_master(xlsx)
        _downgrade_to_old_format(xlsx)
        wb_before = load_workbook(filename=str(xlsx))
        assert "aforos" not in wb_before.sheetnames
        assert "margin_config" not in wb_before.sheetnames

        migrate(xlsx)
        wb_after = load_workbook(filename=str(xlsx))
        assert "aforos" in wb_after.sheetnames
        assert "margin_config" in wb_after.sheetnames
        # Aforos debe tener data
        ws = wb_after["aforos"]
        assert ws.cell(row=5, column=1).value == "CLASS"
        assert ws.cell(row=5, column=2).value == "BOND_AR"
        # Margin config con ibkr default
        ws = wb_after["margin_config"]
        assert ws.cell(row=5, column=1).value == "ibkr"
        print(f"  ✓ aforos + margin_config creadas con defaults")


def test_3_migration_adds_columns():
    print("\n[mig 3] columnas nuevas en cuentas y funding:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "wm.xlsx"
        build_master(xlsx)
        _downgrade_to_old_format(xlsx)
        wb = load_workbook(filename=str(xlsx))
        assert not _has_column(wb["cuentas"], 4, "Investible")

        migrate(xlsx)
        wb = load_workbook(filename=str(xlsx))
        assert _has_column(wb["cuentas"], 4, "Investible")
        assert _has_column(wb["cuentas"], 4, "Cash Purpose")
        assert _has_column(wb["funding"], 4, "Linked Trade ID")
        print(f"  ✓ Investible, Cash Purpose, Linked Trade ID presentes")

        # Verificar valores generados:
        # external_income debe ser NO
        ws = wb["cuentas"]
        col_inv = _column_index(ws, 4, "Investible")
        for r in range(5, 30):
            code = ws.cell(row=r, column=1).value
            inv = ws.cell(row=r, column=col_inv).value
            if code == "external_income":
                assert inv == "NO", f"external_income debería ser NO, got {inv}"
            elif code == "cocos":
                assert inv == "YES", f"cocos debería ser YES, got {inv}"
        print(f"  ✓ Investible auto-resuelto: técnicas=NO, reales=YES")


def test_4_migration_adds_row_ids():
    print("\n[mig 4] Row IDs generados en hojas de eventos:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "wm.xlsx"
        build_master(xlsx)
        _downgrade_to_old_format(xlsx)

        migrate(xlsx)
        wb = load_workbook(filename=str(xlsx))
        # Blotter tiene 2 ejemplos, deberían tener BL-0001 y BL-0002
        ws = wb["blotter"]
        col_id = _column_index(ws, 4, "Row ID")
        assert col_id > 0
        ids = [ws.cell(row=r, column=col_id).value for r in (5, 6)]
        assert ids == ["BL-0001", "BL-0002"], f"got {ids}"
        print(f"  ✓ blotter Row IDs: {ids}")

        # Asientos contables: 4 filas pero 2 grupos (OPEN-001, ADJ-002)
        ws = wb["asientos_contables"]
        col_id = _column_index(ws, 4, "Row ID")
        ids = [ws.cell(row=r, column=col_id).value for r in (5, 6, 7, 8)]
        # Filas 5+6 son OPEN-001 → mismo Row ID; filas 7+8 son ADJ-002 → otro
        assert ids[0] == ids[1], f"OPEN-001 debería compartir Row ID: {ids[:2]}"
        assert ids[2] == ids[3], f"ADJ-002 debería compartir Row ID: {ids[2:]}"
        assert ids[0] != ids[2], "Grupos distintos deben tener IDs distintos"
        print(f"  ✓ asientos_contables agrupados por Event ID: {ids}")


def test_5_migration_preserves_data():
    print("\n[mig 5] datos originales NO se perdieron:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "wm.xlsx"
        build_master(xlsx)
        _downgrade_to_old_format(xlsx)

        # Capturar datos clave antes
        wb = load_workbook(filename=str(xlsx))
        blotter_before = []
        ws = wb["blotter"]
        for r in range(5, 7):
            blotter_before.append({
                "Trade ID": ws.cell(row=r, column=1).value,
                "Ticker": ws.cell(row=r, column=6).value,
                "Side": ws.cell(row=r, column=7).value,
                "Qty": ws.cell(row=r, column=8).value,
            })

        migrate(xlsx)

        # Ahora capturar datos después y comparar
        wb = load_workbook(filename=str(xlsx))
        ws = wb["blotter"]
        blotter_after = []
        for r in range(5, 7):
            blotter_after.append({
                "Trade ID": ws.cell(row=r, column=1).value,
                "Ticker": ws.cell(row=r, column=6).value,
                "Side": ws.cell(row=r, column=7).value,
                "Qty": ws.cell(row=r, column=8).value,
            })

        assert blotter_before == blotter_after, \
            f"Datos del blotter cambiaron:\n  before: {blotter_before}\n  after: {blotter_after}"
        print(f"  ✓ blotter intacto: {len(blotter_after)} trades preservados")


def test_6_migration_idempotent():
    print("\n[mig 6] re-migrar es idempotente:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "wm.xlsx"
        build_master(xlsx)
        _downgrade_to_old_format(xlsx)

        migrate(xlsx)
        # Capturar estado tras 1ra migración
        wb1 = load_workbook(filename=str(xlsx))
        sheets_1 = wb1.sheetnames[:]
        n_cuentas_cols_1 = wb1["cuentas"].max_column

        # Re-migrar
        migrate(xlsx)
        wb2 = load_workbook(filename=str(xlsx))
        sheets_2 = wb2.sheetnames[:]
        n_cuentas_cols_2 = wb2["cuentas"].max_column

        assert sheets_1 == sheets_2, f"Sheets cambian: {sheets_1} vs {sheets_2}"
        assert n_cuentas_cols_1 == n_cuentas_cols_2, \
            f"Cols cuentas cambian: {n_cuentas_cols_1} → {n_cuentas_cols_2}"
        print(f"  ✓ 2da migración no agrega columnas/hojas duplicadas")


def test_7_importer_works_after_migration():
    print("\n[mig 7] importer funciona tras migración:")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "wm.xlsx"
        db = tmp / "wm.db"
        build_master(xlsx)
        _downgrade_to_old_format(xlsx)
        migrate(xlsx)

        stats = import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        assert stats["blotter"] == 2
        assert stats["aforos"] >= 7, f"aforos no cargados: {stats.get('aforos')}"
        assert stats["margin_config"] >= 1
        print(f"  ✓ importer corre OK: {stats}")

        # Verificar que cocos sigue siendo investible y cash_reserva no
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row
        cur = conn.execute("SELECT code, investible, cash_purpose FROM accounts WHERE code IN ('cocos','cash_reserva','external_income')")
        rows = {r["code"]: r for r in cur.fetchall()}
        assert rows["cocos"]["investible"] == 1
        assert rows["cash_reserva"]["investible"] == 0
        assert rows["external_income"]["investible"] == 0
        print(f"  ✓ Investible flags correctos en DB tras migración")
        conn.close()


if __name__ == "__main__":
    tests = [
        test_1_migration_creates_backup,
        test_2_migration_adds_new_sheets,
        test_3_migration_adds_columns,
        test_4_migration_adds_row_ids,
        test_5_migration_preserves_data,
        test_6_migration_idempotent,
        test_7_importer_works_after_migration,
    ]
    failed = []
    for t in tests:
        try:
            t()
        except Exception as e:
            print(f"  ✗ FAIL: {e}")
            import traceback; traceback.print_exc()
            failed.append(t.__name__)
    print("\n" + "=" * 70)
    if failed:
        print(f"✗ {len(failed)}/{len(tests)} tests FALLARON: {failed}")
        sys.exit(1)
    else:
        print(f"✓ Todos los {len(tests)} tests pasaron")
    print("=" * 70)
