# -*- coding: utf-8 -*-
"""
test_engine.py

Tests end-to-end del motor de wealth management.

USO:
    python3 test_engine.py
"""

from __future__ import annotations

import sqlite3
import sys
import tempfile
from datetime import date
from pathlib import Path

# Permitir import desde la carpeta padre
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from build_master import build_master
from engine.importer import import_all
from engine.liabilities import card_snapshot, all_card_snapshots
from engine.schema import init_db, AccountKind


def test_1_build_master():
    """build_master genera archivo con 15 hojas."""
    print("\nTest 1 (build_master):")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = Path(f.name)
    try:
        build_master(path)
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(path))
        assert len(wb.sheetnames) >= 14, f"Esperaba ≥14 hojas, got {len(wb.sheetnames)}"
        assert "blotter" in wb.sheetnames
        assert "gastos" in wb.sheetnames
        assert "cuentas" in wb.sheetnames
        assert "asientos_contables" in wb.sheetnames
        print(f"  ✓ {len(wb.sheetnames)} hojas: {', '.join(wb.sheetnames)}")
    finally:
        path.unlink()


def test_2_import_round_trip():
    """build_master → import → verificar contadores."""
    print("\nTest 2 (import round-trip):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        stats = import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        # Las hojas con ejemplos pre-cargados:
        assert stats["monedas"] >= 12
        assert stats["cuentas"] >= 16
        assert stats["especies"] >= 12
        assert stats["blotter"] == 2  # 2 trades de TXMJ9
        assert stats["transferencias_cash"] == 3
        assert stats["ingresos"] == 2
        assert stats["gastos"] == 3
        assert stats["recurrentes"] >= 1
        assert stats["asientos_contables"] == 2  # 2 grupos
        print(f"  ✓ stats: {stats}")


def test_3_blotter_balance():
    """Test que el TXMJ9 trade neteado da qty=0 (compra+venta misma cantidad)."""
    print("\nTest 3 (blotter qty net):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row
        cur = conn.execute(
            """SELECT SUM(qty) AS net FROM movements
               WHERE account='cocos' AND asset='TXMJ9'"""
        )
        net = cur.fetchone()["net"]
        # BUY 250M + SELL 250M = 0
        assert abs(net) < 1e-3, f"Expected ~0, got {net}"
        print(f"  ✓ TXMJ9 net qty: {net}")

        # PnL del trade: SELL 0.835 - BUY 0.80 = +0.035 × 250M = 8.75M
        cur = conn.execute(
            """SELECT SUM(qty) AS cash_net FROM movements
               WHERE account='cocos' AND asset='ARS'
               AND event_id IN (SELECT event_id FROM events WHERE event_type='TRADE')"""
        )
        cash_net = cur.fetchone()["cash_net"]
        # BUY: -200M, SELL: +208.75M → net +8.75M
        expected = 250000000 * (0.835 - 0.80)
        assert abs(cash_net - expected) < 1, f"Expected {expected}, got {cash_net}"
        print(f"  ✓ TXMJ9 cash net: ARS {cash_net:,.0f} (esperado {expected:,.0f})")
        conn.close()


def test_4_cuotas_expansion():
    """Test que 6 cuotas → 6 events CARD_INSTALLMENT."""
    print("\nTest 4 (cuotas expansion):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        cur = conn.execute(
            """SELECT COUNT(*) AS n FROM events WHERE event_type='CARD_INSTALLMENT'"""
        )
        n = cur.fetchone()["n"]
        # El vuelo es 1 fila con 6 cuotas → 6 CARD_INSTALLMENT
        assert n == 6, f"Esperaba 6 cuotas, got {n}"
        print(f"  ✓ {n} CARD_INSTALLMENT events generados")

        # Verificar que cada cuota es 200 USD
        cur = conn.execute(
            """SELECT m.qty FROM movements m
               JOIN events e ON e.event_id = m.event_id
               WHERE e.event_type='CARD_INSTALLMENT'
               AND m.account='galicia_visa_usd'"""
        )
        qtys = [row["qty"] for row in cur.fetchall()]
        assert all(abs(q - 200) < 1e-3 for q in qtys), f"Cuotas no son 200: {qtys}"
        print(f"  ✓ Cada cuota = USD 200")
        conn.close()


def test_5_card_snapshot_galicia_usd():
    """Galicia Visa USD: 1 cuota cayó en mayo 1, próximo cierre 28 mayo."""
    print("\nTest 5 (card_snapshot Galicia USD):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        snap = card_snapshot(conn, "galicia_visa_usd", ref_date=date(2026, 5, 2))
        assert snap is not None
        # 2 mayo: ya pasó 1 cuota (1 mayo) → saldo actual = 200
        assert abs(snap.saldo_actual - 200) < 1e-3, \
            f"saldo actual: {snap.saldo_actual}"
        # Último cierre: 28 abril → 0 (la cuota cayó después del 28/4)
        assert abs(snap.saldo_ultimo_resumen) < 1e-3, \
            f"último resumen: {snap.saldo_ultimo_resumen}"
        # Próximo cierre: 28 mayo
        assert snap.fecha_proximo_cierre == "2026-05-28", \
            f"próximo cierre: {snap.fecha_proximo_cierre}"
        # Próximo vto: la 1 cuota va al cierre de 28/5 → USD 200
        assert abs(snap.saldo_proximo_vto - 200) < 1e-3, \
            f"próximo vto: {snap.saldo_proximo_vto}"
        print(f"  ✓ Saldo actual: USD {snap.saldo_actual}")
        print(f"  ✓ Último resumen ({snap.fecha_ultimo_cierre}): USD {snap.saldo_ultimo_resumen}")
        print(f"  ✓ Próximo vto ({snap.fecha_proximo_vto}): USD {snap.saldo_proximo_vto}")
        conn.close()


def test_6_card_snapshot_galicia_ars_post_cierre():
    """Si pongo ref_date después del cierre del 28/5, el resumen ya cerró."""
    print("\nTest 6 (card snapshot post-cierre):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        # Fecha = 1 junio 2026 (después del cierre 28/5)
        import_all(db, xlsx, fecha_corte=date(2026, 6, 1))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        snap = card_snapshot(conn, "galicia_visa_ars", ref_date=date(2026, 6, 1))
        assert snap is not None
        # Restaurante de 35k cargado el 3/5 → cae en cierre 28/5 → último resumen
        assert abs(snap.saldo_ultimo_resumen - 35000) < 1e-3, \
            f"último resumen: {snap.saldo_ultimo_resumen}"
        # Saldo actual igual al último (no hay más gastos después)
        assert abs(snap.saldo_actual - 35000) < 1e-3, \
            f"saldo actual: {snap.saldo_actual}"
        print(f"  ✓ Saldo actual al 1/6: ARS {snap.saldo_actual:,.0f}")
        print(f"  ✓ Último resumen ({snap.fecha_ultimo_cierre}): ARS {snap.saldo_ultimo_resumen:,.0f}")
        conn.close()


def test_7_recurrentes_expansion():
    """Recurrentes: sueldo enero 2026 hasta hoy = 5 meses."""
    print("\nTest 7 (recurrentes expansion):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        # Sueldo: enero a mayo (5 meses, día 1 cada mes)
        cur = conn.execute(
            """SELECT COUNT(*) AS n FROM events
               WHERE event_type='INCOME'
               AND description='Sueldo regular'"""
        )
        n_sueldos = cur.fetchone()["n"]
        # Desde 1/1/2026 hasta 1/5/2026 = 5 ocurrencias el día 1
        assert n_sueldos == 5, f"Sueldos: {n_sueldos}"
        print(f"  ✓ Sueldos generados: {n_sueldos}")

        # Verificar que cada uno mueve 5M ARS a galicia_caja_ars
        cur = conn.execute(
            """SELECT SUM(m.qty) AS total FROM movements m
               JOIN events e ON e.event_id=m.event_id
               WHERE e.description='Sueldo regular'
               AND m.account='galicia_caja_ars'"""
        )
        total = cur.fetchone()["total"]
        # 5 sueldos × 5M = 25M
        assert abs(total - 25_000_000) < 1, f"Total sueldos: {total}"
        print(f"  ✓ Total ingresos: ARS {total:,.0f}")
        conn.close()


def test_8_asientos_apertura():
    """Asientos contables: apertura de 1000 AL30D balancea con opening_balance."""
    print("\nTest 8 (asientos apertura):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        # Verificar que cocos tiene 1000 AL30D
        cur = conn.execute(
            """SELECT SUM(qty) AS bal FROM movements
               WHERE account='cocos' AND asset='AL30D'"""
        )
        bal = cur.fetchone()["bal"]
        assert abs(bal - 1000) < 1e-3, f"Cocos AL30D bal: {bal}"
        print(f"  ✓ Saldo Cocos AL30D: {bal}")

        # Verificar que opening_balance tiene -1000
        cur = conn.execute(
            """SELECT SUM(qty) AS bal FROM movements
               WHERE account='opening_balance' AND asset='AL30D'"""
        )
        bal = cur.fetchone()["bal"]
        assert abs(bal + 1000) < 1e-3, f"opening_balance AL30D bal: {bal}"
        print(f"  ✓ Saldo opening_balance AL30D: {bal}")

        # Total AL30D en sistema = 0 (balance contable)
        cur = conn.execute(
            "SELECT SUM(qty) AS total FROM movements WHERE asset='AL30D'"
        )
        total = cur.fetchone()["total"]
        assert abs(total) < 1e-3, f"AL30D total no balancea: {total}"
        print(f"  ✓ AL30D balance global: {total}")
        conn.close()


def test_9_v_balances():
    """La vista v_balances filtra los saldos = 0."""
    print("\nTest 9 (v_balances view):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)
        import_all(db, xlsx, fecha_corte=date(2026, 5, 2))
        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        cur = conn.execute("SELECT * FROM v_balances ORDER BY account, asset")
        rows = list(cur.fetchall())
        # No debería haber saldos exactamente 0 (la vista los filtra)
        assert all(abs(r["balance"]) >= 1e-6 for r in rows)
        print(f"  ✓ {len(rows)} balances (todos != 0):")
        for r in rows[:10]:
            print(f"    {r['account']:<22} {r['asset']:<10} {r['balance']:>16,.4f}")
        if len(rows) > 10:
            print(f"    ... ({len(rows)-10} más)")
        conn.close()


def test_10_fx_module():
    """Tests del módulo fx: get_rate, convert con cross-rates."""
    print("\nTest 10 (FX module):")
    from engine.fx import get_rate, convert, FxError
    from engine.schema import init_db

    with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as f:
        dbpath = Path(f.name)
    try:
        conn = init_db(dbpath, drop_existing=True)
        # Cargar FX manualmente: USB y USD el 2026-04-30
        conn.execute(
            "INSERT INTO fx_rates (fecha, moneda, rate, base, source) VALUES (?,?,?,?,?)",
            ("2026-04-30", "USB", 1180.0, "ARS", "test"),
        )
        conn.execute(
            "INSERT INTO fx_rates (fecha, moneda, rate, base, source) VALUES (?,?,?,?,?)",
            ("2026-04-30", "USD", 1380.0, "ARS", "test"),
        )
        conn.commit()

        # Caso 1: get_rate exacto
        assert abs(get_rate(conn, "2026-04-30", "USB") - 1180.0) < 1e-6
        print(f"  ✓ get_rate exacto: USB/ARS @ 2026-04-30 = 1180.0")

        # Caso 2: moneda == base
        assert get_rate(conn, "2026-04-30", "ARS") == 1.0
        print(f"  ✓ ARS/ARS = 1.0")

        # Caso 3: fallback a día anterior
        rate = get_rate(conn, "2026-05-02", "USB", fallback_days=7)
        assert abs(rate - 1180.0) < 1e-6
        print(f"  ✓ Fallback 2 días: USB/ARS @ 2026-05-02 (no hay) → 2026-04-30 = 1180.0")

        # Caso 4: convert ARS → USB (1180 ARS = 1 USB)
        result = convert(conn, 11800, "ARS", "USB", "2026-04-30")
        assert abs(result - 10) < 1e-6
        print(f"  ✓ Convert: 11800 ARS → USB @ 1180 = 10 USB")

        # Caso 5: convert USB → ARS
        result = convert(conn, 10, "USB", "ARS", "2026-04-30")
        assert abs(result - 11800) < 1e-6
        print(f"  ✓ Convert: 10 USB → ARS @ 1180 = 11800 ARS")

        # Caso 6: cross-rate USB → USD
        # 1 USB = 1180 ARS, 1 USD = 1380 ARS, entonces 1 USB = 1180/1380 USD = 0.855 USD
        result = convert(conn, 1, "USB", "USD", "2026-04-30")
        expected = 1180.0 / 1380.0
        assert abs(result - expected) < 1e-6
        print(f"  ✓ Cross-rate: 1 USB → USD = {result:.4f} (esperado {expected:.4f})")

        # Caso 7: FX faltante levanta FxError
        try:
            convert(conn, 100, "EUR", "ARS", "2026-04-30")
            assert False, "Esperaba FxError"
        except FxError as e:
            print(f"  ✓ Faltante levanta FxError: {e}")

        conn.close()
    finally:
        dbpath.unlink(missing_ok=True)


def test_11_blotter_with_fx_conversion():
    """Trade en moneda distinta a la nativa: motor convierte unit_price con FX."""
    print("\nTest 11 (blotter FX conversion):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        # CSV de FX con USB y USD
        fxcsv = tmp / "fx_historico.csv"
        fxcsv.write_text(
            "fecha,moneda,rate,base,source\n"
            "2026-04-28,USB,1170.0,ARS,test\n"
            "2026-04-30,USB,1180.0,ARS,test\n"
            "2026-04-30,USD,1380.0,ARS,test\n",
            encoding="utf-8",
        )
        build_master(xlsx)

        # Modificar el master: cambiar el ticker de prueba a uno donde
        # moneda trade != currency. Vamos a usar especies AL30D (currency=USB)
        # y agregar un trade en ARS de AL30D para forzar conversión.
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(xlsx))
        # Hoja blotter: agregar fila de prueba al final de los ejemplos
        ws = wb["blotter"]
        # Las filas existentes son 5 y 6 (TXMJ9). Agrego en fila 7
        # Trade en ARS de AL30D (que es USB) para forzar conversion
        ws.cell(row=7, column=1).value = "T9999"        # Trade ID
        ws.cell(row=7, column=2).value = date(2026, 4, 30)  # Trade Date
        ws.cell(row=7, column=3).value = date(2026, 4, 30)  # Settle Date
        ws.cell(row=7, column=4).value = "cocos"        # Cuenta
        ws.cell(row=7, column=5).value = "BH"            # Strategy
        ws.cell(row=7, column=6).value = "AL30D"         # Ticker (currency=USB)
        ws.cell(row=7, column=7).value = "BUY"           # Side
        ws.cell(row=7, column=8).value = 1000             # Qty
        ws.cell(row=7, column=9).value = 850             # Precio EN ARS
        ws.cell(row=7, column=10).value = "ARS"          # Moneda Trade ARS!=USB
        ws.cell(row=7, column=11).value = "cocos"        # Cuenta Cash
        ws.cell(row=7, column=12).value = 0              # Comisión
        ws.cell(row=7, column=13).value = "ARS"          # Moneda Com
        ws.cell(row=7, column=14).value = "BUY 1000 AL30D pagando ARS"
        ws.cell(row=7, column=15).value = "FX test"
        wb.save(str(xlsx))

        # Importar con fx_csv_path explícito
        stats = import_all(db, xlsx, fecha_corte=date(2026, 5, 2),
                           fx_csv_path=fxcsv)
        assert stats.get("fx_rates", 0) >= 3, f"FX cargado: {stats.get('fx_rates')}"
        print(f"  ✓ FX cargado: {stats['fx_rates']} filas")

        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        # Verificar el trade T9999: el cash sale en ARS (850000)
        cur = conn.execute(
            """SELECT m.* FROM movements m
               JOIN events e ON e.event_id=m.event_id
               WHERE e.external_id='T9999' AND m.account='cocos'"""
        )
        rows = list(cur.fetchall())
        assert len(rows) == 2, f"Esperaba 2 movements, got {len(rows)}"

        # Movement del activo (AL30D):
        asset_mov = [r for r in rows if r["asset"] == "AL30D"][0]
        # qty = +1000
        assert abs(asset_mov["qty"] - 1000) < 1e-3
        # unit_price convertido a USB: 850 ARS / 1180 ARS-por-USB = 0.7203 USB
        expected_native = 850 / 1180
        assert abs(asset_mov["unit_price"] - expected_native) < 1e-4, \
            f"unit_price native: {asset_mov['unit_price']} vs esperado {expected_native}"
        # price_currency = USB (la nativa)
        assert asset_mov["price_currency"] == "USB", \
            f"price_currency: {asset_mov['price_currency']}"
        print(f"  ✓ Activo AL30D: qty=1000, unit_price={asset_mov['unit_price']:.4f} USB")

        # Movement del cash (ARS):
        cash_mov = [r for r in rows if r["asset"] == "ARS"][0]
        # cash sale: -850 * 1000 = -850000 ARS
        assert abs(cash_mov["qty"] + 850000) < 1
        print(f"  ✓ Cash ARS: qty={cash_mov['qty']:,.0f}")

        # Verificación contable:
        # - inventario AL30D: aumenta en 1000 a precio ~0.72 USB
        # - cash ARS: cae en 850000 ARS
        # - 850000 ARS / 1180 = 720.34 USB (que es 1000 × 0.7203 USB)
        # cost_basis del asset:
        expected_cost_basis = 1000 * (850 / 1180)
        assert abs(asset_mov["cost_basis"] - expected_cost_basis) < 1e-2
        print(f"  ✓ Cost basis AL30D: {asset_mov['cost_basis']:.4f} USB "
              f"(= 850000 ARS / 1180)")
        conn.close()


def test_12_blotter_without_fx_skips():
    """Trade en moneda distinta sin FX cargado: registra el trade con
    price_currency = moneda Trade (cash igual va correcto) y reporta el
    fallback en stats['blotter_fx_failed']."""
    print("\nTest 12 (blotter sin FX → fallback con warning):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        xlsx = tmp / "test.xlsx"
        db = tmp / "test.db"
        build_master(xlsx)

        # Hoja blotter: agregar trade en moneda distinta SIN cargar FX
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(xlsx))
        ws = wb["blotter"]
        # Trade de AL30D pagando en USDT (sin FX cargado de USDT)
        ws.cell(row=7, column=1).value = "T_NOFX"
        ws.cell(row=7, column=2).value = date(2026, 4, 30)
        ws.cell(row=7, column=3).value = date(2026, 4, 30)
        ws.cell(row=7, column=4).value = "binance"
        ws.cell(row=7, column=5).value = "BH"
        ws.cell(row=7, column=6).value = "AL30D"
        ws.cell(row=7, column=7).value = "BUY"
        ws.cell(row=7, column=8).value = 100
        ws.cell(row=7, column=9).value = 0.85
        ws.cell(row=7, column=10).value = "USDT"  # No hay FX USDT/ARS cargado
        ws.cell(row=7, column=11).value = "binance"
        ws.cell(row=7, column=12).value = 0
        ws.cell(row=7, column=13).value = "USDT"
        ws.cell(row=7, column=14).value = "BUY USDT sin FX"
        ws.cell(row=7, column=15).value = ""
        wb.save(str(xlsx))

        # Importar sin FX disponible
        stats = import_all(db, xlsx, fecha_corte=date(2026, 5, 2),
                           fx_csv_path=tmp / "noexiste.csv")
        # Nuevo behavior: el trade se registra igual (cash leg correcto)
        # con price_currency = moneda Trade. El fallback queda flagueado en
        # stats['blotter_fx_failed'] para que la UI lo muestre.
        assert stats["blotter"] == 3, f"blotter trades: {stats['blotter']}"
        assert stats.get("blotter_fx_failed"), (
            "Esperaba blotter_fx_failed con la fila sin FX"
        )
        assert any(f["ticker"] == "AL30D"
                   for f in stats["blotter_fx_failed"])
        print(f"  ✓ Trade sin FX: registrado con fallback "
              f"({len(stats['blotter_fx_failed'])} fila(s) flagueada(s))")

        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row
        cur = conn.execute(
            "SELECT COUNT(*) AS n FROM events WHERE external_id='T_NOFX'"
        )
        n = cur.fetchone()["n"]
        assert n == 1, f"T_NOFX debería existir como evento (n={n})"
        # El movement del activo debe quedar en moneda del trade (USDT),
        # no en la moneda nativa (USB) porque el FX falló.
        cur = conn.execute(
            """SELECT m.price_currency, m.qty FROM movements m
               JOIN events e ON e.event_id = m.event_id
               WHERE e.external_id='T_NOFX' AND m.asset='AL30D'"""
        )
        row = cur.fetchone()
        assert row["price_currency"] == "USDT", (
            f"price_currency debe ser USDT (sin convertir), got {row['price_currency']}"
        )
        # Cash leg: cuenta binance perdió USDT
        cur = conn.execute(
            """SELECT m.qty, m.asset FROM movements m
               JOIN events e ON e.event_id = m.event_id
               WHERE e.external_id='T_NOFX' AND m.asset='USDT'"""
        )
        cash = cur.fetchone()
        assert cash is not None and cash["qty"] < 0, (
            "Cash leg debe existir y restarle USDT a la cuenta"
        )
        print(f"  ✓ T_NOFX: evento creado, cash leg correcto, "
              f"price_currency=USDT (fallback)")
        conn.close()


def test_13_bond_prices_scaled_by_100():
    """Bonos BYMA: el feed los devuelve en % del par (ej 65.50). El motor
    los divide por 100 al importar para que mv = qty × price quede en
    moneda real. Esto aplica a las 3 clases de bonos: BOND_AR, BOND_CORP_AR
    (ONs argentinas) y BOND_US.

    Regression test: BOND_CORP_AR quedaba sin escalar y el user veía valores
    inflados ~100x.
    """
    import csv
    from engine.prices import import_prices_csv
    from engine.schema import init_db

    print("\nTest 13 (bond prices: las 3 clases escalan /100):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        db = tmp / "test.db"
        init_db(str(db), drop_existing=True)

        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row
        # Seed assets con las 3 clases de bonos + 1 equity (no escala)
        conn.executemany(
            "INSERT INTO assets (ticker, name, asset_class, currency) VALUES (?,?,?,?)",
            [
                ("AL30D", "Bonar 30 USD", "BOND_AR", "USB"),
                ("YPFCO", "ON YPF Corp", "BOND_CORP_AR", "USB"),
                ("T2X3", "Bond US", "BOND_US", "USD"),
                ("GGAL", "Galicia", "EQUITY_AR", "ARS"),
            ],
        )
        conn.commit()

        # CSV con precios en base 100 (BYMA convention)
        csv_path = tmp / "precios.csv"
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            w.writerow(["fecha", "ticker", "price", "currency", "source"])
            w.writerow(["2026-05-08", "AL30D", "65.50", "USB", "BYMA"])
            w.writerow(["2026-05-08", "YPFCO", "98.20", "USB", "BYMA"])
            w.writerow(["2026-05-08", "T2X3",  "102.5", "USD", "BYMA"])
            w.writerow(["2026-05-08", "GGAL",  "2500",  "ARS", "BYMA"])  # NO escala

        import_prices_csv(conn, str(csv_path))

        rows = {r[0]: r[1] for r in
                conn.execute("SELECT ticker, price FROM prices")}
        # Los 3 bonos deben quedar en decimal
        assert abs(rows["AL30D"] - 0.655) < 1e-6, f"AL30D = {rows['AL30D']}"
        assert abs(rows["YPFCO"] - 0.982) < 1e-6, (
            f"YPFCO = {rows['YPFCO']} — BOND_CORP_AR debe escalar /100 también"
        )
        assert abs(rows["T2X3"] - 1.025) < 1e-6, f"T2X3 = {rows['T2X3']}"
        # Equity NO escala
        assert abs(rows["GGAL"] - 2500) < 1e-6, f"GGAL = {rows['GGAL']}"
        print(f"  ✓ AL30D=0.655, YPFCO=0.982, T2X3=1.025 (los 3 escalados), "
              f"GGAL=2500 (sin escalar)")
        conn.close()


def test_14_fx_fallback_preserves_currency_for_cost_basis():
    """Regression: comprar USB bond contra ARS sin FX cargado.

    Pasa: importer guarda cost_basis con price_currency=ARS (porque el FX
    ARS→USB falló al importar).  _calc_position re-intenta el FX y si
    sigue fallando, deja avg_cost en RAW (1137.50).

    Antes del fix, el caller etiquetaba ese avg_cost como native_ccy=USB
    para el fallback de market price → mv = qty × 1137.50 USB → patrimonio
    inflado 1500x.

    Ahora se devuelve avg_cost_currency=ARS para que el caller use la
    moneda real en cost_basis_fallback. mv termina siendo qty × 1137.50
    ARS, que convertido a anchor es ≈ el cash que se gastó → net ≈ 0.
    """
    from engine.holdings import _calc_position
    from engine.schema import init_db

    print("\nTest 14 (FX fallback preserva currency real del avg_cost):")
    with tempfile.TemporaryDirectory() as tmp:
        tmp = Path(tmp)
        db = tmp / "test.db"
        init_db(str(db), drop_existing=True)

        conn = sqlite3.connect(str(db))
        conn.row_factory = sqlite3.Row

        # Seed: currencies (FK)
        conn.executemany(
            "INSERT INTO currencies (code, name) VALUES (?, ?)",
            [("ARS", "Pesos"), ("USB", "USD MEP")],
        )
        # Seed: asset USB + cuenta
        conn.execute(
            "INSERT INTO assets (ticker, name, asset_class, currency) "
            "VALUES (?, ?, ?, ?)",
            ("AE38D", "Bonar 38 USD", "BOND_AR", "USB"),
        )
        conn.execute(
            "INSERT INTO accounts (code, name, kind, currency, investible) "
            "VALUES (?, ?, ?, ?, ?)",
            ("cocos", "Cocos", "CASH_BROKER", "ARS", 1),
        )
        # Insert un event TRADE + movement con unit_price en ARS
        # (simula FX-falló al importar — el price_currency quedó en ARS,
        # no en USB).
        cur = conn.execute(
            "INSERT INTO events (event_type, event_date, description) "
            "VALUES (?, ?, ?)",
            ("TRADE", "2026-05-05", "BUY AE38D contra ARS"),
        )
        event_id = cur.lastrowid
        conn.execute(
            "INSERT INTO movements (event_id, account, asset, qty, "
            "unit_price, price_currency, cost_basis) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (event_id, "cocos", "AE38D", 3700, 1137.50, "ARS", 3700 * 1137.50),
        )
        conn.commit()

        # NO seedeamos fx_rates ARS→USB → el FX falla en _calc_position
        pos = _calc_position(conn, "cocos", "AE38D", "2026-05-12",
                              native_ccy="USB")
        assert pos is not None, "Debería haber una posición"
        assert pos["qty"] == 3700
        # Sin FX, avg_cost queda en valor RAW (1137.50)
        assert abs(pos["avg_cost"] - 1137.50) < 1e-6, (
            f"avg_cost = {pos['avg_cost']}"
        )
        # KEY: avg_cost_currency debe ser ARS (la del trade), NO USB
        # (la nativa) — sino el caller meta el cost basis como USB y
        # mv = 3700 × 1137.50 ≈ 4.2M USB ≈ 6 billones ARS.
        assert pos["avg_cost_currency"] == "ARS", (
            f"avg_cost_currency = {pos['avg_cost_currency']!r} — "
            f"esperaba 'ARS' porque el FX no resolvió"
        )
        assert pos["has_fx_fallback"] is True, (
            "has_fx_fallback debería ser True"
        )
        print(f"  ✓ avg_cost={pos['avg_cost']} avg_cost_currency=ARS "
              f"has_fx_fallback=True")

        # Ahora con FX cargado: la conversión debe funcionar
        conn.execute(
            "INSERT INTO fx_rates (fecha, moneda, base, rate, source) "
            "VALUES (?, ?, ?, ?, ?)",
            ("2026-05-05", "ARS", "USB", 0.000615, "test"),
        )
        conn.commit()
        pos2 = _calc_position(conn, "cocos", "AE38D", "2026-05-12",
                                native_ccy="USB")
        # Con FX, avg_cost convertido: 1137.50 × 0.000615 ≈ 0.699
        assert abs(pos2["avg_cost"] - 0.6995625) < 1e-4, (
            f"avg_cost con FX = {pos2['avg_cost']}"
        )
        assert pos2["avg_cost_currency"] == "USB", (
            f"avg_cost_currency con FX = {pos2['avg_cost_currency']!r}"
        )
        assert pos2["has_fx_fallback"] is False
        print(f"  ✓ con FX cargado: avg_cost={pos2['avg_cost']:.4f} "
              f"avg_cost_currency=USB has_fx_fallback=False")
        conn.close()


# =============================================================================
# Main
# =============================================================================

if __name__ == "__main__":
    tests = [
        test_1_build_master,
        test_2_import_round_trip,
        test_3_blotter_balance,
        test_4_cuotas_expansion,
        test_5_card_snapshot_galicia_usd,
        test_6_card_snapshot_galicia_ars_post_cierre,
        test_7_recurrentes_expansion,
        test_8_asientos_apertura,
        test_9_v_balances,
        test_10_fx_module,
        test_11_blotter_with_fx_conversion,
        test_12_blotter_without_fx_skips,
        test_13_bond_prices_scaled_by_100,
        test_14_fx_fallback_preserves_currency_for_cost_basis,
    ]
    for t in tests:
        t()
    print("\n" + "=" * 70)
    print(f"✓ Todos los {len(tests)} tests pasaron")
    print("=" * 70)
